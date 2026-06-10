# -*- coding: utf-8 -*-
"""Thin bridge between the ERP package and the live bot.py module.

`B` is bot.py's module object (set once by finance.mount). Everything the ERP
reuses from the monolith — auth, STATE_DIR stores, Daftra import, dup shield,
custody math, expenses V4, owner-report math — is reached as `B.<name>` so the
data layer stays single-sourced and untouched.
"""

import json
import uuid
from calendar import monthrange
from decimal import Decimal
from datetime import datetime

from aiohttp import web

from . import statements as ST

B = None  # bot.py module object — set by finance.mount()


def attach(botmod):
    global B
    B = botmod


def jres(data, status=200):
    """JSON response that keeps Arabic readable (no \\uXXXX escapes)."""
    return web.json_response(
        data, status=status, dumps=lambda o: json.dumps(o, ensure_ascii=False))


# ---------------- auth / roles (server-side, reusing bot.py's system) ----------------

def authed(request):
    """Same login the dashboard uses (DASHBOARD_TOKEN or session token)."""
    try:
        return bool(B and B._dash_auth(request))
    except Exception:
        return False


def can(request, tab, action="read"):
    """Role check via bot.py's _user_can (legacy token = super-admin)."""
    try:
        return bool(B and B._user_can(request, tab, action))
    except Exception:
        return False


def role(request):
    return B._req_role(request)


def actor(request):
    return B._req_actor(request)


def can_finance(request):
    """admin or accountant — the two roles allowed inside the ERP."""
    return bool(B and B._fb_can_finance(request))


def can_high(request):
    """May approve >= 3000 SAR (Faisal tier: admin/owner; legacy token = admin)."""
    return bool(B and B.can_finance_approve_high_value(request))


def is_admin(request):
    return B._req_role(request) == "admin"


# ---------------- helpers ----------------

def _days_since(iso):
    try:
        d = datetime.fromisoformat(str(iso)[:19]).date()
        return max(0, (datetime.now(B.TZ).date() - d).days)
    except Exception:
        return None


def _amt(x):
    """Bank txn signed view: debit = money out, credit = money in."""
    deb = B._fb_money(x.get("debit"))
    cred = B._fb_money(x.get("credit"))
    if deb > 0:
        return B._fb_money_str(deb), "out"
    return B._fb_money_str(cred), "in"


_DUP_SUGGESTED = ("possible_duplicate", "strong_possible_duplicate", "needs_manual_review")


# ---------------- Today: the one prioritized work queue ----------------

async def work_queue(request):
    """Aggregate of everything that needs a human, in priority order.
    Reads existing state only — no live external calls (fast + honest)."""
    ov_resp = await B._api_fb_overview(request)
    ov = json.loads(ov_resp.body)
    if not ov.get("ok"):
        return {"error": ov.get("error") or "overview_failed"}

    ib = B._fb_inbox({})

    # 1) >= 3000 awaiting Faisal (largest first — that's where the risk is)
    approvals = [i for i in ib["items"] if i.get("lane") == "needs_faisal_approval"]
    approvals.sort(key=lambda i: float(B._fb_money(i.get("amount"))), reverse=True)
    g_approvals = [{
        "id": i["id"], "kind": i.get("kind"), "date": i.get("date"),
        "desc": (i.get("description") or "")[:140], "amount": i.get("amount"),
        "direction": i.get("direction"), "apartment": i.get("apartment") or "",
        "category": i.get("category") or "", "chip_ar": i.get("reason_chip_ar") or "",
        "chip_en": i.get("reason_chip_en") or "",
    } for i in approvals[:50]]

    # 2) unclassified bank txns (same definition as the bank summary: status=needs_review)
    uncls = [x for x in B._fb_bank.values() if x.get("status") == "needs_review"]
    uncls.sort(key=lambda x: x.get("date") or "", reverse=True)
    g_uncls = []
    for x in uncls[:25]:
        amount, dirn = _amt(x)
        g_uncls.append({"id": x["id"], "date": x.get("date"),
                        "desc": (x.get("description") or "")[:140],
                        "amount": amount, "dir": dirn,
                        "card": x.get("card_last4") or "",
                        "category": x.get("category") or "unknown"})

    # 3) suggested matches from the dup shield, awaiting a human verdict
    sugg = [x for x in B._fb_bank.values()
            if x.get("daftra_duplicate_status") in _DUP_SUGGESTED
            and x.get("match_status") != "matched"]
    sugg.sort(key=lambda x: int(x.get("daftra_duplicate_confidence") or 0), reverse=True)
    g_sugg = []
    for x in sugg[:25]:
        amount, dirn = _amt(x)
        g_sugg.append({"id": x["id"], "date": x.get("date"),
                       "desc": (x.get("description") or "")[:140],
                       "amount": amount, "dir": dirn,
                       "conf": int(x.get("daftra_duplicate_confidence") or 0),
                       "reason_ar": x.get("daftra_duplicate_reason_ar") or "",
                       "reason_en": x.get("daftra_duplicate_reason_en") or "",
                       "journal_no": x.get("matched_daftra_number")})

    # 4) contracts with no Daftra cost-center link (breaks unit profitability)
    g_contracts = [{"key": p.get("id") or k, "name": p.get("apartment_name") or k,
                    "owner": p.get("owner_name") or ""}
                   for k, p in B._fb_contracts.items() if not p.get("daftra_cost_center_id")]

    # 5) stale / failed imports
    g_imports = []
    for r in list(B._fb_runs)[-60:]:
        if r.get("status") == "failed":
            age = _days_since(r.get("finished_at") or r.get("started_at"))
            if age is not None and age <= 14:
                g_imports.append({"id": r.get("id"), "source": r.get("source"),
                                  "status": "failed", "when": r.get("finished_at") or r.get("started_at"),
                                  "file": r.get("filename") or "", "error": (r.get("error_summary") or "")[:200]})
    bank_run = (ov.get("bank") or {}).get("last_run") or {}
    bank_age = _days_since(bank_run.get("finished_at")) if bank_run else None
    if bank_age is not None and bank_age >= 3:
        g_imports.append({"id": "stale-bank", "source": "bank", "status": "stale",
                          "when": bank_run.get("finished_at"), "file": bank_run.get("filename") or "",
                          "error": ""})

    steps = ov.get("steps") or []
    done = sum(1 for s in steps if s.get("done"))
    health = int(round(100.0 * done / len(steps))) if steps else 0

    g_budget = budget_alerts_now()

    groups = [
        {"key": "approvals", "count": len(approvals), "items": g_approvals},
        {"key": "budget", "count": len(g_budget), "items": g_budget[:12]},
        {"key": "unclassified", "count": len(uncls), "items": g_uncls},
        {"key": "suggested", "count": len(sugg), "items": g_sugg},
        {"key": "contracts", "count": len(g_contracts), "items": g_contracts[:25]},
        {"key": "imports", "count": len(g_imports), "items": g_imports[:10]},
    ]
    next_key = next((g["key"] for g in groups if g["count"]), None)

    return {
        "ok": True,
        "role": ov.get("role"),
        "can_high": can_high(request),
        "health": {"pct": health, "done": done, "total": len(steps), "steps": steps},
        "bank_age_days": bank_age,
        "bank_last_run": {"when": bank_run.get("finished_at"), "file": bank_run.get("filename")} if bank_run else None,
        "next_best": next_key,
        "groups": groups,
        "counters": counters_snapshot(),
    }


def counters_snapshot():
    """Light recount used by mutation responses so the UI patches counters
    without re-rendering (R1/R1b): queue membership derives from store state."""
    ib = B._fb_inbox({})
    uncls = sum(1 for x in B._fb_bank.values() if x.get("status") == "needs_review")
    sugg = sum(1 for x in B._fb_bank.values()
               if x.get("daftra_duplicate_status") in _DUP_SUGGESTED
               and x.get("match_status") != "matched")
    contracts = sum(1 for p in B._fb_contracts.values() if not p.get("daftra_cost_center_id"))
    return {"approvals": ib["counts"].get("needs_faisal", 0),
            "unclassified": uncls, "suggested": sugg, "contracts": contracts}


def approve(request, body):
    """Faisal decision on one item. Server-side permission split:
    >= 3000 needs the admin tier; accountant may decide below it."""
    item_id = (body.get("id") or "").strip()
    decision = (body.get("decision") or "").strip()
    reason = (body.get("reason") or "").strip()
    res = B._fb_do_approval(item_id, decision, actor(request), reason,
                            can_high(request), can_finance(request))
    if res.get("error"):
        return res, int(res.get("code") or 400)
    res["counters"] = counters_snapshot()
    return res, 200


# ====================== البنك Bank workspace ======================

def _acct_records():
    return [r for r in B._fb_external.values() if r.get("source_type") == "accounts"]


def _cc_records():
    return [r for r in B._fb_external.values() if r.get("source_type") == "cost_centers"]


def _acct_by_id():
    return {str(r.get("source_id")): r for r in _acct_records()}


def accounts_payload():
    """The REAL Daftra chart (دليل الحسابات) as imported — classification options.
    Free text is rejected at classify time; this list is the whole universe."""
    accs = sorted(_acct_records(), key=lambda r: ((r.get("code") or ""), (r.get("display_name") or "")))
    ccs = sorted(_cc_records(), key=lambda r: (r.get("display_name") or ""))
    units = sorted({(p.get("apartment_name") or "").strip()
                    for p in B._fb_contracts.values() if (p.get("apartment_name") or "").strip()})
    return {"ok": True,
            "accounts": [{"id": str(r.get("source_id")), "code": r.get("code") or "",
                          "name": r.get("display_name") or ""} for r in accs],
            "cost_centers": [{"id": str(r.get("source_id")), "code": r.get("code") or "",
                              "name": r.get("display_name") or ""} for r in ccs],
            "units": units,
            "counts": {"accounts": len(accs), "cost_centers": len(ccs)}}


def _bank_row(x):
    """Compact row view + the pipeline chip states (مُصنّف → مُتحقق → مُرحّل)."""
    amount, dirn = _amt(x)
    ec = x.get("erp_class") or {}
    dq = x.get("daftra_duplicate_status") or "not_checked"
    migrated = bool(x.get("matched_daftra_id")) and dq in ("already_in_daftra_verified", "linked_existing")
    verified = dq in ("already_in_daftra_verified", "linked_existing", "not_found_after_full_check")
    return {"id": x["id"], "date": x.get("date") or "", "desc": x.get("description") or "",
            "amount": amount, "dir": dirn, "ref": x.get("ref") or "", "card": x.get("card_last4") or "",
            "category": x.get("category") or "unknown", "status": x.get("status") or "needs_review",
            "match_status": x.get("match_status") or "unmatched",
            "dup": dq, "dup_conf": x.get("daftra_duplicate_confidence"),
            "journal_no": x.get("matched_daftra_number"),
            "cls": {"account_id": str(ec.get("daftra_account_id") or ""),
                    "code": ec.get("account_code") or "", "name": ec.get("account_name") or "",
                    "cost_center_id": str(ec.get("cost_center_id") or ""),
                    "cost_center": ec.get("cost_center_name") or "",
                    "counterparty": ec.get("counterparty") or "", "unit": ec.get("unit") or "",
                    "rule_id": ec.get("rule_id") or "", "auto": bool(ec.get("auto"))},
            "classified": bool(ec.get("daftra_account_id")) or (x.get("status") == "reviewed"),
            "verified": verified, "migrated": migrated}


def _sim_index():
    """description-key -> [(account_id, uses)] built from already erp-classified txns."""
    idx = {}
    for x in B._fb_bank.values():
        ec = x.get("erp_class") or {}
        aid = str(ec.get("daftra_account_id") or "")
        if not aid:
            continue
        key = B._fb_similar_key(x.get("description") or "")
        if not key:
            continue
        idx.setdefault(key, {})
        idx[key][aid] = idx[key].get(aid, 0) + 1
    return {k: sorted(v.items(), key=lambda kv: kv[1], reverse=True) for k, v in idx.items()}


def _bank_suggestions(x, acct_idx, sim_idx):
    """Top-3 account suggestions: category mapping first, then same-description history."""
    out, seen = [], set()
    m = B._fb_mappings.get(x.get("category") or "")
    if m and m.get("daftra_account_id"):
        aid = str(m["daftra_account_id"])
        acc = acct_idx.get(aid)
        if acc:
            out.append({"account_id": aid, "code": acc.get("code") or "",
                        "name": acc.get("display_name") or "",
                        "why_ar": "ربط فئة «" + (x.get("category") or "") + "»",
                        "why_en": "category mapping"})
            seen.add(aid)
    key = B._fb_similar_key(x.get("description") or "")
    for aid, n in (sim_idx.get(key) or []):
        if len(out) >= 3:
            break
        if aid in seen:
            continue
        acc = acct_idx.get(aid)
        if not acc:
            continue
        out.append({"account_id": aid, "code": acc.get("code") or "",
                    "name": acc.get("display_name") or "",
                    "why_ar": "استُخدم " + str(n) + "× لوصف مشابه",
                    "why_en": "used " + str(n) + "x for similar"})
        seen.add(aid)
    return out


def bank_register(params):
    """Server-side pagination over the FULL register (no 400-row cap) + filter counts."""
    f = (params.get("f") or "all").strip()
    q = (params.get("q") or "").strip().lower()
    dfrom = (params.get("from") or "").strip()
    dto = (params.get("to") or "").strip()
    try:
        page = max(1, int(params.get("p") or 1))
    except Exception:
        page = 1
    try:
        ps = min(200, max(10, int(params.get("ps") or 50)))
    except Exception:
        ps = 50

    def in_range(x):
        d = (x.get("date") or "")[:10]
        if dfrom and d and d < dfrom:
            return False
        if dto and d and d > dto:
            return False
        if q:
            hay = ((x.get("description") or "") + " " + (x.get("ref") or "") + " " +
                   (x.get("debit") or "") + " " + (x.get("credit") or "")).lower()
            if q not in hay:
                return False
        return True

    base = [x for x in B._fb_bank.values() if in_range(x)]

    def is_done(x):
        return x.get("status") == "reviewed"

    def ge3000(x):
        return float(B._fb_money(x.get("debit"))) >= 3000.0

    counts = {"all": len(base),
              "needs_review": sum(1 for x in base if x.get("status") == "needs_review"),
              "done": sum(1 for x in base if is_done(x)),
              "unmatched": sum(1 for x in base if (x.get("match_status") or "unmatched") == "unmatched"),
              "ge3000": sum(1 for x in base if ge3000(x))}

    if f == "needs_review":
        sel = [x for x in base if x.get("status") == "needs_review"]
    elif f == "done":
        sel = [x for x in base if is_done(x)]
    elif f == "unmatched":
        sel = [x for x in base if (x.get("match_status") or "unmatched") == "unmatched"]
    elif f == "ge3000":
        sel = [x for x in base if ge3000(x)]
    else:
        sel = base

    sel.sort(key=lambda x: ((x.get("date") or ""), str(x.get("row") or "")), reverse=True)
    total = len(sel)
    pages = max(1, (total + ps - 1) // ps)
    page = min(page, pages)
    chunk = sel[(page - 1) * ps: (page - 1) * ps + ps]

    acct_idx = _acct_by_id()
    sim_idx = _sim_index()
    rows = []
    for x in chunk:
        r = _bank_row(x)
        if x.get("status") == "needs_review":
            r["suggestions"] = _bank_suggestions(x, acct_idx, sim_idx)
        rows.append(r)
    return {"ok": True, "rows": rows, "total": total, "page": page, "pages": pages,
            "page_size": ps, "counts": counts}


def bank_classify(request, body):
    """v2 classification: ONLY a real imported Daftra account is accepted.
    Stores {daftra_account_id, account_code, cost_center, counterparty, unit} on the
    txn (erp_class), flips status to reviewed, audits, saves. Bulk via ids[]."""
    ids = body.get("ids") or ([body.get("id")] if body.get("id") else [])
    ids = [str(i) for i in ids if i]
    if not ids:
        return {"error": "no_ids"}, 400
    clear = bool(body.get("clear"))
    acct_idx = _acct_by_id()
    acc = cc = None
    aid = ccid = ""
    if not clear:
        aid = str(body.get("account_id") or "").strip()
        acc = acct_idx.get(aid)
        if not acc:
            return {"error": "account_not_in_chart",
                    "message_ar": "اختر حسابًا من دليل دافترة المستورد — النص الحر مرفوض.",
                    "message_en": "Pick an account from the imported Daftra chart — free text is rejected."}, 422
        ccid = str(body.get("cost_center_id") or "").strip()
        if ccid:
            cc = next((r for r in _cc_records() if str(r.get("source_id")) == ccid), None)
            if not cc:
                return {"error": "cost_center_unknown",
                        "message_ar": "مركز التكلفة غير موجود في المستورد من دافترة.",
                        "message_en": "Cost center not found in the imported Daftra data."}, 422
    now = datetime.now(B.TZ).isoformat(timespec="seconds")
    who = actor(request)
    rows, missing = [], []
    for i in ids:
        x = B._fb_bank.get(i)
        if not x:
            missing.append(i)
            continue
        before = {"status": x.get("status"), "erp_class": x.get("erp_class")}
        if clear:
            x.pop("erp_class", None)
            x["status"] = "needs_review"
        else:
            x["erp_class"] = {"daftra_account_id": aid, "account_code": acc.get("code") or "",
                              "account_name": acc.get("display_name") or "",
                              "cost_center_id": ccid,
                              "cost_center_name": (cc or {}).get("display_name") or "",
                              "counterparty": str(body.get("counterparty") or "").strip()[:120],
                              "unit": str(body.get("unit") or "").strip()[:80],
                              "by": who, "at": now}
            x["status"] = "reviewed"
            x["reviewed_by"] = who
            x["reviewed_at"] = now
            if (body.get("unit") or "").strip():
                x["apartment"] = str(body.get("unit")).strip()[:80]
        B._fb_audit_add(who, "erp_unclassify" if clear else "erp_classify", "bank", i,
                        before=before, after=x.get("erp_class"))
        rows.append(_bank_row(x))
    if rows:
        B._fb_save("finance_bank_transactions.json", B._fb_bank)
    return {"ok": True, "rows": rows, "missing": missing, "counters": counters_snapshot()}, 200


# ====================== Rules engine (Slice 3) ======================
# Rules are NEW v2 data (erp_rules.json in STATE_DIR) — they never mutate the
# sacred stores except by writing erp_class on needs_review bank txns, exactly
# like a human classification (and ALWAYS leaving the >=3000 Faisal approval
# lane untouched: classification never bypasses approval).

_RULES_FILE = "erp_rules.json"
_rules_cache = {"v": None}


def rules():
    if _rules_cache["v"] is None:
        _rules_cache["v"] = B._load_json(_RULES_FILE, []) or []
    return _rules_cache["v"]


def _rules_save():
    B._save_json(_RULES_FILE, _rules_cache["v"])


def _norm(s):
    try:
        return B._fb_ar_norm(s or "")
    except Exception:
        return (s or "").lower()


def rule_matches(rule, x):
    m = rule.get("matcher") or {}
    c = m.get("desc_contains") or ""
    if c and _norm(c) not in _norm(x.get("description")):
        return False
    deb = B._fb_money(x.get("debit"))
    cred = B._fb_money(x.get("credit"))
    d = m.get("direction") or "any"
    if d == "out" and not deb > 0:
        return False
    if d == "in" and not cred > 0:
        return False
    amt = deb if deb > 0 else cred
    try:
        if m.get("amount_min") not in (None, "") and float(amt) < float(m["amount_min"]):
            return False
        if m.get("amount_max") not in (None, "") and float(amt) > float(m["amount_max"]):
            return False
    except (TypeError, ValueError):
        return False
    return True


def _rule_apply_to_txn(rule, x, who, now):
    s = rule.get("set") or {}
    x["erp_class"] = {"daftra_account_id": s.get("account_id") or "",
                      "account_code": s.get("account_code") or "",
                      "account_name": s.get("account_name") or "",
                      "cost_center_id": s.get("cost_center_id") or "",
                      "cost_center_name": s.get("cost_center_name") or "",
                      "counterparty": s.get("counterparty") or "",
                      "unit": s.get("unit") or "",
                      "by": who, "at": now, "rule_id": rule["id"], "auto": True}
    x["status"] = "reviewed"
    x["reviewed_by"] = who
    x["reviewed_at"] = now
    if s.get("unit"):
        x["apartment"] = s.get("unit")


def rules_apply_pending(who, only_rule=None):
    """Run enabled rules over ALL needs_review txns (idempotent — only fills
    unclassified rows). Returns the changed rows. >=3000 approval unaffected."""
    now = datetime.now(B.TZ).isoformat(timespec="seconds")
    active = [r for r in rules()
              if r.get("enabled", True) and int(r.get("strength") or 0) > 0
              and (only_rule is None or r["id"] == only_rule)]
    if not active:
        return []
    changed = []
    for x in B._fb_bank.values():
        if x.get("status") != "needs_review":
            continue
        for r in active:
            if rule_matches(r, x):
                _rule_apply_to_txn(r, x, who, now)
                r["hits"] = int(r.get("hits") or 0) + 1
                r["last_hit_at"] = now
                changed.append(_bank_row(x))
                break
    if changed:
        B._fb_save("finance_bank_transactions.json", B._fb_bank)
        _rules_save()
    return changed


def rules_list(request):
    return {"ok": True, "rules": rules(), "is_admin": is_admin(request)}


def rule_create(request, body):
    """Create a rule from a classification («طبّق على المشابهة») and optionally
    apply it now to every matching unclassified txn."""
    aid = str(body.get("account_id") or "").strip()
    acc = _acct_by_id().get(aid)
    if not acc:
        return {"error": "account_not_in_chart",
                "message_ar": "القاعدة لازم تشير لحساب من دليل دافترة.",
                "message_en": "A rule must target an imported Daftra account."}, 422
    contains = str(body.get("contains") or "").strip()
    if len(contains) < 2:
        return {"error": "matcher_too_weak",
                "message_ar": "حدد نص مطابقة أطول (حرفين على الأقل).",
                "message_en": "Give the matcher at least 2 characters."}, 422
    ccid = str(body.get("cost_center_id") or "").strip()
    cc = None
    if ccid:
        cc = next((r for r in _cc_records() if str(r.get("source_id")) == ccid), None)
        if not cc:
            return {"error": "cost_center_unknown"}, 422
    now = datetime.now(B.TZ).isoformat(timespec="seconds")
    rule = {"id": "rule-" + uuid.uuid4().hex[:10],
            "matcher": {"desc_contains": contains,
                        "direction": (body.get("direction") or "any"),
                        "amount_min": body.get("amount_min"),
                        "amount_max": body.get("amount_max")},
            "set": {"account_id": aid, "account_code": acc.get("code") or "",
                    "account_name": acc.get("display_name") or "",
                    "cost_center_id": ccid,
                    "cost_center_name": (cc or {}).get("display_name") or "",
                    "counterparty": str(body.get("counterparty") or "").strip()[:120],
                    "unit": str(body.get("unit") or "").strip()[:80]},
            "enabled": True, "hits": 0, "strength": 3, "weakened": 0,
            "created_by": actor(request), "created_at": now, "last_hit_at": ""}
    rules().insert(0, rule)
    _rules_save()
    B._fb_audit_add(actor(request), "erp_rule_create", "rule", rule["id"], after=rule)
    applied = rules_apply_pending(actor(request), only_rule=rule["id"]) if body.get("apply_now") else []
    return {"ok": True, "rule": rule, "rows": applied, "applied": len(applied),
            "counters": counters_snapshot()}, 200


def rule_toggle(request, body):
    rid = str(body.get("id") or "")
    r = next((x for x in rules() if x["id"] == rid), None)
    if not r:
        return {"error": "rule_not_found"}, 404
    r["enabled"] = bool(body.get("enabled"))
    if r["enabled"] and int(r.get("strength") or 0) <= 0:
        r["strength"] = 1          # re-enabling revives a weakened-out rule at minimum strength
    _rules_save()
    B._fb_audit_add(actor(request), "erp_rule_toggle", "rule", rid, after={"enabled": r["enabled"]})
    return {"ok": True, "rule": r}, 200


def rule_delete(request, body):
    if not is_admin(request):
        return {"error": "forbidden", "message_ar": "حذف القواعد للأدمن فقط.",
                "message_en": "Deleting rules is admin-only."}, 403
    rid = str(body.get("id") or "")
    before = len(rules())
    _rules_cache["v"] = [x for x in rules() if x["id"] != rid]
    if len(rules()) == before:
        return {"error": "rule_not_found"}, 404
    _rules_save()
    B._fb_audit_add(actor(request), "erp_rule_delete", "rule", rid)
    return {"ok": True, "id": rid}, 200


def rule_undo(request, body):
    """One-click undo of an auto-classification: clears the txn back to
    needs_review and WEAKENS the rule (strength-1; 0 disables it)."""
    txid = str(body.get("txn_id") or "")
    x = B._fb_bank.get(txid)
    if not x:
        return {"error": "txn_not_found"}, 404
    ec = x.get("erp_class") or {}
    rid = ec.get("rule_id")
    if not (rid and ec.get("auto")):
        return {"error": "not_auto_classified"}, 400
    x.pop("erp_class", None)
    x["status"] = "needs_review"
    r = next((q for q in rules() if q["id"] == rid), None)
    if r:
        r["strength"] = max(0, int(r.get("strength") or 0) - 1)
        r["weakened"] = int(r.get("weakened") or 0) + 1
        if r["strength"] == 0:
            r["enabled"] = False
        _rules_save()
    B._fb_save("finance_bank_transactions.json", B._fb_bank)
    B._fb_audit_add(actor(request), "erp_rule_undo", "bank", txid,
                    after={"rule_id": rid, "strength": (r or {}).get("strength")})
    return {"ok": True, "row": _bank_row(x), "rule": r, "counters": counters_snapshot()}, 200


def rules_precision():
    """Replay every rule against txns that carry a NON-auto (human) classification —
    the measured precision the slice proof requires. No writes."""
    ground = [x for x in B._fb_bank.values()
              if (x.get("erp_class") or {}).get("daftra_account_id")
              and not (x.get("erp_class") or {}).get("auto")]
    out, tot_match, tot_correct = [], 0, 0
    for r in rules():
        m = [x for x in ground if rule_matches(r, x)]
        correct = sum(1 for x in m
                      if str((x.get("erp_class") or {}).get("daftra_account_id")) ==
                         str((r.get("set") or {}).get("account_id")))
        pending = sum(1 for x in B._fb_bank.values()
                      if x.get("status") == "needs_review" and rule_matches(r, x))
        tot_match += len(m)
        tot_correct += correct
        out.append({"id": r["id"], "contains": (r.get("matcher") or {}).get("desc_contains"),
                    "account": (r.get("set") or {}).get("account_name"),
                    "enabled": r.get("enabled", True), "hits": r.get("hits", 0),
                    "matched_human": len(m), "agree": correct,
                    "precision": (round(100.0 * correct / len(m)) if m else None),
                    "would_apply_now": pending})
    overall = (round(100.0 * tot_correct / tot_match) if tot_match else None)
    return {"ok": True, "rules": out, "ground_truth_rows": len(ground),
            "overall_precision": overall}


# ====================== المطابقة Matching (Slice 4) ======================
# ONE queue across four engines. Engine 1 (بنك↔دافترة) REUSES the existing dup
# machinery wholesale (suggestions + link/link_distributed/not_duplicate via
# delegation). Engines 2-4 (مصاريف/مؤسس وبطاقات/Hostaway) are new scorers that
# write additive fields only. EVERY decision is appended to erp_match_log.json
# with who/when/what-was-suggested.

_MATCH_LOG_FILE = "erp_match_log.json"
_DISMISS_KEY = "erp_match_dismissed"


def match_log_add(request, txn_id, engine, action, detail):
    log = B._load_json(_MATCH_LOG_FILE, []) or []
    log.append({"at": datetime.now(B.TZ).isoformat(timespec="seconds"),
                "by": actor(request), "txn": txn_id, "engine": engine,
                "action": action, "detail": detail})
    if len(log) > 8000:
        del log[:len(log) - 8000]
    B._save_json(_MATCH_LOG_FILE, log)


def match_log_recent(limit=80):
    log = B._load_json(_MATCH_LOG_FILE, []) or []
    return list(reversed(log[-limit:]))


def _pdate(s):
    try:
        return datetime.fromisoformat(str(s)[:10]).date()
    except Exception:
        return None


def _exp_candidates(x, limit=3):
    """Engine 2: bank debit ↔ OJ-EXP expense records (amount tight, ±7d, text sim)."""
    deb = float(B._fb_money(x.get("debit")))
    if deb <= 0:
        return []
    d0 = _pdate(x.get("date"))
    out = []
    for ex in B._expenses.values():
        try:
            amt = float(ex.get("amount") or 0)
        except (TypeError, ValueError):
            continue
        if amt <= 0 or abs(amt - deb) > max(2.0, deb * 0.01):
            continue
        if ex.get("bank_txn_id"):
            continue                       # already consumed by another bank txn
        d1 = _pdate(ex.get("expense_date") or ex.get("date"))
        dd = abs((d1 - d0).days) if (d0 and d1) else None
        if dd is None or dd > 7:
            continue
        txt = " ".join(str(ex.get(k) or "") for k in ("concept", "apartment", "category"))
        try:
            sim = B._fb_text_sim(x.get("description") or "", txt)
        except Exception:
            sim = 0.0
        score = 58 + (20 if dd <= 1 else 12 if dd <= 3 else 6) + min(18, int(sim * 40)) \
                + (3 if abs(amt - deb) < 0.01 else 0)
        out.append({"engine": "exp", "key": str(ex.get("id")), "score": min(score, 99),
                    "label": (ex.get("concept") or ex.get("category") or "OJ-EXP"),
                    "sub": ex.get("apartment") or "",
                    "date": ex.get("expense_date") or "", "amount": B._fb_money_str(amt)})
    out.sort(key=lambda c: -c["score"])
    return out[:limit]


def _hostaway_candidates(x, res_list, limit=3):
    """Engine 4: channel_payout credits ↔ cached Hostaway reservations (تقريبي —
    payout ≈ totalPrice minus 0–4.5% channel fee, departure within ±10d)."""
    cred = float(B._fb_money(x.get("credit")))
    if cred <= 0 or (x.get("category") or "") != "channel_payout":
        return []
    d0 = _pdate(x.get("date"))
    out = []
    for r in res_list or []:
        if (r.get("status") or "") not in ("new", "modified"):
            continue
        try:
            tp = float(r.get("totalPrice") or 0)
        except (TypeError, ValueError):
            continue
        if tp <= 0 or not (tp * 0.955 - 1 <= cred <= tp + 1):
            continue
        d1 = _pdate(r.get("departureDate")) or _pdate(r.get("arrivalDate"))
        dd = abs((d1 - d0).days) if (d0 and d1) else None
        if dd is None or dd > 10:
            continue
        closeness = 1.0 - min(1.0, abs(tp - cred) / max(tp, 1.0) / 0.045)
        score = int(50 + 25 * closeness + (15 if dd <= 2 else 8 if dd <= 5 else 3))
        out.append({"engine": "hostaway", "key": str(r.get("id")), "score": min(score, 95),
                    "label": r.get("guestName") or "حجز", "sub": str(r.get("listingMapId") or ""),
                    "date": r.get("departureDate") or "", "amount": B._fb_money_str(tp),
                    "approx": True})
    out.sort(key=lambda c: -c["score"])
    return out[:limit]


def _founder_candidates(x):
    """Engine 3: card settlements (via the card→employee registry) + founder/
    partner/fit-out transfer suggestions from the existing keyword sets."""
    out = []
    card = (x.get("card_last4") or "").strip()
    if card:
        m = (B._fb_cards or {}).get(card)
        if m and m.get("employee"):
            out.append({"engine": "founder", "key": "card:" + card, "score": 84, "kind": "card",
                        "label": "تسوية بطاقة — " + m["employee"], "label_en": "Card settlement — " + m["employee"],
                        "sub": m["employee"], "flow": "employee_advance_settlement",
                        "employee": m["employee"]})
    if float(B._fb_money(x.get("credit"))) > 0:
        desc = _norm(x.get("description"))
        try:
            if any(_norm(k) in desc for k in B._FB_FITOUT_KW):
                out.append({"engine": "founder", "key": "flow:fitout", "score": 64, "kind": "flow",
                            "label": "وارد تجهيز من مالك", "label_en": "Owner fit-out funding",
                            "sub": "", "flow": "cash_in_owner_fitout_funding"})
            elif any(_norm(k) in desc for k in (tuple(B._FB_PARTNER_KW) + tuple(B._FB_FUNDING_KW))):
                out.append({"engine": "founder", "key": "flow:partner", "score": 66, "kind": "flow",
                            "label": "تمويل من جاري الشريك/المؤسس", "label_en": "Partner/founder funding",
                            "sub": "", "flow": "cash_in_partner_funding"})
        except Exception:
            pass
    return out


_DAFTRA_SUGGESTED = ("possible_duplicate", "strong_possible_duplicate",
                     "needs_manual_review", "already_in_daftra_verified")


def _daftra_candidate(x):
    dq = x.get("daftra_duplicate_status")
    if dq not in _DAFTRA_SUGGESTED or x.get("match_status") == "matched":
        return None
    return {"engine": "daftra", "key": str(x.get("matched_daftra_id") or ""),
            "score": int(x.get("daftra_duplicate_confidence") or 0),
            "label": ("قيد #" + str(x.get("matched_daftra_number"))) if x.get("matched_daftra_number") else "قيد دافترة",
            "label_en": ("Journal #" + str(x.get("matched_daftra_number"))) if x.get("matched_daftra_number") else "Daftra journal",
            "sub": x.get("daftra_duplicate_reason_ar") or "",
            "sub_en": x.get("daftra_duplicate_reason_en") or "",
            "needs_drawer": True}


def _match_item_for(x, res_list):
    dismissed = x.get(_DISMISS_KEY) or {}
    cands = []
    dc = None if dismissed.get("daftra") else _daftra_candidate(x)
    if dc:
        cands.append(dc)
    if not dismissed.get("exp"):
        cands.extend(_exp_candidates(x))
    if not dismissed.get("founder"):
        cands.extend(_founder_candidates(x))
    if not dismissed.get("hostaway"):
        cands.extend(_hostaway_candidates(x, res_list))
    cands.sort(key=lambda c: -c["score"])
    amount, dirn = _amt(x)
    return {"id": x["id"], "date": x.get("date") or "", "desc": (x.get("description") or "")[:160],
            "amount": amount, "dir": dirn, "category": x.get("category") or "unknown",
            "card": x.get("card_last4") or "", "dup": x.get("daftra_duplicate_status") or "not_checked",
            "cands": cands, "dismissed": dismissed}


def match_queue(params):
    """The ONE matching queue. Heavy-ish (scans expenses × unmatched txns) —
    the handler runs it in a thread."""
    engine = (params.get("engine") or "all").strip()
    try:
        page = max(1, int(params.get("p") or 1))
    except Exception:
        page = 1
    ps = 30
    res_list = []
    need_hostaway = any(
        (x.get("category") or "") == "channel_payout"
        and float(B._fb_money(x.get("credit"))) > 0
        and (x.get("match_status") or "unmatched") == "unmatched"
        for x in B._fb_bank.values())
    if need_hostaway:
        try:
            res_list = B.get_reservations_cached() or []
        except Exception:
            res_list = []
    items = []
    for x in B._fb_bank.values():
        if (x.get("match_status") or "unmatched") == "matched":
            continue
        if x.get("daftra_duplicate_status") == "ignored" or x.get("ledger_entry_id"):
            continue
        items.append(_match_item_for(x, res_list))
    counts = {"all": len(items),
              "daftra": sum(1 for i in items if any(c["engine"] == "daftra" for c in i["cands"])),
              "exp": sum(1 for i in items if any(c["engine"] == "exp" for c in i["cands"])),
              "founder": sum(1 for i in items if any(c["engine"] == "founder" for c in i["cands"])),
              "hostaway": sum(1 for i in items if any(c["engine"] == "hostaway" for c in i["cands"])),
              "none": sum(1 for i in items if not i["cands"])}
    if engine in ("daftra", "exp", "founder", "hostaway"):
        items = [i for i in items if any(c["engine"] == engine for c in i["cands"])]
    elif engine == "none":
        items = [i for i in items if not i["cands"]]
    items.sort(key=lambda i: (-(i["cands"][0]["score"] if i["cands"] else -1), i["date"]), reverse=False)
    total = len(items)
    pages = max(1, (total + ps - 1) // ps)
    page = min(page, pages)
    return {"ok": True, "items": items[(page - 1) * ps: (page - 1) * ps + ps],
            "total": total, "page": page, "pages": pages, "counts": counts,
            "hostaway_cache": len(res_list)}


def match_accept(request, body):
    """Accept a non-Daftra candidate (Daftra accepts go through the delegated
    dup endpoint so verification semantics stay identical)."""
    x = B._fb_bank.get(str(body.get("id") or ""))
    if not x:
        return {"error": "txn_not_found"}, 404
    engine = (body.get("engine") or "").strip()
    key = str(body.get("key") or "")
    now = datetime.now(B.TZ).isoformat(timespec="seconds")
    who = actor(request)
    if engine == "exp":
        ex = B._expenses.get(key) or next(
            (e for e in B._expenses.values() if str(e.get("id")) == key), None)
        if not ex:
            return {"error": "expense_not_found"}, 404
        if ex.get("bank_txn_id") and ex.get("bank_txn_id") != x["id"]:
            return {"error": "expense_already_matched",
                    "message_ar": "هذا المصروف مرتبط بحركة بنك ثانية.",
                    "message_en": "This expense is already matched to another bank txn."}, 409
        x["match_status"] = "matched"
        x["erp_match"] = {"engine": "exp", "key": str(ex.get("id")), "at": now, "by": who,
                          "label": ex.get("concept") or "", "amount": str(ex.get("amount") or "")}
        ex["bank_txn_id"] = x["id"]
        ex["bank_match_at"] = now
        B._save_json("expenses.json", B._expenses)
    elif engine == "hostaway":
        x["match_status"] = "matched"
        x["erp_match"] = {"engine": "hostaway", "key": key, "at": now, "by": who,
                          "label": body.get("label") or "", "approx": True}
    elif engine == "founder":
        flow = (body.get("flow") or "").strip()
        if flow not in ("employee_advance_settlement", "cash_in_owner_fitout_funding",
                        "cash_in_partner_funding"):
            return {"error": "bad_flow"}, 400
        x["daftra_flow_type"] = flow
        x["flow_manual"] = True
        if flow == "employee_advance_settlement":
            if body.get("employee"):
                x["custody_employee"] = str(body.get("employee"))[:80]
            if (x.get("category") or "unknown") == "unknown":
                x["category"] = "employee_card_settlement"
        elif flow == "cash_in_owner_fitout_funding":
            if (x.get("category") or "unknown") == "unknown":
                x["category"] = "owner_fitout_funding"
        else:
            if (x.get("category") or "unknown") == "unknown":
                x["category"] = "founder_funding"
        if x.get("status") == "needs_review":
            x["status"] = "reviewed"
            x["reviewed_by"] = who
            x["reviewed_at"] = now
        x["match_status"] = "matched"
        x["erp_match"] = {"engine": "founder", "key": key, "at": now, "by": who, "flow": flow}
    else:
        return {"error": "bad_engine"}, 400
    B._fb_save("finance_bank_transactions.json", B._fb_bank)
    B._fb_audit_add(who, "erp_match_accept", "bank", x["id"],
                    after={"engine": engine, "key": key})
    match_log_add(request, x["id"], engine, "accept",
                  {"key": key, "suggested": body.get("suggested") or []})
    return {"ok": True, "row": _bank_row(x), "counters": counters_snapshot()}, 200


def match_reject(request, body):
    """Dismiss an engine's candidates for this txn (it leaves that engine's queue)."""
    x = B._fb_bank.get(str(body.get("id") or ""))
    if not x:
        return {"error": "txn_not_found"}, 404
    engine = (body.get("engine") or "").strip()
    if engine not in ("daftra", "exp", "founder", "hostaway"):
        return {"error": "bad_engine"}, 400
    dis = x.get(_DISMISS_KEY) or {}
    dis[engine] = True
    x[_DISMISS_KEY] = dis
    B._fb_save("finance_bank_transactions.json", B._fb_bank)
    match_log_add(request, x["id"], engine, "reject",
                  {"suggested": body.get("suggested") or []})
    return {"ok": True, "dismissed": dis, "counters": counters_snapshot()}, 200


# ============ المصاريف Expenses (V4 re-shell) + العهد Custody (Slice 5) ============
# The V4 approval-center flow is REUSED via delegation (overview/detail/approve/
# reject/edit/export/recheck are bot.py handlers). v2 only decorates rows with
# the bank-match chip (bank_txn_id written by the Matching engine).

def exp_attach_bank(payload):
    for r in payload.get("rows") or []:
        ex = B._expenses.get(str(r.get("expense_id")))
        if ex is not None:
            r["bank_txn_id"] = ex.get("bank_txn_id") or ""
    return payload


def custody_payload():
    """Per-employee advances from the existing Decimal-correct balance math."""
    return {"ok": True, **B._fb_custody_balances()}


# ========== Slice 7: القوائم المالية + الإقفال والترحيل + الميزانية ==========

def _month_key_or_now(v):
    v = (v or "").strip()[:7]
    if len(v) == 7 and v[:2] == "20" and v[4] == "-":
        return v
    return datetime.now(B.TZ).date().isoformat()[:7]


def _month_bounds_iso(mkey):
    y, m = int(mkey[:4]), int(mkey[5:7])
    return "%04d-%02d-01" % (y, m), "%04d-%02d-%02d" % (y, m, monthrange(y, m)[1])


def _prior_month(mkey):
    y, m = int(mkey[:4]), int(mkey[5:7]) - 1
    if m == 0:
        y, m = y - 1, 12
    return "%04d-%02d" % (y, m)


def _bank_account_ids():
    return {str((v or {}).get("daftra_account_id"))
            for v in (B._fb_bankmap or {}).values() if (v or {}).get("daftra_account_id")}


def _bank_register_delta(start_iso, end_iso):
    tot = Decimal(0)
    for x in B._fb_bank.values():
        d = (x.get("date") or "")[:10]
        if start_iso <= d <= end_iso:
            tot += B._fb_money(x.get("credit")) - B._fb_money(x.get("debit"))
    return str(tot)


def _daftra_sync_stamp():
    run = B._fb_last_run("daftra") or {}
    return run.get("finished_at") or run.get("started_at") or "—"


def stmts_payload(params):
    mkey = _month_key_or_now(params.get("m"))
    start, end = _month_bounds_iso(mkey)
    pm = _prior_month(mkey)
    ps, pe = _month_bounds_iso(pm)
    res = ST.build_statements(B._fb_djournals, _acct_records(), start, end, ps, pe,
                              _bank_account_ids(), _bank_register_delta(start, end))
    stamp = _daftra_sync_stamp()
    res.update({"ok": True, "month": mkey, "prior_month": pm,
                "journals_count": len(B._fb_djournals),
                "provenance_ar": "محسوبة من قيود دافترة المستوردة حتى " + str(stamp) + " — دافترة هي المصدر الرسمي",
                "provenance_en": "Computed from imported Daftra journals as of " + str(stamp) + " — Daftra is the official ledger"})
    return res


def stmts_account_lines(params):
    mkey = _month_key_or_now(params.get("m"))
    start, end = _month_bounds_iso(mkey)
    return {"ok": True,
            "rows": ST.account_lines(B._fb_djournals, params.get("id") or "", start, end)}


def stmts_type_probe():
    """Diagnostic for «verify Daftra's type field live»: what each imported
    account's payload actually carries."""
    rows = []
    keys_hist = {}
    for rec in _acct_records():
        typ, key, raw = ST.detect_type(rec.get("source_payload") or {})
        if key:
            keys_hist[key] = keys_hist.get(key, 0) + 1
        rows.append({"account_id": str(rec.get("source_id")), "code": rec.get("code") or "",
                     "name": rec.get("display_name") or "", "type": typ,
                     "probe_key": key, "probe_raw": raw})
    typed = sum(1 for r in rows if r["type"])
    return {"ok": True, "rows": rows, "keys_histogram": keys_hist,
            "typed": typed, "total": len(rows)}


def stmts_xlsx(payload):
    """Excel export — 4 sheets, plain and audit-friendly."""
    import io
    import openpyxl
    wb = openpyxl.Workbook()

    def sheet(title, rows):
        ws = wb.create_sheet(title)
        for r in rows:
            ws.append(r)
        return ws

    wb.remove(wb.active)
    bs = payload["balance_sheet"]
    rows = [["قائمة المركز المالي — " + payload["month"]], []]
    for sec, label in (("asset", "الأصول"), ("liability", "الخصوم"), ("equity", "حقوق الملكية"), ("untyped", "غير مصنّف النوع")):
        rows.append([label])
        for r in bs["rows"][sec]:
            rows.append([r["code"], r["name"], r["amount"], r.get("prior")])
        rows.append([])
    tt = bs["totals"]
    rows += [["الإجمالي", "الأصول", tt["assets"]], ["", "الخصوم", tt["liabilities"]],
             ["", "حقوق الملكية", tt["equity"]], ["", "أرباح جارية", tt["current_earnings"]],
             ["", "الفجوة", tt["gap"]]]
    sheet("المركز المالي", rows)

    inc = payload["income"]
    rows = [["قائمة الدخل — " + payload["month"]], [], ["الإيرادات"]]
    rows += [[r["code"], r["name"], r["amount"], r.get("prior")] for r in inc["income_rows"]]
    rows += [[], ["المصروفات"]]
    rows += [[r["code"], r["name"], r["amount"], r.get("prior")] for r in inc["expense_rows"]]
    rows += [[], ["صافي الدخل", "", inc["totals"]["net"], inc["prior_totals"]["net"]]]
    rows += [[], ["حسب مركز التكلفة"]]
    rows += [[c["name"] or c["cost_center_id"], c["income"], c["expense"], c["net"]] for c in inc["by_cost_center"]]
    sheet("الدخل", rows)

    eq = payload["equity"]
    sheet("حقوق الملكية", [
        ["قائمة التغير في حقوق الملكية — " + payload["month"]], [],
        ["الرصيد الافتتاحي", eq["opening"]], ["صافي الدخل", eq["net_income"]],
        ["إضافات الملاك", eq["contributions"]], ["مسحوبات", eq["withdrawals"]],
        ["الرصيد الختامي", eq["closing"]], ["مطابقة مع المركز المالي", "نعم" if eq["ties_to_balance_sheet"] else ("فجوة " + str(eq["gap"]))]])

    cf = payload["cash_flow"]
    g = cf["groups"]
    sheet("التدفقات النقدية", [
        ["قائمة التدفقات النقدية (مباشرة) — " + payload["month"]], [],
        ["من الإيرادات", g["income"]], ["مصروفات", g["expense"]],
        ["أصول", g["asset"]], ["التزامات", g["liability"]], ["حقوق/مالك", g["equity"]],
        ["غير مصنّف", g["untyped"]], [],
        ["صافي التدفق", cf["net_cash"]],
        ["نقد افتتاحي", cf["opening_cash"]], ["نقد ختامي", cf["closing_cash"]],
        ["حركة سجل البنك", cf["bank_register_delta"]],
        ["مطابقة سجل البنك", {True: "نعم", False: "لا", None: "—"}[cf["ties_bank_register"]]]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def stmts_pdf(payload):
    """PDF export. Hard-fails loudly (PdfFontError) if the Arabic font can't load."""
    from fpdf import FPDF
    import arabic_reshaper
    from bidi.algorithm import get_display

    font_path = B._pdf_font()        # raises B.PdfFontError when unavailable

    def ar(s):
        try:
            return get_display(arabic_reshaper.reshape(str(s)))
        except Exception:
            return str(s)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_font("ar", "", font_path, uni=True)

    def h1(txt):
        pdf.set_font("ar", size=15)
        pdf.cell(0, 9, ar(txt), align="R")
        pdf.ln(11)

    def row(a, b, c=""):
        pdf.set_font("ar", size=10)
        pdf.cell(38, 6.5, str(c), align="L")
        pdf.cell(38, 6.5, str(b), align="L")
        pdf.cell(0, 6.5, ar(a), align="R")
        pdf.ln(6.5)

    pdf.add_page()
    h1("القوائم المالية — " + payload["month"] + " · عوجا")
    pdf.set_font("ar", size=8.5)
    pdf.cell(0, 5, ar(payload["provenance_ar"]), align="R")
    pdf.ln(8)

    bs = payload["balance_sheet"]
    h1("قائمة المركز المالي")
    for sec, label in (("asset", "الأصول"), ("liability", "الخصوم"), ("equity", "حقوق الملكية"), ("untyped", "غير مصنّف النوع")):
        if not bs["rows"][sec]:
            continue
        pdf.set_font("ar", size=11)
        pdf.cell(0, 7, ar(label), align="R")
        pdf.ln(7.5)
        for r in bs["rows"][sec][:40]:
            row(r["name"], "%.2f" % r["amount"], r["code"])
    tt = bs["totals"]
    row("الفجوة" if not bs["balanced"] else "متوازنة", "%.2f" % tt["gap"], "")
    pdf.ln(4)

    inc = payload["income"]
    pdf.add_page()
    h1("قائمة الدخل")
    for r in inc["income_rows"][:35]:
        row(r["name"], "%.2f" % r["amount"], r["code"])
    for r in inc["expense_rows"][:35]:
        row(r["name"], "-%.2f" % r["amount"], r["code"])
    row("صافي الدخل", "%.2f" % inc["totals"]["net"], "")

    eq = payload["equity"]
    pdf.ln(4)
    h1("التغير في حقوق الملكية")
    row("الرصيد الافتتاحي", "%.2f" % eq["opening"], "")
    row("صافي الدخل", "%.2f" % eq["net_income"], "")
    row("إضافات الملاك", "%.2f" % eq["contributions"], "")
    row("مسحوبات", "%.2f" % eq["withdrawals"], "")
    row("الرصيد الختامي", "%.2f" % eq["closing"], "")

    cf = payload["cash_flow"]
    pdf.ln(4)
    h1("التدفقات النقدية (مباشرة)")
    labels = {"income": "من الإيرادات", "expense": "مصروفات", "asset": "أصول",
              "liability": "التزامات", "equity": "حقوق/مالك", "untyped": "غير مصنّف"}
    for k, v in cf["groups"].items():
        row(labels.get(k, k), "%.2f" % v, "")
    row("صافي التدفق", "%.2f" % cf["net_cash"], "")
    row("حركة سجل البنك", str(cf["bank_register_delta"]), "")

    out = pdf.output(dest="S")
    return out.encode("latin-1") if isinstance(out, str) else bytes(out)


# ---------------- month close (gated checklist → immutable snapshot) ----------------

_CLOSE_FILE = "erp_month_close.json"
_MIG_FILE = "erp_migrations.json"


def _close_store():
    return B._load_json(_CLOSE_FILE, {}) or {}


def close_checks(mkey):
    start, end = _month_bounds_iso(mkey)
    in_month = [x for x in B._fb_bank.values() if start <= (x.get("date") or "")[:10] <= end]
    bank_uncls = sum(1 for x in in_month if x.get("status") == "needs_review")
    unmatched = sum(1 for x in in_month
                    if (x.get("match_status") or "unmatched") == "unmatched"
                    and x.get("daftra_duplicate_status") != "ignored"
                    and not x.get("ledger_entry_id"))
    exp_pending = 0
    for e in B._expenses.values():
        d = (e.get("expense_date") or "")[:10]
        if not (start <= d <= end):
            continue
        try:
            st = B._exp4_approval_status(e)
        except Exception:
            st = e.get("approval_status") or ""
        if st in ("pending_approval", "needs_edit"):
            exp_pending += 1
    owners_unbalanced = []
    owners = sorted({(r.get("owner") or "").strip() for r in _registry_rows() if (r.get("owner") or "").strip()})
    for o in owners:
        try:
            rep = B._owner_month_report(o, mkey)
        except Exception:
            rep = None
        if rep and not (rep.get("reconciliation") or {}).get("balanced"):
            owners_unbalanced.append(o)
    needs_faisal = B._fb_inbox({})["counts"].get("needs_faisal", 0)
    checks = [
        {"key": "bank_classified", "ok": bank_uncls == 0, "count": bank_uncls},
        {"key": "matching_done", "ok": unmatched == 0, "count": unmatched},
        {"key": "expenses_approved", "ok": exp_pending == 0, "count": exp_pending},
        {"key": "owners_balanced", "ok": not owners_unbalanced, "count": len(owners_unbalanced),
         "owners": owners_unbalanced[:12]},
        {"key": "approvals_clear", "ok": needs_faisal == 0, "count": needs_faisal},
    ]
    return checks, all(c["ok"] for c in checks)


def close_get(params):
    mkey = _month_key_or_now(params.get("m"))
    checks, all_ok = close_checks(mkey)
    snap = _close_store().get(mkey)
    migs = (B._load_json(_MIG_FILE, {}) or {}).get(mkey) or []
    prev = migrate_entry_ids(mkey)
    return {"ok": True, "month": mkey, "checks": checks, "all_ok": all_ok,
            "closed": bool(snap), "snapshot": snap,
            "post_enabled": bool(getattr(B, "DAFTRA_POST_ENABLED", False)),
            "migrations": migs, "migratable_entries": len(prev)}


def close_do(request, body):
    if not is_admin(request):
        return {"error": "forbidden", "message_ar": "إقفال الشهر للأدمن فقط.",
                "message_en": "Month close is admin-only."}, 403
    mkey = _month_key_or_now(body.get("month"))
    store = _close_store()
    if store.get(mkey):
        return {"error": "already_closed", "snapshot": store[mkey],
                "message_ar": "الشهر مقفول من قبل — الإقفال نهائي.",
                "message_en": "Month already closed — snapshots are immutable."}, 409
    checks, all_ok = close_checks(mkey)
    if not all_ok:
        return {"error": "checklist_incomplete", "checks": checks,
                "message_ar": "ما نقدر نقفل — فيه بنود ناقصة بالقائمة.",
                "message_en": "Cannot close — checklist items remain."}, 422
    start, end = _month_bounds_iso(mkey)
    stm = stmts_payload({"m": mkey})
    snap = {"month": mkey, "closed_at": datetime.now(B.TZ).isoformat(timespec="seconds"),
            "closed_by": actor(request), "checks": checks,
            "income_totals": stm["income"]["totals"],
            "balance_totals": stm["balance_sheet"]["totals"],
            "bank_register_delta": _bank_register_delta(start, end)}
    store[mkey] = snap
    B._save_json(_CLOSE_FILE, store)
    B._fb_audit_add(actor(request), "erp_month_close", "close", mkey, after=snap)
    return {"ok": True, "snapshot": snap}, 200


# ---------------- الترحيل إلى دافترة (idempotent migration) ----------------

def migrate_entry_ids(mkey):
    start, end = _month_bounds_iso(mkey)
    ids = []
    for e in B._fb_ledger.values():
        d = (e.get("date") or "")[:10]
        if start <= d <= end and e.get("status") in ("approved", "ready_to_post"):
            ids.append(e["id"])
    return sorted(ids)


def migrate_preview(params):
    mkey = _month_key_or_now(params.get("m"))
    ids = migrate_entry_ids(mkey)
    draft = B._fb_journal_draft(ids) if ids else {"lines": [], "total": "0.00", "n": 0,
                                                  "issues": [], "ready": False}
    closed = bool(_close_store().get(mkey))
    return {"ok": True, "month": mkey, "entry_ids": ids, "draft": draft,
            "closed": closed, "post_enabled": bool(getattr(B, "DAFTRA_POST_ENABLED", False))}


def migrate_run(request, body):
    if not is_admin(request):
        return {"error": "forbidden", "message_ar": "الترحيل للأدمن فقط.",
                "message_en": "Migration is admin-only."}, 403
    mkey = _month_key_or_now(body.get("month"))
    if not _close_store().get(mkey):
        return {"error": "month_not_closed",
                "message_ar": "اقفل الشهر أولًا — الترحيل يفتح بعد الإقفال.",
                "message_en": "Close the month first — migration unlocks after close."}, 409
    ids = migrate_entry_ids(mkey)
    if not body.get("confirm"):
        # DRY-RUN: builds the preview, writes NOTHING anywhere.
        draft = B._fb_journal_draft(ids) if ids else {"lines": [], "n": 0, "issues": []}
        return {"ok": True, "dry_run": True, "month": mkey, "entry_ids": ids, "draft": draft}, 200
    if not ids:
        return {"ok": True, "month": mkey, "posted": 0, "entry_ids": [],
                "message_ar": "ما فيه قيود جاهزة للترحيل (الترحيل السابق أخذها — التشغيل المزدوج ما ينشئ شي).",
                "message_en": "No entries ready (a previous run consumed them — double-run creates nothing)."}, 200
    res = B._fb_journal_post(ids, actor(request))
    log = B._load_json(_MIG_FILE, {}) or {}
    log.setdefault(mkey, []).append({
        "at": datetime.now(B.TZ).isoformat(timespec="seconds"), "by": actor(request),
        "entry_ids": ids, "result_ok": bool(res.get("ok")),
        "disabled": bool(res.get("disabled")),
        # revert map: every source entry and (once verified) its Daftra journal id
        "revert_map": [{"entry_id": i,
                        "daftra_journal_id": (B._fb_ledger.get(i) or {}).get("daftra_journal_id")}
                       for i in ids]})
    B._save_json(_MIG_FILE, log)
    B._fb_audit_add(actor(request), "erp_migrate", "close", mkey,
                    after={"n": len(ids), "ok": bool(res.get("ok"))})
    status = 200 if res.get("ok") or res.get("disabled") else 502
    return {"ok": bool(res.get("ok")), "month": mkey, "result": res,
            "entry_ids": ids, "disabled": bool(res.get("disabled"))}, status


# ---------------- الميزانية budgets ----------------

_BUDGET_FILE = "erp_budgets.json"


def _budget_store():
    return B._load_json(_BUDGET_FILE, {}) or {}


def budget_actuals(mkey):
    """Actuals per Daftra account from CLASSIFIED bank txns in the month
    (debits − credits — spend-positive)."""
    start, end = _month_bounds_iso(mkey)
    out = {}
    for x in B._fb_bank.values():
        d = (x.get("date") or "")[:10]
        if not (start <= d <= end):
            continue
        aid = str(((x.get("erp_class") or {}).get("daftra_account_id")) or "")
        if not aid:
            continue
        out[aid] = out.get(aid, Decimal(0)) + B._fb_money(x.get("debit")) - B._fb_money(x.get("credit"))
    return out


def budget_get(params):
    mkey = _month_key_or_now(params.get("m"))
    store = _budget_store()
    month = store.get(mkey) or {"accounts": {}, "versions": []}
    actuals = budget_actuals(mkey)
    acct_idx = _acct_by_id()
    rows = []
    for aid, cfg in sorted(month["accounts"].items()):
        acc = acct_idx.get(aid) or {}
        st = ST.budget_row(cfg.get("amount") or 0, actuals.get(aid, 0))
        rows.append({"account_id": aid, "code": acc.get("code") or "",
                     "name": acc.get("display_name") or cfg.get("name") or aid,
                     "weekly": cfg.get("weekly") or [],
                     "week_start": cfg.get("week_start") or "sun", **st})
    # suggestions: last month's budget + 3-month average of actuals
    last = store.get(_prior_month(mkey)) or {}
    avg3 = {}
    mk = mkey
    for _ in range(3):
        mk = _prior_month(mk)
        for aid, v in budget_actuals(mk).items():
            avg3.setdefault(aid, []).append(v)
    sugg = {aid: ST.fnum(sum(vals, Decimal(0)) / len(vals)) for aid, vals in avg3.items() if vals}
    alerts = [r for r in rows if r["alert"]]
    return {"ok": True, "month": mkey, "rows": rows,
            "alerts": [{"account_id": r["account_id"], "name": r["name"], "pct": r["pct"],
                        "alert": r["alert"]} for r in alerts],
            "last_month_budget": {aid: (cfg.get("amount") or 0)
                                  for aid, cfg in (last.get("accounts") or {}).items()},
            "avg3_actuals": sugg, "versions": (month.get("versions") or [])[-10:]}


def budget_set(request, body):
    mkey = _month_key_or_now(body.get("month"))
    store = _budget_store()
    month = store.setdefault(mkey, {"accounts": {}, "versions": []})
    action = (body.get("action") or "set").strip()
    now = datetime.now(B.TZ).isoformat(timespec="seconds")
    acct_idx = _acct_by_id()
    if action == "copy_last":
        last = store.get(_prior_month(mkey)) or {}
        if not last.get("accounts"):
            return {"error": "no_last_month", "message_ar": "ما فيه ميزانية الشهر الماضي.",
                    "message_en": "No last-month budget to copy."}, 404
        month["versions"].append({"at": now, "by": actor(request),
                                  "accounts": json.loads(json.dumps(month["accounts"]))})
        month["accounts"] = json.loads(json.dumps(last["accounts"]))
    elif action == "delete":
        aid = str(body.get("account_id") or "")
        if aid not in month["accounts"]:
            return {"error": "not_budgeted"}, 404
        month["versions"].append({"at": now, "by": actor(request),
                                  "accounts": json.loads(json.dumps(month["accounts"]))})
        month["accounts"].pop(aid, None)
    else:
        aid = str(body.get("account_id") or "")
        acc = acct_idx.get(aid)
        if not acc:
            return {"error": "account_not_in_chart",
                    "message_ar": "الميزانية تكون على حساب من دليل دافترة فقط.",
                    "message_en": "Budgets attach to imported Daftra accounts only."}, 422
        try:
            amount = float(body.get("amount") or 0)
        except (TypeError, ValueError):
            return {"error": "bad_amount"}, 400
        weekly = body.get("weekly")
        if weekly:
            if not ST.weekly_sums_ok(weekly, amount):
                return {"error": "weekly_sum_mismatch",
                        "message_ar": "مجموع الأسابيع لازم يساوي مبلغ الشهر بالضبط.",
                        "message_en": "Weekly amounts must sum exactly to the month."}, 422
        else:
            weekly = ST.split_weekly(amount, weeks=int(body.get("weeks") or 4))
        month["versions"].append({"at": now, "by": actor(request),
                                  "accounts": json.loads(json.dumps(month["accounts"]))})
        month["accounts"][aid] = {"amount": amount, "weekly": weekly,
                                  "week_start": (body.get("week_start") or "sun"),
                                  "name": acc.get("display_name") or "",
                                  "updated_at": now, "updated_by": actor(request)}
    if len(month["versions"]) > 24:
        del month["versions"][:len(month["versions"]) - 24]
    B._save_json(_BUDGET_FILE, store)
    B._fb_audit_add(actor(request), "erp_budget_" + action, "budget", mkey)
    return budget_get({"m": mkey}), 200


def budget_alerts_now():
    """Current-month 90%/100% alerts — surfaced in اليوم."""
    try:
        mkey = datetime.now(B.TZ).date().isoformat()[:7]
        store = _budget_store()
        month = store.get(mkey)
        if not month or not month.get("accounts"):
            return []
        actuals = budget_actuals(mkey)
        acct_idx = _acct_by_id()
        out = []
        for aid, cfg in month["accounts"].items():
            st = ST.budget_row(cfg.get("amount") or 0, actuals.get(aid, 0))
            if st["alert"]:
                acc = acct_idx.get(aid) or {}
                out.append({"account_id": aid, "name": acc.get("display_name") or cfg.get("name") or aid,
                            "pct": st["pct"], "alert": st["alert"],
                            "budget": st["budget"], "actual": st["actual"]})
        out.sort(key=lambda r: -(r["pct"] or 0))
        return out
    except Exception:
        return []


# ====================== الملاك Owners (Slice 6) ======================

def _registry_rows():
    reg = getattr(B, "_owner_registry", None) or {}
    return list(reg.values()) if isinstance(reg, dict) else list(reg)


def owner_apartments(owner):
    """(apartment names, listing ids) belonging to one owner — the receipt-proxy scope."""
    apts, lids = set(), set()
    for r in _registry_rows():
        if (r.get("owner") or "").strip() == (owner or "").strip():
            if (r.get("apartment") or "").strip():
                apts.add(r.get("apartment").strip())
            if r.get("lid") not in (None, ""):
                lids.add(str(r.get("lid")))
    return apts, lids


def owners_payload():
    """Faisal's owners list: units + link state (آخر فتح، عدد الفتحات) per owner."""
    by_owner = {}
    for r in _registry_rows():
        o = (r.get("owner") or "").strip()
        if not o:
            continue
        d = by_owner.setdefault(o, {"owner": o, "apartments": [], "mgmt_pcts": set()})
        if (r.get("apartment") or "").strip():
            d["apartments"].append(r.get("apartment").strip())
        if r.get("mgmt_pct") is not None:
            d["mgmt_pcts"].add(float(r.get("mgmt_pct")))
    links = getattr(B, "_owner_links", None) or {}
    rows = []
    for o, d in sorted(by_owner.items()):
        lk = links.get(o) or {}
        rows.append({"owner": o, "units": len(d["apartments"]),
                     "apartments": sorted(d["apartments"])[:12],
                     "mgmt_pct": (sorted(d["mgmt_pcts"])[0] if len(d["mgmt_pcts"]) == 1
                                  else (list(sorted(d["mgmt_pcts"])) or None)),
                     "link": {"exists": bool(lk.get("token")),
                              "active": bool(lk.get("active")),
                              "url": ("/fin/o/" + lk["token"]) if lk.get("token") else "",
                              "created_at": lk.get("created_at") or "",
                              "opened_at": lk.get("opened_at") or "",
                              "opens": int(lk.get("opens") or 0)}})
    return {"ok": True, "rows": rows, "total": len(rows)}


# ====================== Contracts linking (Setup) ======================

def contracts_list():
    rows = []
    for k, p in B._fb_contracts.items():
        rows.append({"key": p.get("id") or k, "name": p.get("apartment_name") or k,
                     "owner": p.get("owner_name") or "",
                     "cc_id": str(p.get("daftra_cost_center_id") or ""),
                     "cc_name": p.get("daftra_cost_center_name") or "",
                     "status": p.get("validation_status") or "incomplete"})
    rows.sort(key=lambda r: (bool(r["cc_id"]), r["name"]))
    return {"ok": True, "rows": rows,
            "unlinked": sum(1 for r in rows if not r["cc_id"])}


def contract_link(request, body):
    key = str(body.get("key") or "")
    p = B._fb_contracts.get(key) or next(
        (v for v in B._fb_contracts.values() if (v.get("id") or "") == key), None)
    if not p:
        return {"error": "contract_not_found"}, 404
    ccid = str(body.get("cost_center_id") or "").strip()
    if not ccid:
        return {"error": "missing_cost_center"}, 400
    cc = next((r for r in _cc_records() if str(r.get("source_id")) == ccid), None)
    if not cc:
        return {"error": "cost_center_unknown",
                "message_ar": "مركز التكلفة مو موجود في المستورد من دافترة.",
                "message_en": "Cost center not in the imported Daftra data."}, 422
    before = {"daftra_cost_center_id": p.get("daftra_cost_center_id")}
    p["daftra_cost_center_id"] = ccid
    p["daftra_cost_center_name"] = cc.get("display_name") or ""
    try:
        p["validation_status"], p["validation_issues"] = B._fb_contract_validate(p)
    except Exception:
        pass
    B._fb_save("finance_contract_profiles.json", B._fb_contracts)
    B._fb_audit_add(actor(request), "erp_contract_link", "contract", key,
                    before=before, after={"daftra_cost_center_id": ccid})
    return {"ok": True, "row": {"key": key, "name": p.get("apartment_name") or key,
                                "owner": p.get("owner_name") or "", "cc_id": ccid,
                                "cc_name": p.get("daftra_cost_center_name"),
                                "status": p.get("validation_status") or ""},
            "counters": counters_snapshot()}, 200
