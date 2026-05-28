#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ouja Turnover Bot
=================
Watches Hostaway for upcoming checkouts and, 12 hours before each one, opens a
Discord channel named after the unit's internal name. The channel shows the
guest name + checkout time and has a "✅ Cleaning Done" button. When a cleaner
taps it, the channel is deleted instantly.

Runs 24/7 (e.g. on Railway). All secrets come from environment variables —
never hard-code them in this file.

Required environment variables:
  HOSTAWAY_ACCOUNT_ID   your Hostaway account id (e.g. 147296)
  HOSTAWAY_API_KEY      your Hostaway API key
  DISCORD_TOKEN         the bot token from the Discord Developer Portal
  DISCORD_GUILD_ID      your Discord server id (right-click server -> Copy Server ID)

Optional (sensible defaults):
  HOURS_AHEAD           how many hours before checkout to open the channel (default 12)
  POLL_MINUTES          how often to check Hostaway (default 15)
  TIMEZONE              default "Asia/Riyadh"
  CATEGORY_NAME         Discord category to group channels (default "🧹 Turnovers")
  DEFAULT_CHECKOUT_HOUR used if a reservation has no checkout time (default 12)
"""

import os
import re
import io
import json
import time
import random
import asyncio
from collections import defaultdict, deque, OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, time as dt_time
from zoneinfo import ZoneInfo

import requests
import discord
from discord.ext import commands, tasks

try:
    from aiohttp import web        # for the Hostaway webhook server (Stage 2)
    _HAS_AIOHTTP = True
except Exception:
    _HAS_AIOHTTP = False

# ---------------- config ----------------
HOSTAWAY_ACCOUNT_ID = os.environ.get("HOSTAWAY_ACCOUNT_ID", "")
HOSTAWAY_API_KEY    = os.environ.get("HOSTAWAY_API_KEY", "")
DISCORD_TOKEN       = os.environ.get("DISCORD_TOKEN", "")
GUILD_ID            = int(os.environ.get("DISCORD_GUILD_ID", "0") or "0")

HOURS_AHEAD          = int(os.environ.get("HOURS_AHEAD", "12"))
POLL_MINUTES         = int(os.environ.get("POLL_MINUTES", "15"))
TZ                   = ZoneInfo(os.environ.get("TIMEZONE", "Asia/Riyadh"))
CATEGORY_NAME        = os.environ.get("CATEGORY_NAME", "🧹 Turnovers")
# Category for the guest-assistant + escalations channels (your existing ops category).
ASSISTANT_CATEGORY   = os.environ.get("ASSISTANT_CATEGORY", "Operations")
DEFAULT_CHECKOUT_HOUR = int(os.environ.get("DEFAULT_CHECKOUT_HOUR", "12"))

# ---- last-minute tiered discount (all Riyadh time) ----
# Tier 1 fires at midnight, Tier 2 deepens it at noon, Tier 3 deepens again at 4 PM.
DISCOUNT_TIER1_PERCENT = float(os.environ.get("DISCOUNT_TIER1_PERCENT", "15"))
DISCOUNT_TIER1_HOUR    = int(os.environ.get("DISCOUNT_TIER1_HOUR", "0"))    # 00:00 = midnight
DISCOUNT_TIER2_PERCENT = float(os.environ.get("DISCOUNT_TIER2_PERCENT", "20"))
DISCOUNT_TIER2_HOUR    = int(os.environ.get("DISCOUNT_TIER2_HOUR", "12"))   # 12:00 = noon
DISCOUNT_TIER3_PERCENT = float(os.environ.get("DISCOUNT_TIER3_PERCENT", "30"))
DISCOUNT_TIER3_HOUR    = int(os.environ.get("DISCOUNT_TIER3_HOUR", "18"))   # 18:00 = 6 PM
DISCOUNT_DRY_RUN = os.environ.get("DISCOUNT_DRY_RUN", "1") not in ("0", "false", "False", "no")
DISCOUNT_FLOOR   = float(os.environ.get("DISCOUNT_FLOOR", "0") or "0")      # 0 = no floor
DISCOUNT_CHANNEL = os.environ.get("DISCOUNT_CHANNEL", "pricing-log")        # summary channel
DISCOUNT_STATE_FILE_NAME = "discount_state.json"                            # path resolved via _state_path() below (survives redeploys)
# Diagnostics: after each live write, re-read the day and log requested vs actual price.
DISCOUNT_VERIFY  = os.environ.get("DISCOUNT_VERIFY", "0") in ("1", "true", "True", "yes")
# Set to a percent (e.g. "15") to run one tier immediately on startup for testing.
DISCOUNT_TEST    = os.environ.get("DISCOUNT_TEST", "0")

# ---- 9 PM heads-up: preview tomorrow's still-empty units (3h before the midnight tier) ----
HEADS_UP_HOUR    = int(os.environ.get("HEADS_UP_HOUR", "21"))               # 21:00 = 9 PM Riyadh
HEADS_UP_CHANNEL = os.environ.get("HEADS_UP_CHANNEL", "discount-heads-up")  # where the preview is posted
HEADS_UP_TEST    = os.environ.get("HEADS_UP_TEST", "0") in ("1", "true", "True", "yes")  # post once on startup
# ---- Weekly Revenue Report (recommend-only; optimizes for max total revenue) ----
REVENUE_CHANNEL     = os.environ.get("REVENUE_CHANNEL", "revenue-report")
REVENUE_REPORT_DOW  = int(os.environ.get("REVENUE_REPORT_DOW", "6"))    # 0=Mon … 6=Sun
REVENUE_REPORT_HOUR = int(os.environ.get("REVENUE_REPORT_HOUR", "9"))   # Riyadh hour to post
REVENUE_MAX_PAGES   = int(os.environ.get("REVENUE_MAX_PAGES", "60"))    # 100 reservations/page
REVENUE_WINDOW_DAYS = int(os.environ.get("REVENUE_WINDOW_DAYS", "90"))  # trailing perf window
REVENUE_TEST        = os.environ.get("REVENUE_TEST", "0") in ("1", "true", "True", "yes")  # post once on startup
REVENUE_DEBUG       = os.environ.get("REVENUE_DEBUG", "0") in ("1", "true", "True", "yes")
# ---- Per-date pricing opportunities (#price-opportunities) ----
PRICE_OPP_CHANNEL   = os.environ.get("PRICE_OPP_CHANNEL", "price-opportunities")
PRICE_OPP_HORIZON   = int(os.environ.get("PRICE_OPP_HORIZON", "45"))    # days ahead to price
PRICE_OPP_DOW       = int(os.environ.get("PRICE_OPP_DOW", "6"))         # 0=Mon … 6=Sun
PRICE_OPP_HOUR      = int(os.environ.get("PRICE_OPP_HOUR", "10"))
PRICE_OPP_TEST      = os.environ.get("PRICE_OPP_TEST", "0") in ("1", "true", "True", "yes")
# When you click ✅ apply, the bot writes the new price to your Hostaway calendar.
# Set PRICE_APPLY_DRYRUN=1 to TEST safely (logs what it would do, changes nothing).
PRICE_APPLY_DRYRUN  = os.environ.get("PRICE_APPLY_DRYRUN", "0") in ("1", "true", "True", "yes")
# Dynamic pricing strategy: after you Apply, the bot keeps optimizing that unit's open nights
# toward the best numbers (holds high when far out; steps down as a night stays empty & nears).
PRICING_STRATEGY_ENABLED = os.environ.get("PRICING_STRATEGY_ENABLED", "1") in ("1", "true", "True", "yes")
PRICING_STRATEGY_MIN     = int(os.environ.get("PRICING_STRATEGY_MIN", "10"))   # re-optimize interval (min)
PRICE_OPP_MAX_CARDS = int(os.environ.get("PRICE_OPP_MAX_CARDS", "15"))  # action cards to post
# ---- Last-week performance review (#last-week-review) ----
WEEKLY_REVIEW_CHANNEL = os.environ.get("WEEKLY_REVIEW_CHANNEL", "last-week-review")
WEEKLY_REVIEW_DOW   = int(os.environ.get("WEEKLY_REVIEW_DOW", "6"))     # 0=Mon … 6=Sun
WEEKLY_REVIEW_HOUR  = int(os.environ.get("WEEKLY_REVIEW_HOUR", "8"))
WEEKLY_REVIEW_TEST  = os.environ.get("WEEKLY_REVIEW_TEST", "0") in ("1", "true", "True", "yes")

# ---- weekend rule (Thu/Fri): no midnight/noon tiers — a single softer discount at 5:30 PM ----
WEEKEND_DAYS = set(int(x) for x in os.environ.get("WEEKEND_DAYS", "3,4").split(",")
                   if x.strip().isdigit())                                  # 3=Thu, 4=Fri (Mon=0)
WEEKEND_DISCOUNT_PERCENT = float(os.environ.get("WEEKEND_DISCOUNT_PERCENT", "20"))
WEEKEND_DISCOUNT_HOUR    = int(os.environ.get("WEEKEND_DISCOUNT_HOUR", "17"))
WEEKEND_DISCOUNT_MINUTE  = int(os.environ.get("WEEKEND_DISCOUNT_MINUTE", "30"))  # 17:30 = 5:30 PM

# ---- cleaning reminders: nag the open turnover channels until ✅ Cleaning Done ----
REMINDER_START_HOUR = int(os.environ.get("REMINDER_START_HOUR", "12"))  # start at 12 PM
REMINDER_FAST_HOUR  = int(os.environ.get("REMINDER_FAST_HOUR", "15"))   # speed up at 3 PM
REMINDER_END_HOUR   = int(os.environ.get("REMINDER_END_HOUR", "23"))    # stop at 11 PM
REMINDER_SLOW_MIN   = int(os.environ.get("REMINDER_SLOW_MIN", "30"))    # 12 PM–3 PM: every 30 min
REMINDER_FAST_MIN   = int(os.environ.get("REMINDER_FAST_MIN", "15"))    # after 3 PM: every 15 min
OPERATION_ROLE_ID   = os.environ.get("OPERATION_ROLE_ID", "")           # role to ping if no cleaner
OPERATION_ROLE_NAME = os.environ.get("OPERATION_ROLE_NAME", "operation")

# ---- AI guest-message assistant (Claude drafts, a human approves, then it sends) ----
ANTHROPIC_API_KEY  = os.environ.get("ANTHROPIC_API_KEY", "")
CLAUDE_MODEL       = os.environ.get("CLAUDE_MODEL", "claude-haiku-4-5-20251001")
# Smarter model for the few high-touch, personality-heavy messages (manager script +
# empathetic repeat-escalation acks). Low volume, so the extra cost is tiny.
CLAUDE_MODEL_PREMIUM = os.environ.get("CLAUDE_MODEL_PREMIUM", "claude-sonnet-4-6")
ASSISTANT_ENABLED  = os.environ.get("ASSISTANT_ENABLED", "0") in ("1", "true", "True", "yes")
ASSISTANT_CHANNEL  = os.environ.get("ASSISTANT_CHANNEL", "guest-assistant")
ASSISTANT_POLL_MIN = float(os.environ.get("ASSISTANT_POLL_MIN", "0.5"))   # check inbox every N min (float ok)
ASSISTANT_SCAN     = int(os.environ.get("ASSISTANT_SCAN", "30"))      # how many recent convos to scan
# ---- Stage 2: Hostaway webhooks (instant replies). 100% optional/backward-compatible:
# if not set up in Hostaway, the bot just keeps polling as before. ----
WEBHOOKS_ENABLED = os.environ.get("WEBHOOKS_ENABLED", "1") in ("1", "true", "True", "yes")
WEBHOOK_SECRET   = os.environ.get("WEBHOOK_SECRET", "ouja-hook")      # the secret path segment
WEB_PORT         = int(os.environ.get("PORT", "8080"))               # Railway provides PORT
# ---- Web dashboard (served by the same bot at /dashboard). Protected by a token. ----
DASHBOARD_ENABLED = os.environ.get("DASHBOARD_ENABLED", "1") in ("1", "true", "True", "yes")
DASHBOARD_TOKEN   = os.environ.get("DASHBOARD_TOKEN", "")            # REQUIRED to view (set your own)
DASH_TTL          = int(os.environ.get("DASH_TTL", "600"))          # cache computed analytics (sec)
DASH_REFRESH_MIN  = int(os.environ.get("DASH_REFRESH_MIN", "7"))    # background pre-compute interval (min)
# Safety: replies are NEVER sent automatically unless you explicitly turn this on later.
ASSISTANT_AUTOSEND = os.environ.get("ASSISTANT_AUTOSEND", "0") in ("1", "true", "True", "yes")
ASSISTANT_DEBUG    = os.environ.get("ASSISTANT_DEBUG", "0") in ("1", "true", "True", "yes")
# Skip the startup "mark everything seen" baseline so the latest unanswered guest
# messages get drafted right away (handy for a first test without sending a new message).
ASSISTANT_TEST     = os.environ.get("ASSISTANT_TEST", "0") in ("1", "true", "True", "yes")
# On startup, messages newer than this many minutes are NOT baselined, so a fresh guest
# message still gets a card even if the bot just restarted/redeployed. Set 0 to baseline all.
ASSISTANT_BASELINE_GRACE_MIN = int(os.environ.get("ASSISTANT_BASELINE_GRACE_MIN", "15"))
# Also draft a reply when the only thing sent AFTER the guest's question is an automated
# welcome/booking message (so guest questions buried under auto-messages still get answered).
# A real human/bot reply after the guest still counts as "answered" and is skipped.
ASSISTANT_ANSWER_PAST_AUTO = os.environ.get("ASSISTANT_ANSWER_PAST_AUTO", "1") in ("1", "true", "True", "yes")
# Outbound messages containing any of these (case-insensitive, '|'-separated) are treated as
# automated, not a real answer. Add your Hostaway welcome/automation phrases here.
AUTO_REPLY_MARKERS = [m.strip() for m in os.environ.get(
    "AUTO_REPLY_MARKERS", "truly delighted|we are truly|delighted by your|we've prepared|we have prepared").split("|") if m.strip()]
# Always draft a card for the guest's latest message, even if the team/automation already
# replied after it. Set 0 to only draft when the guest is unanswered.
ASSISTANT_ALWAYS_DRAFT = os.environ.get("ASSISTANT_ALWAYS_DRAFT", "1") in ("1", "true", "True", "yes")
# Knowledge base: a Discord channel the assistant reads as facts about Ouja. Anyone can add
# facts by typing in it, and the 🧠 Teach button on a card saves corrections here.
KNOWLEDGE_CHANNEL     = os.environ.get("KNOWLEDGE_CHANNEL", "knowledge")
KNOWLEDGE_REFRESH_MIN = int(os.environ.get("KNOWLEDGE_REFRESH_MIN", "5"))
KNOWLEDGE_MAX         = int(os.environ.get("KNOWLEDGE_MAX", "300"))   # facts (messages) to load
# When on, logs each listing's active-status + whether an Airbnb link was found (for tuning).
CATALOG_DEBUG = os.environ.get("CATALOG_DEBUG", "0") in ("1", "true", "True", "yes")
# Pull a real "starting from" nightly price from each unit's calendar (reflects dynamic pricing).
CATALOG_CALENDAR_PRICES = os.environ.get("CATALOG_CALENDAR_PRICES", "1") in ("1", "true", "True", "yes")
# Auto-send: when ON, very simple/safe replies (action="auto") go straight to the guest;
# anything needing approval or a human still posts a card. Default OFF — everything waits
# for approval/edit until you've judged the assistant's quality. Set to 1 to enable later.
ASSISTANT_AUTO     = os.environ.get("ASSISTANT_AUTO", "0") in ("1", "true", "True", "yes")
ASSISTANT_AUTO_CONF = float(os.environ.get("ASSISTANT_AUTO_CONF", "0.85"))  # Stage 1: auto-send at/above this confidence
ESCALATE_BELOW     = float(os.environ.get("ESCALATE_BELOW", "0.55"))       # Stage 3: escalate below this confidence
# Signature appended to every guest-facing message.
ASSISTANT_SIGNATURE_AR = os.environ.get("ASSISTANT_SIGNATURE_AR", "الدعم الفني - مساعد 🤍")
ASSISTANT_SIGNATURE_EN = os.environ.get("ASSISTANT_SIGNATURE_EN", "Technical Support - Musaid 🤍")
# When a chat escalates to a human, auto-send the guest a holding message.
ASSISTANT_ESC_ACK  = os.environ.get("ASSISTANT_ESC_ACK", "1") in ("1", "true", "True", "yes")
ASSISTANT_ACK_AR   = os.environ.get("ASSISTANT_ACK_AR",
    "حياك الله 🤍 رفعنا موضوعك للقسم المختص، وبيتواصل معك الفريق في أقرب وقت.")
ASSISTANT_ACK_EN   = os.environ.get("ASSISTANT_ACK_EN",
    "Thank you 🤍 We've escalated this to our specialized team and someone will contact you shortly.")
# Escalations go to their own channel, ping the operation team, and re-ping until claimed.
ESCALATION_CHANNEL    = os.environ.get("ESCALATION_CHANNEL", "escalations")
AUTO_REPLY_CHANNEL    = os.environ.get("AUTO_REPLY_CHANNEL", "auto-replies")  # audit log of Stage-1 auto-sends
ESCALATION_REPING_MIN = int(os.environ.get("ESCALATION_REPING_MIN", "10"))   # re-ping every N min
ESCALATION_MAX_PINGS  = int(os.environ.get("ESCALATION_MAX_PINGS", "12"))    # stop after this many re-pings
CLAIM_NAMES = [n.strip() for n in os.environ.get(
    "CLAIM_NAMES", "اسيل,فيصل,ماثر,نوره,ناصر,محمد").split(",") if n.strip()]
# When someone claims an escalation, DM them a ready-to-send reply in the owner's warm style.
MANAGER_SCRIPT = os.environ.get("MANAGER_SCRIPT", "1") in ("1", "true", "True", "yes")

BASE = "https://api.hostaway.com/v1"
GOLD = 0xC8A24B

# ---------------- bounded helpers ----------------
class _BoundedSet:
    """Insertion-ordered set with a hard cap. Oldest entries auto-evicted.
    Used for `_assistant_seen` so we don't (a) grow forever in memory or
    (b) lose recent message IDs when an unordered set is sliced on save."""
    __slots__ = ("_d", "maxlen")

    def __init__(self, items=(), maxlen=20000):
        self._d = OrderedDict()
        self.maxlen = maxlen
        for x in items:
            self.add(x)

    def add(self, item):
        if item in self._d:
            self._d.move_to_end(item)
        else:
            self._d[item] = None
            if len(self._d) > self.maxlen:
                self._d.popitem(last=False)

    def __contains__(self, item):
        return item in self._d

    def __len__(self):
        return len(self._d)

    def __iter__(self):
        return iter(self._d)

    def __bool__(self):
        return bool(self._d)

def _bounded_cache_put(cache, key, value, maxlen):
    """Trim an OrderedDict cache to `maxlen` after inserting. Caller passes an OrderedDict."""
    if key in cache:
        cache.move_to_end(key)
    cache[key] = value
    while len(cache) > maxlen:
        cache.popitem(last=False)

# ---------------- Persistent state (Railway Volume) ----------------
# Mount a Railway Volume at this path so state survives restarts AND redeploys.
STATE_DIR = os.environ.get("STATE_DIR", "/data")

def _state_path(name):
    return os.path.join(STATE_DIR, name)

def _save_json(name, obj):
    try:
        os.makedirs(STATE_DIR, exist_ok=True)
        tmp = _state_path(name + ".tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False)
        os.replace(tmp, _state_path(name))   # atomic write
    except Exception as e:
        print(f"state save error ({name}):", e)

def _load_json(name, default):
    try:
        with open(_state_path(name), encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default

def _state_persistent():
    """True if STATE_DIR looks like a mounted volume (writable & exists)."""
    try:
        os.makedirs(STATE_DIR, exist_ok=True)
        return os.path.isdir(STATE_DIR)
    except Exception:
        return False

HANDLED_FILE = _state_path("handled.json")

# ---------------- Hostaway ----------------
_token = {"value": None}

def get_token(force=False):
    if _token["value"] and not force:
        return _token["value"]
    r = requests.post(
        f"{BASE}/accessTokens",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        data={"grant_type": "client_credentials", "client_id": HOSTAWAY_ACCOUNT_ID,
              "client_secret": HOSTAWAY_API_KEY, "scope": "general"},
        timeout=30,
    )
    r.raise_for_status()
    _token["value"] = r.json()["access_token"]
    return _token["value"]

# Status codes that mean "your bearer token is no longer accepted — refresh once and retry."
# Hostaway has been observed to use 403; spec says 401 — accept both to be safe.
_AUTH_RETRY_CODES = (401, 403)

def api_get(path, params=None, _retry=0):
    token = get_token()
    r = requests.get(
        f"{BASE}{path}",
        headers={"Authorization": f"Bearer {token}", "Cache-control": "no-cache"},
        params=params or {}, timeout=60,
    )
    if r.status_code in _AUTH_RETRY_CODES and _retry == 0:   # token expired -> refresh once
        get_token(force=True)
        return api_get(path, params, _retry + 1)
    if r.status_code == 429:                          # rate limited -> brief backoff
        time.sleep(10)
        if _retry < 3:
            return api_get(path, params, _retry + 1)
    r.raise_for_status()
    return r.json()

def api_post(path, body, _retry=0):
    token = get_token()
    r = requests.post(
        f"{BASE}{path}",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json",
                 "Cache-control": "no-cache"},
        json=body, timeout=60,
    )
    if r.status_code in _AUTH_RETRY_CODES and _retry == 0:
        get_token(force=True)
        return api_post(path, body, _retry + 1)
    if r.status_code == 429:
        time.sleep(10)
        if _retry < 3:
            return api_post(path, body, _retry + 1)
    r.raise_for_status()
    return r.json()


def api_put(path, body, _retry=0):
    token = get_token()
    r = requests.put(
        f"{BASE}{path}",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json",
                 "Cache-control": "no-cache"},
        json=body, timeout=60,
    )
    if r.status_code in _AUTH_RETRY_CODES and _retry == 0:
        get_token(force=True)
        return api_put(path, body, _retry + 1)
    if r.status_code == 429:
        time.sleep(10)
        if _retry < 3:
            return api_put(path, body, _retry + 1)
    r.raise_for_status()
    return r.json()

_listings = {"map": {}, "ts": 0}

def get_listings_map():
    """listingMapId -> internal name (cached for 1 hour)."""
    if _listings["map"] and time.time() - _listings["ts"] < 3600:
        return _listings["map"]
    m, limit, offset = {}, 100, 0
    while True:
        data = api_get("/listings", params={"limit": limit, "offset": offset})
        batch = data.get("result", []) or []
        for l in batch:
            m[l.get("id")] = l.get("internalListingName") or l.get("name") or f"unit-{l.get('id')}"
        if len(batch) < limit:
            break
        offset += limit
    _listings["map"], _listings["ts"] = m, time.time()
    return m

def parse_hour(v, default):
    if v is None:
        return default
    s = str(v)
    if ":" in s:
        try: return int(s.split(":")[0])
        except: return default
    try: return int(float(s))
    except: return default

SKIP_STATUSES = {"cancelled", "declined", "expired", "inquiry", "inquirydenied", "inquirytimedout"}

def fetch_upcoming_checkouts():
    """Reservations whose checkout falls within the next HOURS_AHEAD hours."""
    listings = get_listings_map()
    now = datetime.now(TZ)
    window_end = now + timedelta(hours=HOURS_AHEAD)
    today = now.date().isoformat()
    tomorrow = (now.date() + timedelta(days=1)).isoformat()

    out, limit, offset, pages = [], 100, 0, 0
    while pages < 10:
        data = api_get("/reservations", params={
            "departureStartDate": today, "departureEndDate": tomorrow,
            "limit": limit, "offset": offset, "includeResources": 0,
        })
        batch = data.get("result", []) or []
        if not batch:
            break
        for res in batch:
            if (res.get("status") or "").lower() in SKIP_STATUSES:
                continue
            dep = res.get("departureDate")
            if not dep:
                continue
            hour = parse_hour(res.get("checkOutTime"), DEFAULT_CHECKOUT_HOUR)
            try:
                checkout = datetime.strptime(dep[:10], "%Y-%m-%d").replace(
                    hour=min(hour, 23), tzinfo=TZ)
            except ValueError:
                continue
            if now <= checkout <= window_end:
                lm = res.get("listingMapId")
                out.append({
                    "res_id": str(res.get("id")),
                    "listing": listings.get(lm) or res.get("listingName") or f"unit-{lm}",
                    "guest": res.get("guestName") or res.get("guestFirstName") or "Guest",
                    "checkout": checkout,
                })
        if len(batch) < limit:
            break
        offset += limit
        pages += 1
    return out

def _load_discount_state():
    # _load_json returns the default on any error and reads from STATE_DIR (the volume),
    # so tonight's "original price" snapshot survives redeploys.
    return _load_json(DISCOUNT_STATE_FILE_NAME, {})

def _save_discount_state(st):
    _save_json(DISCOUNT_STATE_FILE_NAME, st)

def apply_discount_tier(pct):
    """Set tonight's price to `pct`% off the ORIGINAL price, for every still-empty unit.
    The original is captured once per day (state file + Hostaway note) so the two tiers
    never compound. Returns (changes, today). Honors DRY-RUN."""
    factor = (100.0 - pct) / 100.0
    today = datetime.now(TZ).date().isoformat()
    state = _load_discount_state()
    state = {today: state.get(today, {})}      # keep only today's record
    day_orig = state[today]
    listings = get_listings_map()
    changes = []
    for lid, name in listings.items():
        lid_s = str(lid)
        if is_unit_skipped(lid):
            continue                                       # owner asked to hold price on this unit
        try:
            cal = api_get(f"/listings/{lid}/calendar",
                          params={"startDate": today, "endDate": today})
            days = cal.get("result", []) or []
            if not days:
                continue
            d = days[0]
            available = int(d.get("isAvailable", 0) or 0) == 1
            booked = bool(d.get("reservationId"))
            current = d.get("price")
            if not available or booked or not current:
                continue                                   # booked / blocked / no price
            current = float(current)
            # recover tonight's original: remembered state -> Hostaway note -> current
            original = day_orig.get(lid_s)
            if original is None:
                m = re.search(r"ouja-orig:(\d+(?:\.\d+)?)", str(d.get("note") or ""))
                original = float(m.group(1)) if m else current
                day_orig[lid_s] = original
            new_price = round(original * factor)
            if DISCOUNT_FLOOR and new_price < DISCOUNT_FLOOR:
                new_price = int(DISCOUNT_FLOOR)
            if new_price >= current:
                continue                                   # already at/below this level
            if not DISCOUNT_DRY_RUN:
                resp = api_put(f"/listings/{lid}/calendar",
                        {"startDate": today, "endDate": today,
                         "isAvailable": 1, "price": new_price,
                         "note": f"ouja-orig:{int(original)}"})
                if DISCOUNT_VERIFY:
                    try:
                        chk = api_get(f"/listings/{lid}/calendar",
                                      params={"startDate": today, "endDate": today})
                        cd = (chk.get("result") or [{}])[0]
                        actual = cd.get("price")
                        status = resp.get("status") if isinstance(resp, dict) else "?"
                        stuck = "✅ stuck" if str(actual) == str(new_price) else "❌ reverted/ignored"
                        print(f"   VERIFY {name}: requested {new_price}, Hostaway now shows "
                              f"{actual} ({stuck}) · PUT status={status}")
                    except Exception as e:
                        print(f"   VERIFY {name}: read-back failed: {e}")
            changes.append({"name": name, "orig": int(original),
                            "old": int(current), "new": int(new_price)})
        except Exception as e:
            print(f"discount error for {name}: {e}")
    _save_discount_state(state)
    return changes, today

def is_weekend_today():
    return datetime.now(TZ).weekday() in WEEKEND_DAYS

# ====================================================================
# Saudi events / high-demand windows. These boost the bot's reasoning
# about pricing and surface as alerts when the portfolio's forward pace
# for those dates is unusually low. Dates are approximate and easy to
# edit in one place.
# ====================================================================
SAUDI_EVENTS = [
    # name, start_iso, end_iso, demand boost (multiplier hint), kind
    {"name": "يوم التأسيس",  "start": "2026-02-22", "end": "2026-02-22", "boost": 1.40, "kind": "national"},
    {"name": "يوم التأسيس",  "start": "2027-02-22", "end": "2027-02-22", "boost": 1.40, "kind": "national"},
    {"name": "عيد الفطر",     "start": "2026-03-20", "end": "2026-03-25", "boost": 1.60, "kind": "eid"},
    {"name": "عيد الأضحى",    "start": "2026-05-26", "end": "2026-06-01", "boost": 1.55, "kind": "eid"},
    {"name": "عيد الفطر",     "start": "2027-03-09", "end": "2027-03-14", "boost": 1.60, "kind": "eid"},
    {"name": "عيد الأضحى",    "start": "2027-05-16", "end": "2027-05-22", "boost": 1.55, "kind": "eid"},
    {"name": "اليوم الوطني",  "start": "2026-09-22", "end": "2026-09-24", "boost": 1.50, "kind": "national"},
    {"name": "اليوم الوطني",  "start": "2027-09-22", "end": "2027-09-24", "boost": 1.50, "kind": "national"},
    {"name": "موسم الرياض",   "start": "2026-10-23", "end": "2027-03-15", "boost": 1.30, "kind": "season"},
    {"name": "موسم الرياض",   "start": "2027-10-23", "end": "2028-03-15", "boost": 1.30, "kind": "season"},
    # Ramadan = low demand period — encoded as a NEGATIVE boost so the
    # pricing model can pull back rather than push.
    {"name": "رمضان",         "start": "2026-02-17", "end": "2026-03-19", "boost": 0.75, "kind": "ramadan"},
    {"name": "رمضان",         "start": "2027-02-06", "end": "2027-03-08", "boost": 0.75, "kind": "ramadan"},
]

# Owner-added custom events (e.g., Ouja-specific high-demand windows like a
# local conference or a wedding season). Merged with SAUDI_EVENTS at lookup
# time and persisted across redeploys. Each entry has an `id` so we can edit
# or delete it from the dashboard.
_custom_events = []   # list of {id, name, start, end, boost, kind}

def _all_events():
    """Combined view of bundled SAUDI_EVENTS + owner-added _custom_events,
    each tagged with `source` so the UI can show what's editable."""
    out = []
    for e in SAUDI_EVENTS:
        out.append({**e, "source": "default", "id": None})
    for e in _custom_events:
        out.append({**e, "source": "custom"})
    return out

def events_for_date(d):
    """Active events (default + custom) on a given date."""
    if not d:
        return []
    iso = d.isoformat() if hasattr(d, "isoformat") else str(d)[:10]
    return [e for e in _all_events() if e["start"] <= iso <= e["end"]]

def event_boost_for_date(d):
    """Effective multiplier on this date — product of all active events.
    Returns 1.0 if no events apply."""
    boost = 1.0
    for e in events_for_date(d):
        boost *= e["boost"]
    return boost

# ====================================================================
# Forward calendar — for each day in [today, today+N), what's the
# portfolio occupancy, how many units are still empty, what's the avg
# price of those empty units, and which Saudi events are active.
# Powers the dashboard's calendar view + pricing alerts.
# ====================================================================
def compute_forward_calendar(days=60):
    today = datetime.now(TZ).date()
    end_date = today + timedelta(days=days - 1)
    today_iso = today.isoformat()
    end_iso = end_date.isoformat()
    listing_ids = list((get_listings_map() or {}).keys())
    per_listing = {}

    def _fetch(lid):
        try:
            cal = api_get(f"/listings/{lid}/calendar",
                          params={"startDate": today_iso, "endDate": end_iso})
            return lid, (cal.get("result") or [])
        except Exception as e:
            print(f"forward cal error ({lid}):", e)
            return lid, []

    if listing_ids:
        try:
            with ThreadPoolExecutor(max_workers=INTEL_PARALLEL) as ex:
                futs = [ex.submit(_fetch, lid) for lid in listing_ids]
                for f in as_completed(futs):
                    lid, days_data = f.result()
                    per_listing[lid] = {d.get("date"): d for d in days_data if d.get("date")}
        except Exception as e:
            print("forward cal pool error:", e)
    total_units = len(listing_ids) or 1
    out = []
    for i in range(days):
        d = today + timedelta(days=i)
        d_iso = d.isoformat()
        occupied = 0
        avail_prices = []
        for lid in per_listing:
            row = per_listing[lid].get(d_iso)
            if not row:
                continue
            available = int(row.get("isAvailable", 0) or 0) == 1
            booked = bool(row.get("reservationId"))
            if booked or not available:
                occupied += 1
            else:
                price = row.get("price")
                if isinstance(price, (int, float)) and price > 0:
                    avail_prices.append(float(price))
        events = events_for_date(d)
        avg_avail = round(sum(avail_prices) / len(avail_prices)) if avail_prices else None
        out.append({
            "date": d_iso, "weekday": d.weekday(),
            "is_weekend": d.weekday() in WEEKEND_DAYS,
            "occupied": occupied, "available": total_units - occupied,
            "total": total_units,
            "pace_pct": round((occupied / total_units) * 100) if total_units else 0,
            "avg_price": avg_avail,
            "events": [{"name": e["name"], "kind": e["kind"], "boost": e["boost"]} for e in events],
            "event_boost": round(event_boost_for_date(d), 2),
        })
    return out

# Cache it for 20 minutes — calendar fetches across the whole portfolio aren't free.
_forward_cache = {"data": None, "ts": 0}
def get_forward_calendar(days=60, ttl=1200):
    if _forward_cache["data"] is not None and (time.time() - _forward_cache["ts"]) < ttl \
       and len(_forward_cache["data"]) >= days:
        return _forward_cache["data"][:days]
    data = compute_forward_calendar(days=max(days, 60))
    _forward_cache["data"] = data
    _forward_cache["ts"] = time.time()
    return data[:days]

def compute_pricing_alerts():
    """Edge-day attention list. Sees the next 45 days, flags:
       (1) HIGH severity: an active SAUDI event date that's still <50% booked
       (2) MED severity: a non-event weekend that's <30% booked + 14+ days away
       (3) MED severity: a national-event date where the bot's prices aren't elevated
    The Today-page urgent strip surfaces these inline."""
    cal = get_forward_calendar(days=45)
    today = datetime.now(TZ).date()
    alerts = []
    for d in cal:
        d_date = _parse_date(d["date"])
        lead = (d_date - today).days if d_date else 0
        if d["events"] and d["pace_pct"] < 50 and lead >= 2:
            names = "، ".join(e["name"] for e in d["events"])
            alerts.append({
                "kind": "low_pace_event", "severity": "high",
                "date": d["date"], "lead_days": lead,
                "title": f"{names} · {d['date']}",
                "detail": f"pace {d['pace_pct']}% فقط · {d['available']} وحدة فاضية · "
                          f"{'متوسط السعر ' + str(d['avg_price']) + ' ر.س' if d['avg_price'] else 'لا يوجد سعر متوسط'}",
                "action_view": "pricing",
            })
        elif d["is_weekend"] and d["pace_pct"] < 30 and 5 <= lead <= 21 and not d["events"]:
            wd_ar = ["الإثنين","الثلاثاء","الأربعاء","الخميس","الجمعة","السبت","الأحد"][d["weekday"]]
            alerts.append({
                "kind": "low_pace_weekend", "severity": "med",
                "date": d["date"], "lead_days": lead,
                "title": f"نهاية أسبوع {wd_ar} · {d['date']}",
                "detail": f"pace {d['pace_pct']}% · {d['available']} وحدة فاضية بعد {lead} يوم",
                "action_view": "pricing",
            })
    return alerts

# ====================================================================
# Rental-agreement reminder
# --------------------------------------------------------------------
# Hostaway only releases the door code AFTER the rental agreement is
# signed. So if check-in is approaching and we still see no signature on
# the reservation, we re-send the original signing link with a short note
# explaining why their code hasn't arrived yet.
# ====================================================================
def _is_agreement_signed(r):
    """Defensive: Hostaway exposes the signed flag under several names across
    plans/versions. Treat ANY of these being truthy as 'signed'."""
    for k in ("isAccepted", "isSignedRentalAgreement", "rentalAgreementAccepted",
              "isAgreementSigned", "isContractSigned", "isRentalAgreementSigned",
              "agreementAccepted", "agreementSigned"):
        v = r.get(k)
        if v in (1, True, "1", "true", "True", "yes"):
            return True
    if r.get("acceptedDate") or r.get("agreementSignedDate") or r.get("signedAt"):
        return True
    return False

# Phrases that suggest the guest is asking for early check-in. Match in either lang.
_EARLY_CHECKIN_HINTS = [
    # Arabic
    "تشيك ان مبكر", "تشيك-ان مبكر", "تشيك إن مبكر", "تسجيل دخول مبكر", "دخول مبكر",
    "ادخل بدري", "ادخل من الصبح", "ندخل بدري", "ندخل الصبح", "ادخل ابكر", "أدخل أبكر",
    "ادخل قبل الوقت", "ندخل قبل الوقت", "اقدر ادخل قبل", "تقدر تخليني ادخل",
    "ابي ادخل اليوم", "ابي اوصل بدري", "اوصل بدري", "اوصل الصبح", "ابي ادخل الصبح",
    "early checkin", "early check-in", "early check in",
    "check in early", "check-in early", "checking in early",
    "arrive early", "arrival early", "before check-in", "before check in",
]

def _is_early_checkin_request(text):
    t = (text or "").lower()
    return any(h in t for h in _EARLY_CHECKIN_HINTS)

def _calendar_night_free(listing_id, night_date):
    """True iff `night_date` (single date) is not occupied/blocked for `listing_id`."""
    try:
        d_iso = night_date.isoformat()
        cal = api_get(f"/listings/{listing_id}/calendar",
                      params={"startDate": d_iso, "endDate": d_iso})
        days = cal.get("result") or []
        if not days:
            return False
        day = days[0]
        if day.get("reservationId"):
            return False
        return int(day.get("isAvailable", 0) or 0) == 1
    except Exception as e:
        print(f"_calendar_night_free error ({listing_id}, {night_date}):", e)
        return False

def early_checkin_context(reservation_id, listing_id):
    """Heavy helper used only when a guest seems to be requesting early check-in.
    Returns a dict the prompt can use to give Claude a concrete answer:
      - prev_occupied: is the night BEFORE arrival booked for the guest's unit?
      - alternatives: up to 5 active units where the same night IS free (potential
        swaps the team could approve).
    Returns None if we can't determine arrival or have no listing context."""
    if not reservation_id or not listing_id:
        return None
    try:
        rdata = api_get(f"/reservations/{reservation_id}")
        r = rdata.get("result") or {}
        arrival = _parse_date(r.get("arrivalDate"))
        if not arrival:
            return None
    except Exception as e:
        print(f"early_checkin_context arrival fetch error: {e}")
        return None
    prev_night = arrival - timedelta(days=1)
    prev_occupied = not _calendar_night_free(listing_id, prev_night)
    alternatives = []
    if prev_occupied and _catalog_units:
        # Pre-rank candidates so we don't burn 70 calendar reads if we don't need to:
        # prefer the same bedroom count + area as the guest's current unit when known.
        current = next((u for u in _catalog_units if u.get("id") == listing_id), {})
        want = {"beds": current.get("beds"), "area": current.get("area"),
                "tags": list(current.get("tags") or [])[:3]}
        candidates = sorted(
            [u for u in _catalog_units if u.get("id") and u["id"] != listing_id],
            key=lambda u: -_unit_match_score(u, want),
        )[:18]   # check at most ~18 so the loop stays fast
        for u in candidates:
            if _calendar_night_free(u["id"], prev_night) and _calendar_night_free(u["id"], arrival):
                alternatives.append({
                    "id": u["id"], "name": u["name"],
                    "beds": u.get("beds"), "area": u.get("area") or u.get("neighbourhood"),
                    "link": u.get("link"), "price": u.get("price"),
                })
                if len(alternatives) >= 5:
                    break
    return {
        "prev_occupied": prev_occupied,
        "prev_night": prev_night.isoformat(),
        "arrival": arrival.isoformat(),
        "alternatives": alternatives,
    }

_AGREEMENT_URL_HINTS = ("signing.hostaway", "hostaway.com/signing", "/signing/",
                        "/rental-agreement", "rental_agreement", "agreement-sign",
                        "esign", "docusign", "signing-link", "signnow",
                        "hostaway.com/contracts", "hostawayintegrations.com")

def find_agreement_url(conversation_id):
    """Scan past HOST (outbound) messages in the conversation for a signing-link URL.
    Picks the most-recent matching URL so a re-sent link wins."""
    if not conversation_id:
        return None
    try:
        data = api_get(f"/conversations/{conversation_id}/messages")
        msgs = sorted((data.get("result") or []), key=_msg_sort_key)
    except Exception as e:
        print(f"find_agreement_url fetch error ({conversation_id}):", e)
        return None
    candidate = None
    for m in msgs:
        if _msg_is_inbound(m):
            continue                                       # we want OUR earlier messages
        body = (m.get("body") or "")
        if not body:
            continue
        low = body.lower()
        for url in re.findall(r"https?://[^\s)>\]\"']+", body):
            ul = url.lower()
            if any(h in ul for h in _AGREEMENT_URL_HINTS):
                candidate = url       # keep updating -> ends on most recent
    return candidate

def _reminder_message_ar(link):
    return (
        f"مرحباً 🤍\n"
        f"لاحظنا أن العقد الإلكتروني لحجزك لم يتم التوقيع عليه بعد. هذا السبب اللي بسببه "
        f"رمز الدخول للوحدة ما وصلك للحين — يُرسَل تلقائياً فور توقيع العقد.\n\n"
        f"رابط التوقيع:\n{link}\n\n"
        f"بعد التوقيع راح يوصلك رمز الدخول مباشرة. إذا واجهت أي مشكلة في فتح الرابط أو "
        f"التوقيع، رد علينا هنا ونساعدك على طول."
    )

def _reminder_message_en(link):
    return (
        f"Hi 🤍\n"
        f"Just a quick note — the rental agreement for your booking hasn't been signed yet, "
        f"which is why your access code hasn't arrived. The code is released automatically "
        f"as soon as the agreement is signed.\n\n"
        f"Sign here:\n{link}\n\n"
        f"Once signed, your access code will reach you straight away. If the link doesn't "
        f"open or anything's unclear, reply here and we'll help right away."
    )

def _check_agreement_for_one(r, now):
    """Single-reservation evaluation. Returns True if we sent a reminder."""
    res_id = r.get("id")
    if not res_id or res_id in _agreement_reminded:
        return False
    if (r.get("status") or "").lower() not in CONFIRMED_STATUSES:
        return False
    arrival = _parse_date(r.get("arrivalDate"))
    if not arrival or arrival != now.date():
        return False                                       # only act on TODAY's arrivals
    if _is_agreement_signed(r):
        return False                                       # already signed, nothing to do
    hour = parse_hour(r.get("checkInTime"), 15)
    checkin_dt = datetime(arrival.year, arrival.month, arrival.day,
                          min(hour, 23), 0, tzinfo=TZ)
    hours_until = (checkin_dt - now).total_seconds() / 3600.0
    if not (0 < hours_until <= AGREEMENT_REMINDER_LEAD_HOURS):
        return False
    cid = r.get("conversationId")
    link = find_agreement_url(cid)
    if not link:
        print(f"  agreement-reminder: res {res_id} not signed, but no signing URL found in conversation")
        return False
    is_ar = _has_arabic((r.get("guestName") or "")) or True   # default Arabic for KSA guests
    body = _reminder_message_ar(link) if is_ar else _reminder_message_en(link)
    try:
        send_guest_message(cid, body, "email")
        _agreement_reminded.add(res_id)
        log_event("guest", f"تذكير توقيع العقد · {r.get('guestName','ضيف')} · check-in بعد {hours_until:.1f}س")
        print(f"  agreement-reminder: nudged res {res_id} (check-in in {hours_until:.1f}h)")
        return True
    except Exception as e:
        print(f"agreement-reminder send error (res {res_id}):", e)
        return False

def check_agreement_reminders():
    """Top-level: pull today's reservations and run _check_agreement_for_one on each."""
    if not AGREEMENT_REMINDER_ENABLED:
        return
    today_iso = datetime.now(TZ).date().isoformat()
    try:
        data = api_get("/reservations", params={
            "arrivalStartDate": today_iso, "arrivalEndDate": today_iso,
            "limit": 200, "includeResources": 0,
        })
    except Exception as e:
        print("agreement-reminder fetch error:", e)
        return
    rows = data.get("result", []) or []
    now = datetime.now(TZ)
    n = sum(1 for r in rows if _check_agreement_for_one(r, now))
    if n:
        print(f"agreement-reminder: sent {n} nudge(s)")

# ====================================================================
# Deep-clean scheduler
# ====================================================================
def _dc_init(lid):
    if lid in _deep_clean_state:
        return
    _deep_clean_state[lid] = {
        "last_done": DEEPCLEAN_DEFAULT_LAST,
        "next_scheduled": None,
        "next_status": "unscheduled",   # unscheduled | scheduled | blocked | pushed | done
        "history": [],
        "notes": "",
    }

def _dc_calendar_day_free(lid, d_iso):
    try:
        cal = api_get(f"/listings/{lid}/calendar",
                      params={"startDate": d_iso, "endDate": d_iso})
        days = cal.get("result") or []
        if not days:
            return False
        d = days[0]
        return int(d.get("isAvailable", 0) or 0) == 1 and not d.get("reservationId")
    except Exception as e:
        print(f"deepclean cal check error ({lid}, {d_iso}):", e)
        return False

def _dc_scheduled_dates():
    return {s.get("next_scheduled") for s in _deep_clean_state.values() if s.get("next_scheduled")}

def _dc_find_next_date(lid, after_date=None):
    """Pick the earliest valid date for this unit's next deep clean.
    Constraints: weekday not in DEEPCLEAN_AVOID_WD, not already scheduled for another unit,
    within the 45-60 day window (extended to 75 if needed). Does NOT do per-date
    Hostaway lookups (too expensive) — caller verifies at confirm time."""
    if lid not in _deep_clean_state:
        _dc_init(lid)
    s = _deep_clean_state[lid]
    last = _parse_date(s.get("last_done")) or (datetime.now(TZ).date() - timedelta(days=30))
    today = datetime.now(TZ).date()
    earliest = max(last + timedelta(days=DEEPCLEAN_MIN_DAYS),
                   (after_date or today) + timedelta(days=1))
    horizon = max(last + timedelta(days=DEEPCLEAN_MAX_DAYS + 30),
                  earliest + timedelta(days=30))
    taken = _dc_scheduled_dates()
    d = earliest
    while d <= horizon:
        if d.weekday() not in DEEPCLEAN_AVOID_WD and d.isoformat() not in taken:
            return d.isoformat()
        d += timedelta(days=1)
    return None

def schedule_deep_cleans():
    """Assign next_scheduled for every unit that doesn't have one."""
    if not DEEPCLEAN_ENABLED:
        return
    listings = get_listings_map() or {}
    for lid in listings:
        if lid not in _deep_clean_state:
            _dc_init(lid)
        s = _deep_clean_state[lid]
        if s.get("next_scheduled"):
            continue
        nd = _dc_find_next_date(lid)
        if nd:
            s["next_scheduled"] = nd
            s["next_status"] = "scheduled"
            print(f"deepclean: scheduled {listings[lid]} ({lid}) for {nd}")
    return True

def confirm_tomorrow_deepcleans():
    """At 9pm: lock in tomorrow's planned deep cleans.
       - if the unit is still free → block the calendar (isAvailable=0) so the cleaner has the day
       - if a guest booked it last-minute → push to the next available slot"""
    if not DEEPCLEAN_ENABLED:
        return
    tomorrow = datetime.now(TZ).date() + timedelta(days=1)
    iso = tomorrow.isoformat()
    listings = get_listings_map() or {}
    confirmed, pushed = 0, 0
    for lid, s in list(_deep_clean_state.items()):
        if s.get("next_scheduled") != iso:
            continue
        name = listings.get(lid, str(lid))
        if not _dc_calendar_day_free(lid, iso):
            # guest beat us to it; reschedule
            s["next_scheduled"] = None
            s["next_status"] = "pushed"
            nd = _dc_find_next_date(lid, after_date=tomorrow)
            if nd:
                s["next_scheduled"] = nd
                s["next_status"] = "scheduled"
            pushed += 1
            log_event("pricing", f"تنظيف عميق · {name}: انحجزت ليوم {iso}، نُقل إلى {s.get('next_scheduled','—')}")
            continue
        # block the day in Hostaway
        try:
            api_put(f"/listings/{lid}/calendar",
                    {"startDate": iso, "endDate": iso,
                     "isAvailable": 0, "note": "deep-clean"})
            s["next_status"] = "blocked"
            confirmed += 1
            log_event("pricing", f"تنظيف عميق · {name}: مؤكد ومحجوز ليوم {iso}")
        except Exception as e:
            print(f"deepclean block error ({lid}, {iso}):", e)
    if confirmed or pushed:
        print(f"deepclean: confirmed={confirmed} pushed={pushed} for {iso}")

def _make_feedback_token(lid, res_id):
    """Compact opaque token combining unit + reservation. Random suffix so the
    URL can't be guessed."""
    import secrets
    return f"{int(lid)}-{int(res_id)}-{secrets.token_urlsafe(6)}"

def send_cleaning_feedback_request(reservation_row):
    """For one arrival, queue a rating-request message to the guest. Called by
    the cleaning_feedback_loop. Safe to call repeatedly — _cleaning_feedback_sent
    short-circuits duplicates."""
    if not CLEAN_FEEDBACK_ENABLED:
        return False
    r = reservation_row
    res_id = r.get("id")
    if not res_id or res_id in _cleaning_feedback_sent:
        return False
    if (r.get("status") or "").lower() not in CONFIRMED_STATUSES:
        return False
    cid = r.get("conversationId")
    if not cid:
        return False
    lid = r.get("listingMapId")
    if not lid:
        return False
    # Only ask if the unit had a recent deep clean (within last 30 days)
    s = _deep_clean_state.get(int(lid)) or {}
    last = _parse_date(s.get("last_done"))
    if not last:
        return False
    if (datetime.now(TZ).date() - last).days > 30:
        return False
    token = _make_feedback_token(lid, res_id)
    unit_name = (get_listings_map() or {}).get(lid) or f"unit-{lid}"
    base_url = (os.environ.get("PUBLIC_BASE_URL") or "").rstrip("/")
    url = f"{base_url}/clean-feedback?id={token}" if base_url else f"/clean-feedback?id={token}"
    is_ar = _has_arabic(r.get("guestName") or "") or True
    body = (
        f"حياك الله 🤍\nخدمة سريعة عشانك: كيف لقيت نظافة الشقة لما دخلت؟ تقييمك يفرق معنا "
        f"وبيوصل لشركة التنظيف. ضغطة وحدة فقط:\n{url}\n\nشكراً 🤍"
        if is_ar else
        f"Hi 🤍\nQuick favor: how did you find the apartment's cleanliness on check-in? "
        f"Your rating goes straight to our cleaning team. One tap:\n{url}\n\nThank you 🤍"
    )
    try:
        send_guest_message(cid, body, "email")
        _cleaning_feedback[token] = {
            "lid": int(lid), "reservation_id": int(res_id),
            "unit": unit_name, "guest": r.get("guestName", ""),
            "ts_sent": datetime.now(TZ).isoformat(timespec="minutes"),
            "score": None, "comment": "", "ts_done": None,
            "last_clean_date": s.get("last_done"),
        }
        _cleaning_feedback_sent.add(res_id)
        log_event("guest", f"تقييم نظافة · أُرسل للضيف {r.get('guestName','')} · {unit_name}")
        return True
    except Exception as e:
        print("send_cleaning_feedback_request error:", e)
        return False

def check_cleaning_feedback_requests():
    """Walk today's arrivals; for each one whose check-in passed >= N hours ago
    and the unit was recently deep-cleaned, dispatch the feedback request once."""
    if not CLEAN_FEEDBACK_ENABLED:
        return
    today_iso = datetime.now(TZ).date().isoformat()
    yest_iso = (datetime.now(TZ).date() - timedelta(days=1)).isoformat()
    try:
        data = api_get("/reservations", params={
            "arrivalStartDate": yest_iso, "arrivalEndDate": today_iso,
            "limit": 200, "includeResources": 0,
        })
    except Exception as e:
        print("cleaning feedback fetch error:", e); return
    rows = data.get("result", []) or []
    now = datetime.now(TZ)
    sent = 0
    for r in rows:
        arr = _parse_date(r.get("arrivalDate"))
        if not arr: continue
        ci_hour = parse_hour(r.get("checkInTime"), 15)
        ci_dt = datetime(arr.year, arr.month, arr.day, min(ci_hour, 23), 0, tzinfo=TZ)
        hours_after_ci = (now - ci_dt).total_seconds() / 3600
        if hours_after_ci < CLEAN_FEEDBACK_DELAY_HOURS or hours_after_ci > 36:
            continue
        if send_cleaning_feedback_request(r):
            sent += 1
    if sent:
        print(f"cleaning-feedback: queued {sent} request(s)")

def cleaning_quality_summary():
    """Aggregate stats: avg score per unit, recent scores, distribution."""
    by_unit = defaultdict(lambda: {"name":"", "lid":None, "scores":[], "comments":[]})
    for tok, fb in _cleaning_feedback.items():
        if fb.get("score") is None:
            continue
        u = by_unit[fb["lid"]]
        u["name"] = fb["unit"]; u["lid"] = fb["lid"]
        u["scores"].append({"ts": fb.get("ts_done"), "score": fb["score"], "guest": fb.get("guest","")})
        if (fb.get("comment") or "").strip():
            u["comments"].append({"ts": fb.get("ts_done"), "comment": fb["comment"], "score": fb["score"]})
    units = []
    for lid, u in by_unit.items():
        scores = [x["score"] for x in u["scores"]]
        avg = round(sum(scores) / len(scores), 2) if scores else None
        units.append({"lid": lid, "name": u["name"],
                      "avg": avg, "count": len(scores),
                      "recent": u["scores"][-5:], "comments": u["comments"][-5:]})
    units.sort(key=lambda x: (x["avg"] if x["avg"] is not None else 99))
    sent = len(_cleaning_feedback)
    done = sum(1 for fb in _cleaning_feedback.values() if fb.get("score") is not None)
    overall_avg = None
    if done:
        all_scores = [fb["score"] for fb in _cleaning_feedback.values() if fb.get("score") is not None]
        overall_avg = round(sum(all_scores)/len(all_scores), 2)
    return {
        "units": units,
        "stats": {"sent": sent, "responded": done,
                  "response_rate": round(done/sent*100) if sent else 0,
                  "overall_avg": overall_avg},
    }

def mark_deep_clean_done(lid, date_iso=None, notes=""):
    if lid not in _deep_clean_state:
        _dc_init(lid)
    s = _deep_clean_state[lid]
    done_date = date_iso or datetime.now(TZ).date().isoformat()
    s["history"].append({"date": done_date,
                         "ts": datetime.now(TZ).isoformat(timespec="minutes"),
                         "notes": notes})
    s["history"] = s["history"][-20:]
    s["last_done"] = done_date
    s["next_scheduled"] = None
    s["next_status"] = "unscheduled"
    # try to free the calendar date in case we'd blocked it
    try:
        api_put(f"/listings/{lid}/calendar",
                {"startDate": done_date, "endDate": done_date, "isAvailable": 1})
    except Exception:
        pass
    return True

def compute_tonight_empty():
    """For each unit empty TONIGHT, return its current price + the discount schedule the
    bot will apply if the unit stays empty (tier 1 at midnight, 2 at noon, 3 at 6 PM, or
    a single weekend drop on Thu/Fri at 5:30 PM). Also includes the owner-set skip status.

    Returned shape (one entry per empty unit):
      {lid, name, price, t1, t2, t3, w, weekend, skipped_until, paused_global,
       tier_times: [{label, hour, minute, pct, price}], next: {label, when_iso, price}}
    """
    today = datetime.now(TZ).date()
    today_iso = today.isoformat()
    weekend = today.weekday() in WEEKEND_DAYS
    listings = get_listings_map()
    now = datetime.now(TZ)
    paused_global = is_discount_paused()

    f1 = (100.0 - DISCOUNT_TIER1_PERCENT) / 100.0
    f2 = (100.0 - DISCOUNT_TIER2_PERCENT) / 100.0
    f3 = (100.0 - DISCOUNT_TIER3_PERCENT) / 100.0
    fw = (100.0 - WEEKEND_DISCOUNT_PERCENT) / 100.0

    items = []
    for lid, name in listings.items():
        try:
            cal = api_get(f"/listings/{lid}/calendar",
                          params={"startDate": today_iso, "endDate": today_iso})
            days = cal.get("result", []) or []
            if not days:
                continue
            d = days[0]
            available = int(d.get("isAvailable", 0) or 0) == 1
            booked = bool(d.get("reservationId"))
            price = d.get("price")
            if not available or booked or not price:
                continue
            price = float(price)
            # recover the original (anchor) so the tier prices we show match what the
            # discount loop would actually write.
            m = re.search(r"ouja-orig:(\d+(?:\.\d+)?)", str(d.get("note") or ""))
            original = float(m.group(1)) if m else price

            tier_times = []
            if weekend:
                tier_times.append({"label": "Weekend",
                                   "hour": WEEKEND_DISCOUNT_HOUR, "minute": WEEKEND_DISCOUNT_MINUTE,
                                   "pct": int(WEEKEND_DISCOUNT_PERCENT),
                                   "price": int(round(original * fw))})
            else:
                tier_times.append({"label": "T1", "hour": DISCOUNT_TIER1_HOUR, "minute": 0,
                                   "pct": int(DISCOUNT_TIER1_PERCENT),
                                   "price": int(round(original * f1))})
                tier_times.append({"label": "T2", "hour": DISCOUNT_TIER2_HOUR, "minute": 0,
                                   "pct": int(DISCOUNT_TIER2_PERCENT),
                                   "price": int(round(original * f2))})
                tier_times.append({"label": "T3", "hour": DISCOUNT_TIER3_HOUR, "minute": 0,
                                   "pct": int(DISCOUNT_TIER3_PERCENT),
                                   "price": int(round(original * f3))})

            # which tier is "next" today (still upcoming)?
            next_tier = None
            for tt in tier_times:
                fire = now.replace(hour=tt["hour"], minute=tt["minute"], second=0, microsecond=0)
                if fire > now:
                    next_tier = {"label": tt["label"], "when_iso": fire.isoformat(timespec="minutes"),
                                 "pct": tt["pct"], "price": tt["price"]}
                    break

            items.append({
                "lid": lid, "name": name,
                "price": int(round(price)), "original": int(round(original)),
                "t1": int(round(original * f1)), "t2": int(round(original * f2)),
                "t3": int(round(original * f3)), "w": int(round(original * fw)),
                "weekend": weekend, "tier_times": tier_times, "next": next_tier,
                "skipped_until": unit_skip_until_iso(lid),
                "paused_global": paused_global,
            })
        except Exception as e:
            print(f"tonight-empty error for {name}: {e}")
    items.sort(key=lambda x: x["name"])
    return items

def compute_arrivals_with_status(window_hours=36):
    """For each upcoming arrival inside `window_hours`, attach the agreement
    signing status, the conversation id, and how far away check-in is. Powers
    the new home-page arrivals timeline."""
    now = datetime.now(TZ)
    cutoff = now + timedelta(hours=window_hours)
    today_iso = now.date().isoformat()
    tomorrow_iso = (now.date() + timedelta(days=1)).isoformat()
    try:
        data = api_get("/reservations", params={
            "arrivalStartDate": today_iso, "arrivalEndDate": tomorrow_iso,
            "limit": 200, "includeResources": 0,
        })
    except Exception as e:
        print("compute_arrivals_with_status fetch error:", e)
        return []
    rows = data.get("result", []) or []
    listings = get_listings_map()
    out = []
    for r in rows:
        if (r.get("status") or "").lower() not in CONFIRMED_STATUSES:
            continue
        arrival = _parse_date(r.get("arrivalDate"))
        if not arrival:
            continue
        hour = parse_hour(r.get("checkInTime"), 15)
        ci_dt = datetime(arrival.year, arrival.month, arrival.day, min(hour, 23), 0, tzinfo=TZ)
        if not (now - timedelta(hours=4) <= ci_dt <= cutoff):
            continue
        lid = r.get("listingMapId")
        unit_name = listings.get(lid) or r.get("listingName") or f"unit-{lid}"
        hrs = (ci_dt - now).total_seconds() / 3600.0
        out.append({
            "reservation_id": r.get("id"),
            "guest": r.get("guestName") or r.get("guestFirstName") or "Guest",
            "unit": unit_name, "listing_id": lid,
            "checkin_iso": ci_dt.isoformat(timespec="minutes"),
            "checkin_label": ci_dt.strftime("%a %H:%M"),
            "hours_until": round(hrs, 1),
            "nights": _res_nights(r),
            "signed": _is_agreement_signed(r),
            "conversation_id": r.get("conversationId"),
            "total_price": r.get("totalPrice"),
        })
    out.sort(key=lambda x: x["hours_until"])
    return out

def compute_urgent_now():
    """Top operational items the owner needs to see at a glance:
       - escalations open + age
       - pending replies + age
       - upcoming arrivals with unsigned agreements
       - empty units tonight (with a price-status hint)
    """
    now = time.time()
    items = []
    # ---- escalations ----
    for mid, e in _escalations.items():
        if e.get("claimed_by"):
            continue
        age_min = int((now - (e.get("last_ping") or now)) / 60)
        items.append({
            "kind": "escalation", "severity": "high",
            "id": str(mid), "title": e.get("guest", "ضيف"),
            "subtitle": e.get("unit", ""),
            "detail": (e.get("reason") or "")[:140],
            "age_min": max(0, age_min),
            "action_view": "inbox",
        })
    # ---- pending replies (>15 min counts as aging) ----
    for mid, d in _pending_replies.items():
        item = d.get("item", {})
        last_time = item.get("last_time", "")
        age_min = None
        try:
            if last_time:
                dt = _parse_msg_dt(last_time)
                if dt:
                    age_min = int((datetime.now(TZ) - dt).total_seconds() / 60)
        except Exception:
            pass
        items.append({
            "kind": "pending_reply",
            "severity": "med" if (age_min or 0) > 15 else "low",
            "id": str(mid), "title": item.get("guest", "ضيف"),
            "subtitle": item.get("unit", ""),
            "detail": (item.get("guest_text") or "")[:140],
            "age_min": age_min,
            "action_view": "inbox",
        })
    # ---- arrivals with unsigned agreements (next 24h) ----
    try:
        arrivals = compute_arrivals_with_status(window_hours=24)
        for a in arrivals:
            if a["signed"] or a["hours_until"] < 0:
                continue
            sev = "high" if a["hours_until"] <= 2 else "med"
            items.append({
                "kind": "unsigned_agreement", "severity": sev,
                "id": str(a["reservation_id"]),
                "title": a["guest"], "subtitle": a["unit"],
                "detail": f"وصول بعد {a['hours_until']} ساعة · العقد غير موقّع",
                "age_min": None, "action_view": "home",
                "checkin": a["checkin_label"],
            })
    except Exception as e:
        print("urgent arrivals error:", e)
    # ---- pricing alerts on edge days (Saudi events + low-pace weekends) ----
    try:
        for pa in compute_pricing_alerts():
            items.append({
                "kind": pa["kind"], "severity": pa["severity"],
                "id": pa["date"],
                "title": pa["title"], "subtitle": "",
                "detail": pa["detail"],
                "age_min": None,
                "action_view": pa.get("action_view", "pricing"),
                "checkin": pa["date"],
            })
    except Exception as e:
        print("pricing alerts error:", e)
    # severity order, then age
    sev_order = {"high": 0, "med": 1, "low": 2}
    items.sort(key=lambda x: (sev_order.get(x["severity"], 9), -(x.get("age_min") or 0)))
    return items

def get_inbox_item_detail(item_id):
    """Build a rich detail view for a pending reply or open escalation, including the
    full conversation history (refetched from Hostaway) and the booking context."""
    try:
        item_id = int(item_id)
    except Exception:
        return None
    src = None; kind = None
    if item_id in _pending_replies:
        src = _pending_replies[item_id]; kind = "reply"
    elif item_id in _escalations:
        src = _escalations[item_id]; kind = "escalation"
    if not src:
        return None

    if kind == "reply":
        item = src.get("item", {})
        cid = item.get("conversation_id")
        guest = item.get("guest"); unit = item.get("unit"); lid = item.get("listing_id")
        res_id = item.get("reservation_id")
        checkin = item.get("checkin"); checkout = item.get("checkout")
        draft = src.get("draft", ""); confirmed = src.get("confirmed", False)
        reason = ""; intent = src.get("intent", ""); conf = src.get("confidence")
        sentiment = src.get("sentiment", "")
    else:   # escalation
        cid = src.get("conversation_id")
        guest = src.get("guest"); unit = src.get("unit"); lid = None
        res_id = None; checkin = None; checkout = None
        draft = ""; confirmed = False
        reason = src.get("reason", ""); intent = ""; conf = None; sentiment = ""

    # refetch the live thread so the dashboard shows the latest
    thread = []
    if cid:
        try:
            data = api_get(f"/conversations/{cid}/messages")
            msgs = sorted((data.get("result") or []), key=_msg_sort_key)
            for m in msgs[-30:]:
                thread.append({
                    "from": "guest" if _msg_is_inbound(m) else "host",
                    "text": (m.get("body") or "").strip(),
                    "ts": _msg_time(m),
                    "automated": (not _msg_is_inbound(m)) and _looks_automated(m.get("body") or ""),
                })
        except Exception as e:
            print(f"detail fetch thread error ({cid}):", e)

    # booking context
    nights = None; total = None; status = ""
    if res_id:
        try:
            data = api_get(f"/reservations/{res_id}")
            r = data.get("result") or {}
            nights = _res_nights(r) or None
            total = r.get("totalPrice") or None
            status = (r.get("status") or "").lower()
            checkin = checkin or r.get("arrivalDate")
            checkout = checkout or r.get("departureDate")
        except Exception as e:
            print(f"detail fetch reservation error ({res_id}):", e)

    return {
        "kind": kind, "id": item_id, "guest": guest or "", "unit": unit or "",
        "listing_id": lid, "conversation_id": cid, "reservation_id": res_id,
        "status": status, "confirmed": confirmed,
        "checkin": checkin, "checkout": checkout, "nights": nights, "total_price": total,
        "draft": draft, "reason": reason, "intent": intent, "confidence": conf,
        "sentiment": sentiment, "thread": thread,
    }

def compute_headsup():
    """List units still empty for TOMORROW night, with the discount they'll actually get."""
    f1 = (100.0 - DISCOUNT_TIER1_PERCENT) / 100.0
    f2 = (100.0 - DISCOUNT_TIER2_PERCENT) / 100.0
    f3 = (100.0 - DISCOUNT_TIER3_PERCENT) / 100.0
    fw = (100.0 - WEEKEND_DISCOUNT_PERCENT) / 100.0
    tomorrow_date = datetime.now(TZ).date() + timedelta(days=1)
    tomorrow = tomorrow_date.isoformat()
    weekend = tomorrow_date.weekday() in WEEKEND_DAYS
    listings = get_listings_map()
    items = []
    for lid, name in listings.items():
        try:
            cal = api_get(f"/listings/{lid}/calendar",
                          params={"startDate": tomorrow, "endDate": tomorrow})
            days = cal.get("result", []) or []
            if not days:
                continue
            d = days[0]
            available = int(d.get("isAvailable", 0) or 0) == 1
            booked = bool(d.get("reservationId"))
            price = d.get("price")
            if not available or booked or not price:
                continue
            price = float(price)
            items.append({"name": name, "price": int(price),
                          "t1": int(round(price * f1)), "t2": int(round(price * f2)),
                          "t3": int(round(price * f3)), "w": int(round(price * fw))})
        except Exception as e:
            print(f"headsup error for {name}: {e}")
    return items, tomorrow, weekend

# ---------------- handled-set persistence ----------------
def load_handled():
    try:
        return set(json.load(open(HANDLED_FILE)))
    except Exception:
        return set()

def save_handled(s):
    try:
        json.dump(list(s), open(HANDLED_FILE, "w"))
    except Exception:
        pass

handled = load_handled()

# ---------------- Discord ----------------
# Message Content Intent is enabled in the Discord Developer Portal (Bot ->
# Privileged Gateway Intents). Declaring it here lets the bot read message
# contents in #knowledge (and any other channel) for users who type facts
# directly without @-mentioning the bot, AND silences the startup warning.
intents = discord.Intents.default()
intents.message_content = True
bot = commands.Bot(command_prefix="!ouja ", intents=intents)

class CleaningDoneView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)   # persistent across restarts

    @discord.ui.button(label="✅ Cleaning Done", style=discord.ButtonStyle.success,
                       custom_id="ouja_cleaning_done")
    async def done(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_message(
            f"✅ Marked done by {interaction.user.mention}. Closing this channel…",
            ephemeral=False)
        await asyncio.sleep(2)
        try:
            await interaction.channel.delete(reason=f"Cleaning completed by {interaction.user}")
        except Exception as e:
            print("delete error:", e)

def channel_name(internal_name):
    # Discord channel names must be lowercase, no spaces/symbols.
    # "Ouja | B209" -> "ouja-b209". The full name is shown inside the channel.
    return re.sub(r"[^a-z0-9]+", "-", internal_name.lower()).strip("-")[:90] or "unit"

# ---- responsible-person lookup (from assignments.json built off the Excel) ----
try:
    ASSIGNMENTS = json.load(open("assignments.json", encoding="utf-8"))
except Exception as e:
    print("Could not load assignments.json:", e)
    ASSIGNMENTS = {"by_day": {}, "discord_ids": {}}

# Python weekday(): Mon=0 ... Sun=6
ARABIC_DAYS = ["الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]

def norm_unit(s):
    s = str(s).strip().lower().replace("ouja", "").replace("|", "")
    return re.sub(r"[\s\-_]+", " ", s).strip()

def _clip(s, n=18):
    """Trim a unit name so the table columns stay aligned."""
    s = str(s).strip()
    return s if len(s) <= n else s[:n - 1] + "…"

def _fmt_hour(h):
    """24h int -> friendly label, e.g. 0->'12 AM', 12->'12 PM', 18->'6 PM'."""
    suffix = "AM" if h < 12 else "PM"
    return f"{h % 12 or 12} {suffix}"

def parse_topic_res(topic):
    """Reservation id stored in a turnover channel's topic."""
    m = re.search(r"hostaway-res:(\d+)", topic or "")
    return m.group(1) if m else None

def parse_topic_did(topic):
    """Responsible cleaner's Discord id stored in the topic (None if unassigned)."""
    m = re.search(r"did:(\d+)", topic or "")
    return m.group(1) if m else None

def find_operation_role(guild):
    """The role to ping when a unit has no assigned cleaner."""
    if OPERATION_ROLE_ID.isdigit():
        r = guild.get_role(int(OPERATION_ROLE_ID))
        if r:
            return r
    name = OPERATION_ROLE_NAME.lower()
    for r in guild.roles:
        if r.name.lower() == name:
            return r
    return None

def responsible_for(internal_name, checkout_dt):
    """Returns (employee_name, discord_id, arabic_day). Any may be None if no match."""
    day = ARABIC_DAYS[checkout_dt.weekday()]
    emp = ASSIGNMENTS.get("by_day", {}).get(day, {}).get(norm_unit(internal_name))
    did = ASSIGNMENTS.get("discord_ids", {}).get(emp) if emp else None
    return emp, did, day

async def get_category(guild):
    for cat in guild.categories:
        if cat.name == CATEGORY_NAME:
            return cat
    return await guild.create_category(CATEGORY_NAME)

async def get_assistant_category(guild):
    """Category for guest-assistant + escalations. Reuses your existing Operations category."""
    want = ASSISTANT_CATEGORY.lower()
    for cat in guild.categories:
        if cat.name.lower() == want:
            return cat
    return await guild.create_category(ASSISTANT_CATEGORY)

async def ensure_channel(guild, name, category=None):
    """Find a text channel by name (create it if missing) and keep it under `category`."""
    ch = discord.utils.get(guild.text_channels, name=name)
    if ch is None:
        try:
            ch = await guild.create_text_channel(name, category=category)
        except Exception as e:
            print(f"channel create error ({name}):", e)
            return None
    elif category is not None and ch.category_id != category.id:
        try:
            await ch.edit(category=category)      # move existing channel under the category
        except Exception as e:
            print(f"channel move error ({name}):", e)
    return ch

async def sync_checkouts():
    guild = bot.get_guild(GUILD_ID)
    if guild is None:
        print("Guild not found — check DISCORD_GUILD_ID and that the bot was invited.")
        return
    category = await get_category(guild)

    # reservation ids that already have a live channel
    existing = set()
    for ch in category.text_channels:
        rid = parse_topic_res(ch.topic)
        if rid:
            existing.add(rid)
    handled.update(existing)

    items = await asyncio.to_thread(fetch_upcoming_checkouts)
    for it in items:
        if it["res_id"] in handled or it["res_id"] in existing:
            continue
        try:
            emp, did, day = responsible_for(it["listing"], it["checkout"])
            # topic carries the reservation id + responsible discord id (for reminders)
            topic = f"hostaway-res:{it['res_id']} did:{did or ''}"
            ch = await guild.create_text_channel(
                channel_name(it["listing"]), category=category, topic=topic)
            embed = discord.Embed(title="🧹 Turnover Ready", color=GOLD)
            embed.add_field(name="Unit", value=it["listing"], inline=False)
            embed.add_field(name="Guest", value=it["guest"], inline=True)
            embed.add_field(name="Checkout",
                            value=it["checkout"].strftime("%a %d %b · %I:%M %p"), inline=True)

            if emp:
                embed.add_field(name="مسؤول التنظيف",
                                value=(f"<@{did}> ({emp})" if did else emp), inline=False)
            embed.set_footer(text="Tap the button below once cleaning is complete to close this channel.")

            content = f"<@{did}> 🧹 وحدة جاهزة للتنظيف" if did else None
            await ch.send(content=content, embed=embed, view=CleaningDoneView(),
                          allowed_mentions=discord.AllowedMentions(users=True, roles=True))
            handled.add(it["res_id"])
            save_handled(handled)
            note = f" -> {emp}" if emp else " (no responsible match — check unit name)"
            print(f"Opened channel for {it['listing']} (res {it['res_id']}){note}")
        except Exception as e:
            print("create error:", e)

@tasks.loop(minutes=POLL_MINUTES)
async def poll_loop():
    try:
        await sync_checkouts()
    except Exception as e:
        print("poll error:", e)

# remembers when each open channel was last pinged (resets on restart, which is fine)
_last_reminder = {}

@tasks.loop(minutes=1)
async def reminder_loop():
    """Nag every open turnover channel until cleaning is marked done.
    12 PM–3 PM: every 30 min. After 3 PM: every 15 min. Quiet outside those hours."""
    now = datetime.now(TZ)
    hour = now.hour
    if hour < REMINDER_START_HOUR or hour >= REMINDER_END_HOUR:
        return
    interval = REMINDER_SLOW_MIN if hour < REMINDER_FAST_HOUR else REMINDER_FAST_MIN
    guild = bot.get_guild(GUILD_ID)
    if guild is None:
        return
    category = discord.utils.get(guild.categories, name=CATEGORY_NAME)
    if category is None:
        return
    op_role = find_operation_role(guild)
    live_ids = set()
    for ch in list(category.text_channels):
        if not parse_topic_res(ch.topic):
            continue                      # only turnover channels
        live_ids.add(ch.id)
        last = _last_reminder.get(ch.id)
        if last and (now - last) < timedelta(minutes=interval):
            continue
        did = parse_topic_did(ch.topic)
        if did:
            mention = f"<@{did}>"
        elif op_role:
            mention = op_role.mention
        else:
            mention = f"@{OPERATION_ROLE_NAME}"
        try:
            await ch.send(
                f"{mention} ⏰ تذكير: هالوحدة لسه تحتاج تنظيف — اضغط ✅ Cleaning Done لما تخلص.",
                allowed_mentions=discord.AllowedMentions(users=True, roles=True))
            _last_reminder[ch.id] = now
        except Exception as e:
            print(f"reminder error ({ch.name}):", e)
    # drop timers for channels that no longer exist (cleaning done)
    for cid in list(_last_reminder.keys()):
        if cid not in live_ids:
            _last_reminder.pop(cid, None)

async def post_discount_summary(changes, today, pct, label):
    if not changes:
        return
    guild = bot.get_guild(GUILD_ID)
    if guild is None:
        return
    category = await get_category(guild)
    channel = await ensure_channel(guild, DISCOUNT_CHANNEL, category)
    if channel is None:
        return
    state = ("🧪 معاينة فقط — ما تغيّر أي سعر"
             if DISCOUNT_DRY_RUN else f"✅ تم تطبيق خصم {int(pct)}٪ من السعر الأصلي")
    lines = "\n".join(f"• {c['name']}: {c['orig']} → {c['new']} ر.س" for c in changes[:60])
    embed = discord.Embed(title=f"💰 {label} · خصم {int(pct)}٪ · {today}",
                          description=lines, color=GOLD)
    embed.set_footer(text=f"{state} · {len(changes)} وحدة فاضية")
    await channel.send(embed=embed)

# Owner-controlled pause for the discount tiers. 0 = not paused; otherwise unix-timestamp
# until which all tier loops skip. Persisted in state so a redeploy doesn't accidentally
# resume discounts the owner wanted off.
_discount_paused_until = 0

# Per-unit discount skip: {listing_id (int) -> unix-timestamp until which discounts skip
# for this specific unit only}. Lets the owner say "hold price on this apartment" without
# pausing all discounts globally. Persisted.
_unit_discount_skip = {}

def is_unit_skipped(lid):
    try:
        ts = _unit_discount_skip.get(int(lid), 0)
        return float(ts) > time.time()
    except Exception:
        return False

def unit_skip_until_iso(lid):
    try:
        ts = _unit_discount_skip.get(int(lid), 0)
        if float(ts) <= time.time():
            return ""
        return datetime.fromtimestamp(float(ts), TZ).isoformat(timespec="minutes")
    except Exception:
        return ""

def is_discount_paused():
    return _discount_paused_until > time.time()

def discount_pause_status():
    """{paused, until_ts, until_iso}"""
    if not is_discount_paused():
        return {"paused": False, "until_ts": 0, "until_iso": ""}
    return {"paused": True, "until_ts": _discount_paused_until,
            "until_iso": datetime.fromtimestamp(_discount_paused_until, TZ).isoformat(timespec="minutes")}

async def _run_tier(pct, label):
    if is_discount_paused():
        print(f"[{label}] skipped — discounts paused by owner until "
              f"{datetime.fromtimestamp(_discount_paused_until, TZ).strftime('%a %H:%M')}")
        return
    try:
        changes, today = await asyncio.to_thread(apply_discount_tier, pct)
        mode = "DRY-RUN" if DISCOUNT_DRY_RUN else "LIVE"
        print(f"[{label} {today}] {mode} {pct:.0f}%: {len(changes)} empty units")
        for c in changes:
            print(f"   {c['name']}: orig {c['orig']} -> {c['new']}")
        await post_discount_summary(changes, today, pct, label)
    except Exception as e:
        print(f"{label} error:", e)

@tasks.loop(time=dt_time(hour=DISCOUNT_TIER1_HOUR, tzinfo=TZ))
async def discount_tier1_loop():
    if is_weekend_today():
        print("[Tier 1] Thu/Fri — skipping midnight discount (weekend rule)")
        return
    await _run_tier(DISCOUNT_TIER1_PERCENT, "Tier 1 (midnight)")

@tasks.loop(time=dt_time(hour=DISCOUNT_TIER2_HOUR, tzinfo=TZ))
async def discount_tier2_loop():
    if is_weekend_today():
        print("[Tier 2] Thu/Fri — skipping noon discount (weekend rule)")
        return
    await _run_tier(DISCOUNT_TIER2_PERCENT, "Tier 2 (noon)")

@tasks.loop(time=dt_time(hour=DISCOUNT_TIER3_HOUR, tzinfo=TZ))
async def discount_tier3_loop():
    if is_weekend_today():
        print("[Tier 3] Thu/Fri — skipping evening discount (weekend rule)")
        return
    await _run_tier(DISCOUNT_TIER3_PERCENT, "Tier 3 (6 PM)")

@tasks.loop(time=dt_time(hour=WEEKEND_DISCOUNT_HOUR, minute=WEEKEND_DISCOUNT_MINUTE, tzinfo=TZ))
async def discount_weekend_loop():
    if not is_weekend_today():
        return                       # only fires on Thu/Fri
    await _run_tier(WEEKEND_DISCOUNT_PERCENT, "Weekend (Thu/Fri 5:30 PM)")

def _headsup_table(items, weekend):
    """Build a clean monospace table + an explanatory note for the heads-up preview."""
    if weekend:
        wp = int(WEEKEND_DISCOUNT_PERCENT)
        c1, c2 = f"-{wp}%", "SAVE"
        header = f"{'UNIT':<16}{'NOW':>6}{c1:>6}{c2:>6}"
        rows = [f"{_clip(it['name'], 16):<16}{it['price']:>6}{it['w']:>6}"
                f"{it['price'] - it['w']:>6}" for it in items]
        note = f"Single drop if still empty: **-{wp}% at 5:30 PM** (weekend rule)."
    else:
        p1, p2, p3 = (int(DISCOUNT_TIER1_PERCENT), int(DISCOUNT_TIER2_PERCENT),
                      int(DISCOUNT_TIER3_PERCENT))
        c1, c2, c3 = f"-{p1}%", f"-{p2}%", f"-{p3}%"
        header = f"{'UNIT':<16}{'NOW':>6}{c1:>6}{c2:>6}{c3:>6}"
        rows = [f"{_clip(it['name'], 16):<16}{it['price']:>6}{it['t1']:>6}"
                f"{it['t2']:>6}{it['t3']:>6}" for it in items]
        note = (f"Prices auto-drop if still empty: **-{p1}% at {_fmt_hour(DISCOUNT_TIER1_HOUR)}** → "
                f"**-{p2}% at {_fmt_hour(DISCOUNT_TIER2_HOUR)}** → "
                f"**-{p3}% at {_fmt_hour(DISCOUNT_TIER3_HOUR)}**.")
    sep = "─" * len(header)
    # Prefix each line with a zero-width LEFT-TO-RIGHT MARK so an Arabic unit
    # name can't flip the line's direction and scramble the number columns.
    lines = [header, sep, *rows]
    return note, "\n".join("\u200e" + ln for ln in lines)

async def post_headsup(items, tomorrow, weekend):
    guild = bot.get_guild(GUILD_ID)
    if guild is None:
        return
    category = await get_category(guild)
    channel = await ensure_channel(guild, HEADS_UP_CHANNEL, category)
    if channel is None:
        return
    try:
        date_str = datetime.strptime(tomorrow, "%Y-%m-%d").strftime("%a · %d %b %Y")
    except ValueError:
        date_str = tomorrow
    if not items:
        await channel.send(embed=discord.Embed(
            title="📋 Tomorrow's Open Units",
            description=f"**{date_str}**\n\nAll units are booked for tomorrow — nothing to discount. 🎉",
            color=GOLD))
        return
    shown = items[:45]
    extra = len(items) - len(shown)
    note, table = _headsup_table(shown, weekend)
    desc = f"**{date_str}** · **{len(items)} units** still open\n{note}\n```\n{table}\n```"
    if extra > 0:
        desc += f"\n*+{extra} more not shown.*"
    embed = discord.Embed(title="📋 Tomorrow's Open Units", description=desc, color=GOLD)
    embed.set_footer(text="All prices in SAR · preview only — nothing changes until the scheduled time.")
    await channel.send(embed=embed)

@tasks.loop(time=dt_time(hour=HEADS_UP_HOUR, tzinfo=TZ))
async def headsup_loop():
    try:
        items, tomorrow, weekend = await asyncio.to_thread(compute_headsup)
        print(f"[heads-up {tomorrow}] {len(items)} units open for tomorrow "
              f"({'weekend' if weekend else 'weekday'})")
        await post_headsup(items, tomorrow, weekend)
    except Exception as e:
        print("headsup error:", e)

# ====================== AI guest-message assistant ======================
# Claude drafts a reply, a human approves it in Discord, and only then is it sent.
ASSISTANT_RULES = """You are "فيصل", the front-desk assistant for Ouja Residence, a premium short-term \
rental company in Riyadh, Saudi Arabia. You draft replies to guest messages. A human teammate reviews \
every draft before it is sent, so be helpful but stay strictly within your lane.

TONE
- Warm, friendly, professional — Ritz-Carlton hospitality standard.
- Reply in the SAME language the guest used, switching immediately if they switch.
- For Arabic: ALWAYS Saudi/Gulf — either Najdi (preferred) or clean "white" Saudi \
that any Gulf reader hears as one of their own. NEVER Levantine, Egyptian, Iraqi, \
or Maghrebi. Avoid heavy/folksy slang too — aim for how a polished Saudi host talks: \
simple, elegant, respectful.

⛔ HARD-BANNED non-Saudi words (a single occurrence = wrong reply, rewrite before sending):

【مصرية / Egyptian】
دلوقتي→الحين · إيه→إيش/وش · إزيك→كيفك · إزاي→كيف · إيوه→إي · لأ→لا · ازاي→كيف ·
مش→مو/مب · عايز/عاوز/عاوزة→أبي/أبغى · كده/كدا→كذا · فين→وين · إمتى→متى · ليه→ليش ·
بتاع/بتاعك/بتاعي/بتاعنا→حق/حقك/حقي/حقنا (أو مال/مالي) · مفيش→ما فيه · معلش→ما يخالف/ما عليه ·
حاجة/حاجات→شي/أشياء · ولا حاجة→ولا شي · خالص (بمعنى أبداً)→أبداً · أوي/قوي (بمعنى جداً)→مرة/كثير ·
جدا (بنطق مصري للجيم)→جداً/مرة · هـ+فعل (هتعمل/هاجي/هاروح)→بـ+فعل (بتسوي/بجي/بروح) ·
بقى (سياقاً)→صار/خلاص · معايا/معاك/معاه→معي/معك/معه · ربنا→الله · حضرتك→حياك/تكفى ·
طب/طب ماشي→طيب/تمام · ساعتها→وقتها · بصراحة كده→بصراحة · لسه/لسة→لسا (في الخليج نادرة، تجنّبها) ·
قاعد + فعل (مصري نقيب)→نفس البنية موجودة سعودياً، بس لا تستخدم نطق "إنت قاعد تعمل ايه" ·
ايوه يا فندم→إي حياك · صباح الفل/الورد→صبحك بالخير

【شامية / Levantine】
شو→وش/إيش · بدك/بدّك/بدي/بدنا/بدّو/بدّها→تبي/تبغى/أبي/نبي/يبي/تبي · هلق/هلأ→الحين ·
هيك→كذا · كتير→كثير/مرة · منيح/منيحة→زين/كويس · لشو→ليش · شو رأيك→وش رايك/إيش رايك ·
شلونك (عراقية/شامية)→كيفك · متل/متلك/متل ما→مثل/مثلك/مثل ما · كرمالك/كرمال→عشانك/علشان ·
ولاي/والك→والله · عنجد→فعلاً/بصراحة · منشان/مشان→عشان/علشان · بلكي/يمكن أنو→يمكن/ممكن ·
لوين→وين · بكير→بدري · إجا/إجت→جا/جت · رح + فعل (رح يجي/رح أعمل)→بـ+فعل (بيجي/بسوي) ·
كرمى لخاطرك→عشان خاطرك · هلا فيك→هلا والله/حياك · شو القصة→وش القصة/وش فيه ·
في شي (شامية)→فيه شي · ما في/ما حدا→ما فيه/ما حد · هاد/هادي (بنطق شامي)→هذا/هذي ·
طيار (سيارة طيارة بشامي)→تجنّب الالتباس · لساتو→لسا

【عراقية / Iraqi】
شكو→وش فيه/إيش فيه · ماكو→ما فيه · اكو→فيه · هسة/هسه→الحين · شوكت→متى ·
هواية→كثير · خوش (بمعنى حلو)→زين/كويس · بيش (بمعنى بكم)→بكم · شلون→كيف ·
ها (في بداية الجملة كنداء)→تجنّب · ريّس→أستاذ/يا غالي

【مغربية / تونسية / Maghrebi】
كيفاش→كيف · واش (لو مغربية)→إيش/وش · بزاف→كثير/مرة · غادي→بـ/راح ·
كاين→فيه · ماكاينش→ما فيه · ديالي/ديالك→حقي/حقك (أو مالي/مالك) · نتا/نتي→أنت/إنتي ·
واخا→تمام/ماشي · لاباس→كويس/بخير · بصح (جزائرية)→بس/لكن

【MSA متكلّف يُخفّف لكاجوال خليجي】
الآن→الحين · ربما→يمكن/ممكن · غداً→بكرة · لذلك→عشان كذا · إذن→إذاً/يعني ·
كذلك→وكذلك (مقبولة) أو "وبعد" · فضلاً عن→وبعد · جداً→مرة/كثير (في الكاجوال)

✅ استخدم هذه السعودية الطبيعية بحرية:
حياك الله · هلا والله · يا هلا · مرحباً · أبشر · تم · ما يخالف · إن شاء الله ·
بإذن الله · ولا يهمك · يعطيك العافية · الله يعطيك العافية · يسعدك · الله يسعدك ·
عشان/علشان · بس · طيب · تمام · ماشي · زين · كويس · الحين · بدري · بكرة ·
الصبح · المغرب · العصر · وش · إيش · ليش · متى · وين · كيف · كم · مين ·
تبي/تبغى · أبي/أبغى · تكفى · لو سمحت · يا غالي · يا طويل العمر · حياك ·
حقي/حقك/حقنا · عندي/عندك · مو/مب · ما · لا · إي · أكيد · أكيد طبعاً

🔍 Mental check قبل كل رسالة: "لو قرأها سعودي في الرياض، يحسّها طبيعية ١٠٠٪ ولا يطلع له خيط
شامي/مصري؟" لو فيها كلمة تخوّن (مثل: شو، بدك، دلوقتي، إيه، عايز، إزاي، مش، فين، كده،
هلق، هيك، كتير)، أعد صياغة الجملة.

- Keep replies short, human, and to the point. Never reveal you are an AI unless asked.

CHOOSE ONE OF THREE ACTIONS
- "auto" → a VERY simple, safe, low-risk reply you are highly confident about, where a wrong \
answer would do no harm: greetings, thanks, reassurance, simple confirmations, and basic facts \
that are obvious or that you were explicitly given (e.g. a friendly "حياك الله", "شكراً لك", \
"تم، بالتوفiق"). These get sent to the guest automatically, so only use "auto" when you are sure.
- "reply" → a helpful reply you CAN draft, but a human should approve it first because it is more \
substantive, or you are not fully certain (most amenity/directions/check-in answers fall here).
- "escalate" → matches the MUST-escalate list below. Draft no reply.

WHEN IN DOUBT between "auto" and "reply", pick "reply" — a human approves it before it sends, so it is \
always safe. Only choose "escalate" when the request matches the MUST-escalate list (complaint, dispute, \
refund, booking change, upset guest, security info). A question you simply don't have the answer to \
(like live availability) is NOT a reason to escalate — suggest what you can and point the guest to the \
Airbnb link. Never gamble on "auto".

YOU MAY draft replies (auto or reply) about
- Unit amenities (wifi, parking, pool, kitchen, facilities)
- Check-in / check-out TIMES and the self-entry PROCESS (never the actual code)
- Directions, location, nearby restaurants and areas
- House-rules clarifications
- Greetings, thanks, general hospitality

ARRIVAL GUIDE LINK (important)
- When a guest asks about the location, the address, directions, where to park, the outer/building \
door number, or any arrival detail, point them to the unit's arrival-guide link (it is provided to \
you in the context as "Arrival-guide link for this unit"). Tell them all the details are there.
- Always include the exact link given to you. NEVER invent a link.
- If the link is "NOT AVAILABLE", do NOT make one up — set action to "escalate" instead.
- If the guest says they can't open it or nothing happens, explain it opens an Airbnb warning page \
("you're leaving Airbnb"): they should tap "متابعة" (or "Continue" if their app is in English), \
then tap the guide/directions option. (This is "auto"/"reply" — just a helpful instruction.)

PRE-BOOKING PRIVACY (critical)
- The context gives you "Booking status". If it is NOT CONFIRMED (an inquiry or pre-booking), you \
MUST NOT share the exact location, address, building/door number, or the arrival-guide link — even \
if the guest asks directly. Politely tell them the full location and arrival details are sent right \
after the booking is confirmed. You may still talk about the general area, amenities, and price.
- Only share location and the arrival-guide link when Booking status is CONFIRMED.

SUGGESTING A UNIT / AVAILABILITY  (do NOT escalate these — suggest instead)
- If the guest asks for another/different unit, a bigger/cheaper/smaller one, a unit in another area, \
a unit with a certain feature (balcony, pool, terrace, X bedrooms, etc.), OR simply "do you have a \
unit available today / for these dates?": this is a SUGGESTION task. Use action "reply" (a human \
approves it). It is NOT an escalation.
- FIRST, if you don't already know what they want, ask briefly: preferred area, number of bedrooms, \
and any must-have or budget. If they already told you, skip the questions and suggest right away.
- Then, from the "قائمة وحدات عوجا" in the context, suggest 1-3 matches. For each: name, bedrooms, \
area, the "starting from" nightly price, and the Airbnb link if it is in the list. NEVER invent a link \
or a detail. If nothing matches exactly, suggest the closest and state the differences honestly.
- AVAILABILITY: you do NOT have live availability. Never promise a unit is free. Instead, suggest the \
options and tell the guest to check live availability and book directly from the Airbnb link (the link \
always shows what is open for their dates). Not knowing availability is NEVER a reason to escalate — \
suggest + send them to the link.
- A FEATURE you're not sure about (e.g. whether a unit allows smoking, or has a specific view): if it \
is not in your provided info, suggest the closest units, be honest that you'd confirm that specific \
detail with the team, and keep it action "reply" so a human checks. Do NOT invent the feature.
- Whenever you mention a price, add a short note that prices are approximate and BEFORE tax and the \
platform service fee.
- MIXED message: if a guest asks several things and you can help with some (suggest units, directions, \
amenities) but one part truly needs a human (e.g. extending checkout, a refund), DRAFT a "reply" that \
handles what you can and says you'll check the rest with the team — do not escalate the whole message.

OPTIONAL EXTRAS / UPSELLS (offer gently — never push, never auto-promise)
- You may mention Ouja's optional paid extras when it naturally fits and would genuinely help the guest, \
phrased as a friendly offer, not a hard sell. Good moments: a guest arriving early or asking about \
arrival time → early check-in; a guest asking about checkout or a late flight → late checkout; a guest \
asking about airport pickup or transport → the Sawari Al Musafir chauffeur service; a long stay → an \
extra mid-stay cleaning.
- These are subject to availability and a fee, so you CANNOT confirm them yourself. Offer it, say the \
team will confirm availability and the price, and set action to "reply" (a human approves). Never promise \
a time or a price you weren't given, and never invent a fee. One soft offer is enough — don't repeat it.
- If the guest says no or ignores it, drop it immediately and don't bring it up again.

YOU MUST NOT do these — instead set action to "escalate"
- Confirm, modify, cancel, or refund a booking
- Offer any discount, comp, or price change
- Share the door/entry CODE or any building security info
- Share other guests', owners', or internal/financial information
- Handle any complaint, dispute, damage claim, or an upset guest
- Promise late checkout, early check-in, or extra services you cannot verify
- Give legal, medical, or financial advice
- Discuss anything outside hosting / off-topic
- State any fact about the unit you were not given — never invent details
- Anything you are unsure about — when in doubt, escalate; never guess

OUTPUT — respond with ONLY a JSON object, nothing else:
{"action": "auto" | "reply" | "escalate",
 "reply": "the drafted message to the guest IN THE GUEST'S OWN LANGUAGE; empty string if escalating",
 "intent": "short label IN ARABIC, e.g. واي فاي، اتجاهات، تسجيل دخول، تسعير، شكوى، تعديل حجز",
 "sentiment": "ok" | "upset",
 "reason": "one short line IN ARABIC for the human reviewer explaining your choice (the team reads Arabic)",
 "confidence": 0.0-1.0}

IMPORTANT: "reply" stays in the guest's language (Najdi Arabic if they wrote Arabic). But "intent" and \
"reason" must ALWAYS be written in Arabic, because the Ouja team reading these cards speaks Arabic."""

def _msg_is_inbound(m):
    return int(m.get("isIncoming", m.get("incoming", 0)) or 0) == 1

def _msg_time(m):
    return str(m.get("date") or m.get("insertedOn") or m.get("latestMessageDate") or "")

def _msg_sort_key(m):
    """Sortable timestamp so the guest's *latest* message is reliably the newest.
    (Within one conversation the tz skew is uniform, so relative order stays correct.)"""
    return _parse_msg_dt(_msg_time(m)) or datetime.min.replace(tzinfo=TZ)

def _parse_msg_dt(s):
    """Parse a Hostaway message timestamp as Riyadh local time. Returns aware dt or None.
    Parsing as TZ is safe for baselining: a tz mismatch can only make a message look OLDER
    (so it gets baselined like before), never wrongly resurrect an old one."""
    s = str(s).strip().replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s[:19], fmt).replace(tzinfo=TZ)
        except Exception:
            pass
    return None

def _looks_automated(body):
    """True if an outbound message looks like an automated welcome/booking message
    (i.e. not a real answer to the guest)."""
    b = (body or "").lower()
    return any(mk.lower() in b for mk in AUTO_REPLY_MARKERS)

_guide_cache = {}

def get_guide_url(listing_id):
    """Pull the arrival-guide link stored in the listing's Custom Fields. Cached per listing."""
    if listing_id in _guide_cache:
        return _guide_cache[listing_id]
    url = None
    try:
        data = api_get(f"/listings/{listing_id}", params={"includeResources": 1})
        listing = data.get("result") or {}
        # 1) look through the listing's custom-field values for a URL
        for key in ("listingCustomFieldValues", "customFieldValues", "customFields"):
            for cf in (listing.get(key) or []):
                v = str(cf.get("value") or "").strip()
                if v.startswith("http"):
                    url = v
                    break
            if url:
                break
        # 2) fallback: scan the whole listing for the guide domain
        if not url:
            blob = json.dumps(listing, ensure_ascii=False)
            m = (re.search(r"https?://[^\s\"']*oujaguide[^\s\"']*", blob)
                 or re.search(r"https?://[^\s\"']*netlify[^\s\"']*", blob))
            if m:
                url = m.group(0)
    except Exception as e:
        print(f"guide url error for listing {listing_id}: {e}")
    _guide_cache[listing_id] = url
    return url

# A booking only counts as "confirmed" (safe to share location/guide) for these statuses.
CONFIRMED_STATUSES = {"new", "modified"}

def get_reservation_status(reservation_id):
    """Return the reservation status (lowercased), or '' if unknown. Not cached — it can change."""
    if not reservation_id:
        return ""
    try:
        data = api_get(f"/reservations/{reservation_id}")
        return (data.get("result", {}) or {}).get("status", "").lower()
    except Exception as e:
        print(f"reservation status error ({reservation_id}):", e)
        return ""

# Knowledge base loaded from the #knowledge Discord channel; injected into every draft.
_knowledge_text = ""
# Per-apartment raw facts pulled from #knowledge. Messages mentioning an
# apartment name go HERE (scoped); messages without a name go into
# _knowledge_text (general). Re-derived from the channel on every
# load_knowledge() call, so the channel is the source of truth.
_knowledge_apartment_facts = {}    # lid (int) -> [fact strings]

# Units catalog (name · bedrooms · area · base price · Airbnb link) pulled from Hostaway.
_catalog_text = ""
_catalog_ts = 0
_catalog_units = []        # structured: [{id, name, beds, area, price, link}]
# availability/price cache: (listing_id, checkin, checkout) -> (result|None, ts)
# Bounded so a busy day can't grow this forever (each unique date range is one entry).
_avail_cache = OrderedDict()
_AVAIL_CACHE_MAX = 2000
INTEL_CACHE_MIN  = int(os.environ.get("INTEL_CACHE_MIN", "20"))    # cache calendar lookups
INTEL_MAX_CHECKS = int(os.environ.get("INTEL_MAX_CHECKS", "80"))   # max units to date-check per msg (parallel — covers full portfolio)
INTEL_PARALLEL = int(os.environ.get("INTEL_PARALLEL", "12"))         # concurrent calendar lookups

# Model for guest-facing drafts. Always premium by default — Haiku produces too many
# generic/templated replies that ignore the actual context (e.g. saying "code arrives
# 5 days before check-in" when check-in is today). Cost increase is modest at this volume.
GUEST_DRAFT_MODEL = os.environ.get("GUEST_DRAFT_MODEL", CLAUDE_MODEL_PREMIUM)

def _listing_active(L):
    """False if the listing looks inactive/unlisted (the red 🚫 in Hostaway) — skip those."""
    v = L.get("status")
    if isinstance(v, str) and v.lower() in (
            "inactive", "disabled", "unlisted", "delisted", "deleted", "draft", "paused", "off"):
        return False
    if v in (0, "0", False):
        return False
    for key in ("isActive", "listed", "active", "isListed"):
        if key in L and L.get(key) in (0, "0", False, "false", "False"):
            return False
    return True

def _airbnb_link(L):
    """Best-effort Airbnb listing URL from the listing's channel data."""
    blob = json.dumps(L, ensure_ascii=False)
    m = (re.search(r"https?://[^\s\"']*airbnb\.[^\s\"']*/rooms/\d+", blob)
         or re.search(r"https?://[^\s\"']*airbnb\.[^\s\"']*", blob))
    if m:
        return m.group(0)
    m2 = re.search(r'"airbnb[A-Za-z]*[Ii]d"\s*:\s*"?(\d{5,})"?', blob)
    if m2:
        return f"https://www.airbnb.com/rooms/{m2.group(1)}"
    return ""

def _nightly_from(listing_id):
    """A representative 'starting from' nightly price from the calendar (next ~30 days)."""
    try:
        today = datetime.now(TZ).date()
        data = api_get(f"/listings/{listing_id}/calendar",
                       params={"startDate": today.isoformat(),
                               "endDate": (today + timedelta(days=30)).isoformat()})
        prices = [d.get("price") for d in (data.get("result", []) or [])
                  if d.get("isAvailable", 1) and isinstance(d.get("price"), (int, float))
                  and d.get("price") > 0]
        return min(prices) if prices else None
    except Exception as e:
        print(f"price fetch error ({listing_id}):", e)
        return None

def _extract_amenities(L):
    """Pull a list of human-readable amenity names from a Hostaway listing object.
    Hostaway uses different shapes across endpoints (listingAmenities / amenities /
    amenityIds) — try the common ones, skip blanks."""
    out = []
    for key in ("listingAmenities", "amenities", "listing_amenities"):
        for a in (L.get(key) or []):
            if isinstance(a, dict):
                n = a.get("amenityName") or a.get("name") or a.get("amenity")
                if n: out.append(str(n).strip())
            elif isinstance(a, str) and a.strip():
                out.append(a.strip())
    # dedupe, preserve order
    seen, dedup = set(), []
    for a in out:
        k = a.lower()
        if k not in seen:
            seen.add(k); dedup.append(a)
    return dedup

# Common amenity keywords -> normalized tag we can match guest hints against
_AMENITY_TAGS = {
    "wifi": ["wifi", "wi-fi", "internet", "واي فاي", "واي-فاي", "انترنت", "إنترنت"],
    "pool": ["pool", "swimming", "مسبح", "حمام سباحه", "حمام سباحة"],
    "parking": ["parking", "garage", "موقف", "باركن", "موقف خاص", "موقف سيارات"],
    "kitchen": ["kitchen", "kitchenette", "مطبخ"],
    "balcony": ["balcony", "terrace", "patio", "بلكون", "بلكونة", "تراس"],
    "smoking": ["smoking", "smoking allowed", "تدخين", "مسموح التدخين"],
    "gym": ["gym", "fitness", "جيم", "نادي رياضي", "صالة رياضية"],
    "elevator": ["elevator", "lift", "مصعد"],
    "washer": ["washer", "washing machine", "غسالة", "غسالة ملابس"],
    "ac": ["air conditioning", "ac", "تكييف", "مكيف"],
    "family": ["family-friendly", "kid friendly", "child friendly", "عائلي", "مناسب للعوائل"],
    "workspace": ["workspace", "desk", "مكتب", "غرفة مكتب"],
}

def _amenity_tags(amenity_names):
    """Reduce raw amenity strings to a normalized tag set (wifi/pool/parking/...)."""
    blob = " ; ".join(amenity_names).lower()
    tags = set()
    for tag, kws in _AMENITY_TAGS.items():
        if any(kw in blob for kw in kws):
            tags.add(tag)
    return sorted(tags)

def load_catalog(force=False):
    """Build a units catalog from Hostaway listings (cached 1h). Used to suggest alternatives.
    Skips inactive/unlisted listings (the red 🚫).

    For each active unit we pull: name, bedrooms, bathrooms, max guests, area,
    starting price, Airbnb link, raw amenity list + normalized tag set
    (wifi/pool/parking/...). The richer set powers smarter matching in
    enrich_catalog_for_dates() and clearer suggestion blurbs in claude_draft()."""
    global _catalog_text, _catalog_ts, _catalog_units
    if not force and _catalog_text and (time.time() - _catalog_ts) < 3600:
        return
    try:
        data = api_get("/listings", params={"limit": 100, "includeResources": 1})
        rows, units, skipped = [], [], 0
        for L in (data.get("result", []) or []):
            name = (L.get("internalListingName") or L.get("name") or "").strip()
            if not name:
                continue
            if not _listing_active(L):
                skipped += 1
                if CATALOG_DEBUG:
                    print(f"  catalog SKIP (inactive): {name} · status={L.get('status')!r}")
                continue
            beds = L.get("bedroomsNumber")
            baths = L.get("bathroomsNumber")
            capacity = L.get("personCapacity") or L.get("guestsIncluded") or L.get("maxGuests")
            area = (L.get("city") or L.get("address") or "").strip()
            neighbourhood = (L.get("neighbourhood") or L.get("neighborhood") or "").strip()
            ptype = (L.get("propertyTypeName") or L.get("propertyType") or "").strip()
            price = (_nightly_from(L.get("id")) if CATALOG_CALENDAR_PRICES else None) or L.get("price")
            link = _airbnb_link(L)
            amen_raw = _extract_amenities(L)
            tags = _amenity_tags(amen_raw)
            parts = [name]
            if beds:    parts.append(f"{beds} غرفة نوم")
            if baths:   parts.append(f"{baths} حمام")
            if capacity:parts.append(f"يستوعب {capacity}")
            if ptype:   parts.append(ptype)
            if area or neighbourhood:
                parts.append((neighbourhood + " · " + area).strip(" ·"))
            if price:   parts.append(f"تبدأ من ~{round(price)} ر.س/الليلة")
            if tags:    parts.append("مرافق: " + ", ".join(tags))
            if link:    parts.append(link)
            rows.append(" · ".join(parts))
            units.append({"id": L.get("id"), "name": name, "beds": beds,
                          "baths": baths, "capacity": capacity,
                          "area": area, "neighbourhood": neighbourhood, "ptype": ptype,
                          "price": round(price) if price else None,
                          "link": link, "amenities": amen_raw[:30], "tags": tags})
            if CATALOG_DEBUG:
                print(f"  catalog OK: {name} · beds={beds} · cap={capacity} · "
                      f"tags={','.join(tags)} · link={'y' if link else 'n'}")
        _catalog_text = "\n".join(rows)[:8000]
        _catalog_units = units
        _catalog_ts = time.time()
        print(f"catalog: loaded {len(rows)} active units (skipped {skipped} inactive)")
    except Exception as e:
        print("catalog load error:", e)

def _parse_date(s):
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except Exception:
        return None

def unit_availability_price(listing_id, checkin, checkout):
    """Check the calendar for ONE unit over a specific stay. Returns
    {available, nights, total, avg} (total/avg before tax & platform fee) or None
    if unknown. Cached for INTEL_CACHE_MIN minutes."""
    if not listing_id or not checkin or not checkout:
        return None
    ci, co = _parse_date(checkin), _parse_date(checkout)
    if not ci or not co or co <= ci:
        return None
    key = (listing_id, ci.isoformat(), co.isoformat())
    hit = _avail_cache.get(key)
    if hit and (time.time() - hit[1]) < INTEL_CACHE_MIN * 60:
        return hit[0]
    last_night = co - timedelta(days=1)          # checkout night is not charged
    nights = (co - ci).days
    result = None
    try:
        data = api_get(f"/listings/{listing_id}/calendar",
                       params={"startDate": ci.isoformat(), "endDate": last_night.isoformat()})
        days = data.get("result", []) or []
        if days:
            available = all(int(d.get("isAvailable", 0) or 0) == 1 for d in days)
            prices = [d.get("price") for d in days
                      if isinstance(d.get("price"), (int, float)) and d.get("price") > 0]
            total = round(sum(prices)) if len(prices) == len(days) and prices else None
            avg = round(total / nights) if total else None
            result = {"available": available, "nights": nights, "total": total, "avg": avg}
    except Exception as e:
        print(f"availability fetch error ({listing_id}):", e)
        result = None
    _bounded_cache_put(_avail_cache, key, (result, time.time()), _AVAIL_CACHE_MAX)
    return result

def _unit_match_score(u, want):
    """Score a unit against a desired-criteria dict so we date-check the most-likely
    matches FIRST. `want` keys: beds, capacity, area, tags (set of normalized tags).
    Score is purely a sort key; nothing is filtered out."""
    s = 0
    if want.get("beds") and u.get("beds"):
        diff = abs(int(u["beds"]) - int(want["beds"]))
        s += max(0, 10 - diff * 4)            # exact = +10, off by 1 = +6, etc.
    if want.get("capacity") and u.get("capacity"):
        if int(u["capacity"]) >= int(want["capacity"]):
            s += 5
    if want.get("area"):
        a = (want["area"] or "").lower()
        if a and a in ((u.get("area") or "") + " " + (u.get("neighbourhood") or "")).lower():
            s += 6
    wtags = set(want.get("tags") or [])
    if wtags:
        match = wtags & set(u.get("tags") or [])
        s += 4 * len(match)
    if u.get("price"):
        s += 1     # prefer units we actually know the price of
    if u.get("link"):
        s += 1
    return s

def enrich_catalog_for_dates(checkin, checkout, exclude_id=None, want=None):
    """Build a catalog block annotated with live availability + real total for the
    guest's dates. With parallel calendar fetches (INTEL_PARALLEL workers) we can
    now check the FULL portfolio in ~3-5 s instead of ~30 s sequential, so the
    default cap (INTEL_MAX_CHECKS) is large enough to cover everything. `want`
    still orders the output by relevance for the guest."""
    if not _catalog_units or not checkin or not checkout:
        return _catalog_text
    want = want or {}
    ranked = sorted(
        [u for u in _catalog_units if u.get("id") and u["id"] != exclude_id],
        key=lambda u: -_unit_match_score(u, want),
    )
    targets = ranked[:INTEL_MAX_CHECKS]
    # ---- parallel calendar lookups (the old loop was sequential = ~30s for 69 units) ----
    info_map = {}
    if targets:
        try:
            with ThreadPoolExecutor(max_workers=INTEL_PARALLEL) as ex:
                futures = {ex.submit(unit_availability_price, u["id"], checkin, checkout): u["id"]
                           for u in targets}
                for fut in as_completed(futures):
                    lid = futures[fut]
                    try:
                        info_map[lid] = fut.result()
                    except Exception as e:
                        print(f"enrich parallel error ({lid}):", e)
                        info_map[lid] = None
        except Exception as e:
            print("enrich pool error:", e)
    lines = []
    for u in ranked:
        base = [u["name"]]
        if u.get("beds"):     base.append(f"{u['beds']} غرفة نوم")
        if u.get("baths"):    base.append(f"{u['baths']} حمام")
        if u.get("capacity"): base.append(f"يستوعب {u['capacity']}")
        loc = " · ".join(filter(None, [u.get("neighbourhood"), u.get("area")])).strip(" ·")
        if loc: base.append(loc)
        info = info_map.get(u["id"])
        if info and info.get("total") is not None:
            tag = "✅ متاحة لتواريخه" if info["available"] else "❌ غير متاحة لتواريخه"
            base.append(f"{tag} · {info['nights']} ليلة ≈ {info['total']} ر.س (متوسط {info['avg']}/ليلة)")
        elif info and info.get("available") is True:
            base.append("✅ متاحة لتواريخه")
        elif info and info.get("available") is False:
            base.append("❌ غير متاحة لتواريخه")
        elif u.get("price"):
            base.append(f"تبدأ من ~{u['price']} ر.س/الليلة")
        if u.get("tags"):  base.append("مرافق: " + ", ".join(u["tags"]))
        if u.get("link"):  base.append(u["link"])
        lines.append(" · ".join(base))
    return "\n".join(lines)[:8000]

def _criteria_from_text(text):
    """Heuristic: pull what the guest seems to want from a short inquiry.
    Returns {beds, capacity, area, tags}. All optional. Used to rank suggestions."""
    t = (text or "").lower()
    want = {}
    # bedrooms
    bed_words = {1: ["غرفة وحدة", "وحدة وحدة", "غرفة واحدة", "studio", "استوديو", "1br", "1 br", "1-bedroom", "غرفه واحده"],
                 2: ["غرفتين", "غرفتان", "2br", "2 br", "two bedroom", "two-bedroom"],
                 3: ["ثلاث غرف", "ثلاث غرفة", "3br", "3 br", "three bedroom", "three-bedroom"],
                 4: ["اربع غرف", "أربع غرف", "4br", "4 br", "four bedroom"]}
    for n, words in bed_words.items():
        if any(w in t for w in words):
            want["beds"] = n; break
    # capacity
    m = re.search(r"(?:عدد|كم|نحن|احنا|نكون)?\s*(\d{1,2})\s*(?:شخص|اشخاص|أشخاص|ضيف|ضيوف|persons?|people|guests?)", t)
    if m:
        try: want["capacity"] = int(m.group(1))
        except: pass
    # area hints (Riyadh neighbourhoods)
    for area in ["الملقا", "النرجس", "العارض", "النفل", "حطين", "الياسمين", "الربيع",
                 "قرطبة", "العقيق", "القيروان", "التعاون", "عرقه", "الماجدية", "الملقى",
                 "malqa", "narjis", "qurtuba", "yasmin", "rabie"]:
        if area in t:
            want["area"] = area; break
    # amenity tags
    tags = set()
    for tag, kws in _AMENITY_TAGS.items():
        if any(kw in t for kw in kws):
            tags.add(tag)
    if tags: want["tags"] = list(tags)
    return want

# PHRASES (not single words) that strongly indicate the guest is asking for a
# DIFFERENT unit — not asking about their current one. The previous bare-word list
# triggered on "متاح" inside "الواي فاي متاح؟" which caused the bot to launch into
# "tell me your dates and bedrooms" — completely wrong for an amenity question.
# These phrases require a unit-noun + intent so they only fire on real alt-asking.
_ALT_PHRASES = [
    # Arabic — explicit "another / alternative / different unit"
    "شقة ثاني", "شقه ثاني", "شقة ثانية", "شقه ثانيه",
    "وحدة ثاني", "وحده ثاني", "وحدة ثانية", "وحده ثانيه",
    "شقة بديل", "وحدة بديل", "شقه بديل", "وحده بديل",
    "بديل", "بدائل", "غيرها", "غيره",
    "خيار ثاني", "خيارات ثاني", "خيار آخر", "خيارات اخرى", "خيارات أخرى",
    # bigger/smaller/cheaper relative requests
    "اكبر من", "أكبر من", "ارخص من", "أرخص من", "اصغر من", "أصغر من",
    "ارخص شقة", "أرخص شقة", "اكبر شقة", "أكبر شقة",
    "في شقة اكبر", "فيه شقة اكبر", "وحدة اكبر",
    # explicit "do you have a unit" — must include a unit-noun
    "عندكم شقة", "عندكم شقه", "عندكم وحدة", "عندكم وحده",
    "عندكم استوديو", "فيه شقة", "فيه شقه", "فيه وحدة", "فيه وحده",
    "فيه استوديو", "تتوفر شقة", "تتوفر وحدة", "متوفر شقة", "متوفر وحدة",
    "متوفره شقه", "متوفره وحده",
    # multi-bedroom shopping (clearly looking around)
    "غرفتين", "ثلاث غرف", "ثلاث غرفه", "اربع غرف", "أربع غرف",
    "studio apartment", "two bedroom", "three bedroom", "2 bedroom", "3 bedroom",
    # English alternatives
    "another apartment", "another unit", "different apartment", "different unit",
    "alternative", "any other", "any apartment", "any unit",
    "bigger apartment", "smaller apartment", "cheaper apartment",
    "do you have a unit", "do you have an apartment", "do you have any",
    "do you have other", "do you have another",
]

def _is_asking_alternatives(text):
    """Strict catalog trigger. Returns True only on clear 'I want a different unit'
    phrasing — won't false-fire on amenity questions ('is the wifi available?')."""
    t = (text or "").lower()
    return any(p in t for p in _ALT_PHRASES)

# Late checkout requests. Mirror of early check-in. We check if the NEXT night
# (after the guest's departure) is free — if yes, late checkout is easy and team
# just needs to approve; if no, very tight (cleaner needs the room) and the
# escalation context warns the team.
_LATE_CHECKOUT_HINTS = [
    # Arabic
    "تشيك اوت متاخر", "تشيك آوت متأخر", "تشيك اوت متأخر", "خروج متاخر", "خروج متأخر",
    "اطلع متاخر", "أطلع متأخر", "اطلع بعد", "أطلع بعد", "اطلع الساعة", "أطلع الساعة",
    "اخر موعد للخروج", "آخر موعد للخروج", "اطول وقت", "أطول وقت",
    "ابقى لين", "أبقى لين", "نبقى لين", "ابقى الى", "أبقى إلى",
    "خروج متاخر شوي", "تأخير الخروج", "تاخير الخروج", "ممكن اتاخر بالخروج",
    "checkout بعد", "check out بعد", "تشيك أوت بعد", "تشيك اوت بعد",
    # English
    "late checkout", "late check-out", "late check out",
    "check out late", "checkout late", "leave later",
    "stay until", "stay till", "checkout extension", "extend checkout",
    "later checkout time",
]
def _is_late_checkout_request(text):
    t = (text or "").lower()
    return any(h in t for h in _LATE_CHECKOUT_HINTS)

def late_checkout_context(reservation_id, listing_id):
    """Mirror of early_checkin_context but for LATE checkout. Checks whether the
    night AFTER departure is occupied — if yes, late checkout is tight; if no,
    easy (team just needs to approve)."""
    if not reservation_id or not listing_id:
        return None
    try:
        rdata = api_get(f"/reservations/{reservation_id}")
        r = rdata.get("result") or {}
        departure = _parse_date(r.get("departureDate"))
        if not departure:
            return None
    except Exception as e:
        print(f"late_checkout_context departure fetch error: {e}")
        return None
    # The "next night" in calendar terms IS the departure date (hostaway represents
    # the departure date as the last night NOT booked by this reservation).
    next_night_occupied = not _calendar_night_free(listing_id, departure)
    return {
        "next_occupied": next_night_occupied,
        "departure": departure.isoformat(),
        "next_night": departure.isoformat(),
    }

# Door-code / access questions. When the guest asks where their code is, we must
# answer with REAL context (hours until check-in + signing status), not a template.
_CODE_Q_HINTS = [
    "الكود", "كود الدخول", "كود البوابة", "كود الباب", "رمز الدخول", "رمز البوابة",
    "كلمة المرور", "الباسوورد", "باسوورد", "باسورد",
    "ما وصلني الكود", "ما جاني الكود", "وين الكود", "وين رمز", "ما وصل الكود",
    "ما جاء الكود", "لم يصلني", "لم يصل الكود",
    "access code", "door code", "key code", "passcode", "passcode for",
    "where is the code", "where's the code", "code didn't arrive",
    "didn't receive the code", "haven't received the code", "no code",
    "didn't get the code", "haven't got the code",
]
def _is_code_question(text):
    t = (text or "").lower()
    return any(h in t for h in _CODE_Q_HINTS)

# Kept for back-compat; nothing reads it directly anymore (replaced by _is_asking_alternatives).
_ALT_HINTS = _ALT_PHRASES

# Hints that the guest is asking about price / total cost (to compute their real total).
_PRICE_HINTS = ["سعر", "السعر", "كم", "بكم", "كام", "تكلفة", "التكلفة", "المبلغ", "اجمالي", "الاجمالي",
                "الإجمالي", "price", "cost", "how much", "total", "rate", "nightly"]

def claude_draft(guest_name, unit, history_text, guide_url=None, confirmed=False,
                 dates=None, listing_id=None, reservation_id=None, profile_key=None):
    """Call Claude to draft a reply. Returns parsed dict or None on failure.
    If profile_key is provided AND the profile has prior stays / summaries,
    the bot greets them as a returning guest and references past context."""
    if not ANTHROPIC_API_KEY:
        print("assistant: ANTHROPIC_API_KEY not set")
        return None
    low = history_text.lower()
    status_line = ("Booking status: CONFIRMED" if confirmed
                   else "Booking status: NOT CONFIRMED (inquiry / pre-booking)")
    guide_line = (f"Arrival-guide link for this unit: {guide_url}"
                  if (guide_url and confirmed)
                  else "Arrival-guide link for this unit: NOT AVAILABLE / do not share")
    dates_line = (f"\nتواريخ الحجز: {dates[0]} إلى {dates[1]}"
                  if dates and dates[0] else "")
    facts = _knowledge_text.strip()
    facts_block = (f"معلومات معتمدة عن عوجا (استخدمها كمصدر الحقيقة وصحّح أي تعارض):\n{facts}\n\n"
                   if facts else "")
    # Apartment-scoped facts the team posted in #knowledge mentioning this unit
    # by name. These take PRECEDENCE over general facts and over learned summaries
    # when they conflict — they're explicit, team-authored, and unit-specific.
    if listing_id:
        apt_kn = _knowledge_apartment_facts.get(int(listing_id), [])
        if apt_kn:
            facts_block += (
                "حقائق محدّدة عن هذه الوحدة (كتبها الفريق في قناة المعرفة — "
                "اعتبرها أعلى مصدر ثقة عن هذه الشقة بالذات):\n"
                + "\n".join("- " + f for f in apt_kn[-20:]) + "\n\n"
            )
    # ---- learned summaries (distilled by Claude from past team-approved replies) ----
    # General first (cross-portfolio patterns), then apartment-specific if we have a lid.
    gen_learn = (_general_learnings.get("summary") or "").strip()
    if gen_learn:
        facts_block += ("دروس عامة استخلصها النظام من ردود الفريق السابقة (طبّقها بشكل افتراضي إلا "
                        "إذا تعارضت مع معلومة محدّدة لهذه الوحدة):\n" + gen_learn + "\n\n")
    if listing_id:
        apt_learn = (_apartment_learnings.get(int(listing_id)) or {}).get("summary", "").strip()
        if apt_learn:
            facts_block += ("دروس خاصة بهذه الوحدة بالذات (تجمعت من تفاعلات سابقة عليها — اعتبرها "
                            "مصدر حقيقة قوي عن هذه الشقة تحديداً):\n" + apt_learn + "\n\n")
    want_catalog = bool(_catalog_text) and _is_asking_alternatives(history_text)
    # ---- real pricing for the guest's OWN unit (when dates known + a price/availability question) ----
    own_price_line = ""
    if (dates and dates[0] and listing_id
            and (want_catalog or any(h in low for h in _PRICE_HINTS))):
        info = unit_availability_price(listing_id, dates[0], dates[1])
        if info and info.get("total") is not None:
            avail = "متاحة" if info["available"] else "غير متاحة حالياً"
            own_price_line = (
                f"\nالتسعير الفعلي لوحدة الضيف ({unit}) لتواريخه (قبل الضريبة ورسوم المنصة): "
                f"{info['nights']} ليالي ≈ {info['total']} ر.س (متوسط {info['avg']}/ليلة) · الوحدة {avail}.")
        elif info and info.get("available") is not None:
            own_price_line = (f"\nوحدة الضيف ({unit}) "
                              f"{'متاحة' if info['available'] else 'غير متاحة'} لتواريخه.")
    # ---- early check-in detection + pre-computation ----
    # Done HERE (in claude_draft) rather than in the system prompt because the bot
    # can't call functions on its own — we precompute prev-night occupancy +
    # alternative units the team could swap to, then inject the result as facts
    # the model quotes with certainty.
    early_block = ""
    if listing_id and _is_early_checkin_request(history_text):
        ec = early_checkin_context(reservation_id, listing_id) or \
             {"prev_occupied": None, "prev_night": "", "arrival": "", "alternatives": []}
        if ec.get("prev_occupied") is True:
            alts_txt = "\n".join([
                f"  • {a['name']}"
                + (f" · {a['beds']} غرفة" if a.get('beds') else "")
                + (f" · {a['area']}" if a.get('area') else "")
                + (f" · يبدأ من ~{a['price']} ر.س" if a.get('price') else "")
                + (f" · {a['link']}" if a.get('link') else "")
                for a in ec["alternatives"]
            ]) or "  (ما لقيت بدائل واضحة قريبة من معايير وحدته)"
            early_block = (
                f"\n\nطلب تشيك-إن مبكر — السياق المحسوب مسبقاً:\n"
                f"- الليلة السابقة لوصول الضيف ({ec.get('prev_night')}) محجوزة في وحدته الحالية، "
                f"فالتشيك-إن المبكر **غير ممكن في وحدته الأصلية**.\n"
                f"- وحدات بديلة كانت ليلتها السابقة + ليلة الوصول كلاهما فاضية (يمكن تحويله إليها بموافقة الفريق):\n{alts_txt}\n\n"
                f"خطوات الرد المطلوبة:\n"
                f"1) أخبر الضيف بصراحة إن وحدته الحالية الليلة قبل وصوله محجوزة، فالتشيك-إن المبكر فيها غير ممكن.\n"
                f"2) اعرض عليه خيار التحويل لوحدة بديلة (لو فيه بدائل فوق): اسأله إذا يبيك تتأكد من توفّر "
                f"وحدة ثانية مناسبة للدخول المبكر، ولو قال نعم اسأله المعايير المهمة له (كم غرفة، الميزانية، حي معيّن).\n"
                f"3) action='reply' لأن الموافقة النهائية على التحويل تحتاج قسم المختص — قول له إن الفريق "
                f"بيتأكد ويرد عليه.\n"
                f"4) لا تَعِد بشيء قبل موافقة الفريق."
            )
        elif ec.get("prev_occupied") is False:
            early_block = (
                f"\n\nطلب تشيك-إن مبكر — السياق المحسوب مسبقاً:\n"
                f"- الليلة السابقة لوصوله ({ec.get('prev_night')}) **فاضية** في وحدته، فالتشيك-إن المبكر ممكن "
                f"من ناحية التوفّر.\n"
                f"- لكن التأكيد النهائي يحتاج موافقة قسم المختص.\n\n"
                f"خطوات الرد: action='reply'. قل للضيف إن طلبه ممكن من ناحية التوفّر، وإن الفريق بيراجع "
                f"الموافقة النهائية ويتواصل معه بأقرب وقت. لا تؤكّد له ساعة دخول بعينها قبل موافقة الفريق."
            )

    # ---- alternatives: use a live-availability catalog when we know the dates ----
    want = _criteria_from_text(history_text) if want_catalog else {}
    has_dates = bool(dates and dates[0] and dates[1])
    knows_what = bool(want.get("beds") or want.get("capacity") or want.get("area") or want.get("tags"))
    must_ask = want_catalog and (not has_dates or not knows_what)

    if want_catalog and has_dates:
        catalog_data = enrich_catalog_for_dates(dates[0], dates[1], exclude_id=listing_id, want=want)
        avail_note = ("- التواريخ معروفة. القائمة فوق مرتّبة حسب الأقرب لطلب الضيف، وتبيّن التوفّر الفعلي "
                      "✅/❌ والإجمالي الحقيقي لتواريخه. **اقترح فقط الوحدات ✅ المتاحة** واذكر الإجمالي "
                      "والمتوسط/الليلة كما هو. لا تقترح وحدة ❌.\n")
    else:
        catalog_data = _catalog_text
        avail_note = ("- إنت ما تعرف التوفّر المباشر لتواريخه — اعرض الخيارات ووجّهه يتأكد ويحجز من رابط "
                      "Airbnb. **السؤال عن التوفّر مو سبب للتصعيد إطلاقاً**.\n")

    # Build a tiny "what we already know about this guest" line so the bot doesn't re-ask.
    known_bits = []
    if has_dates: known_bits.append(f"التواريخ: {dates[0]} → {dates[1]}")
    if want.get("beds"): known_bits.append(f"عدد الغرف المطلوب: {want['beds']}")
    if want.get("capacity"): known_bits.append(f"عدد الضيوف: {want['capacity']}")
    if want.get("area"): known_bits.append(f"المنطقة المفضّلة: {want['area']}")
    if want.get("tags"): known_bits.append("مرافق مطلوبة: " + ", ".join(want["tags"]))
    known_line = ("معلومات الضيف المستخلصة من المحادثة:\n- " + "\n- ".join(known_bits) + "\n\n") if known_bits else ""

    # When the inquiry is short and we don't yet know dates/beds/capacity/area,
    # force the bot to ASK first instead of dumping a catalog.
    ask_block = ""
    if must_ask:
        missing = []
        if not has_dates: missing.append("التواريخ (الوصول والمغادرة)")
        if not want.get("beds") and not want.get("capacity"): missing.append("عدد الضيوف أو عدد الغرف")
        if not want.get("area"): missing.append("الحي/المنطقة المفضّلة (أو لا يهم)")
        ask_block = (
            "⚠️ المهمّة الأولى الآن: قبل ما تقترح أي وحدة، اسأل الضيف عن: "
            + " · ".join(missing) + ". "
            "اسأل بأسلوب لطيف ومختصر (سؤال أو سؤالين بحد أقصى)، ولا تقترح وحدات بعد لأن "
            "أي اقتراح بدون هالمعلومات بيكون عشوائي. لو الضيف قال 'لا يهم' أو 'أي شي' لشي منهم، "
            "اعتبره معروف وكمّل بالاقتراح. action لازم يكون 'reply' (يراجعه إنسان).\n\n"
        )

    catalog_block = (
        known_line + ask_block +
        "قائمة وحدات عوجا للاقتراح:\n" + catalog_data + "\n\n"
        "تعليمات الاقتراح (لما تكون جاهز):\n"
        "- اقترح **٢-٣ خيارات بالضبط**، أفضل ما يطابق طلب الضيف.\n"
        "- لكل خيار اذكر بالترتيب: الاسم · عدد الغرف · عدد الحمامات · سعة الضيوف · المنطقة "
        "(الحي + المدينة) · الإجمالي الحقيقي لتواريخه (لا تذكر فقط 'تبدأ من' لما يكون الإجمالي معروف) "
        "· رابط Airbnb للحجز المباشر. لا تخترع رابط أو رقم.\n"
        "- لو فيه مرافق مطلوبة (مسبح، باركن، تدخين، بلكون...) أكّد أيها متوفر في كل خيار من قائمة "
        "'مرافق:' المعطاة لك. لو غير متأكد من ميزة، قل بصراحة 'بتأكد من الفريق' وخلها 'reply'.\n"
        "- لو ما فيه مطابق ١٠٠٪، اقترح أقرب خيار ووضّح الفروقات بصراحة (مثلاً: 'الأقرب لطلبك "
        "غرفتين بدل ثلاث').\n"
        + avail_note +
        "- دائماً ختام: 'الأسعار تقريبية، قبل الضريبة ورسوم المنصة. التوفّر النهائي يتأكد من رابط "
        "Airbnb عند الحجز.'\n\n"
        ) if want_catalog else ""
    # ---- Late-checkout: check whether the next night is occupied so the bot
    # tells the team whether it's an easy yes or a tight ask.
    late_block = ""
    if listing_id and _is_late_checkout_request(history_text):
        lc = late_checkout_context(reservation_id, listing_id) or {}
        if lc.get("next_occupied") is True:
            late_block = (
                f"\n\nطلب تشيك-آوت متأخّر — السياق المحسوب:\n"
                f"- الليلة بعد مغادرة الضيف ({lc.get('next_night')}) محجوزة في نفس الوحدة، فالتأخير "
                f"مقيّد بساعات قليلة (المنظّف يحتاج وقت قبل دخول الضيف الجاي).\n\n"
                f"كيف ترد: action='reply'. قل للضيف إن الطلب وارد بس مقيّد، الفريق بيحاول يعطيه "
                f"أطول وقت ممكن (عادةً ساعة-ساعتين بعد الموعد الأصلي) وبيتأكد ويرد. لا تَعِد بساعة "
                f"بعينها قبل موافقة الفريق."
            )
        elif lc.get("next_occupied") is False:
            late_block = (
                f"\n\nطلب تشيك-آوت متأخّر — السياق المحسوب:\n"
                f"- الليلة بعد مغادرة الضيف ({lc.get('next_night')}) **فاضية** في وحدته، "
                f"فالتأخير ممكن لساعات طويلة (الجدول مفتوح).\n\n"
                f"كيف ترد: action='reply'. قل للضيف إن الطلب ممكن من ناحية الجدول، والفريق "
                f"بيراجع ويؤكّد ساعة الخروج. لا تؤكّد ساعة بعينها قبل موافقة الفريق."
            )

    # ---- Code/access question: inject real-time context (hours until check-in,
    # signing status) so the bot stops giving generic "code arrives 5 days before"
    # answers when check-in is today.
    code_block = ""
    if reservation_id and _is_code_question(history_text):
        try:
            rdata = api_get(f"/reservations/{reservation_id}")
            r = rdata.get("result") or {}
            arrival = _parse_date(r.get("arrivalDate"))
            signed = _is_agreement_signed(r)
            if arrival:
                now = datetime.now(TZ)
                hour = parse_hour(r.get("checkInTime"), 15)
                checkin_dt = datetime(arrival.year, arrival.month, arrival.day,
                                      min(hour, 23), 0, tzinfo=TZ)
                hrs = (checkin_dt - now).total_seconds() / 3600.0
                if hrs < -2:
                    when = "موعد التشيك-إن **مضى فعلاً** — الضيف كان لازم يكون داخل الوحدة"
                elif hrs < 0:
                    when = "موعد التشيك-إن **حلّ الحين**"
                elif hrs < 6:
                    when = f"التشيك-إن **اليوم بعد {hrs:.1f} ساعة فقط** — قريب جداً"
                elif hrs < 24:
                    when = f"التشيك-إن **اليوم** (بعد {int(hrs)} ساعة)"
                elif hrs < 48:
                    when = "التشيك-إن **بكرة**"
                elif hrs < 168:
                    when = f"التشيك-إن بعد **{int(hrs/24)} يوم**"
                else:
                    when = f"التشيك-إن بعد **{int(hrs/24)} يوم** — لسه فيه وقت طويل"
                code_block = (
                    f"\n\n⚠ الضيف يسأل عن كود الدخول. الواقع المحسوب:\n"
                    f"- {when}.\n"
                    f"- العقد {'موقّع ✅' if signed else 'غير موقّع ❌'}.\n\n"
                    f"كيف ترد (حسب هذي البيانات بالضبط):\n"
                )
                if not signed and hrs < 48:
                    code_block += (
                        "- العقد غير موقّع، فالكود مايوصل إلا بعد التوقيع. اشرح له ذلك بلطف، "
                        "وقول إن الكود يُرسَل تلقائياً فور توقيعه. لو الرابط في رسالة سابقة "
                        "اطلب منه يفتحها، أو قل إن الفريق راح يعيد إرساله. action='reply'."
                    )
                elif not signed:
                    code_block += (
                        "- العقد لسه غير موقّع لكن التشيك-إن بعيد. ذكّره بلطف إنه يحتاج "
                        "يوقّع قبل الدخول، والكود بيوصله بعد التوقيع."
                    )
                elif hrs < 0:
                    code_block += (
                        "- العقد موقّع والوقت مضى. الكود كان لازم وصله. اطلب منه يفحص "
                        "البريد بما فيها Spam، ورسائل Airbnb/Hostaway. لو ما لقاه action='reply' "
                        "وقل إن الفريق يتحقّق فوراً."
                    )
                elif hrs < 6:
                    code_block += (
                        "- العقد موقّع والتشيك-إن قريب جداً. اطلب منه يفحص البريد + Spam، "
                        "وإذا ما وصله بعد، action='reply' عشان الفريق يتأكد."
                    )
                elif hrs < 48:
                    code_block += (
                        "- العقد موقّع، الكود يوصل عادةً قبل التشيك-إن ببضع ساعات. طمّنه "
                        "إنه بيوصله في وقته قبل دخوله. **لا تقول له رقم ثابت مثل '٥ أيام'**."
                    )
                else:
                    code_block += (
                        "- التشيك-إن بعيد لسه، فالكود ما يُرسل الآن. طمّنه إنه بيوصله قبل "
                        "دخوله ببضع ساعات بإذن الله. **لا تخترع رقم ثابت — لا تقول '٥ أيام قبل' "
                        "ولا '٤٨ ساعة قبل' — قول 'قبل دخولك ببضع ساعات' فقط**."
                    )
        except Exception as e:
            print(f"code_question_context error: {e}")

    # ---- Returning-guest / VIP context ----
    # If the bot has talked to this person before AND we have prior context,
    # surface it so the reply feels personal ("welcome back", references a
    # past stay, etc.). Does nothing for first-time guests.
    profile_block = ""
    if profile_key and profile_key in _guest_profiles:
        p = _guest_profiles[profile_key]
        stays = len(p.get("reservations", []))
        if stays >= 2 or p.get("vip"):
            past = p.get("reservations", [])[-3:]
            past_summary = "; ".join(
                (r.get("unit","") + " " + (r.get("checkin","") or "")) for r in past if r.get("unit")
            ) or "—"
            summaries = (p.get("summaries", []) or [])[-3:]
            sum_text = "\n".join("- " + (s.get("text","") or "")[:280] for s in summaries if s.get("text"))
            vip_line = "⭐ ضيف مميّز (VIP) — سبق له " + str(stays) + " إقامة معنا." if p.get("vip") else "ضيف عائد — سبق له " + str(stays) + " إقامة."
            profile_block = (
                f"\n\nملف الضيف (لا تذكر أنه ملف — استخدمه طبيعياً):\n"
                f"- {vip_line}\n"
                f"- إقاماته السابقة: {past_summary}\n"
                + (f"- ملخصات محادثات سابقة:\n{sum_text}\n" if sum_text else "")
                + "- ابدأ ردك بترحيب يعكس عودته (مثل 'حياك الله مرة ثانية' / 'نوّرتنا تاني'). "
                + "لا تطلب منه معلومات يفترض أنها معروفة لنا من الإقامات السابقة. "
                + "لو ذكر تفضيلاً سابقاً في الملخصات، استخدمه."
            )

    # ---- Hard unit-context guard: the bot has been asking 'which apartment?' even
    # when the inquiry is clearly about the booked unit. Reinforce that the Unit
    # field below IS the unit they're inquiring about, unless they explicitly
    # asked for an alternative.
    unit_guard = (
        f"⚠ السياق الثابت: الضيف يسأل عن وحدته **{unit}** (هذي هي الشقة المرتبطة بحجزه/استفساره). "
        f"لا تسأله أبداً 'أي شقة تقصد' أو 'أي وحدة'. الاقتراحات البديلة تظهر فقط لما يطلبها صراحةً.\n\n"
        if unit else ""
    )
    user = (f"{unit_guard}{facts_block}{catalog_block}{profile_block}Guest name: {guest_name}\nUnit: {unit}\n"
            f"{status_line}\n{guide_line}{dates_line}{own_price_line}{early_block}{late_block}{code_block}\n\n"
            f"Conversation so far (oldest first, last line is the guest's new message):\n"
            f"{history_text}\n\nDraft your reply as the JSON object.")
    # Always use the premium model for guest drafts — Haiku produces too many
    # context-blind replies (the "code arrives 5 days before" template being the
    # canonical example). Can be overridden via GUEST_DRAFT_MODEL env var.
    model = GUEST_DRAFT_MODEL
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": ANTHROPIC_API_KEY, "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": model, "max_tokens": 700, "system": ASSISTANT_RULES,
                  "messages": [{"role": "user", "content": user}]},
            timeout=60,
        )
        r.raise_for_status()
        blocks = r.json().get("content", []) or []
        text = "".join(b.get("text", "") for b in blocks if b.get("type") == "text").strip()
        text = text.replace("```json", "").replace("```", "").strip()
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            # The model sometimes wraps the JSON in a sentence ("Here's the reply: {...}").
            # Pull out the first balanced {...} block and try again instead of dropping the
            # draft entirely.
            m = re.search(r"\{.*\}", text, re.DOTALL)
            if m:
                try:
                    return json.loads(m.group(0))
                except Exception:
                    pass
            print("claude_draft: could not parse JSON · raw=", text[:300])
            return None
    except Exception as e:
        print("claude_draft error:", e)
        return None

def claude_text(system, user, max_tokens=600, model=None):
    """Plain-text Claude call (no JSON). Returns the text or None."""
    if not ANTHROPIC_API_KEY:
        return None
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": ANTHROPIC_API_KEY, "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": model or CLAUDE_MODEL, "max_tokens": max_tokens, "system": system,
                  "messages": [{"role": "user", "content": user}]},
            timeout=60,
        )
        r.raise_for_status()
        blocks = r.json().get("content", []) or []
        return "".join(b.get("text", "") for b in blocks if b.get("type") == "text").strip() or None
    except Exception as e:
        print("claude_text error:", e)
        return None

# Reusable dialect lock — appended to every Arabic-generating system prompt so
# we get consistent Najdi/white-Saudi output across all paths (drafts, distill,
# acks, manager scripts). Centralising it keeps the rules in one place.
_DIALECT_LOCK = (
    "\n\n⛔ قيد لهجة صارم — اكتب عربي سعودي/نجدي فقط. أي كلمة من القوائم التالية = خطأ، استبدلها:\n"
    "【مصرية】 دلوقتي→الحين · إيه→إيش/وش · إزيك→كيفك · إزاي→كيف · إيوه→إي · مش→مو/مب · "
    "عايز/عاوز→أبي/أبغى · كده→كذا · فين→وين · إمتى→متى · ليه→ليش · بتاع/بتاعك→حق/حقك · "
    "مفيش→ما فيه · معلش→ما يخالف · حاجة→شي · خالص→أبداً · أوي/قوي→مرة · هـ+فعل (هتعمل)→بـ+فعل (بتسوي) · "
    "بقى→صار · معايا/معاك→معي/معك · ربنا→الله · حضرتك→حياك · طب→طيب\n"
    "【شامية】 شو→وش/إيش · بدك/بدّك/بدي/بدنا→تبي/تبغى/أبي/نبي · هلق/هلأ→الحين · هيك→كذا · "
    "كتير→كثير/مرة · منيح→زين/كويس · لشو→ليش · شلونك→كيفك · متل/متلك→مثل/مثلك · "
    "كرمالك→عشانك · ولاي→والله · عنجد→فعلاً · منشان→عشان · بلكي→يمكن · لوين→وين · "
    "بكير→بدري · رح+فعل→بـ+فعل · هاد/هادي→هذا/هذي · لساتو→لسا\n"
    "【عراقية】 شكو→وش فيه · ماكو→ما فيه · اكو→فيه · هسة→الحين · شوكت→متى · هواية→كثير · "
    "خوش→زين · شلون→كيف\n"
    "【مغربية】 كيفاش→كيف · واش→إيش · بزاف→كثير · غادي→بـ/راح · كاين→فيه · ديالي/ديالك→حقي/حقك · "
    "نتا→أنت · واخا→تمام\n"
    "✅ سعودي طبيعي: حياك الله · هلا والله · أبشر · تم · ما يخالف · ولا يهمك · يعطيك العافية · "
    "إن شاء الله · عشان/علشان · بس · طيب · ماشي · الحين · بدري · بكرة · وش/إيش · ليش · متى · "
    "وين · كيف · تبي/تبغى · أبي/أبغى · تكفى · حقي/حقك · مو/مب · إي · يا غالي · يا طويل العمر\n"
    "🔍 قبل ما تكتب: \"لو قرأها سعودي في الرياض يحسّها طبيعية ولا يكتشف لهجة ثانية؟\" لو لا، أعد الصياغة."
)

def claude_escalation_ack(guest, unit, history, guest_text):
    """An empathetic, problem-specific holding message for a repeat escalation."""
    sys = ("أنت تكتب رسالة طمأنة قصيرة لضيف في عوجا تصعّد موضوعه مرة ثانية وهو لا يزال ينتظر أو منزعج. "
           "اكتب بأسلوب سعودي نجدي دافئ وراقٍ: اعترف بمشكلته تحديداً، تفهّم شعوره، اعتذر بصدق، "
           "وطمّنه إن الفريق المختص يشتغل على موضوعه الحين وبيتواصل معه قريب جداً. "
           "لا تكتبها كأنها قالب آلي مكرر. لو الضيف يكتب إنجليزي رد بالإنجليزي. اكتب نص الرسالة فقط."
           + _DIALECT_LOCK)
    user = (f"الوحدة: {unit}\nالضيف: {guest}\nالمحادثة:\n{history}\n\n"
            f"آخر رسالة من الضيف: {guest_text}")
    return claude_text(sys, user, 500, model=CLAUDE_MODEL_PREMIUM) or claude_text(sys, user, 500)

def claude_manager_script(guest, unit, history, guest_text, reason, manager_name):
    """A ready-to-send reply in the owner's warm/charismatic voice for whoever claims."""
    sys = (f"أنت تكتب رسالة جاهزة بينسخها {manager_name} (مدير العقار في عوجا) ويرسلها للضيف مباشرة. "
           "الأسلوب: سعودي نجدي دافئ، واثق، وكاريزما عالية مثل صاحب البيت اللي يهتم بضيفه شخصياً. "
           f"الرسالة لازم: تحيي الضيف بحرارة، تعرّف إن {manager_name} مدير العقار وتولّى الموضوع بنفسه، "
           "تذكر إن الشباب (الفريق) بلّغوه عن مشكلته، تلخّص مشكلته باختصار عشان يحس إنه مسموع، "
           "تتفهّم وتعتذر بصدق، وتطمّنه بعبارات دافئة (مثل: أبشر بالي يسرّ خاطرك يا طويل العمر) إن الموضوع بينحل. "
           "طبيعية مهب قالب جامد. لو الضيف يكتب إنجليزي رد بالإنجليزي. اكتب نص الرسالة فقط بدون أي شرح."
           + _DIALECT_LOCK)
    user = (f"الوحدة: {unit}\nالضيف: {guest}\nسبب التصعيد: {reason}\nالمحادثة:\n{history}\n\n"
            f"آخر رسالة من الضيف: {guest_text}")
    return claude_text(sys, user, 700, model=CLAUDE_MODEL_PREMIUM) or claude_text(sys, user, 700)

# ========================================================================
# Self-learning helpers
# ========================================================================
def _text_diff_ratio(a, b):
    """0.0 = identical, 1.0 = totally different. Cheap approximate via SequenceMatcher."""
    if not a and not b:
        return 0.0
    if not a or not b:
        return 1.0
    try:
        from difflib import SequenceMatcher
        return 1.0 - SequenceMatcher(None, a, b).ratio()
    except Exception:
        return 0.0

def record_learning(item, original_draft, final_reply, via, approver=None):
    """Append one send/edit event to _learning_log.
       `item` is the message item dict (guest, unit, listing_id, conversation_id, guest_text).
       `via` is one of: discord_send | discord_edit | dashboard_send | auto | escalation_ack."""
    try:
        if not item:
            return
        diff = _text_diff_ratio((original_draft or "").strip(), (final_reply or "").strip())
        entry = {
            "ts": datetime.now(TZ).isoformat(timespec="seconds"),
            "conversation_id": item.get("conversation_id"),
            "listing_id": item.get("listing_id"),
            "unit": (item.get("unit") or "")[:80],
            "guest": (item.get("guest") or "")[:60],
            "guest_question": (item.get("guest_text") or "")[:900],
            "bot_draft": (original_draft or "")[:1400],
            "final_reply": (final_reply or "")[:1400],
            "diff_ratio": round(diff, 2),
            "was_edited": diff >= 0.20,                # human reshaped the bot >20%
            "via": via,
            "approver": approver or "",
        }
        _learning_log.append(entry)
        # ---- daily metrics ----
        metric_bump("replies_total")
        if via == "auto":           metric_bump("replies_auto")
        elif via == "discord_send": metric_bump("replies_manual")
        elif via == "discord_edit": metric_bump("replies_edited")
        elif via == "dashboard_send":
            metric_bump("replies_dashboard")
            if entry["was_edited"]: metric_bump("replies_edited")
            else:                   metric_bump("replies_manual")
        metric_record_apartment(item.get("unit"))
    except Exception as e:
        print("record_learning error:", e)

_DISTILL_SYSTEM = (
    "أنت تحلّل آخر تفاعلات بوت إدارة عقارات «عوجا» في الرياض لاستخراج درس مكثّف يساعد البوت "
    "يجاوب بشكل أصحّ في المستقبل. كل تفاعل يحتوي على: سؤال الضيف، مسوّدة البوت الأصلية، "
    "والرد النهائي الذي أرسله الفريق (مع إشارة إن كان مُعدّل أو مُرسل كما هو).\n\n"
    "ركّز على:\n"
    "- حقائق متكرّرة عن الوحدة/الشقة بالتحديد (موقع، مدخل، رقم العمارة، باركن، WiFi، أي خصوصية)\n"
    "- الأسئلة الشائعة وكيف يردّ عليها الفريق فعلياً\n"
    "- الحالات التي عدّل فيها الفريق المسوّدة بشكل كبير = البوت كان يخطئ، استخرج الدرس\n"
    "- الأمور المحلية (مطاعم، مسافة المطار، مولات قريبة) لو ذُكرت أكثر من مرة\n\n"
    "تجاهل أسماء الضيوف وأي معلومات شخصية. اكتب الملخص بعربي **سعودي/نجدي فقط** — لا تستخدم "
    "أبداً كلمات شامية أو مصرية مثل: شو، بدك، دلوقتي، عايز، إزيك، كده، فين، بتاع، ايه، معلش، "
    "هلق. استبدلها بـ: وش/إيش، تبي/تبغى، الحين، أبي، كيفك، كذا، وين، حق/مال، إي، ما يخالف. "
    "هذا الملخص بيُحقن في كل مسوّدة جاية، فلو تسرّبت لهجة غير سعودية هنا راح تنسخها كل ردود البوت. "
    "اكتب مرتّب بـbullet points تحت عناوين قصيرة. أقصى ٥٠٠ كلمة. لو فيه ملخص سابق، ابنِ عليه "
    "(احتفظ بما لا يزال صحيحاً، أضف الجديد، احذف ما يناقضه التفاعل الأخير)."
    + _DIALECT_LOCK
)

def _format_entries_for_distill(entries):
    """Compact textual rendering of recent send events for the distillation prompt."""
    out = []
    for i, e in enumerate(entries, 1):
        tag = "✏️ مُعدّل" if e.get("was_edited") else "✅ كما هو"
        out.append(
            f"[{i}] {tag}\n"
            f"سؤال الضيف: {e.get('guest_question','')[:500]}\n"
            f"مسوّدة البوت: {(e.get('bot_draft') or '')[:500]}\n"
            f"الرد النهائي: {(e.get('final_reply') or '')[:500]}"
        )
    return "\n\n".join(out)

def _distill_apartment(lid, entries, prior_summary):
    """Ask Claude to distill recent entries for one apartment. Returns summary string or None."""
    if not entries:
        return None
    name = entries[-1].get("unit") or f"unit-{lid}"
    sample = _format_entries_for_distill(entries[-LEARNING_SAMPLE_PER_APT:])
    user = (
        f"الوحدة: {name}\n\n"
        f"الملخص السابق:\n{prior_summary or '(لا يوجد بعد)'}\n\n"
        f"آخر التفاعلات لهذه الوحدة:\n{sample}\n\n"
        f"اكتب الملخص المحدّث لهذه الوحدة فقط."
    )
    return claude_text(_DISTILL_SYSTEM, user, max_tokens=900, model=CLAUDE_MODEL_PREMIUM) \
        or claude_text(_DISTILL_SYSTEM, user, max_tokens=900)

def _distill_general(entries, prior_summary):
    """Distill cross-apartment learnings (tone, common questions, escalation patterns)."""
    if not entries:
        return None
    sample = _format_entries_for_distill(entries[-60:])
    user = (
        f"تفاعلات حديثة من عدة وحدات (للأنماط العامة عبر المحفظة):\n\n"
        f"الملخص العام السابق:\n{prior_summary or '(لا يوجد بعد)'}\n\n"
        f"آخر التفاعلات:\n{sample}\n\n"
        f"اكتب الملخص العام المحدّث (أنماط لغوية، أسئلة شائعة، طريقة الفريق في الرد، "
        f"تجاهل التفاصيل الخاصة بشقة واحدة فقط)."
    )
    return claude_text(_DISTILL_SYSTEM, user, max_tokens=900, model=CLAUDE_MODEL_PREMIUM) \
        or claude_text(_DISTILL_SYSTEM, user, max_tokens=900)

def bootstrap_learnings_from_history(limit_conversations=300, min_pairs_per_apt=3):
    """One-shot: walk the most-recent N Hostaway conversations, extract (guest_question
    → team_reply) pairs grouped by listing, then distill a per-apartment summary for
    each apartment that has enough material. Designed to seed the assistant with the
    team's historical voice and unit-specific knowledge — so it starts at ~20% instead
    of from zero. Intended to run as a one-time background job, not in a loop.

    Returns {conversations_scanned, pairs_extracted, apartments_distilled}."""
    # ---- Step 1: pull recent conversations (paginated) ----
    convos, offset, page = [], 0, 100
    while len(convos) < limit_conversations:
        try:
            data = api_get("/conversations",
                           params={"limit": page, "offset": offset, "includeResources": 1})
        except Exception as e:
            print(f"bootstrap: /conversations fetch error at offset {offset}: {e}")
            break
        batch = data.get("result", []) or []
        if not batch:
            break
        convos.extend(batch)
        if len(batch) < page:
            break
        offset += page
    convos = convos[:limit_conversations]
    print(f"bootstrap: pulled {len(convos)} conversation(s)")
    log_event("guest", f"بدأ التعلّم التاريخي · {len(convos)} محادثة")

    # ---- Step 2: build (question → team-reply) pairs grouped by apartment ----
    listings = get_listings_map()
    by_apt = defaultdict(list)
    scanned, pairs = 0, 0
    for c in convos:
        cid = c.get("id")
        lid = c.get("listingMapId")
        if not cid or not lid:
            continue
        try:
            data = api_get(f"/conversations/{cid}/messages")
            msgs = sorted(data.get("result", []) or [], key=_msg_sort_key)
        except Exception as e:
            print(f"bootstrap: convo {cid} fetch error: {e}")
            continue
        scanned += 1
        unit_name = listings.get(lid) or c.get("listingName") or f"unit-{lid}"
        # Pair each inbound message with the NEXT non-automated outbound reply
        for i, m in enumerate(msgs):
            if not _msg_is_inbound(m):
                continue
            qtext = (m.get("body") or "").strip()
            if not qtext or len(qtext) < 6:
                continue
            for j in range(i + 1, len(msgs)):
                m2 = msgs[j]
                if _msg_is_inbound(m2):
                    break                                      # unanswered before next inbound
                body2 = (m2.get("body") or "").strip()
                if not body2 or _looks_automated(body2) or len(body2) < 10:
                    continue
                by_apt[int(lid)].append({
                    "ts": _msg_time(m2),
                    "conversation_id": cid,
                    "listing_id": int(lid),
                    "unit": unit_name,
                    "guest_question": qtext[:900],
                    "bot_draft": "",                            # no bot draft existed historically
                    "final_reply": body2[:1400],
                    "diff_ratio": 1.0,
                    "was_edited": True,                         # team-authored by definition
                    "via": "history",
                    "approver": "(historical)",
                })
                pairs += 1
                break
    print(f"bootstrap: extracted {pairs} (question, reply) pair(s) across {len(by_apt)} apt(s)")

    # ---- Step 3: distill per apartment + general ----
    distilled = 0
    for lid, entries in by_apt.items():
        if len(entries) < min_pairs_per_apt:
            continue
        existing = _apartment_learnings.get(lid, {})
        summary = _distill_apartment(lid, entries, existing.get("summary", ""))
        if summary and summary.strip():
            _apartment_learnings[lid] = {
                "summary": summary.strip(),
                "last_distilled": time.time(),
                "examples_count": len(entries),
                "unit": entries[-1].get("unit", ""),
            }
            distilled += 1
            print(f"bootstrap: distilled {entries[-1].get('unit','')} ({len(entries)} pairs)")
    # general summary across a sampled subset (cap to keep prompt size sane)
    all_sample = []
    for entries in by_apt.values():
        all_sample.extend(entries[-15:])
    if all_sample:
        g = _distill_general(all_sample, _general_learnings.get("summary", ""))
        if g and g.strip():
            _general_learnings["summary"] = g.strip()
            _general_learnings["last_distilled"] = time.time()
            _general_learnings["examples_count"] = len(all_sample)
            print(f"bootstrap: distilled general ({len(all_sample)} pairs)")
    print(f"bootstrap: COMPLETE · {distilled} apartment(s) distilled")
    log_event("guest", f"اكتمل التعلّم التاريخي · {distilled} شقة من {scanned} محادثة")
    return {"conversations_scanned": scanned, "pairs_extracted": pairs,
            "apartments_distilled": distilled}

def distill_learnings():
    """Walk _learning_log, distill per-apartment + general summaries when there's
       enough new material. Designed to be cheap when nothing changed."""
    if not _learning_log:
        return
    # group by apartment
    by_apt = defaultdict(list)
    for e in list(_learning_log):
        lid = e.get("listing_id")
        if lid:
            by_apt[int(lid)].append(e)
    # per-apartment
    distilled = 0
    for lid, entries in by_apt.items():
        existing = _apartment_learnings.get(lid, {})
        prior_count = existing.get("examples_count", 0)
        if len(entries) - prior_count < LEARNING_MIN_NEW_EXAMPLES and existing.get("summary"):
            continue                                    # not enough new material to re-spend tokens
        summary = _distill_apartment(lid, entries, existing.get("summary", ""))
        if summary and summary.strip():
            _apartment_learnings[lid] = {
                "summary": summary.strip(),
                "last_distilled": time.time(),
                "examples_count": len(entries),
                "unit": entries[-1].get("unit", ""),
            }
            distilled += 1
            log_event("guest", f"تعلّم محدّث · {entries[-1].get('unit','')} ({len(entries)} مثال)")
    # general
    all_entries = list(_learning_log)
    prior_g = _general_learnings.get("examples_count", 0)
    if len(all_entries) - prior_g >= LEARNING_MIN_NEW_EXAMPLES or not _general_learnings.get("summary"):
        g_summary = _distill_general(all_entries, _general_learnings.get("summary", ""))
        if g_summary and g_summary.strip():
            _general_learnings["summary"] = g_summary.strip()
            _general_learnings["last_distilled"] = time.time()
            _general_learnings["examples_count"] = len(all_entries)
            distilled += 1
    if distilled:
        print(f"learnings: distilled {distilled} summary block(s)")

def _conv_to_item(c, listings, seen, debug=False):
    """Turn ONE conversation object into a new-guest-message item, or None.
    Shared by the full scan and the single-conversation webhook path."""
    cid = c.get("id")
    if not cid:
        return None
    try:
        md = api_get(f"/conversations/{cid}/messages")
        msgs = md.get("result", []) or []
    except Exception as e:
        if debug:
            print(f"  conv {cid}: messages fetch error: {e}")
        return None
    if not msgs:
        return None
    msgs = sorted(msgs, key=_msg_sort_key)
    # the guest's most recent (inbound) message
    guest_idx = next((i for i in range(len(msgs) - 1, -1, -1)
                      if _msg_is_inbound(msgs[i])), None)
    if guest_idx is None:
        return None                                     # guest never messaged
    guest_msg = msgs[guest_idx]
    mid = str(guest_msg.get("id"))
    after = msgs[guest_idx + 1:]                         # anything sent after the guest spoke
    # "answered" = a real reply exists after the guest (not just an automated welcome)
    answered = bool(after) and not all(
        _looks_automated(m.get("body") or "") for m in after)
    if debug:
        print(f"  conv {cid}: {len(msgs)} msgs · guest_last_id={mid} · after={len(after)} · "
              f"answered={answered} · body={(guest_msg.get('body') or '')[:50]!r}")
    if mid in seen:
        return None                                     # already drafted for this message
    if not ASSISTANT_ALWAYS_DRAFT:
        if answered:
            return None                                 # a real human/bot reply already exists
        if after and not ASSISTANT_ANSWER_PAST_AUTO:
            return None                                 # only auto-replies after, feature off
    lm = c.get("listingMapId")
    unit = listings.get(lm) or c.get("listingName") or f"unit-{lm}"
    guest = c.get("recipientName") or c.get("guestName") or "Guest"
    res = c.get("reservation") or {}
    history = "\n".join(
        f"{'Guest' if _msg_is_inbound(m) else 'Host'}: {(m.get('body') or '').strip()}"
        for m in msgs[-8:] if (m.get("body") or "").strip())
    return {
        "conversation_id": cid, "message_id": mid, "guest": guest, "unit": unit,
        "listing_id": lm,
        "reservation_id": c.get("reservationId") or res.get("id"),
        "res_status": (res.get("status") or "").lower(),
        "comm_type": guest_msg.get("communicationType") or "email",
        "guest_text": (guest_msg.get("body") or "").strip(), "history": history,
        "last_time": _msg_time(guest_msg),
        "checkin": res.get("arrivalDate"), "checkout": res.get("departureDate"),
    }

def fetch_new_guest_messages(seen, debug=False):
    """Scan recent conversations, fetch each one's messages, and return new inbound
    guest messages (the guest's latest message that hasn't been answered/seen)."""
    listings = get_listings_map()
    out = []
    try:
        data = api_get("/conversations", params={"limit": ASSISTANT_SCAN, "includeResources": 1})
    except Exception as e:
        print("assistant fetch error:", e)
        return out
    convos = data.get("result", []) or []
    if debug:
        print(f"assistant DEBUG: /conversations returned {len(convos)} conversations")
    for c in convos:
        it = _conv_to_item(c, listings, seen, debug)
        if it:
            out.append(it)
    if debug:
        print(f"assistant DEBUG: {len(out)} new inbound guest message(s) to draft")
    return out

def fetch_conversation_item(conversation_id, seen):
    """Fetch ONE conversation (for the webhook path) and return its new-message item or None."""
    listings = get_listings_map()
    try:
        data = api_get(f"/conversations/{conversation_id}", params={"includeResources": 1})
    except Exception as e:
        print(f"webhook: conversation {conversation_id} fetch error:", e)
        return None
    c = data.get("result") or {}
    if not c:
        return None
    if not c.get("id"):
        c["id"] = conversation_id
    return _conv_to_item(c, listings, seen)

SIGNATURE_VARY = os.environ.get("SIGNATURE_VARY", "1") in ("1", "true", "True", "yes")

# 50 varied Najdi/warm sign-offs so no two messages end the same way.
SIGNATURES_AR = [
    "تحياتي،\nمساعد - فريق عوجا 🤍", "كامل التوفيق،\nأخوك مساعد - فريق عوجا",
    "في أمان الله،\nمساعد من عوجا 🤍", "تسلم،\nمساعد - عوجا",
    "أي خدمة ثانية أنا حاضر،\nمساعد - فريق عوجا 🤍", "سعدنا بخدمتك،\nمساعد من عوجا",
    "لا تتردد تطلبني،\nأخوك مساعد - عوجا 🤍", "نتشرف فيك دايم،\nفريق عوجا - مساعد",
    "حياك الله،\nمساعد - فريق عوجا 🤍", "بالخدمة دايماً،\nمساعد من عوجا",
    "تحياتي القلبية،\nمساعد - عوجا 🤍", "يومك سعيد،\nأخوك مساعد - فريق عوجا",
    "تواصل معي أي وقت،\nمساعد - عوجا 🤍", "شاكر لك،\nمساعد من فريق عوجا",
    "دمت بخير،\nمساعد - عوجا 🤍", "خدمتك شرف لنا،\nفريق عوجا",
    "أبشر بكل اللي يسعدك،\nمساعد - عوجا 🤍", "تقبل تحياتي،\nمساعد من عوجا",
    "نحن في خدمتك،\nفريق عوجا - مساعد 🤍", "الله يحييك،\nأخوك مساعد - عوجا",
    "أي استفسار ثاني أنا موجود،\nمساعد - فريق عوجا 🤍", "سرّنا تواصلك،\nمساعد من عوجا",
    "بالتوفيق،\nمساعد - عوجا 🤍", "تستاهل كل خير،\nفريق عوجا - مساعد",
    "حاضرين لك،\nمساعد - فريق عوجا 🤍", "تحياتي وتقديري،\nمساعد من عوجا",
    "نوّرتنا،\nمساعد - عوجا 🤍", "أمرك،\nأخوك مساعد - فريق عوجا",
    "خليك على تواصل،\nمساعد - عوجا 🤍", "سعيد بخدمتك،\nمساعد من فريق عوجا",
    "الله يسعدك،\nمساعد - عوجا 🤍", "تحت أمرك دايم،\nفريق عوجا - مساعد",
    "ودّي وتحياتي،\nمساعد من عوجا 🤍", "عساك على القوة،\nأخوك مساعد - عوجا",
    "أي شي تحتاجه أنا جاهز،\nمساعد - فريق عوجا 🤍", "شكراً لثقتك،\nمساعد من عوجا",
    "بكل سرور،\nمساعد - عوجا 🤍", "اعتبرني أخوك،\nمساعد - فريق عوجا",
    "دايم بالخدمة،\nمساعد من عوجا 🤍", "تحياتي الحارة،\nفريق عوجا - مساعد",
    "الله يوفقك،\nمساعد - عوجا 🤍", "نتمنى لك إقامة سعيدة،\nمساعد من فريق عوجا",
    "حياك في أي وقت،\nمساعد - عوجا 🤍", "تسلم وما تقصّر،\nأخوك مساعد - عوجا",
    "خدمتك أولوية عندنا،\nفريق عوجا 🤍", "بانتظار خدمتك،\nمساعد - عوجا",
    "مع خالص الود،\nمساعد من عوجا 🤍", "سعدنا فيك،\nمساعد - فريق عوجا",
    "كلنا بالخدمة،\nمساعد - عوجا 🤍", "الله يتمّم لك على خير،\nأخوك مساعد - فريق عوجا",
]
SIGNATURES_EN = [
    "Best regards,\nMusaid – Ouja Team 🤍", "At your service,\nMusaid – Ouja",
    "Always happy to help,\nMusaid – Ouja Team 🤍", "Warm regards,\nMusaid from Ouja",
    "Reach out anytime,\nMusaid – Ouja Team 🤍", "Take care,\nMusaid – Ouja",
    "Glad to assist,\nMusaid – Ouja Team 🤍", "Anything else, I'm here,\nMusaid – Ouja",
    "Wishing you a great stay,\nMusaid – Ouja Team 🤍", "Kind regards,\nMusaid from Ouja",
    "We're here for you,\nOuja Team – Musaid 🤍", "Thanks for reaching out,\nMusaid – Ouja",
    "All the best,\nMusaid – Ouja Team 🤍", "Happy to help anytime,\nMusaid – Ouja",
    "With pleasure,\nMusaid – Ouja Team 🤍", "Don't hesitate to ask,\nMusaid from Ouja",
    "Here whenever you need,\nMusaid – Ouja 🤍", "Cheers,\nMusaid – Ouja Team",
    "Your comfort is our priority,\nOuja Team 🤍", "Talk soon,\nMusaid – Ouja",
]

def _has_arabic(s):
    return any("\u0600" <= ch <= "\u06ff" for ch in str(s))

def _pick_signature(arabic):
    if not SIGNATURE_VARY:
        return ASSISTANT_SIGNATURE_AR if arabic else ASSISTANT_SIGNATURE_EN
    return random.choice(SIGNATURES_AR if arabic else SIGNATURES_EN)

def with_signature(text):
    """Append a (rotating) support signature, language-matched to the message."""
    return f"{str(text).rstrip()}\n\n{_pick_signature(_has_arabic(text))}"

def send_guest_message(conversation_id, body, comm_type="email"):
    return api_post(f"/conversations/{conversation_id}/messages",
                    {"body": with_signature(body), "communicationType": comm_type})

# pending escalations: discord_message_id -> {channel_id, guest, unit, last_ping, attempts, claimed_by}
_escalations = {}
_esc_ack_count = {}     # conversation_id -> how many escalation acks we've sent
_esc_sent_acks = {}     # conversation_id -> [ack bodies we sent] (to tell our msgs from a co-host's)
_claimed_convos = set() # conversation_ids a human has claimed (stop auto-acks)

class NameSelect(discord.ui.Select):
    """The name picker shown after tapping Claim."""
    def __init__(self, channel_id, message_id):
        super().__init__(placeholder="اختر اسمك للاستلام…", min_values=1, max_values=1,
                         options=[discord.SelectOption(label=n) for n in CLAIM_NAMES])
        self.target_channel_id = channel_id
        self.target_message_id = message_id

    async def callback(self, interaction: discord.Interaction):
        name = self.values[0]
        esc = _escalations.get(self.target_message_id)
        if esc and esc.get("claimed_by"):
            await interaction.response.edit_message(
                content=f"⚠️ التصعيد مستلم مسبقاً بواسطة {esc['claimed_by']}.", view=None)
            return
        if esc:
            esc["claimed_by"] = name
            metric_bump("escalations_resolved")
            if esc.get("conversation_id"):
                _claimed_convos.add(esc["conversation_id"])   # stop auto-acks
            log_event("escalation", f"تم استلام تصعيد بواسطة {name} · {esc.get('unit','')}")
        try:
            ch = interaction.client.get_channel(self.target_channel_id)
            msg = await ch.fetch_message(self.target_message_id)
            embed = msg.embeds[0] if msg.embeds else discord.Embed()
            embed.add_field(name="✅ تم الاستلام", value=f"بواسطة **{name}**", inline=False)
            embed.color = 0x3BA55D
            done = ClaimView()
            for c in done.children:
                c.disabled = True
            await msg.edit(embed=embed, view=done)
        except Exception as e:
            print("claim edit error:", e)
        tail = " — جهّزت لك رد جاهز تحت 👇" if (esc and MANAGER_SCRIPT) else ""
        await interaction.response.edit_message(
            content=f"✅ استلمت التصعيد باسم **{name}**.{tail}", view=None)
        # show the claimer a ready-to-send reply (private to them, right here in the channel)
        if esc and MANAGER_SCRIPT:
            script = await asyncio.to_thread(
                claude_manager_script, esc.get("guest"), esc.get("unit"),
                esc.get("history", ""), esc.get("guest_text", ""), esc.get("reason", ""), name)
            if script:
                dm = (f"🎯 رد جاهز للضيف **{esc.get('guest')} · {esc.get('unit')}** "
                      f"(انسخ والصق ثم أرسل):\n\n{script}")
            else:
                dm = ("⚠️ تعذّر توليد الرد الجاهز — تواصل مع الضيف مباشرة.")
            await interaction.followup.send(dm, ephemeral=True)

class ClaimView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)   # persistent across restarts

    @discord.ui.button(label="🙋 أخذ المهمة / Claim", style=discord.ButtonStyle.primary,
                       custom_id="ouja_claim")
    async def claim(self, interaction: discord.Interaction, button: discord.ui.Button):
        esc = _escalations.get(interaction.message.id)
        if esc and esc.get("claimed_by"):
            await interaction.response.send_message(
                f"⚠️ مستلم مسبقاً بواسطة {esc['claimed_by']}.", ephemeral=True)
            return
        picker = discord.ui.View(timeout=300)
        picker.add_item(NameSelect(interaction.channel.id, interaction.message.id))
        await interaction.response.send_message("اختر اسمك للاستلام:", view=picker, ephemeral=True)

class EditModal(discord.ui.Modal, title="تعديل الرد قبل الإرسال"):
    def __init__(self, item, draft, message_id=None):
        super().__init__()
        self.item = item
        self.message_id = message_id
        self._original_draft = draft or ""    # keep for learning capture below
        self.box = discord.ui.TextInput(label="الرد للضيف", style=discord.TextStyle.paragraph,
                                        default=draft, max_length=1800)
        self.add_item(self.box)

    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(thinking=True)   # ack now; sending can be slow
        text = str(self.box.value).strip()
        try:
            await asyncio.to_thread(send_guest_message, self.item["conversation_id"], text,
                                    self.item["comm_type"])
            # learning: capture the team's edited reply as a strong correction signal
            record_learning(self.item, self._original_draft, text,
                            via="discord_edit", approver=str(interaction.user))
            try:
                msg = await interaction.channel.fetch_message(self.message_id)
                done = ApproveView()
                for c in done.children:
                    c.disabled = True
                await msg.edit(view=done)
            except Exception:
                pass
            _pending_replies.pop(self.message_id, None)
            _replied_msgs.add(self.message_id)
            await interaction.followup.send(
                f"✅ تم الإرسال (بعد التعديل) بواسطة {interaction.user.mention} "
                f"للضيف **{self.item['guest']}**.")
        except Exception as e:
            await interaction.followup.send(f"⚠️ فشل الإرسال: {e}", ephemeral=True)

# pending approval cards: discord_message_id -> {"item": item, "draft": draft}
_pending_replies = {}

def _recover_from_embed(message):
    """Rebuild (item, draft) from the card itself, so buttons work even after a full redeploy."""
    try:
        if not message.embeds:
            return None, None
        emb = message.embeds[0]
        foot = (emb.footer.text or "") if emb.footer else ""
        m = re.search(r"#(\d+)·(\S+)", foot)
        if not m:
            return None, None
        title = emb.title or ""
        guest = "الضيف"
        if "💬" in title:
            t = title.split("💬", 1)[1].strip()
            guest = (t.split("·", 1)[0].strip() or guest)
        draft = None
        for f in emb.fields:
            if "الرد المقترح" in (f.name or ""):
                draft = (f.value or "").strip()
        if not draft or draft == "—":
            return None, None
        return {"conversation_id": int(m.group(1)), "comm_type": m.group(2), "guest": guest}, draft
    except Exception as e:
        print("embed recover error:", e)
        return None, None

def _detect_apartment_in_text(text):
    """Return (lid, unit_name) if `text` mentions one of our catalog apartments by
    name, else (None, None). Uses normalized substring match — picks the LONGEST
    matching name so "12B HTN" wins over "12B" and "A12 - النرجس" wins over "A12".
    Only triggers on names ≥ 2 normalized chars to avoid spurious matches."""
    if not text or not _catalog_units:
        return None, None
    norm_text = " " + norm_unit(text) + " "
    best = (None, None, 0)   # lid, name, score (length)
    for u in _catalog_units:
        unit_name = u.get("name", "")
        if not unit_name:
            continue
        unit_norm = norm_unit(unit_name)
        if len(unit_norm) < 2:
            continue
        # Require a word-ish boundary: pad with spaces, then check
        if (" " + unit_norm + " ") in norm_text or norm_text.endswith(" " + unit_norm) \
           or norm_text.startswith(unit_norm + " ") or unit_norm in norm_text:
            if len(unit_norm) > best[2]:
                best = (int(u["id"]), unit_name, len(unit_norm))
    return (best[0], best[1]) if best[0] else (None, None)

async def load_knowledge(guild):
    """Read the #knowledge channel and rebuild BOTH:
      - _knowledge_text (general facts, injected into every draft)
      - _knowledge_apartment_facts[lid] (scoped facts, injected only when
        drafting for that apartment)
    A message is treated as apartment-scoped if it mentions a catalog
    apartment by name, otherwise it's general."""
    global _knowledge_text, _knowledge_apartment_facts
    ch = discord.utils.get(guild.text_channels, name=KNOWLEDGE_CHANNEL)
    if ch is None:
        return
    try:
        general_facts = []
        apt_facts = {}
        total = 0
        async for m in ch.history(limit=KNOWLEDGE_MAX, oldest_first=True):
            body = (m.content or "").strip()
            if not body or body.startswith(("/", "!")):
                continue
            total += 1
            lid, _name = _detect_apartment_in_text(body)
            if lid:
                apt_facts.setdefault(lid, []).append(body)
            else:
                general_facts.append(f"- {body}")
        _knowledge_text = "\n".join(general_facts)[:8000]
        _knowledge_apartment_facts = {k: v[-30:] for k, v in apt_facts.items()}
        scoped = sum(len(v) for v in apt_facts.values())
        print(f"knowledge: loaded {total} fact(s) from #{KNOWLEDGE_CHANNEL} "
              f"({len(general_facts)} general, {scoped} apt-scoped across {len(apt_facts)} units)")
    except Exception as e:
        print("knowledge load error:", e)

async def save_fact(guild, text):
    """Append a fact to the #knowledge channel (creating it if needed) and refresh."""
    ch = await ensure_channel(guild, KNOWLEDGE_CHANNEL, await get_assistant_category(guild))
    if ch is None:
        return False
    try:
        await ch.send(text[:1900])
        await load_knowledge(guild)
        return True
    except Exception as e:
        print("save fact error:", e)
        return False

class TeachModal(discord.ui.Modal, title="🧠 علّم المساعد معلومة"):
    def __init__(self, message_id, item=None):
        super().__init__()
        self.message_id = message_id
        self.item = item
        self.topic = discord.ui.TextInput(label="الموضوع (اختياري)", required=False,
                                          placeholder="مثال: واي فاي وحدة C08 / موقف السيارة",
                                          max_length=120)
        self.fact = discord.ui.TextInput(label="المعلومة الصحيحة", style=discord.TextStyle.paragraph,
                                         placeholder="اكتب المعلومة الصح اللي تبي المساعد يعرفها ويستخدمها.",
                                         max_length=900)
        self.add_item(self.topic)
        self.add_item(self.fact)

    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer(thinking=True, ephemeral=True)
        topic = str(self.topic.value).strip()
        fact = str(self.fact.value).strip()
        line = f"**{topic}**: {fact}" if topic else fact
        ok = await save_fact(interaction.guild, line)
        if not ok:
            await interaction.followup.send("⚠️ تعذّر حفظ المعلومة.", ephemeral=True)
            return
        # regenerate this card's draft using the new knowledge (if we still have context)
        data = _pending_replies.get(self.message_id)
        item = (data or {}).get("item") or self.item
        regen = False
        if item and item.get("history"):
            try:
                result = await asyncio.to_thread(
                    claude_draft, item["guest"], item["unit"], item["history"],
                    (data or {}).get("guide"), (data or {}).get("confirmed", False))
                reply = (result or {}).get("reply", "").strip()
                if reply:
                    msg = await interaction.channel.fetch_message(self.message_id)
                    emb = msg.embeds[0]
                    fields = emb.fields
                    emb.clear_fields()
                    for f in fields:
                        emb.add_field(name=f.name,
                                      value=(reply[:1024] if "الرد المقترح" in f.name else f.value),
                                      inline=f.inline)
                    await msg.edit(embed=emb, view=ApproveView())
                    if data:
                        data["draft"] = reply
                    else:
                        _pending_replies[self.message_id] = {"item": item, "draft": reply,
                                                             "guide": (data or {}).get("guide"),
                                                             "confirmed": (data or {}).get("confirmed", False)}
                    regen = True
            except Exception as e:
                print("teach regen error:", e)
        note = "✅ حفظت المعلومة" + (" وأعدت صياغة الرد بها 👆" if regen else " — بتنطبق على الردود الجاية.")
        await interaction.followup.send(note, ephemeral=True)

class ConfirmActionView(discord.ui.View):
    """Ephemeral 'are you sure?' shown before a Send or Reject actually happens."""
    def __init__(self, action, channel_id, message_id):
        super().__init__(timeout=120)
        self.action = action            # "send" or "reject"
        self.channel_id = channel_id
        self.message_id = message_id

    async def _disable_card(self, interaction):
        try:
            ch = interaction.client.get_channel(self.channel_id)
            card = await ch.fetch_message(self.message_id)
            done = ApproveView()
            for c in done.children:
                c.disabled = True
            await card.edit(view=done)
        except Exception as e:
            print("confirm card edit error:", e)

    @discord.ui.button(label="✅ نعم، أكمل", style=discord.ButtonStyle.success)
    async def yes(self, interaction: discord.Interaction, button: discord.ui.Button):
        if self.action == "send":
            data = _pending_replies.get(self.message_id)
            item = (data or {}).get("item")
            draft = (data or {}).get("draft")
            if not item:
                try:
                    ch = interaction.client.get_channel(self.channel_id)
                    card = await ch.fetch_message(self.message_id)
                    item, draft = _recover_from_embed(card)
                except Exception:
                    item = None
            if not item:
                await interaction.response.edit_message(
                    content="⚠️ هذا الكرت قديم — تعامل مع الرسالة يدوياً.", view=None)
                return
            await interaction.response.edit_message(content="⏳ جاري الإرسال…", view=None)
            try:
                await asyncio.to_thread(send_guest_message, item["conversation_id"], draft,
                                        item["comm_type"])
                # learning: send-as-is means the team approved the draft verbatim
                record_learning(item, draft, draft,
                                via="discord_send", approver=str(interaction.user))
                await self._disable_card(interaction)
                _pending_replies.pop(self.message_id, None)
                _replied_msgs.add(self.message_id)
                await interaction.followup.send(
                    f"✅ تم الإرسال بواسطة {interaction.user.mention} للضيف **{item['guest']}**.")
            except Exception as e:
                await interaction.followup.send(f"⚠️ فشل الإرسال: {e}", ephemeral=True)
        else:   # reject
            await self._disable_card(interaction)
            _pending_replies.pop(self.message_id, None)
            _replied_msgs.add(self.message_id)
            await interaction.response.edit_message(content="🗑️ تم التجاهل.", view=None)
            await interaction.followup.send(f"🗑️ تم التجاهل بواسطة {interaction.user.mention}.")

    @discord.ui.button(label="✖️ تراجع", style=discord.ButtonStyle.secondary)
    async def no(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.edit_message(
            content="تمام، ما سويت شي. الكرت زي ما هو 👍", view=None)

class PriceConfirmView(discord.ui.View):
    """Ephemeral 'are you sure?' before writing prices to the calendar."""
    def __init__(self, card_message_id):
        super().__init__(timeout=300)
        self.card_message_id = card_message_id

    @discord.ui.button(label="✅ نعم، طبّق", style=discord.ButtonStyle.success)
    async def yes(self, interaction: discord.Interaction, button: discord.ui.Button):
        opp = _price_opps.get(self.card_message_id)
        if not opp:
            await interaction.response.edit_message(
                content="انتهت صلاحية هالبطاقة — اطلب تقرير جديد.", view=None)
            return
        await interaction.response.edit_message(content="⏳ جاري تطبيق الأسعار…", view=None)
        applied, skipped, _results = await asyncio.to_thread(
            apply_price_changes, opp["listing_id"], opp["changes"])
        _price_opps.pop(self.card_message_id, None)
        tail = (" — (تجربة DRY-RUN، ما تغيّر شي فعلي)" if PRICE_APPLY_DRYRUN else "")
        skip_txt = f" · تخطّيت {skipped} (محجوزة/تغيّرت)" if skipped else ""
        result = f"✅ طبّقت {applied} ليلة على **{opp['name']}**{skip_txt}{tail}"
        log_event("pricing", f"طبّق {applied} سعر على {opp['name']} بواسطة {interaction.user.display_name}"
                  + (" (DRY-RUN)" if PRICE_APPLY_DRYRUN else ""))
        await interaction.followup.send(result, ephemeral=True)
        try:
            card = await interaction.channel.fetch_message(self.card_message_id)
            emb = card.embeds[0] if card.embeds else discord.Embed()
            emb.color = 0x2ECC71
            emb.set_footer(text=f"{result} · بواسطة {interaction.user.display_name}")
            await card.edit(embed=emb, view=None)
        except Exception as e:
            print("price card update error:", e)

    @discord.ui.button(label="✖️ لا", style=discord.ButtonStyle.secondary)
    async def no(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.edit_message(content="تمام، ما غيّرت شي 👍", view=None)

class PriceApplyView(discord.ui.View):
    """Persistent Apply/Skip buttons under each unit's pricing card."""
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(label="✅ طبّق", style=discord.ButtonStyle.success,
                       custom_id="ouja_price_apply")
    async def apply(self, interaction: discord.Interaction, button: discord.ui.Button):
        opp = _price_opps.get(interaction.message.id)
        if not opp:
            await interaction.response.send_message(
                "هالبطاقة قديمة (البوت اتحدّث) — اطلب تقرير جديد بـ PRICE_OPP_TEST=1.",
                ephemeral=True)
            return
        n = len(opp["changes"])
        warn = ("\n⚠️ وضع التجربة شغّال — ما راح يتغيّر شي فعلي." if PRICE_APPLY_DRYRUN
                else "\n‼️ بيتغيّر السعر فعلياً في تقويمك (Airbnb/Hostaway).")
        await interaction.response.send_message(
            f"متأكد تبي تطبّق **{n}** تغيير سعر على **{opp['name']}**؟{warn}",
            view=PriceConfirmView(interaction.message.id), ephemeral=True)

    @discord.ui.button(label="❌ تجاهل", style=discord.ButtonStyle.secondary,
                       custom_id="ouja_price_skip")
    async def skip(self, interaction: discord.Interaction, button: discord.ui.Button):
        _price_opps.pop(interaction.message.id, None)
        emb = interaction.message.embeds[0] if interaction.message.embeds else discord.Embed()
        emb.color = 0x95A5A6
        emb.set_footer(text=f"❌ تم التجاهل بواسطة {interaction.user.display_name}")
        await interaction.response.edit_message(embed=emb, view=None)

class ApproveView(discord.ui.View):
    def __init__(self, item=None, draft=None):
        super().__init__(timeout=None)   # persistent — survives bot restarts
        self.item = item
        self.draft = draft

    def _resolve(self, interaction):
        data = _pending_replies.get(interaction.message.id)
        if data:
            return data["item"], data["draft"]
        if self.item:
            return self.item, self.draft     # same-session fallback
        return _recover_from_embed(interaction.message)   # rebuild after a redeploy

    @discord.ui.button(label="✅ إرسال", style=discord.ButtonStyle.success, custom_id="ouja_send")
    async def send(self, interaction: discord.Interaction, button: discord.ui.Button):
        if interaction.message.id in _replied_msgs:
            await interaction.response.send_message(
                "✅ تم التعامل مع هذا الرد مسبقاً (من اللوحة).", ephemeral=True)
            return
        item, draft = self._resolve(interaction)
        if not item:
            await interaction.response.send_message(
                "⚠️ هذا الكرت قديم (انعاد تشغيل البوت) — تجاهله وتعامل مع الرسالة يدوياً.",
                ephemeral=True)
            return
        preview = (draft or "")[:300]
        await interaction.response.send_message(
            f"📤 متأكد تبي ترسل هذا الرد للضيف **{item['guest']}**؟\n\n>>> {preview}",
            view=ConfirmActionView("send", interaction.channel.id, interaction.message.id),
            ephemeral=True)

    @discord.ui.button(label="✏️ تعديل وإرسال", style=discord.ButtonStyle.primary, custom_id="ouja_edit")
    async def edit(self, interaction: discord.Interaction, button: discord.ui.Button):
        item, draft = self._resolve(interaction)
        if not item:
            await interaction.response.send_message(
                "⚠️ هذا الكرت قديم (انعاد تشغيل البوت) — تجاهله وتعامل يدوياً.", ephemeral=True)
            return
        await interaction.response.send_modal(EditModal(item, draft, interaction.message.id))

    @discord.ui.button(label="🗑️ رفض", style=discord.ButtonStyle.danger, custom_id="ouja_reject")
    async def reject(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_message(
            "🗑️ متأكد تبي ترفض هذا الكرت وتتجاهله؟",
            view=ConfirmActionView("reject", interaction.channel.id, interaction.message.id),
            ephemeral=True)

    @discord.ui.button(label="🧠 علّم", style=discord.ButtonStyle.secondary, custom_id="ouja_teach")
    async def teach(self, interaction: discord.Interaction, button: discord.ui.Button):
        item, _ = self._resolve(interaction)
        await interaction.response.send_modal(TeachModal(interaction.message.id, item))

_assistant_seen = _BoundedSet(maxlen=20000)

_SENTIMENT_AR = {"ok": "عادي", "upset": "غاضب/منزعج"}

async def post_assistant_card(channel, item, result, guide=None, confirmed=False):
    g = item["guest"]
    intent = result.get("intent", "—")
    sentiment = result.get("sentiment", "ok")
    sent_ar = _SENTIMENT_AR.get(sentiment, sentiment)
    conf = float(result.get("confidence", 0) or 0)
    action = result.get("action", "escalate")
    reply = (result.get("reply") or "").strip()
    escalate = action == "escalate" or sentiment == "upset" or conf < ESCALATE_BELOW

    # daily metrics: every draft counts, with confidence + topic + apartment
    metric_bump("drafts_made")
    metric_record_confidence(conf)
    metric_record_topic(intent)
    metric_record_apartment(item.get("unit"))

    # ---- needs a human: tell the guest it's escalated, then alert the team ----
    if escalate:
        metric_bump("escalations_created")
        embed = discord.Embed(title=f"🚨 تصعيد · {g} · {item['unit']}", color=0xD64545)
        embed.add_field(name="📩 الضيف يقول", value=(item["guest_text"] or "—")[:1000], inline=False)
        embed.add_field(name="🔴 يحتاج تدخل بشري",
                        value=result.get("reason", "تم التصعيد — تعامل معه يدوياً."), inline=False)
        cid = item["conversation_id"]
        if cid in _claimed_convos:
            embed.add_field(name="🙋 مستلمة",
                            value="الموضوع مستلم من أحد الفريق — ما أرسلنا رد تلقائي.", inline=False)
        elif ASSISTANT_ESC_ACK:
            _esc_ack_count[cid] = _esc_ack_count.get(cid, 0) + 1
            n = _esc_ack_count[cid]
            is_ar = _has_arabic(item["guest_text"])
            # Off-hours ack: tell the guest exactly what happens next AND when.
            # In-hours ack: the regular static or the empathetic re-ack on repeats.
            offhours_now = bool(item.get("_offhours"))
            if offhours_now and n == 1:
                back = next_work_start()
                same_day = back.date() == datetime.now(TZ).date()
                back_ar = back.strftime("%H:%M") + (" بكرة" if not same_day else "")
                back_en = back.strftime("%H:%M") + (" tomorrow" if not same_day else "")
                ack = (
                    f"حياك الله 🤍 وصلت رسالتك. الفريق خارج ساعات العمل حالياً (نرجع الساعة "
                    f"{back_ar})، لكن لأهمية موضوعك رفعت تنبيهاً للمشرف الحين، وأول ما يبدأ "
                    f"اليوم بيتواصل معك مباشرة. إذا فيه تفاصيل إضافية تساعدنا، اكتبها هنا وراح "
                    f"تكون أمامه أول ما يفتح الجهاز."
                    if is_ar else
                    f"Hi 🤍 your message is in. Our team is off-hours right now (back at "
                    f"{back_en}), but I've flagged this to the supervisor immediately and "
                    f"they'll reach out the moment they start their day. Any extra detail you "
                    f"add here will be the first thing they see when they're back."
                )
            elif n == 1:
                ack = ASSISTANT_ACK_AR if is_ar else ASSISTANT_ACK_EN
            else:                                  # repeat: empathetic, problem-specific
                ack = await asyncio.to_thread(claude_escalation_ack, g, item["unit"],
                                              item["history"], item["guest_text"]) \
                       or (ASSISTANT_ACK_AR if is_ar else ASSISTANT_ACK_EN)
            try:
                await asyncio.to_thread(send_guest_message, cid, ack, item["comm_type"])
                _esc_sent_acks.setdefault(cid, []).append(ack)
                embed.add_field(name="📤 تم إبلاغ الضيف",
                                value=("رسالة طمأنة متعاطفة (متابعة)" if n > 1
                                       else "رسالة طمأنة إنه تم تصعيد طلبه للقسم المختص."),
                                inline=False)
            except Exception as e:
                embed.add_field(name="⚠️ تعذّر إبلاغ الضيف", value=str(e), inline=False)
        embed.set_footer(text=f"النوع: {intent} · المشاعر: {sent_ar} · الثقة: {round(conf*100)}% · "
                              f"يعاد التنبيه كل {ESCALATION_REPING_MIN} دقيقة لين يستلمه أحد")
        # post to the dedicated escalations channel and @mention the operation team
        guild = channel.guild
        esc_channel = await ensure_channel(guild, ESCALATION_CHANNEL,
                                           await get_assistant_category(guild))
        target = esc_channel or channel
        op_role = find_operation_role(guild)
        mention = op_role.mention if op_role else f"@{OPERATION_ROLE_NAME}"
        try:
            msg = await target.send(content=f"{mention} 🚨 تصعيد جديد يحتاج استلام",
                                    embed=embed, view=ClaimView(),
                                    allowed_mentions=discord.AllowedMentions(roles=True))
            _escalations[msg.id] = {"channel_id": target.id, "guest": g, "unit": item["unit"],
                                    "conversation_id": item["conversation_id"],
                                    "guest_text": item["guest_text"],
                                    "reason": result.get("reason", ""), "history": item["history"],
                                    "last_ping": time.time(), "attempts": 0, "claimed_by": None,
                                    "last_msg_id": item.get("message_id") or 0,
                                    "acks": list(_esc_sent_acks.get(item["conversation_id"], []))}
        except Exception as e:
            print("escalation post error:", e)
        return

    # ---- STAGE 1: confident enough -> send automatically, then post an FYI card ----
    # During off-hours we treat auto-send as enabled by default — the team
    # isn't around to approve, and the owner explicitly asked the bot to
    # "answer as much as possible" outside hours.
    offhours = bool(item.get("_offhours"))
    can_auto = (not escalate and bool(reply) and conf >= ASSISTANT_AUTO_CONF and
                (ASSISTANT_AUTO or offhours))
    if can_auto:
        try:
            await asyncio.to_thread(send_guest_message, item["conversation_id"], reply,
                                    item["comm_type"])
            # learning: auto-sent at high confidence (still useful — confirms the
            # bot's wording on simple replies, helps reinforce the pattern)
            record_learning(item, reply, reply, via="auto", approver="(auto)")
            embed = discord.Embed(title=f"⚡ رد تلقائي · {g} · {item['unit']}", color=0x3BA55D)
            embed.add_field(name="📩 الضيف يقول", value=(item["guest_text"] or "—")[:1000], inline=False)
            embed.add_field(name="✅ تم الرد تلقائياً (Stage 1)", value=reply[:1000], inline=False)
            embed.set_footer(text=f"النوع: {intent} · الثقة: {round(conf*100)}% · رد تلقائي للعلم")
            await channel.send(embed=embed)
            log_event("guest", f"رد تلقائي ({round(conf*100)}%) · {g} · {item['unit']}")
            # --- auto-reply audit: keep a record + post to the dedicated audit channel ---
            _auto_replies.appendleft({"ts": datetime.now(TZ).isoformat(timespec="seconds"),
                                      "guest": g, "unit": item["unit"], "conf": round(conf * 100),
                                      "guest_text": (item["guest_text"] or "")[:600], "reply": reply[:1000]})
            try:
                audit = await ensure_channel(channel.guild, AUTO_REPLY_CHANNEL,
                                             await get_assistant_category(channel.guild))
                if audit:
                    a = discord.Embed(title=f"⚡ {g} · {item['unit']}", color=0x3BA55D,
                                      timestamp=datetime.now(TZ))
                    a.add_field(name="📩 الضيف", value=(item["guest_text"] or "—")[:1000], inline=False)
                    a.add_field(name="🤍 رد المساعد (تلقائي)", value=reply[:1000], inline=False)
                    a.set_footer(text=f"ثقة {round(conf*100)}% · أُرسل بدون مراجعة")
                    await audit.send(embed=a)
            except Exception as e:
                print("auto-reply audit post error:", e)
            return
        except Exception as e:
            print("auto-send failed, falling back to approval:", e)
            # fall through to the approval card if the send failed

    # ---- needs approval: draft + buttons ----
    embed = discord.Embed(title=f"💬 {g} · {item['unit']}", color=GOLD)
    embed.add_field(name="📩 الضيف يقول", value=(item["guest_text"] or "—")[:1024], inline=False)
    embed.add_field(name="✍️ الرد المقترح", value=(reply or "—")[:1024], inline=False)
    embed.set_footer(text=f"النوع: {intent} · الثقة: {round(conf*100)}% · راجعه قبل الإرسال · "
                          f"التوقيع يُضاف تلقائياً · #{item['conversation_id']}·{item['comm_type']}")
    sent = await channel.send(embed=embed, view=ApproveView(item, reply))
    _pending_replies[sent.id] = {"item": item, "draft": reply, "guide": guide, "confirmed": confirmed,
                                 "intent": intent, "confidence": round(conf*100), "sentiment": sentiment}

async def process_assistant_item(it, channel):
    """Draft + post a card (or escalate) for ONE guest message. Shared by the poll
    loop and the webhook handler so both behave identically."""
    _assistant_seen.add(it["message_id"])
    if not it["guest_text"]:
        return
    # Update guest profile with this conversation's reservation (silent, fast).
    try:
        prof = record_guest_stay(it)
        if prof:
            it["guest_profile_key"] = prof["key"]
            it["guest_vip"] = prof.get("vip", False)
            it["guest_prior_stays"] = len(prof.get("reservations", []))
    except Exception as e:
        print("guest profile update error:", e)
    # ---- Off-hours flag (used below in the draft path) ----
    # New behavior per owner: outside working hours, answer ANY question the
    # bot is confident about (treat off-hours like ASSISTANT_AUTO=ON with the
    # same confidence threshold). Only escalations and very-low-confidence
    # cases trigger the holding message — and that holding message now
    # promises an escalation + supervisor reminder at the start of the day,
    # instead of just "we'll get back to you".
    it["_offhours"] = (OFFHOURS_AUTOREPLY_ENABLED and not is_within_working_hours())
    status = it.get("res_status") or await asyncio.to_thread(
        get_reservation_status, it.get("reservation_id"))
    confirmed = status in CONFIRMED_STATUSES
    guide = (await asyncio.to_thread(get_guide_url, it.get("listing_id"))
             if (confirmed and it.get("listing_id")) else None)
    result = await asyncio.to_thread(
        claude_draft, it["guest"], it["unit"], it["history"], guide, confirmed,
        (it.get("checkin"), it.get("checkout")), it.get("listing_id"),
        it.get("reservation_id"), it.get("guest_profile_key"))
    if not result:
        return
    try:
        await post_assistant_card(channel, it, result, guide, confirmed)
        act = result.get("action", "reply")
        if act == "escalate" or result.get("sentiment") == "upset":
            log_event("escalation", f"تصعيد · {it['guest']} · {it['unit']}")
        else:
            log_event("guest", f"بطاقة رد ({act}) · {it['guest']} · {it['unit']}")
    except Exception as e:
        print("assistant card error:", e)

async def _assistant_channel():
    guild = bot.get_guild(GUILD_ID)
    if guild is None:
        return None
    return await ensure_channel(guild, ASSISTANT_CHANNEL, await get_assistant_category(guild))

async def run_assistant_scan():
    """One full inbox scan (used by the poll loop and as a webhook fallback)."""
    if not ASSISTANT_ENABLED:
        return
    channel = await _assistant_channel()
    if channel is None:
        return
    try:
        items = await asyncio.to_thread(fetch_new_guest_messages, _assistant_seen, ASSISTANT_DEBUG)
    except Exception as e:
        print("assistant scan fetch error:", e)
        return
    for it in items:
        await process_assistant_item(it, channel)

@tasks.loop(minutes=ASSISTANT_POLL_MIN)
async def assistant_loop():
    try:
        await run_assistant_scan()
    except Exception as e:
        print("assistant_loop error:", e)

# ---------------- Stage 2: Hostaway webhook server ----------------
_web_runner = None

def _extract_conversation_id(payload):
    """Pull a conversationId out of Hostaway's webhook payload, trying common shapes.
    Returns None if we can't find one (caller then falls back to a full scan)."""
    if not isinstance(payload, dict):
        return None
    # top level
    for k in ("conversationId", "conversation_id"):
        if payload.get(k):
            return payload[k]
    body = payload.get("data") or payload.get("object") or payload.get("body") or {}
    if isinstance(body, dict):
        for k in ("conversationId", "conversation_id"):
            if body.get(k):
                return body[k]
        # if the event object itself is a conversation, its id IS the conversation id
        evt = str(payload.get("event") or payload.get("type") or payload.get("object") or "").lower()
        if "conversation" in evt and body.get("id"):
            return body["id"]
    return None

async def _process_conversation_now(conversation_id):
    """Handle a single conversation immediately (triggered by a webhook)."""
    await bot.wait_until_ready()
    if not ASSISTANT_ENABLED:
        return
    channel = await _assistant_channel()
    if channel is None:
        return
    it = await asyncio.to_thread(fetch_conversation_item, conversation_id, _assistant_seen)
    if it:
        print(f"webhook: processing conversation {conversation_id} ({it['guest']})")
        await process_assistant_item(it, channel)

async def _handle_hook(request):
    secret = request.match_info.get("secret", "")
    if secret != WEBHOOK_SECRET:
        return web.Response(status=403, text="forbidden")
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    cid = _extract_conversation_id(payload)
    if cid:
        asyncio.create_task(_process_conversation_now(cid))
    else:
        # couldn't parse a conversation id -> just scan recent inbox (still near-instant)
        asyncio.create_task(run_assistant_scan())
    return web.Response(status=200, text="ok")     # ack fast so Hostaway doesn't retry

async def _handle_health(request):
    return web.Response(status=200, text="Ouja bot is up")

# ---------------- Activity log (feeds the dashboard) ----------------
_activity = deque(maxlen=800)
_auto_replies = deque(maxlen=500)     # audit of Stage-1 auto-sent messages (guest msg + AI reply)
_pricing_strategies = {}              # lid -> active dynamic-pricing strategy state

# ---- Self-learning: capture every send + periodically distill team intent ----
# Every approved/edited reply is appended to _learning_log. A background task
# (learning_distillation_loop) groups the recent entries by listing and asks
# Claude to extract a concise summary of what the team has been telling guests.
# The resulting per-apartment + general summaries are then injected into every
# future draft so the bot becomes more apartment-specific and confident over time.
_learning_log = deque(maxlen=3000)
_apartment_learnings = {}   # lid (int) -> {"summary": str, "last_distilled": ts, "examples_count": int}
_general_learnings = {"summary": "", "last_distilled": 0, "examples_count": 0}

# Daily metrics — one row per calendar date, persisted, used by the Learning page's
# trend charts and the "what improved over time" copy. Every reply send, escalation
# creation, and escalation resolution bumps a counter here.
_daily_metrics = {}     # "YYYY-MM-DD" -> {counters dict, see _new_day_row()}

def _new_day_row():
    return {
        "replies_total": 0,         # any reply that went out (auto + manual + dashboard + edit)
        "replies_auto": 0,           # high-conf auto-sent
        "replies_manual": 0,         # human clicked Send-as-is from Discord
        "replies_edited": 0,         # human edited then sent (correction signal)
        "replies_dashboard": 0,      # sent from the web dashboard
        "escalations_created": 0,
        "escalations_resolved": 0,   # claimed OR auto-resolved
        "drafts_made": 0,            # total drafts the assistant produced (escalated or not)
        "confidence_sum": 0.0,       # for computing avg
        "confidence_count": 0,
        "topics": {},                # intent -> count
        "apartments_touched": [],    # list of unit names that saw activity today
    }

def _today_key():
    return datetime.now(TZ).date().isoformat()

def _day_row(d_key=None):
    k = d_key or _today_key()
    return _daily_metrics.setdefault(k, _new_day_row())

def metric_bump(key, by=1, day=None):
    try:
        row = _day_row(day)
        row[key] = row.get(key, 0) + by
    except Exception:
        pass

def metric_record_confidence(conf):
    try:
        row = _day_row()
        row["confidence_sum"] = row.get("confidence_sum", 0.0) + float(conf or 0)
        row["confidence_count"] = row.get("confidence_count", 0) + 1
    except Exception:
        pass

def metric_record_topic(intent):
    try:
        if not intent:
            return
        row = _day_row()
        topics = row.setdefault("topics", {})
        topics[intent] = topics.get(intent, 0) + 1
    except Exception:
        pass

def metric_record_apartment(unit_name):
    try:
        if not unit_name:
            return
        row = _day_row()
        apt = row.setdefault("apartments_touched", [])
        if unit_name not in apt:
            apt.append(unit_name)
    except Exception:
        pass

# ====================================================================
# WILT — What I Learned Today  (daily journal-style summary)
# ====================================================================
def compose_wilt(date_iso=None):
    """Build the data dict for today's WILT post. Inputs come from the in-memory
    counters + learning log + knowledge cache + cleaning feedback ledger."""
    today = date_iso or datetime.now(TZ).date().isoformat()
    row = _daily_metrics.get(today, _new_day_row())

    # 1) Reply funnel for today
    drafts = row.get("drafts_made", 0)
    total = row.get("replies_total", 0)
    auto = row.get("replies_auto", 0)
    manual = row.get("replies_manual", 0)
    edited = row.get("replies_edited", 0)
    dashboard = row.get("replies_dashboard", 0)
    escs_created = row.get("escalations_created", 0)
    escs_resolved = row.get("escalations_resolved", 0)
    avg_conf = round((row["confidence_sum"] / row["confidence_count"]) * 100) \
        if row.get("confidence_count") else 0
    topics = row.get("topics", {})
    top_topics = sorted(topics.items(), key=lambda kv: -kv[1])[:5]
    apts_touched = row.get("apartments_touched", [])

    # 2) New knowledge added today (from log events tagged 'معرفة')
    knowledge_today = [e for e in _activity
                       if e.get("ts", "").startswith(today) and "معرفة" in e.get("text", "")]
    # 3) Concrete correction examples — top 5 most-edited drafts today
    edits_today = sorted(
        [e for e in _learning_log
         if e.get("ts", "").startswith(today) and e.get("was_edited")],
        key=lambda e: -(e.get("diff_ratio") or 0),
    )[:5]
    # 4) Apartments whose distilled summary refreshed today
    today_dt = datetime.now(TZ).date()
    apt_updated_today = [(lid, v) for lid, v in _apartment_learnings.items()
                         if v.get("last_distilled") and
                         datetime.fromtimestamp(v["last_distilled"], TZ).date() == today_dt]
    # 5) Cleaning quality ratings received today
    quality_today = [fb for fb in _cleaning_feedback.values()
                     if fb.get("ts_done", "").startswith(today)]
    # 6) Net new VIPs today is hard without timestamps; skip — surface running count
    vip_count = sum(1 for p in _guest_profiles.values() if p.get("vip"))

    return {
        "date": today,
        "funnel": {
            "drafts": drafts, "total": total, "auto": auto, "manual": manual,
            "edited": edited, "dashboard": dashboard,
            "auto_rate": round(auto / total * 100) if total else 0,
            "edit_rate": round(edited / max(1, manual + edited + dashboard) * 100)
                if (manual + edited + dashboard) else 0,
            "avg_confidence": avg_conf,
        },
        "escalations": {"created": escs_created, "resolved": escs_resolved},
        "topics": top_topics,
        "apartments_touched": apts_touched,
        "knowledge_added": [{"ts": e["ts"][11:16], "text": e["text"][:160]} for e in knowledge_today],
        "corrections": [
            {"unit": e.get("unit",""), "guest_q": (e.get("guest_question","") or "")[:140],
             "bot_draft": (e.get("bot_draft","") or "")[:160],
             "final": (e.get("final_reply","") or "")[:160],
             "diff": e.get("diff_ratio", 0)}
            for e in edits_today
        ],
        "apartments_distilled_today": [{"lid": l, "unit": v.get("unit","")}
                                       for l, v in apt_updated_today][:10],
        "quality": {"count": len(quality_today),
                    "avg": round(sum(fb["score"] for fb in quality_today) / len(quality_today), 1)
                          if quality_today else None,
                    "comments": [{"unit": fb.get("unit",""), "score": fb["score"],
                                  "text": (fb.get("comment","") or "")[:120]}
                                 for fb in quality_today if fb.get("comment")][:3]},
        "vip_count": vip_count,
    }

def claude_wilt_prose(data):
    """Ask Claude (Sonnet) for a warm 3-4 sentence Najdi summary of the day.
    Returns string or None."""
    if not ANTHROPIC_API_KEY:
        return None
    sys = ("أنت تكتب ملخص يومي قصير دافئ لفريق إدارة عقارات «عوجا». الملخص يكون "
           "بصيغة أنا (المساعد): 3-4 جمل، نجدي طبيعي، صادق، بدون مبالغة. ركّز على: "
           "أهم شي تعلّمته، أبرز شي عدّله الفريق علي، وأي نمط لاحظته. لو اليوم هادي "
           "ما حصل فيه شي مميّز، قول ذلك بوضوح. اكتب نص الفقرة فقط بدون عناوين."
           + _DIALECT_LOCK)
    user = (f"بيانات اليوم {data.get('date')}:\n"
            f"- مسوّدات: {data['funnel']['drafts']}، أرسلت: {data['funnel']['total']} "
            f"(تلقائي {data['funnel']['auto']}, مُعدّلة {data['funnel']['edited']}, "
            f"يدوي {data['funnel']['manual']}, لوحة {data['funnel']['dashboard']}).\n"
            f"- متوسط الثقة: {data['funnel']['avg_confidence']}%.\n"
            f"- تصعيدات: {data['escalations']['created']} جديدة، "
            f"{data['escalations']['resolved']} مُغلقة.\n"
            f"- أكثر المواضيع: {', '.join(k for k,_ in data['topics']) or '—'}.\n"
            f"- معرفة جديدة من الفريق: {len(data['knowledge_added'])}.\n"
            f"- تصحيحات بارزة: {len(data['corrections'])}.\n"
            f"- شقق محدّث ملخصها: {len(data['apartments_distilled_today'])}.\n"
            f"- تقييمات نظافة اليوم: {data['quality']['count']} "
            f"(متوسط {data['quality']['avg']})." if data['quality']['count'] else
            f"- بيانات اليوم {data.get('date')}: قليل من النشاط.\n"
            )
    return claude_text(sys, user, max_tokens=350, model=CLAUDE_MODEL_PREMIUM) \
        or claude_text(sys, user, max_tokens=350)

def _wilt_embed(data, prose=""):
    """Render the WILT data as a Discord embed."""
    f = data["funnel"]
    e = data["escalations"]
    desc_lines = []
    if prose:
        desc_lines.append("*" + prose.strip() + "*\n")
    desc_lines.append(
        f"**🤖 الردود:** {f['total']} مُرسلة من أصل {f['drafts']} مسوّدة "
        f"(تلقائي **{f['auto']}** · يدوي **{f['manual']}** · مُعدّلة **{f['edited']}** · "
        f"لوحة **{f['dashboard']}**)"
    )
    desc_lines.append(
        f"**📊 المؤشرات:** تلقائي {f['auto_rate']}% · تعديل {f['edit_rate']}% · "
        f"ثقة {f['avg_confidence']}% · تصعيدات +{e['created']} / −{e['resolved']}"
    )
    if data["topics"]:
        desc_lines.append("**🏷️ المواضيع:** " + " · ".join(f"{n} ({c})" for n, c in data["topics"]))
    if data["apartments_touched"]:
        desc_lines.append(f"**🏠 شقق نشطة اليوم:** {len(data['apartments_touched'])}")
    embed = discord.Embed(
        title=f"📓 WILT · ما تعلّمته اليوم · {data['date']}",
        description="\n\n".join(desc_lines),
        color=GOLD,
    )
    # New facts the team posted today
    if data["knowledge_added"]:
        embed.add_field(
            name=f"🧠 معرفة جديدة من الفريق ({len(data['knowledge_added'])})",
            value="\n".join(f"`{k['ts']}` {k['text']}" for k in data["knowledge_added"][:6])[:1024],
            inline=False,
        )
    # Top corrections
    if data["corrections"]:
        lines = []
        for c in data["corrections"][:3]:
            unit = c["unit"] or "—"
            lines.append(f"**{unit}** — تعديل {int(c['diff']*100)}%\n"
                         f"  س: _{c['guest_q'][:100]}_\n"
                         f"  أنا قلت: ~~{c['bot_draft'][:100]}~~\n"
                         f"  الفريق صحّحها لـ: **{c['final'][:100]}**")
        embed.add_field(
            name=f"✍️ تصحيحات بارزة ({len(data['corrections'])})",
            value="\n\n".join(lines)[:1024],
            inline=False,
        )
    # Apartments whose distilled brain updated today
    if data["apartments_distilled_today"]:
        names = " · ".join(a["unit"] for a in data["apartments_distilled_today"] if a["unit"])
        embed.add_field(
            name=f"📚 شقق حدّثت ملخصها اليوم ({len(data['apartments_distilled_today'])})",
            value=(names or "—")[:1024], inline=False,
        )
    # Cleaning quality
    if data["quality"]["count"]:
        q = data["quality"]
        cmts = ""
        if q["comments"]:
            cmts = "\n" + "\n".join(f"  • **{c['unit']}** ({c['score']}★): {c['text']}"
                                    for c in q["comments"])
        embed.add_field(
            name=f"⭐ تقييمات نظافة ({q['count']})",
            value=f"المتوسط: **{q['avg']}** ⭐{cmts}"[:1024],
            inline=False,
        )
    embed.set_footer(text=f"إجمالي ضيوف VIP حتى الآن: {data['vip_count']} · "
                          f"ما تعلّمته يُرسَل يومياً الساعة {WILT_HOUR:02d}:00 · "
                          f"غيّر القناة عبر متغيّر WILT_CHANNEL")
    return embed

async def post_wilt():
    """Compose today's WILT, post it to #wilt, ping the team. Idempotent — won't
    repost for the same date thanks to _wilt_last_date."""
    global _wilt_last_date
    if not WILT_ENABLED:
        return
    today = datetime.now(TZ).date()
    if _wilt_last_date == today and not WILT_TEST:
        return
    guild = bot.get_guild(GUILD_ID)
    if guild is None:
        return
    category = await get_assistant_category(guild)
    channel = await ensure_channel(guild, WILT_CHANNEL, category)
    if channel is None:
        return
    data = await asyncio.to_thread(compose_wilt)
    prose = await asyncio.to_thread(claude_wilt_prose, data)
    try:
        await channel.send(embed=_wilt_embed(data, prose or ""))
        _wilt_last_date = today
        log_event("report", f"WILT · ملخص يومي · {data['funnel']['total']} رد، "
                            f"{len(data['corrections'])} تصحيح، {data['vip_count']} VIP")
    except Exception as e:
        print("post_wilt error:", e)
LEARNING_MIN_NEW_EXAMPLES = int(os.environ.get("LEARNING_MIN_NEW_EXAMPLES", "5"))   # don't re-distill until N new entries
LEARNING_DISTILL_MIN = int(os.environ.get("LEARNING_DISTILL_MIN", "30"))           # background distill interval (min)
LEARNING_SAMPLE_PER_APT = int(os.environ.get("LEARNING_SAMPLE_PER_APT", "40"))     # last N entries per apt to feed Claude

# ---- Code-not-received reminder: nudge guests who haven't signed the agreement
# before their check-in, since Hostaway only releases the door code AFTER they sign.
AGREEMENT_REMINDER_ENABLED = os.environ.get("AGREEMENT_REMINDER_ENABLED", "1") in ("1","true","True","yes")
AGREEMENT_REMINDER_LEAD_HOURS = float(os.environ.get("AGREEMENT_REMINDER_LEAD_HOURS", "1"))   # remind when check-in is N hours away
AGREEMENT_REMINDER_POLL_MIN = int(os.environ.get("AGREEMENT_REMINDER_POLL_MIN", "10"))         # how often the loop runs
_agreement_reminded = set()   # set of reservation_ids we've already nudged (persisted)

# ---------------- Deep-clean schedule ----------------
# Every apartment is deep-cleaned every 45-60 days. We pre-schedule, and at 9pm
# the night before we either CONFIRM (and block the day in Hostaway so it's
# unbookable) or PUSH to the next available slot if a guest booked it last-minute.
# Avoid Thu/Fri (peak weekend demand). Only ONE apartment per day.
DEEPCLEAN_ENABLED      = os.environ.get("DEEPCLEAN_ENABLED", "1") in ("1","true","True","yes")
DEEPCLEAN_MIN_DAYS     = int(os.environ.get("DEEPCLEAN_MIN_DAYS", "45"))
DEEPCLEAN_MAX_DAYS     = int(os.environ.get("DEEPCLEAN_MAX_DAYS", "60"))
DEEPCLEAN_DEFAULT_LAST = os.environ.get("DEEPCLEAN_DEFAULT_LAST", "2026-04-27")
DEEPCLEAN_AVOID_WD     = set(int(x) for x in os.environ.get("DEEPCLEAN_AVOID_WD", "3,4").split(",") if x.strip().isdigit())  # 3=Thu, 4=Fri
DEEPCLEAN_CONFIRM_HOUR = int(os.environ.get("DEEPCLEAN_CONFIRM_HOUR", "21"))  # 9pm Riyadh
CLEANING_TOKEN         = os.environ.get("CLEANING_TOKEN", "")   # public link gate

# ---------------- Working hours / off-hours behavior ----------------
# Team is active 11:00–01:30 (next day). Outside this window we:
#   1) Stop counting SLA against the team
#   2) Send a one-time "we're back at HOURS" auto-reply to the guest
WORK_START_HOUR  = int(os.environ.get("WORK_START_HOUR", "11"))
WORK_END_HOUR    = int(os.environ.get("WORK_END_HOUR", "25"))    # 25 == 1am next day (i.e. 24+1)
WORK_END_MIN     = int(os.environ.get("WORK_END_MIN", "30"))     # 30 → 1:30 am
OFFHOURS_AUTOREPLY_ENABLED = os.environ.get("OFFHOURS_AUTOREPLY_ENABLED", "1") in ("1","true","True","yes")
_offhours_acked_convos = set()    # conversation_ids we already auto-replied to in the current off-hours window

# ---------------- WILT: What I Learned Today ----------------
# Once a day at WILT_HOUR Riyadh, post a daily journal-style summary to
# #wilt with: counters, what changed, who corrected what, which apartments
# got new facts, plus a 3-4 sentence Najdi prose summary from Claude.
WILT_ENABLED  = os.environ.get("WILT_ENABLED", "1") in ("1","true","True","yes")
WILT_CHANNEL  = os.environ.get("WILT_CHANNEL", "wilt")
WILT_HOUR     = int(os.environ.get("WILT_HOUR", "23"))      # 11pm Riyadh
WILT_TEST     = os.environ.get("WILT_TEST", "0") in ("1","true","True","yes")
_wilt_last_date = None

# ---------------- Guest profiles ----------------
# A "guest profile" is keyed by best-available stable identifier:
#   1) phone number (normalized) if provided
#   2) lowercase guest email
#   3) lowercase guest name (last resort — risk of merging different people
#      with the same name)
# Each profile aggregates reservations seen, conversation summaries, preferences
# the bot has learned, and a VIP flag toggled when a guest has booked >= 2 times
# or has stayed >= 5 nights total.
_guest_profiles = {}    # key -> {key, names:[], phone, email, reservations:[{id,unit,checkin,checkout,nights,total,ts}],
                        #         summaries:[{ts,conversation_id,text}], notes, vip, first_seen, last_seen, total_nights, total_revenue}
GUEST_VIP_MIN_STAYS  = int(os.environ.get("GUEST_VIP_MIN_STAYS", "2"))
GUEST_VIP_MIN_NIGHTS = int(os.environ.get("GUEST_VIP_MIN_NIGHTS", "5"))
GUEST_SUMMARY_REFRESH_MIN = int(os.environ.get("GUEST_SUMMARY_REFRESH_MIN", "60"))   # how often to re-distill summaries

# ---------------- Cleaning quality feedback ----------------
# After each completed deep-clean, when the NEXT guest checks into that unit
# we send them a one-tap rating link (1-5 stars). The page lives at
# /clean-feedback?id=<token> and writes scores back to a per-unit ledger so
# the owner sees which units are slipping + each cleaning company's avg score.
CLEAN_FEEDBACK_ENABLED = os.environ.get("CLEAN_FEEDBACK_ENABLED", "1") in ("1","true","True","yes")
CLEAN_FEEDBACK_DELAY_HOURS = int(os.environ.get("CLEAN_FEEDBACK_DELAY_HOURS", "3"))   # ask N hours after check-in
_cleaning_feedback = {}     # token -> {lid, unit, guest, ts_sent, score, comment, ts_done}
_cleaning_feedback_sent = set()   # reservation_ids we've already pinged so we don't double-ask

def _normalize_phone(s):
    s = "".join(c for c in (s or "") if c.isdigit())
    if not s:
        return ""
    # KSA: 9665XXXXXXXX or 05XXXXXXXX → unify to +9665XXXXXXXX
    if s.startswith("00"):
        s = s[2:]
    if s.startswith("0") and len(s) == 10:
        s = "966" + s[1:]
    return "+" + s

def _profile_key(name, phone="", email=""):
    p = _normalize_phone(phone)
    if p and len(p) > 7:
        return "ph:" + p
    e = (email or "").strip().lower()
    if e and "@" in e:
        return "em:" + e
    n = (name or "").strip().lower()
    return "nm:" + n if n else ""

def _ensure_profile(key, name="", phone="", email=""):
    if not key:
        return None
    if key not in _guest_profiles:
        _guest_profiles[key] = {
            "key": key, "names": [], "phone": _normalize_phone(phone),
            "email": (email or "").strip().lower(),
            "reservations": [], "summaries": [], "notes": "",
            "vip": False, "first_seen": datetime.now(TZ).isoformat(timespec="minutes"),
            "last_seen": datetime.now(TZ).isoformat(timespec="minutes"),
            "total_nights": 0, "total_revenue": 0.0,
        }
    p = _guest_profiles[key]
    nm = (name or "").strip()
    if nm and nm not in p["names"]:
        p["names"].append(nm)
        p["names"] = p["names"][-5:]      # keep last 5 spellings
    if phone and not p["phone"]:
        p["phone"] = _normalize_phone(phone)
    if email and not p["email"]:
        p["email"] = (email or "").strip().lower()
    p["last_seen"] = datetime.now(TZ).isoformat(timespec="minutes")
    return p

def _recompute_vip(profile):
    stays = len(profile.get("reservations", []))
    nights = profile.get("total_nights", 0) or sum(r.get("nights", 0) for r in profile.get("reservations", []))
    profile["total_nights"] = nights
    profile["vip"] = (stays >= GUEST_VIP_MIN_STAYS) or (nights >= GUEST_VIP_MIN_NIGHTS)

def record_guest_stay(item):
    """Called whenever we encounter a confirmed reservation in conversation handling.
    Builds/updates the profile and writes the reservation into it (dedup by id)."""
    name = item.get("guest") or ""
    phone = item.get("guest_phone") or ""
    email = item.get("guest_email") or ""
    key = _profile_key(name, phone, email)
    if not key:
        return None
    p = _ensure_profile(key, name=name, phone=phone, email=email)
    res_id = item.get("reservation_id")
    if not res_id:
        return p
    # Dedup
    if not any(r.get("id") == res_id for r in p["reservations"]):
        p["reservations"].append({
            "id": res_id,
            "unit": item.get("unit", ""),
            "listing_id": item.get("listing_id"),
            "checkin": item.get("checkin"),
            "checkout": item.get("checkout"),
            "nights": _res_nights({"arrivalDate": item.get("checkin"),
                                    "departureDate": item.get("checkout")}),
            "ts": datetime.now(TZ).isoformat(timespec="minutes"),
        })
        p["reservations"] = p["reservations"][-30:]
    _recompute_vip(p)
    return p

# Periodic Claude-powered conversation summarisation per profile.
def _summarise_conversation_for_profile(p, conv_id, recent_text):
    """Ask Claude for a 2-3 line summary of a single conversation (no PII).
    Skipped silently if no key or text is empty."""
    if not ANTHROPIC_API_KEY or not (recent_text or "").strip():
        return None
    sys = ("لخّص بسطرين أو ثلاثة محادثة ضيف عوجا بحيث يقدر الفريق يتذكر سياقها في الزيارة الجاية: "
           "وش طلب الضيف، وش حصل، إذا فيه تفضيل واضح يستحق التذكّر (مثلاً يحب الطابق العالي، يكره الضوضاء، "
           "عائلة فيها أطفال، طلب تشيك-إن مبكّر سابقاً، إلخ). تجاهل أرقام الهاتف والبريد. عربي سعودي مختصر."
           + _DIALECT_LOCK)
    user = f"المحادثة:\n{recent_text[:6000]}\n\nاكتب الملخص المختصر."
    return claude_text(sys, user, max_tokens=300, model=CLAUDE_MODEL_PREMIUM) \
        or claude_text(sys, user, max_tokens=300)

def is_within_working_hours(dt=None):
    dt = dt or datetime.now(TZ)
    h, m = dt.hour, dt.minute
    minutes = h * 60 + m
    start = WORK_START_HOUR * 60
    end = WORK_END_HOUR * 60 + WORK_END_MIN
    # If end > 24*60, the window wraps past midnight (e.g. 11:00 → 25:30 = 01:30 next day).
    if end <= 24 * 60:
        return start <= minutes < end
    # wrapped window: working if minutes >= start  OR  minutes < (end - 24*60)
    return minutes >= start or minutes < (end - 24 * 60)

def next_work_start(dt=None):
    """Return the next datetime when working hours resume."""
    dt = dt or datetime.now(TZ)
    today_start = dt.replace(hour=WORK_START_HOUR, minute=0, second=0, microsecond=0)
    if dt < today_start:
        return today_start
    # Otherwise tomorrow's start
    return today_start + timedelta(days=1)
# lid (int) -> {last_done, next_scheduled, next_status, history:[{date,ts,notes}], notes}
_deep_clean_state = {}

def log_event(category, text):
    """Record something the bot did, for the dashboard's activity feed."""
    try:
        _activity.append({"ts": datetime.now(TZ).isoformat(timespec="seconds"),
                          "cat": category, "text": str(text)[:300]})
    except Exception:
        pass
    print(f"[{category}] {text}")

# ---------------- Dashboard: auth + cached analytics + API + page ----------------
_dash_cache = {}
_replied_msgs = set()        # guard so a reply isn't sent twice (Discord + dashboard)
_last_price_changes = {}     # lid -> [changes] from the latest pricing compute (for dashboard Apply)
_last_price_detail = {}      # lid -> {name, base, confidence, rows:[per-date decision + why]}

DASHBOARD_HTML = """<!doctype html>
<html lang="ar" dir="rtl" data-theme="auto">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<title>عوجا · Ouja Operations</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@300;400;500;600;700&family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@500;600&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0;-webkit-font-smoothing:antialiased}

:root{
  --font-ar:'IBM Plex Sans Arabic','Inter',-apple-system,BlinkMacSystemFont,'SF Pro Text',system-ui,sans-serif;
  --font-en:'Inter',-apple-system,BlinkMacSystemFont,'SF Pro Text',system-ui,sans-serif;
  --font-mono:'JetBrains Mono','SF Mono','Menlo',monospace;

  --bg:#FAFAF7;
  --surface:#FFFFFF;
  --surface-2:#F5F2EC;
  --surface-3:#EDE8DC;
  --line:#E8E2D5;
  --line-strong:#D4CDB9;

  --text:#1A1815;
  --text-2:#544D43;
  --text-3:#8C8475;
  --mut:#A09989;

  --gold:#A37728;
  --gold-2:#8B6320;
  --gold-soft:#F4EBD5;
  --gold-tint:rgba(163,119,40,0.07);

  --green:#0E9E5F;
  --green-soft:#DCF3E6;
  --red:#C44343;
  --red-soft:#FAE3E3;
  --yellow:#C99617;
  --yellow-soft:#FAEED1;
  --blue:#2F6FD0;
  --blue-soft:#E0EBFA;
  --purple:#6D58C2;
  --purple-soft:#EAE6F8;

  --r-xs:5px; --r-sm:7px; --r:10px; --r-lg:14px;
  --sh-xs:0 1px 2px rgba(26,24,21,0.04);
  --sh-sm:0 2px 5px rgba(26,24,21,0.05),0 1px 2px rgba(26,24,21,0.03);
  --sh-md:0 6px 16px rgba(26,24,21,0.07),0 2px 4px rgba(26,24,21,0.04);
  --sh-lg:0 16px 40px rgba(26,24,21,0.10),0 4px 12px rgba(26,24,21,0.06);
  --sh-drawer:-12px 0 40px rgba(26,24,21,0.10);

  --side:232px;
  --drawer:520px;
  --topbar:56px;
}

html[data-theme="dark"]{
  --bg:#0E0D0C;
  --surface:#18171A;
  --surface-2:#221F1D;
  --surface-3:#2A2722;
  --line:#2C2924;
  --line-strong:#3A362E;

  --text:#F5F0E6;
  --text-2:#C5BEAF;
  --text-3:#928A78;
  --mut:#6E6759;

  --gold:#D4A854;
  --gold-2:#B8893F;
  --gold-soft:#2B2317;
  --gold-tint:rgba(212,168,84,0.08);

  --green:#3ECF8E;
  --green-soft:#0F2E1F;
  --red:#E25C5C;
  --red-soft:#2E1414;
  --yellow:#E9B94A;
  --yellow-soft:#2D2210;
  --blue:#5B9EFF;
  --blue-soft:#0D1F35;
  --purple:#8E78D9;
  --purple-soft:#1B1530;
}
@media (prefers-color-scheme:dark){
  html[data-theme="auto"]{
    --bg:#0E0D0C;--surface:#18171A;--surface-2:#221F1D;--surface-3:#2A2722;
    --line:#2C2924;--line-strong:#3A362E;
    --text:#F5F0E6;--text-2:#C5BEAF;--text-3:#928A78;--mut:#6E6759;
    --gold:#D4A854;--gold-2:#B8893F;--gold-soft:#2B2317;--gold-tint:rgba(212,168,84,0.08);
    --green:#3ECF8E;--green-soft:#0F2E1F;--red:#E25C5C;--red-soft:#2E1414;
    --yellow:#E9B94A;--yellow-soft:#2D2210;--blue:#5B9EFF;--blue-soft:#0D1F35;
    --purple:#8E78D9;--purple-soft:#1B1530;
  }
}

html,body{font-family:var(--font-ar);background:var(--bg);color:var(--text);font-size:13.5px;line-height:1.55;min-height:100vh}
html[lang="en"] body{font-family:var(--font-en)}
body{padding:env(safe-area-inset-top) 0 env(safe-area-inset-bottom)}
button,input,textarea,select{font:inherit;color:inherit}
button{border:none;background:none;cursor:pointer;-webkit-tap-highlight-color:transparent}
input,textarea,select{background:var(--surface);border:1px solid var(--line);color:var(--text);border-radius:var(--r-sm);padding:8px 11px;transition:.15s border-color}
input:focus,textarea:focus,select:focus{outline:none;border-color:var(--gold);box-shadow:0 0 0 3px var(--gold-tint)}
textarea{min-height:80px;resize:vertical;line-height:1.6;font-family:inherit;width:100%}
select{appearance:none;background-image:linear-gradient(45deg,transparent 50%,var(--text-3) 50%),linear-gradient(135deg,var(--text-3) 50%,transparent 50%);background-position:calc(100% - 14px) 14px,calc(100% - 9px) 14px;background-size:5px 5px;background-repeat:no-repeat;padding-inline-end:28px}
html[dir="rtl"] select{background-position:14px 14px,9px 14px}
a{color:inherit;text-decoration:none;cursor:pointer}
.mono{font-family:var(--font-mono);font-variant-numeric:tabular-nums;letter-spacing:-0.02em}

/* ============== LOGIN ============== */
#login{position:fixed;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:18px;padding:24px;background:var(--bg);z-index:1000}
#login .brand-lg{font-size:34px;font-weight:700;color:var(--gold)}
#login .sub{color:var(--mut);font-size:12.5px;margin-top:-10px;letter-spacing:1.5px;text-transform:uppercase}
#login input{max-width:340px;text-align:center;font-size:15px;padding:13px;border-radius:var(--r);box-shadow:var(--sh-sm);width:100%}
#login .err{color:var(--red);font-size:12.5px;min-height:18px}

/* ============== APP SHELL ============== */
#app{display:none}
.shell{display:grid;min-height:100vh;width:100%}

/* Grid-areas decouple visual placement from HTML source order.
   In RTL, CSS-grid auto-reverses columns, so "side main" puts the sidebar
   on the user's right and the main content on the user's left automatically. */
header.mhead{grid-area:head}
main.main{grid-area:main}
aside.side{grid-area:side}

/* Desktop ≥ 1024 */
@media (min-width:1024px){
  .shell{
    grid-template-columns:var(--side) 1fr;
    grid-template-rows:100vh;
    grid-template-areas:"side main";
  }
}
/* Tablet/Mobile */
@media (max-width:1023px){
  .shell{
    grid-template-columns:1fr;
    grid-template-rows:auto 1fr;
    grid-template-areas:"head" "main";
    padding-bottom:64px;
  }
}

/* ============== SIDEBAR (desktop) ============== */
aside.side{display:none;background:var(--surface);border-inline-end:1px solid var(--line);flex-direction:column;padding:18px 12px;position:sticky;top:0;height:100vh;overflow-y:auto}
@media (min-width:1024px){ aside.side{display:flex} }
.side-brand{display:flex;align-items:center;gap:10px;padding:4px 8px 14px;margin-bottom:8px}
.side-brand .logo{width:32px;height:32px;border-radius:8px;background:linear-gradient(135deg,var(--gold),var(--gold-2));display:flex;align-items:center;justify-content:center;color:#fff;font-weight:700;font-size:15px;box-shadow:var(--sh-sm)}
.side-brand .name{font-size:16px;font-weight:700;color:var(--text);line-height:1}
.side-brand .sub{font-size:9.5px;color:var(--mut);margin-top:2px;letter-spacing:1px;text-transform:uppercase}
.side-nav{display:flex;flex-direction:column;gap:1px;flex:1}
.side-nav .item{display:flex;align-items:center;gap:9px;padding:8px 11px;border-radius:var(--r-sm);color:var(--text-2);font-size:13px;font-weight:500;cursor:pointer;transition:.12s;position:relative;user-select:none}
.side-nav .item:hover{background:var(--surface-2);color:var(--text)}
.side-nav .item.on{background:var(--gold-tint);color:var(--gold);font-weight:600}
.side-nav .item.on::before{content:'';position:absolute;inset-inline-start:0;top:8px;bottom:8px;width:3px;background:var(--gold);border-radius:2px}
.side-nav .item .ic{font-size:14px;width:18px;text-align:center;line-height:1}
.side-nav .item .badge{margin-inline-start:auto;background:var(--red);color:#fff;font-size:10px;font-weight:700;padding:1px 6px;border-radius:9px;min-width:17px;text-align:center;line-height:1.3}
.side-nav .item.on .badge{background:var(--gold)}
.side-foot{display:flex;flex-direction:column;gap:6px;padding-top:12px;border-top:1px solid var(--line);margin-top:6px}
.side-status{font-size:10.5px;color:var(--mut);display:flex;align-items:center;gap:6px;padding:0 8px;margin-bottom:6px}
.dot{width:7px;height:7px;border-radius:50%;background:var(--green);box-shadow:0 0 0 3px rgba(14,158,95,.18);flex-shrink:0}
.dot.warm{background:var(--yellow);box-shadow:0 0 0 3px rgba(201,150,23,.18)}
.side-tools{display:flex;gap:5px}
.side-tools .icbtn{flex:1}

/* ============== TOP BAR (mobile/tablet) ============== */
header.mhead{display:none;background:var(--surface);border-bottom:1px solid var(--line);padding:10px 14px;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:50}
@media (max-width:1023px){ header.mhead{display:flex} }
.mhead-brand{display:flex;align-items:center;gap:9px}
.mhead-brand .logo{width:28px;height:28px;border-radius:7px;background:linear-gradient(135deg,var(--gold),var(--gold-2));display:flex;align-items:center;justify-content:center;color:#fff;font-weight:700;font-size:13px}
.mhead-brand .name{font-size:15px;font-weight:700;color:var(--text)}
.mhead-tools{display:flex;gap:5px}

/* ============== ICON BUTTONS ============== */
.icbtn{width:32px;height:32px;border-radius:7px;background:var(--surface);border:1px solid var(--line);color:var(--text-2);font-size:13px;display:inline-flex;align-items:center;justify-content:center;transition:.12s}
.icbtn:hover{color:var(--gold);border-color:var(--gold);background:var(--gold-tint)}

/* ============== BUTTONS ============== */
.btn{padding:7px 13px;border-radius:7px;font-size:12.5px;font-weight:600;display:inline-flex;align-items:center;gap:5px;transition:.12s;border:1px solid transparent;line-height:1.2;white-space:nowrap}
.btn.primary{background:linear-gradient(135deg,var(--gold),var(--gold-2));color:#fff;box-shadow:var(--sh-xs)}
.btn.primary:hover{filter:brightness(1.05);box-shadow:var(--sh-sm)}
.btn.green{background:var(--green-soft);color:var(--green);border-color:rgba(14,158,95,.22)}
.btn.green:hover{background:var(--green);color:#fff;border-color:var(--green)}
.btn.red{background:var(--red-soft);color:var(--red);border-color:rgba(196,67,67,.20)}
.btn.red:hover{background:var(--red);color:#fff;border-color:var(--red)}
.btn.ghost{background:var(--surface);color:var(--text-2);border-color:var(--line)}
.btn.ghost:hover{color:var(--gold);border-color:var(--gold);background:var(--gold-tint)}
.btn.xs{padding:4px 9px;font-size:11.5px;border-radius:6px}
.btn.sm{padding:6px 11px;font-size:12px;border-radius:6px}
.btn:disabled{opacity:.5;cursor:default;filter:none}

/* ============== MAIN ============== */
main.main{padding:20px 24px 48px;overflow-x:hidden;min-width:0;max-width:100%}
@media (max-width:1023px){ main.main{padding:14px 14px 24px} }

.page-head{display:flex;align-items:flex-end;justify-content:space-between;margin-bottom:16px;gap:14px;flex-wrap:wrap}
.page-title{font-size:20px;font-weight:700;color:var(--text);letter-spacing:-.2px;line-height:1.15}
.page-sub{color:var(--mut);font-size:12px;margin-top:2px}
.page-tools{display:flex;gap:7px;align-items:center;flex-wrap:wrap}

/* ============== KPI STRIP ============== */
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:10px;margin-bottom:18px}
@media (max-width:767px){ .kpis{grid-template-columns:repeat(2,1fr);gap:8px} }
.kpi{background:var(--surface);border:1px solid var(--line);border-radius:var(--r-lg);padding:13px 14px;transition:.15s;box-shadow:var(--sh-xs);position:relative;overflow:hidden}
.kpi:hover{box-shadow:var(--sh-sm);border-color:var(--line-strong)}
.kpi-head{display:flex;align-items:center;justify-content:space-between;margin-bottom:7px}
.kpi-ic{width:26px;height:26px;border-radius:7px;background:var(--gold-soft);color:var(--gold);display:flex;align-items:center;justify-content:center;font-size:12.5px}
.kpi-ic.g{background:var(--green-soft);color:var(--green)} .kpi-ic.b{background:var(--blue-soft);color:var(--blue)} .kpi-ic.r{background:var(--red-soft);color:var(--red)} .kpi-ic.y{background:var(--yellow-soft);color:var(--yellow)} .kpi-ic.p{background:var(--purple-soft);color:var(--purple)}
.kpi-val{font-size:23px;font-weight:700;letter-spacing:-.5px;line-height:1.05;color:var(--text);font-family:var(--font-mono)}
.kpi-val.gold{color:var(--gold)} .kpi-val.green{color:var(--green)} .kpi-val.red{color:var(--red)}
.kpi-lbl{color:var(--mut);font-size:11px;margin-top:3px;font-weight:500}
.kpi-delta{font-size:10.5px;font-weight:600;padding:1px 6px;border-radius:5px;font-family:var(--font-mono)}
.kpi-delta.up{background:var(--green-soft);color:var(--green)}
.kpi-delta.dn{background:var(--red-soft);color:var(--red)}

/* ============== CARDS ============== */
.card{background:var(--surface);border:1px solid var(--line);border-radius:var(--r-lg);padding:16px;margin-bottom:14px;box-shadow:var(--sh-xs)}
.card-head{display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;gap:8px}
.card-title{font-size:13.5px;font-weight:700;color:var(--text);display:flex;align-items:center;gap:7px}
.card-sub{color:var(--mut);font-size:11.5px}
.card-actions{display:flex;gap:5px;align-items:center}

.grid2{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px}
.grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:14px}
@media (max-width:1023px){.grid2,.grid3{grid-template-columns:1fr;gap:10px}}

/* ============== PILLS / TAGS ============== */
.pill{display:inline-flex;align-items:center;gap:4px;padding:2px 8px;border-radius:5px;font-size:10.5px;font-weight:600;letter-spacing:.2px;line-height:1.5}
.pill.ok{background:var(--green-soft);color:var(--green)}
.pill.warn{background:var(--yellow-soft);color:var(--yellow)}
.pill.danger{background:var(--red-soft);color:var(--red)}
.pill.info{background:var(--blue-soft);color:var(--blue)}
.pill.purple{background:var(--purple-soft);color:var(--purple)}
.pill.muted{background:var(--surface-2);color:var(--mut)}
.pill.gold{background:var(--gold-soft);color:var(--gold)}

/* ============== FILTER BAR ============== */
.filterbar{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:14px;background:var(--surface);border:1px solid var(--line);padding:10px 12px;border-radius:var(--r-lg);box-shadow:var(--sh-xs)}
.filterbar select,.filterbar input{padding:6px 10px;font-size:12px;height:32px;border-radius:6px;min-width:130px}
.filterbar input{min-width:auto;width:140px}
.tabsfilter{display:flex;background:var(--surface-2);padding:3px;border-radius:7px;gap:1px}
.tabsfilter button{padding:6px 12px;border-radius:5px;font-size:12px;font-weight:500;color:var(--text-2);transition:.12s}
.tabsfilter button:hover{color:var(--text)}
.tabsfilter button.on{background:var(--surface);color:var(--text);font-weight:600;box-shadow:var(--sh-xs)}
.filterbar .clear{margin-inline-start:auto;font-size:11.5px;color:var(--mut);padding:6px 10px}
.filterbar .clear:hover{color:var(--gold)}

/* ============== INBOX LIST ============== */
.inbox-list{display:flex;flex-direction:column;gap:6px}
.ibox{background:var(--surface);border:1px solid var(--line);border-radius:var(--r);transition:.12s;overflow:hidden}
.ibox:hover{border-color:var(--line-strong);box-shadow:var(--sh-xs)}
.ibox.escalation{border-inline-start:3px solid var(--red)}
.ibox.reply{border-inline-start:3px solid var(--gold)}
.ibox-row{display:grid;grid-template-columns:auto 1fr auto auto;gap:12px;padding:10px 13px;align-items:center;cursor:pointer}
@media (max-width:767px){.ibox-row{grid-template-columns:auto 1fr auto;gap:8px;padding:9px 10px}}
.ibox-icon{width:30px;height:30px;border-radius:7px;display:flex;align-items:center;justify-content:center;font-size:13px;flex-shrink:0}
.ibox-icon.esc{background:var(--red-soft);color:var(--red)}
.ibox-icon.rep{background:var(--gold-soft);color:var(--gold)}
.ibox-main{min-width:0}
.ibox-top{display:flex;align-items:center;gap:7px;margin-bottom:2px}
.ibox-who{font-weight:600;font-size:13px;color:var(--text)}
.ibox-unit{font-size:11px;color:var(--mut);background:var(--surface-2);padding:1px 7px;border-radius:5px}
.ibox-preview{font-size:12px;color:var(--text-3);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:100%}
.ibox-meta{display:flex;flex-direction:column;align-items:flex-end;gap:3px}
.ibox-time{font-size:10.5px;color:var(--mut);font-family:var(--font-mono);white-space:nowrap}
.ibox-conf{font-size:10px;font-weight:600;padding:1px 6px;border-radius:4px;font-family:var(--font-mono)}
.ibox-conf.high{background:var(--green-soft);color:var(--green)}
.ibox-conf.mid{background:var(--yellow-soft);color:var(--yellow)}
.ibox-conf.low{background:var(--red-soft);color:var(--red)}
.ibox-expand{color:var(--mut);font-size:14px;transition:.15s transform;flex-shrink:0}
.ibox.open .ibox-expand{transform:rotate(180deg)}
.ibox-body{display:none;border-top:1px solid var(--line);padding:14px}
.ibox.open .ibox-body{display:block;animation:slideDown .18s ease}
@keyframes slideDown{from{opacity:0;transform:translateY(-4px)}to{opacity:1;transform:none}}

/* Conversation thread */
.context-grid{display:grid;grid-template-columns:1.6fr 1fr;gap:14px;margin-bottom:14px}
@media (max-width:900px){.context-grid{grid-template-columns:1fr}}
.thread{display:flex;flex-direction:column;gap:7px;max-height:340px;overflow-y:auto;padding:8px;background:var(--surface-2);border-radius:var(--r);border:1px solid var(--line)}
.bub{display:flex;flex-direction:column;max-width:88%}
.bub.g{align-self:flex-start}
.bub.h{align-self:flex-end}
.bub-meta{font-size:10px;color:var(--mut);margin-bottom:2px;padding:0 4px;font-family:var(--font-mono)}
.bub.h .bub-meta{text-align:end}
.bub-tx{padding:8px 12px;border-radius:11px;font-size:12.5px;line-height:1.55;white-space:pre-wrap;word-wrap:break-word}
.bub.g .bub-tx{background:var(--surface);color:var(--text-2);border-bottom-inline-start-radius:3px;border:1px solid var(--line)}
.bub.h .bub-tx{background:var(--gold-tint);color:var(--text);border-bottom-inline-end-radius:3px;border:1px solid rgba(163,119,40,.18)}
.bub.auto .bub-tx{background:var(--surface-3);color:var(--text-3);font-style:italic}
.bub-auto-tag{font-size:9px;color:var(--mut);text-transform:uppercase;letter-spacing:.5px;padding:0 5px}

.context-box{background:var(--surface-2);border:1px solid var(--line);border-radius:var(--r);padding:11px}
.context-row{display:flex;justify-content:space-between;align-items:center;padding:6px 0;border-bottom:1px solid var(--line);font-size:12px}
.context-row:last-child{border:none}
.context-row .l{color:var(--mut)}
.context-row .v{color:var(--text);font-weight:600;font-family:var(--font-mono)}
.context-h{font-size:11px;color:var(--mut);text-transform:uppercase;letter-spacing:.5px;font-weight:600;margin:10px 0 6px}
.context-h:first-child{margin-top:0}

.reasoning-box{background:var(--purple-soft);border:1px solid rgba(109,88,194,.18);border-radius:var(--r);padding:11px;margin-bottom:12px}
html[data-theme="dark"] .reasoning-box{background:rgba(109,88,194,.10)}
.reasoning-box .h{font-size:11px;color:var(--purple);text-transform:uppercase;letter-spacing:.5px;font-weight:700;margin-bottom:6px;display:flex;align-items:center;gap:5px}
.reasoning-box .reason-txt{font-size:12.5px;color:var(--text);line-height:1.55}
.reasoning-chips{display:flex;flex-wrap:wrap;gap:5px;margin-top:7px}

/* Draft area */
.draft-label{font-size:10.5px;color:var(--gold);text-transform:uppercase;letter-spacing:.5px;font-weight:700;margin:14px 0 6px;display:flex;align-items:center;gap:5px}
.action-row{display:flex;gap:7px;flex-wrap:wrap;margin-top:10px;align-items:center}
.action-row input{flex:1;min-width:130px;font-size:12px;height:32px}

/* Teach inline form */
.teach-form{display:none;background:var(--surface-2);border:1px solid var(--line);border-radius:var(--r);padding:11px;margin-top:10px}
.teach-form.open{display:block}
.teach-form input,.teach-form textarea{font-size:12px;margin-bottom:6px;width:100%}
.teach-form .row{display:flex;gap:6px;justify-content:flex-end}

/* ============== EMPTY-UNITS GRID (today) ============== */
.empty-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:11px}
.eu{background:var(--surface);border:1px solid var(--line);border-radius:var(--r-lg);padding:13px;transition:.12s;position:relative;overflow:hidden}
.eu:hover{border-color:var(--line-strong);box-shadow:var(--sh-sm)}
.eu.skipped{border-color:var(--yellow);background:linear-gradient(135deg,var(--yellow-soft),var(--surface))}
.eu.skipped::after{content:'⏸';position:absolute;top:8px;inset-inline-end:10px;color:var(--yellow);font-size:16px;font-weight:700}
.eu-top{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:9px;gap:8px}
.eu-name{font-size:13.5px;font-weight:700;color:var(--text);min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.eu-now{display:flex;align-items:baseline;gap:5px;margin-bottom:11px}
.eu-now .lbl{font-size:11px;color:var(--mut)}
.eu-now .v{font-size:21px;font-weight:700;color:var(--gold);font-family:var(--font-mono)}
.eu-now .v.struck{text-decoration:line-through;color:var(--mut);font-size:14px}

.timeline{display:flex;align-items:center;gap:4px;margin-bottom:10px}
.tier{flex:1;text-align:center;padding:7px 4px;background:var(--surface-2);border:1px solid var(--line);border-radius:var(--r-sm);font-size:11px;color:var(--text-2);position:relative;transition:.12s}
.tier .tlbl{font-size:9.5px;color:var(--mut);text-transform:uppercase;letter-spacing:.4px;display:block;margin-bottom:2px}
.tier .tprice{font-size:12.5px;font-weight:700;color:var(--text);font-family:var(--font-mono);display:block}
.tier .tpct{font-size:9.5px;color:var(--red);font-weight:600;display:block;margin-top:1px}
.tier.next{background:var(--gold-tint);border-color:var(--gold);color:var(--gold)}
.tier.next .tlbl{color:var(--gold)}
.tier.passed{opacity:.45}
.tline-arrow{font-size:11px;color:var(--mut)}

.eu-actions{display:flex;gap:6px;margin-top:6px}

/* ============== DRAWER (right side, for details) ============== */
.drawer-backdrop{position:fixed;inset:0;background:rgba(26,24,21,.30);z-index:90;opacity:0;pointer-events:none;transition:.22s;backdrop-filter:blur(2px)}
.drawer-backdrop.show{opacity:1;pointer-events:auto}
/* Drawer always slides in from the END side (opposite of the sidebar):
   right edge in LTR, left edge in RTL. inset-inline-end + a single transform
   rule per direction keeps specificity even so .open always wins. */
.drawer{position:fixed;top:0;bottom:0;inset-inline-end:0;width:min(var(--drawer),100vw);background:var(--surface);box-shadow:var(--sh-drawer);z-index:91;display:flex;flex-direction:column;transition:transform .25s ease;transform:translateX(110%)}
html[dir="rtl"] .drawer{transform:translateX(-110%)}
html[dir="ltr"] .drawer.open,
html[dir="rtl"] .drawer.open{transform:translateX(0)}
.drawer-head{padding:14px 18px;border-bottom:1px solid var(--line);display:flex;justify-content:space-between;align-items:center;gap:10px}
.drawer-title{font-size:15px;font-weight:700;color:var(--text);min-width:0;overflow:hidden;text-overflow:ellipsis}
.drawer-sub{font-size:11.5px;color:var(--mut);margin-top:2px}
.drawer-body{flex:1;overflow-y:auto;padding:16px 18px}
.drawer-foot{padding:12px 18px;border-top:1px solid var(--line);display:flex;gap:8px;justify-content:flex-end;background:var(--surface-2)}
@media (max-width:1023px){.drawer{width:100vw}.drawer-body{padding:14px 14px calc(80px + env(safe-area-inset-bottom))}.drawer-head{padding:12px 14px}.drawer-foot{padding:10px 14px}}

/* ============== PRICING DETAIL TABLE ============== */
table.data{width:100%;border-collapse:collapse;font-size:12px}
table.data th{padding:8px 7px;color:var(--mut);font-weight:600;font-size:10.5px;text-align:start;border-bottom:1px solid var(--line);text-transform:uppercase;letter-spacing:.4px;white-space:nowrap;position:sticky;top:0;background:var(--surface);z-index:1}
table.data td{padding:9px 7px;border-bottom:1px solid var(--line);color:var(--text-2);vertical-align:middle}
table.data tr:last-child td{border:none}
table.data tr:hover td{background:var(--surface-2)}
table.data .strong{color:var(--text);font-weight:600}
table.data .num{font-family:var(--font-mono);text-align:end}
.pchange{display:inline-flex;align-items:center;gap:6px;font-family:var(--font-mono);font-weight:600}
.pchange.up{color:var(--green)}.pchange.dn{color:var(--red)}
.pchange .from{color:var(--mut);text-decoration:line-through;font-weight:500}
.pchange .arrow{color:var(--text-3)}

/* ============== STRATEGY DETAIL ============== */
.strat-overview{display:grid;grid-template-columns:repeat(4,1fr);gap:9px;margin-bottom:14px}
@media (max-width:600px){.strat-overview{grid-template-columns:repeat(2,1fr)}}
.stat-mini{background:var(--surface-2);border-radius:var(--r);padding:11px;text-align:center;border:1px solid var(--line)}
.stat-mini .v{font-size:18px;font-weight:700;color:var(--text);font-family:var(--font-mono)}
.stat-mini .v.g{color:var(--green)}.stat-mini .v.r{color:var(--red)}.stat-mini .v.gold{color:var(--gold)}
.stat-mini .l{font-size:10.5px;color:var(--mut);margin-top:3px}

/* ============== CHARTS ============== */
/* Calendar heatmap (forward pace view) */
.calgrid{display:grid;grid-template-columns:repeat(7,1fr);gap:4px}
@media (max-width:600px){.calgrid{gap:2px}}
.calday{position:relative;padding:8px 6px 6px;border-radius:8px;cursor:pointer;min-height:62px;border:1.5px solid transparent;transition:.12s;background:var(--surface-2)}
@media (max-width:600px){.calday{padding:5px 3px 3px;min-height:50px;border-radius:6px}}
@media (max-width:600px){.calday .cd-dnum{font-size:11px}}
@media (max-width:600px){.calday .cd-pct{font-size:8.5px;top:3px;inset-inline-end:3px}}
@media (max-width:600px){.calday .cd-wd{display:none}}
@media (max-width:600px){.calday .cd-evt{font-size:7.5px;bottom:2px;inset-inline-start:2px;inset-inline-end:2px}}
.calday:hover{border-color:var(--gold);transform:translateY(-1px);box-shadow:var(--sh-sm)}
.calday.sel{border-color:var(--gold);background:var(--gold-tint)}
.calday.cal-low{background:#fae3e3}            /* <40% */
.calday.cal-mid{background:#faeed1}            /* 40-69% */
.calday.cal-high{background:#dcf3e6}           /* 70-89% */
.calday.cal-full{background:#a3e0bd}           /* 90%+ */
html[data-theme="dark"] .calday.cal-low{background:#2e1414}
html[data-theme="dark"] .calday.cal-mid{background:#2d2210}
html[data-theme="dark"] .calday.cal-high{background:#0f2e1f}
html[data-theme="dark"] .calday.cal-full{background:#1a4d33}
.calday .cd-dnum{font-family:var(--font-mono);font-size:14px;font-weight:700;color:var(--text);line-height:1}
.calday .cd-pct{position:absolute;top:6px;inset-inline-end:6px;font-family:var(--font-mono);font-size:10px;font-weight:600;color:var(--text-2)}
.calday .cd-wd{font-size:9px;color:var(--mut);margin-top:2px;text-transform:uppercase;letter-spacing:.3px}
.calday .cd-evt{position:absolute;bottom:4px;inset-inline-start:6px;inset-inline-end:6px;font-size:9px;color:var(--gold);font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;text-align:center}
.calday.evt{border-color:var(--gold)}
.calday.weekend .cd-dnum{color:var(--blue)}

.bar-chart{display:flex;gap:4px;align-items:flex-end;height:130px;padding:0 4px;margin-bottom:12px}
.bar-col{flex:1;display:flex;flex-direction:column;align-items:center;gap:4px;min-width:0;cursor:default;position:relative}
.bar-col:hover .bar-tip{opacity:1;transform:translateY(-3px)}
.bar{width:100%;background:linear-gradient(180deg,var(--gold),var(--gold-2));border-radius:4px 4px 0 0;min-height:4px;transition:.12s}
.bar.muted{background:linear-gradient(180deg,var(--surface-3),var(--line-strong))}
.bar-label{color:var(--mut);font-size:9.5px;font-family:var(--font-mono)}
.bar-tip{position:absolute;bottom:100%;background:var(--text);color:var(--bg);padding:3px 7px;border-radius:4px;font-size:10.5px;font-weight:600;white-space:nowrap;opacity:0;pointer-events:none;transition:.15s;margin-bottom:6px;font-family:var(--font-mono);z-index:5}

/* ============== ACTIVITY LOG ============== */
.log-row{display:grid;grid-template-columns:auto auto 1fr;gap:11px;padding:9px 0;border-bottom:1px solid var(--line);font-size:12px;align-items:start}
.log-row:last-child{border:none}
.log-lic{width:22px;height:22px;border-radius:6px;background:var(--surface-2);display:flex;align-items:center;justify-content:center;font-size:11px}
.log-lts{color:var(--mut);font-size:10.5px;font-family:var(--font-mono);white-space:nowrap;padding-top:3px}
.log-ltxt{color:var(--text-2);line-height:1.5}

/* ============== BOTTOM NAV (mobile) ============== */
nav.bnav{display:none;position:fixed;bottom:0;left:0;right:0;background:var(--surface);border-top:1px solid var(--line);padding:5px 6px calc(6px + env(safe-area-inset-bottom));z-index:60}
html[data-theme="dark"] nav.bnav{background-color:rgba(24,23,26,.95);backdrop-filter:blur(12px)}
@media (max-width:1023px){nav.bnav{display:grid;grid-template-columns:repeat(6,1fr);gap:2px}}
@media (max-width:1023px){.bn{padding:6px 2px;font-size:9px}}
@media (max-width:1023px){.bn .ic{font-size:15px}}
.bn{display:flex;flex-direction:column;align-items:center;justify-content:center;gap:2px;padding:7px 4px;border-radius:7px;color:var(--text-3);font-size:9.5px;font-weight:500;transition:.12s;position:relative}
.bn .ic{font-size:16px;line-height:1}
.bn.on{color:var(--gold)}
.bn .badge{position:absolute;top:3px;inset-inline-end:14px;background:var(--red);color:#fff;font-size:9px;font-weight:700;padding:1px 5px;border-radius:7px;min-width:13px;text-align:center;line-height:1.3}

/* ============== VIEWS ============== */
.view{display:none}
.view.on{display:block;animation:fade .18s ease}
@keyframes fade{from{opacity:0;transform:translateY(4px)}to{opacity:1;transform:none}}

/* ============== TOAST ============== */
#toast{position:fixed;bottom:80px;left:50%;transform:translateX(-50%) translateY(20px);background:var(--text);color:var(--bg);padding:10px 18px;border-radius:9px;font-size:12.5px;font-weight:500;opacity:0;transition:.22s;pointer-events:none;z-index:200;box-shadow:var(--sh-lg)}
@media (min-width:1024px){#toast{bottom:24px}}
#toast.show{opacity:1;transform:translateX(-50%) translateY(0)}

/* ============== EMPTY STATES ============== */
.empty{color:var(--mut);text-align:center;padding:32px 14px;font-size:12.5px}
.empty .ic{font-size:28px;display:block;margin-bottom:6px;opacity:.4}
.muted{color:var(--mut);font-size:11.5px}
.sk{background:linear-gradient(90deg,var(--surface-2) 25%,var(--surface-3) 50%,var(--surface-2) 75%);background-size:200% 100%;animation:sk 1.4s infinite;color:transparent!important;border-radius:5px;min-height:14px;display:inline-block;width:60%}
@keyframes sk{0%{background-position:200% 0}100%{background-position:-200% 0}}

/* ============== DISCOUNT STATUS BANNER ============== */
.discount-banner{display:flex;align-items:center;justify-content:space-between;padding:11px 14px;border-radius:var(--r-lg);margin-bottom:14px;border:1px solid var(--line);background:var(--surface);box-shadow:var(--sh-xs);gap:10px;flex-wrap:wrap}
.discount-banner.paused{background:var(--yellow-soft);border-color:rgba(201,150,23,.25)}
.discount-banner .info{display:flex;align-items:center;gap:9px}
.discount-banner .pulse{width:8px;height:8px;border-radius:50%;background:var(--green);box-shadow:0 0 0 3px rgba(14,158,95,.18);animation:pulse 2s infinite}
.discount-banner.paused .pulse{background:var(--yellow);box-shadow:0 0 0 3px rgba(201,150,23,.20);animation:none}
@keyframes pulse{50%{box-shadow:0 0 0 5px rgba(14,158,95,.08)}}
.discount-banner .txt{font-size:13px;color:var(--text-2)}
.discount-banner .txt b{color:var(--text)}

</style>
</head>
<body>

<!-- Login -->
<div id="login">
  <div class="brand-lg">عوجا</div>
  <div class="sub">Ouja Operations</div>
  <input id="tok" type="password" placeholder="رمز الدخول · Access token" autocomplete="off" onkeydown="if(event.key==='Enter')saveTok()">
  <button class="btn primary" onclick="saveTok()" style="padding:12px 26px;font-size:13.5px">دخول · Enter</button>
  <div class="err" id="lerr"></div>
</div>

<div id="app">
  <div class="shell">

    <!-- Top bar (mobile) -->
    <header class="mhead">
      <div class="mhead-brand"><div class="logo">ع</div><div class="name" id="mhead_title">الرئيسية</div></div>
      <div class="mhead-tools">
        <button class="icbtn" onclick="toggleTheme()" id="themeBtn">◐</button>
        <button class="icbtn" onclick="toggleLang()" id="langBtn">EN</button>
        <button class="icbtn" onclick="refresh()" id="refreshBtnM">↻</button>
      </div>
    </header>

    <!-- Main content -->
    <main class="main">

      <!-- ============ HOME VIEW ============ -->
      <section class="view on" id="view_home">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_home">الرئيسية</div>
            <div class="page-sub"><span class="dot" id="dot"></span> <span id="freshness"></span></div>
          </div>
          <div class="page-tools">
            <button class="btn ghost xs" onclick="toggleTheme()" id="dThemeBtn">◐ <span id="t_theme">المظهر</span></button>
            <button class="btn ghost xs" onclick="toggleLang()" id="dLangBtn">EN</button>
            <button class="btn ghost xs" onclick="refresh()" id="refreshBtn">↻ <span id="t_refresh">تحديث</span></button>
          </div>
        </div>

        <div class="kpis" id="kpis"></div>

        <!-- Operational command-center: urgent items at the top (collapsible by severity) -->
        <div id="urgentStrip"></div>

        <div id="needsBanner"></div>

        <!-- Today's arrivals timeline with per-guest signed/code status -->
        <div class="card">
          <div class="card-head">
            <span class="card-title">🛬 <span id="t_arrivals">الوصول القادم</span></span>
            <span class="card-sub" id="arrivalsCount"></span>
          </div>
          <div id="arrivalsTimeline"><div class="empty sk">—</div></div>
        </div>

        <div class="grid2">
          <div class="card">
            <div class="card-head"><span class="card-title">📅 <span id="t_today_h">اليوم</span></span><span class="card-sub" id="t_today_date"></span></div>
            <div id="todayBody"><div class="empty sk">—</div></div>
          </div>
          <div class="card">
            <div class="card-head"><span class="card-title">📈 <span id="t_rev_card">الإيراد الشهري</span></span><span class="card-sub" id="revCardSub"></span></div>
            <div id="revCardBody"><div class="empty sk">—</div></div>
          </div>
        </div>

        <div class="card">
          <div class="card-head">
            <span class="card-title">📋 <span id="t_recent_h">آخر النشاط</span></span>
            <a class="card-sub" style="cursor:pointer;color:var(--gold);font-weight:600" onclick="go('log')" id="t_seeall">عرض الكل ←</a>
          </div>
          <div id="recentBody"><div class="empty sk">—</div></div>
        </div>
      </section>

      <!-- ============ INBOX VIEW ============ -->
      <section class="view" id="view_inbox">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_inbox">صندوق الوارد</div>
            <div class="page-sub" id="t_inbox_sub">كل الردود والتصعيدات في مكان واحد</div>
          </div>
        </div>

        <div class="filterbar">
          <div class="tabsfilter" id="ibFilterTabs"></div>
          <select id="ibFilterUnit" onchange="renderInbox()"><option value="">كل الوحدات</option></select>
          <select id="ibFilterStatus" onchange="renderInbox()">
            <option value="">كل الحالات</option>
            <option value="open">مفتوحة</option>
            <option value="claimed">مستلمة</option>
          </select>
          <input type="search" id="ibFilterSearch" placeholder="بحث · ضيف، رسالة…" oninput="renderInbox()">
          <button class="clear" onclick="clearInboxFilters()" id="t_clear_filt">مسح</button>
        </div>

        <div id="inboxList" class="inbox-list"><div class="empty sk">—</div></div>
      </section>

      <!-- ============ TODAY VIEW (per-empty-unit) ============ -->
      <section class="view" id="view_today">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_today">اليوم · الوحدات الفاضية</div>
            <div class="page-sub" id="t_today_sub">السعر الحالي + جدول الخصومات لكل شقة</div>
          </div>
          <div class="page-tools">
            <button class="btn ghost sm" onclick="loadTodayEmpty()">↻</button>
          </div>
        </div>

        <div id="discountBanner"></div>

        <div id="emptyGridWrap"><div class="empty sk">—</div></div>
      </section>

      <!-- ============ CALENDAR (forward pace + events + bulk apply) ============ -->
      <section class="view" id="view_calendar">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_calendar">📅 تقويم الإيراد · ٦٠ يوم قادمة</div>
            <div class="page-sub" id="t_calendar_sub"></div>
          </div>
          <div class="page-tools">
            <button class="btn ghost sm" onclick="loadForwardCalendar()">↻</button>
          </div>
        </div>
        <div class="card">
          <div id="calendarGrid"><div class="empty sk">—</div></div>
        </div>
        <div class="grid2">
          <div class="card">
            <div class="card-head"><span class="card-title">🗓️ <span id="t_cal_events_legend">المناسبات</span></span></div>
            <div id="calEventsBody"></div>
          </div>
          <div class="card">
            <div class="card-head"><span class="card-title">⚡ <span id="t_bulk_title">تطبيق جماعي على المدى المحدد</span></span></div>
            <div id="bulkForm"></div>
          </div>
        </div>
      </section>

      <!-- ============ PRICING VIEW ============ -->
      <section class="view" id="view_pricing">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_pricing">فرص التسعير</div>
            <div class="page-sub" id="t_pr_sub">توصيات أسعار للـ٤٥ يوم الجاية — اضغط لرؤية تفاصيل كل تاريخ</div>
          </div>
        </div>
        <div class="card">
          <div id="prTotalBody"></div>
        </div>
        <div class="card">
          <div class="card-head"><span class="card-title">📋 <span id="t_pr_list">قائمة الوحدات</span></span><span class="card-sub" id="prListCount"></span></div>
          <div id="prListBody"><div class="empty sk">—</div></div>
        </div>
      </section>

      <!-- ============ STRATEGIES VIEW ============ -->
      <section class="view" id="view_strat">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_strat">الاستراتيجيات</div>
            <div class="page-sub" id="t_strat_sub">الوحدات المتابَعة تلقائياً — اضغط لرؤية قبل/بعد كل ليلة</div>
          </div>
        </div>
        <div class="card">
          <div id="stratListBody"><div class="empty sk">—</div></div>
        </div>
      </section>

      <!-- ============ REVENUE VIEW ============ -->
      <section class="view" id="view_rev">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_rev">الإيرادات والأداء</div>
            <div class="page-sub" id="t_rev_sub">آخر ١٢ شهر + دورة الراتب + أداء الوحدات</div>
          </div>
        </div>

        <!-- Hero KPI strip: this-month vs prior month -->
        <div class="kpis" id="revKpis"></div>

        <!-- Forward pace: next 30 days at a glance -->
        <div class="card">
          <div class="card-head"><span class="card-title">🚀 <span id="t_rev_pace">سرعة الحجز للـ٣٠ يوم القادمة</span></span><span class="card-sub" id="revPaceSub"></span></div>
          <div id="revPaceBody"><div class="empty sk">—</div></div>
        </div>

        <div class="grid2">
          <div class="card">
            <div class="card-head"><span class="card-title">📅 <span id="t_rev_month">الإيراد الشهري</span></span></div>
            <div id="revMonthlyBody"></div>
          </div>
          <div class="card">
            <div class="card-head"><span class="card-title">💵 <span id="t_rev_sal">دورة الراتب</span></span></div>
            <div id="revSalaryBody"></div>
          </div>
        </div>
        <div class="card">
          <div class="card-head"><span class="card-title">🏠 <span id="t_rev_units">أداء الوحدات</span></span></div>
          <div id="revUnitsBody"></div>
        </div>
      </section>

      <!-- ============ LOG VIEW ============ -->
      <section class="view" id="view_log">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_log">سجل النشاط</div>
            <div class="page-sub" id="t_log_sub">كل ما يعمله البوت</div>
          </div>
          <div class="page-tools">
            <select id="logFilter" onchange="renderLog()" style="width:auto;font-size:12px;padding:6px 10px;height:32px"><option value="">الكل</option><option value="guest">الضيوف</option><option value="escalation">تصعيدات</option><option value="pricing">تسعير</option><option value="report">تقارير</option></select>
          </div>
        </div>
        <div class="card"><div id="logBody"><div class="empty sk">—</div></div></div>
      </section>

      <!-- ============ DEEP-CLEAN SCHEDULE ============ -->
      <section class="view" id="view_clean">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_clean">🧹 جدول التنظيف العميق</div>
            <div class="page-sub" id="t_clean_sub"></div>
          </div>
          <div class="page-tools">
            <label class="btn primary sm" style="cursor:pointer">📥 رفع Excel<input type="file" accept=".xlsx" onchange="uploadCleaningXLSX(event)" style="display:none"></label>
            <label class="btn ghost sm" style="cursor:pointer">📄 CSV<input type="file" accept=".csv" onchange="uploadCleaningCSV(event)" style="display:none"></label>
            <button class="btn ghost sm" onclick="loadCleaning()">↻</button>
          </div>
        </div>

        <div id="cleanImportResult" style="display:none"></div>

        <div class="kpis" id="cleanStats"></div>

        <!-- THIS WEEK strip — the supervisor's daily lookup -->
        <div class="card">
          <div class="card-head"><span class="card-title">📆 هذا الأسبوع — ٧ أيام قادمة</span></div>
          <div id="cleanWeek"><div class="empty sk">—</div></div>
        </div>

        <!-- 60-day visual schedule — the actual SCHEDULE the user asked for -->
        <div class="card">
          <div class="card-head"><span class="card-title">🗓️ الجدول البصري · ٦٠ يوم قادمة</span><span class="card-sub">اضغط أي يوم للتفاصيل</span></div>
          <div id="cleanVisualGrid"><div class="empty sk">—</div></div>
        </div>

        <div class="card">
          <div class="card-head"><span class="card-title">🔗 <span id="t_clean_link_title">الرابط لشركة التنظيف</span></span></div>
          <div id="cleanLinkBox"></div>
        </div>

        <div class="card">
          <div class="card-head">
            <span class="card-title">📋 الجدول الكامل</span>
            <div class="card-actions">
              <select id="cleanFilter" onchange="renderCleaningList()" style="width:auto;padding:6px 10px;height:32px;font-size:12px">
                <option value="all" id="cf_all">الكل</option>
                <option value="overdue" id="cf_overdue">متأخّرة</option>
                <option value="soon" id="cf_soon">قريباً (٧ أيام)</option>
                <option value="unscheduled" id="cf_unsch">غير مجدولة</option>
              </select>
              <input id="cleanSearch" placeholder="ابحث…" oninput="renderCleaningList()" style="width:160px;padding:6px 10px;height:32px;font-size:12px">
            </div>
          </div>
          <div id="cleanListBody"><div class="empty sk">—</div></div>
        </div>
      </section>

      <!-- ============ QUALITY (cleaning feedback per unit) ============ -->
      <section class="view" id="view_quality">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_quality">⭐ جودة التنظيف</div>
            <div class="page-sub" id="t_quality_sub"></div>
          </div>
          <div class="page-tools">
            <button class="btn ghost sm" onclick="loadQuality()">↻</button>
          </div>
        </div>
        <div class="kpis" id="qualStats"></div>
        <div class="card">
          <div class="card-head"><span class="card-title">🏠 ترتيب الوحدات</span></div>
          <div id="qualUnitsBody"><div class="empty sk">—</div></div>
        </div>
        <div class="card">
          <div class="card-head"><span class="card-title">💬 آخر التعليقات</span></div>
          <div id="qualCommentsBody"><div class="empty sk">—</div></div>
        </div>
      </section>

      <!-- ============ GUESTS (profiles + VIP + summaries) ============ -->
      <section class="view" id="view_guests">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_guests">👤 سجل الضيوف</div>
            <div class="page-sub" id="t_guests_sub"></div>
          </div>
          <div class="page-tools">
            <button class="btn ghost sm" onclick="loadGuests()">↻</button>
          </div>
        </div>
        <div class="kpis" id="guestStats"></div>
        <div class="card">
          <div class="card-head">
            <span class="card-title">📋 القائمة</span>
            <div class="card-actions">
              <select id="guestFilter" onchange="renderGuestList()" style="width:auto;padding:6px 10px;height:32px;font-size:12px">
                <option value="all" id="gf_all">الكل</option>
                <option value="vip" id="gf_vip">VIP فقط</option>
                <option value="repeat" id="gf_repeat">العائدون</option>
              </select>
              <input id="guestSearch" placeholder="ابحث بالاسم/الهاتف…" oninput="renderGuestList()" style="width:200px;padding:6px 10px;height:32px;font-size:12px">
            </div>
          </div>
          <div id="guestListBody"><div class="empty sk">—</div></div>
        </div>
      </section>

      <!-- ============ LEARNINGS ============ -->
      <section class="view" id="view_learn">
        <div class="page-head">
          <div>
            <div class="page-title" id="t_learn">📚 ما تعلّمه المساعد</div>
            <div class="page-sub" id="t_learn_sub"></div>
          </div>
          <div class="page-tools">
            <button class="btn primary sm" onclick="bootstrapLearnings()" id="learnBootstrapBtn">📥 <span id="t_learn_bootstrap">تعلّم من التاريخ</span></button>
            <button class="btn ghost sm" onclick="distillLearningsNow()" id="learnDistillBtn">↻ <span id="t_learn_distill">تلخيص الآن</span></button>
          </div>
        </div>
        <div id="bootstrapStatus" style="display:none"></div>

        <!-- Stat cards: today + delta vs 7-day average -->
        <div class="kpis" id="learnStats"></div>

        <!-- Trend charts: confidence, auto-send rate, escalation rate -->
        <div class="grid2">
          <div class="card">
            <div class="card-head"><span class="card-title">📈 <span id="t_learn_chart1">نسبة الثقة (٣٠ يوم)</span></span></div>
            <div id="learnChartConf"></div>
          </div>
          <div class="card">
            <div class="card-head"><span class="card-title">⚡ <span id="t_learn_chart2">معدّل الردود التلقائية (٣٠ يوم)</span></span></div>
            <div id="learnChartAuto"></div>
          </div>
        </div>
        <div class="grid2">
          <div class="card">
            <div class="card-head"><span class="card-title">💬 <span id="t_learn_chart3">حجم الردود اليومي</span></span></div>
            <div id="learnChartVol"></div>
          </div>
          <div class="card">
            <div class="card-head"><span class="card-title">🚨 <span id="t_learn_chart4">معدّل التصعيد (٣٠ يوم)</span></span></div>
            <div id="learnChartEsc"></div>
          </div>
        </div>

        <!-- What was learned in the selected window -->
        <div class="card">
          <div class="card-head">
            <span class="card-title">🆕 <span id="t_learn_recent">آخر ما تعلّمه</span></span>
            <div class="card-actions">
              <select id="learnWindow" onchange="loadLearnToday()" style="font-size:12px;padding:6px 10px;height:32px">
                <option value="1">اليوم</option>
                <option value="2">آخر يومين</option>
                <option value="7">آخر أسبوع</option>
                <option value="30">آخر شهر</option>
              </select>
            </div>
          </div>
          <div id="learnRecentBody"><div class="empty sk">—</div></div>
        </div>

        <!-- Distilled summaries (existing) -->
        <div class="card">
          <div class="card-head"><span class="card-title">🌐 <span id="t_learn_general">الملخص العام</span></span><div class="card-actions" id="genActions"></div></div>
          <div id="learnGeneralBody"><div class="empty sk">—</div></div>
        </div>
        <div class="grid2">
          <div class="card">
            <div class="card-head"><span class="card-title">🏠 <span id="t_learn_apt">ملخصات حسب الشقة</span></span></div>
            <input id="learnSearch" placeholder="ابحث عن وحدة…" oninput="renderLearnings()" style="margin-bottom:10px">
            <div id="learnAptList" style="max-height:520px;overflow-y:auto"><div class="empty sk">—</div></div>
          </div>
          <div class="card">
            <div id="learnAptDetail"><div class="empty" id="t_learn_empty_sel">اختر شقة من القائمة</div></div>
          </div>
        </div>
      </section>

      <!-- ============ MORE (mobile only) ============ -->
      <section class="view" id="view_more">
        <div class="page-head"><div><div class="page-title">المزيد</div></div></div>
        <div class="card"><div id="moreNav"></div></div>
      </section>

    </main>

    <!-- Sidebar (desktop) -->
    <aside class="side">
      <div class="side-brand">
        <div class="logo">ع</div>
        <div><div class="name">عوجا</div><div class="sub">Operations</div></div>
      </div>
      <div class="side-nav" id="sideNav"></div>
      <div class="side-foot">
        <div class="side-status"><span class="dot" id="sideDot"></span><span id="sideStatus">…</span></div>
        <div class="side-tools">
          <button class="icbtn" onclick="toggleTheme()" title="theme">◐</button>
          <button class="icbtn" onclick="toggleLang()" id="sLangBtn">EN</button>
          <button class="icbtn" onclick="logout()" title="logout">⎋</button>
        </div>
      </div>
    </aside>

  </div>

  <nav class="bnav" id="bottomNav"></nav>

  <!-- Detail drawer -->
  <div class="drawer-backdrop" id="drawerBg" onclick="closeDrawer()"></div>
  <aside class="drawer" id="drawer">
    <div class="drawer-head">
      <div style="min-width:0">
        <div class="drawer-title" id="drwTitle">—</div>
        <div class="drawer-sub" id="drwSub"></div>
      </div>
      <button class="icbtn" onclick="closeDrawer()">✕</button>
    </div>
    <div class="drawer-body" id="drwBody"></div>
    <div class="drawer-foot" id="drwFoot" style="display:none"></div>
  </aside>
</div>

<div id="toast"></div>

<script>
/* ============================================================
   STATE
   ============================================================ */
const TK='ouja_token', TH='ouja_theme';
const T = {
  ar:{dir:'rtl',
    home:'الرئيسية', inbox:'صندوق الوارد', today:'اليوم', pricing:'فرص التسعير', strat:'الاستراتيجيات', rev:'الإيرادات', learn:'ما تعلّمه', log:'النشاط', more:'المزيد', clean:'التنظيف العميق',
    clean_title:'🧹 جدول التنظيف العميق',
    clean_sub:'كل وحدة تُنظَّف عميق كل ٤٥-٦٠ يوم. الجدول يتجدّد تلقائياً ويتأكّد ٩م الليلة قبل.',
    clean_stat_total:'إجمالي الوحدات', clean_stat_overdue:'متأخّرة', clean_stat_scheduled:'مجدولة', clean_stat_tomorrow:'مؤكدة بكرة',
    clean_link_title:'🔗 الرابط لشركة التنظيف',
    clean_link_copy:'انسخ', clean_link_open:'افتح', clean_link_missing:'⚠ متغيّر CLEANING_TOKEN غير معرّف في Railway — أضفه أولاً لإنشاء الرابط',
    clean_filter_all:'الكل', clean_filter_overdue:'متأخّرة', clean_filter_soon:'قريباً (٧ أيام)', clean_filter_unscheduled:'غير مجدولة',
    clean_unit:'الوحدة', clean_last:'آخر تنظيف', clean_next:'القادم', clean_status:'الحالة', clean_actions:'إجراءات',
    clean_mark_done:'علّم منجز', clean_reschedule:'إعادة جدولة', clean_set_last:'تعديل تاريخ آخر تنظيف',
    clean_status_unscheduled:'غير مجدولة', clean_status_scheduled:'مجدولة', clean_status_blocked:'مؤكدة + مقفلة', clean_status_pushed:'تأجّلت',
    clean_days_ago:'يوم مضى', clean_days_left:'يوم باقي',
    clean_overdue_lbl:'متأخّرة',
    clean_modal_set_last:'تاريخ آخر تنظيف عميق', clean_modal_resched:'تاريخ التنظيف الجديد',
    clean_modal_save:'حفظ', clean_modal_cancel:'إلغاء',
    clean_confirm_done:'تأكيد إنجاز التنظيف لهذي الوحدة اليوم؟',
    copied:'نُسخ ✓',
    guests:'الضيوف', guests_title:'👤 سجل الضيوف',
    guests_sub:'كل من تفاعل معنا — أبرز المتكررين والـVIP',
    guests_stat_total:'الإجمالي', guests_stat_vip:'ضيوف VIP', guests_stat_repeat:'عائدون (٢+)',
    guests_filter_all:'الكل', guests_filter_vip:'VIP فقط', guests_filter_repeat:'العائدون',
    guests_search:'ابحث بالاسم/الهاتف…',
    guest_name:'الاسم', guest_stays:'إقامات', guest_nights:'ليالي', guest_last:'آخر تفاعل',
    guest_no_data:'ما فيه ضيوف بعد', guest_vip_on:'⭐ VIP',
    guest_drw_stays:'الإقامات', guest_drw_summaries:'ملخصات المحادثات',
    guest_drw_notes:'ملاحظات داخلية (لا يراها الضيف)', guest_drw_save:'حفظ',
    guest_drw_toggle_vip:'تبديل VIP',
    quality:'جودة النظافة', quality_title:'⭐ جودة التنظيف',
    quality_sub:'تقييمات الضيوف لنظافة الوحدات — متوسط كل وحدة + التعليقات',
    quality_stat_sent:'طُلب', quality_stat_resp:'استُجيب', quality_stat_rate:'نسبة الاستجابة', quality_stat_avg:'المتوسط العام',
    quality_unit:'الوحدة', quality_avg:'المتوسط', quality_count:'عدد التقييمات', quality_recent:'آخر تقييم',
    quality_empty:'ما فيه تقييمات بعد · أول تقييم بيوصل بعد ما يدخل ضيف بعد التنظيف العميق',
    quality_comments:'تعليقات',
    learn_title:'📚 ما تعلّمه المساعد',
    learn_sub:'الملخصات اللي استخلصها النظام من ردود فريقك — تقدر تعدّل أو تحذف',
    learn_general:'الملخص العام (يطبّق على كل الشقق)', learn_apt:'ملخصات حسب الشقة',
    learn_empty:'ما فيه ملخص بعد · يحتاج تفاعلات أكثر',
    learn_empty_sel:'اختر شقة من اليمين لمشاهدة ملخصها',
    learn_distill:'تلخيص الآن', learn_edit:'تعديل', learn_forget:'حذف', learn_save:'حفظ', learn_cancel:'إلغاء',
    learn_last:'آخر تلخيص', learn_examples:'تفاعل', learn_search:'ابحث عن وحدة…',
    learn_no_apt:'ما فيه شقق فيها ملخص حالياً', learn_confirm_forget:'تحذف الملخص نهائياً؟',
    learn_saved:'تم الحفظ ✅', learn_distilling:'يلخّص الآن… ممكن ياخذ ٣٠ ثانية',
    learn_bootstrap:'تعلّم من التاريخ',
    learn_bootstrap_confirm:'يقرأ آخر ٣٠٠ محادثة من Hostaway ويستخرج منها دروس لكل شقة. يستغرق ١٠-١٥ دقيقة في الخلفية وتكلفته بسيطة من Claude. متأكد؟',
    learn_bootstrap_started:'بدأ التعلّم التاريخي · بيشتغل في الخلفية',
    learn_bootstrap_running:'⏳ يقرأ المحادثات السابقة الحين… ممكن ياخذ ١٠-١٥ دقيقة. لا تغلق اللوحة.',
    learn_bootstrap_done:'✅ اكتمل التعلّم التاريخي',
    learn_bootstrap_scanned:'محادثة مفحوصة', learn_bootstrap_pairs:'سؤال-جواب مستخرج', learn_bootstrap_apts:'شقة تم تلخيصها',
    learn_today:'اليوم', learn_yesterday:'أمس', learn_7d:'٧ أيام', learn_30d:'٣٠ يوم',
    learn_chart1:'نسبة الثقة (٣٠ يوم)', learn_chart2:'معدّل الردود التلقائية (٣٠ يوم)',
    learn_chart3:'حجم الردود اليومي', learn_chart4:'معدّل التصعيد (٣٠ يوم)',
    learn_recent:'آخر ما تعلّمه', learn_recent_empty:'ما فيه نشاط في هذي الفترة',
    learn_stat_replies:'ردود اليوم', learn_stat_auto:'معدّل التلقائي', learn_stat_conf:'متوسط الثقة',
    learn_stat_esc:'تصعيدات اليوم', learn_stat_vs_avg:'مقارنة بمتوسط ٧ أيام',
    learn_event_edited:'مُعدّل', learn_event_sent:'كما هي', learn_event_auto:'تلقائي', learn_event_via:'عبر',
    calendar:'التقويم', calendar_title:'📅 تقويم الإيراد · ٦٠ يوم قادمة',
    calendar_sub:'لون كل يوم = نسبة الإشغال. ضع المؤشر للتفاصيل · اضغط لاختيار مدى للتطبيق الجماعي',
    cal_event:'مناسبة', cal_avg:'متوسط السعر', cal_avail:'فاضي', cal_occ:'مشغول',
    cal_no_events:'لا توجد مناسبات سعودية في هذا النطاق',
    cal_events_legend:'المناسبات',
    cal_pace:'الإشغال', cal_pct:'٪',
    bulk_title:'⚡ تطبيق جماعي على المدى المحدد',
    bulk_from:'من', bulk_to:'إلى', bulk_pct:'النسبة', bulk_action:'الإجراء',
    bulk_raise:'رفع', bulk_lower:'خفض', bulk_apply:'طبّق على الكل',
    bulk_select_range:'اختر مدى من الأيام بالضغط عليها',
    bulk_confirm:'متأكد؟ سيتم تعديل أسعار الليالي المتاحة فقط في كل الوحدات (المحجوزة لن تتأثر).',
    bulk_applied:'تم: {a} ليلة عُدّلت · {s} تُجوهلت',
    rev_pace:'سرعة الحجز للـ٣٠ يوم القادمة', rev_pace_sub:'كم وحدة محجوزة من إجمالي ليالي الـ٣٠ يوم',
    rev_kpi_mtd:'إيراد الشهر للحين', rev_kpi_occ:'إشغال ٣٠ يوم', rev_kpi_adr:'متوسط السعر',
    rev_kpi_pace:'نسبة الحجز للقادم', rev_kpi_vs_prev:'مقارنة بالفترة السابقة',
    rev_pace_total:'إجمالي ليالي ممكنة', rev_pace_booked:'محجوزة', rev_pace_open:'فاضية',
    rev_pace_proj:'إيراد متوقّع', rev_pace_events:'أيام مهمة قادمة',
    arrivals:'الوصول القادم', no_arrivals_window:'ما فيه وصول في الـ٣٦ ساعة الجاية',
    arr_signed:'موقّع', arr_unsigned:'غير موقّع', arr_in:'بعد', arr_hours:'ساعة', arr_minutes:'دقيقة',
    arr_now:'الحين', arr_past:'مضى',
    urgent_title:'🚨 يبيك الحين', urgent_none:'كل شي تحت السيطرة ✓',
    urgent_esc:'تصعيد مفتوح', urgent_pending:'رد بانتظار مراجعتك', urgent_unsigned:'عقد غير موقّع · تشيك-إن قريب',
    urgent_age:'منذ', urgent_open:'افتح',
    refresh:'تحديث', theme:'المظهر', logout:'خروج',
    today_h:'اليوم', today_date_sub:'',
    rev_card:'الإيراد الشهري', recent_h:'آخر النشاط', seeall:'عرض الكل ←',
    inbox_sub:'كل الردود والتصعيدات في مكان واحد',
    today_sub:'السعر الحالي + جدول الخصومات لكل شقة',
    pr_sub:'توصيات أسعار للـ٤٥ يوم الجاية — اضغط لرؤية تفاصيل كل تاريخ',
    strat_sub:'الوحدات المتابَعة تلقائياً — اضغط لرؤية قبل/بعد كل ليلة',
    rev_sub:'آخر ١٢ شهر + دورة الراتب + أداء الوحدات',
    log_sub:'كل ما يعمله البوت',
    all_units:'كل الوحدات', all_status:'كل الحالات', open:'مفتوحة', claimed:'مستلمة',
    f_search:'بحث · ضيف، رسالة…', f_clear:'مسح',
    f_all:'الكل', f_replies:'الردود المعلّقة', f_esc:'التصعيدات', f_auto:'تلقائية',
    ib_no:'ما فيه شي في هالفلتر',
    discount_running:'الخصومات التلقائية شغّالة', discount_paused:'الخصومات متوقفة لين',
    pause_24:'إيقاف ٢٤ ساعة', resume:'استئناف',
    eu_now:'السعر الحالي', eu_skip:'تجاهل الخصم على هالشقة', eu_unskip:'إلغاء التجاهل', eu_skipped_until:'متوقفة لين',
    no_empty_tonight:'ما فيه وحدات فاضية الليلة 🎉',
    tier_t1:'منتصف الليل', tier_t2:'الظهر', tier_t3:'العصر', tier_w:'المساء',
    pr_total:'الإجمالي المتوقع', pr_empty:'ما فيه فرص تسعير حالياً',
    pr_apply:'طبّق', pr_apply_all:'طبّق الكل', pr_confirm:'متأكد؟ بيتغيّر السعر فعلياً في تقويمك.',
    pr_change:'تغيير', pr_uplift:'إيراد إضافي تقديري', pr_conf:'الثقة',
    pr_d_date:'التاريخ', pr_d_day:'اليوم', pr_d_cur:'الحالي', pr_d_new:'المقترح', pr_d_why:'العوامل', pr_d_lead:'باقي',
    pr_why_month:'موسم', pr_why_pay:'راتب', pr_why_week:'يوم الأسبوع', pr_d_days:'يوم',
    st_empty:'ما فيه استراتيجيات بعد. لما تطبّق فرصة تسعير راح تبدأ وحدة هنا.',
    st_running:'شغّالة', st_done:'انتهت', st_stop:'إيقاف', st_booked:'محجوزة', st_open:'مفتوحة', st_changes:'تعديلات',
    st_d_date:'التاريخ', st_d_day:'اليوم', st_d_start:'البداية', st_d_cur:'الحالي', st_d_st:'الحالة', st_d_why:'العوامل', st_d_chg:'تعديلات',
    st_d_booked:'محجوزة ✓', st_d_open:'مفتوحة',
    rev_month:'الإيراد الشهري', rev_sal:'دورة الراتب', rev_units:'أداء الوحدات', rev_no:'ما فيه بيانات بعد',
    log_empty:'لا يوجد نشاط',
    fresh:'آخر تحديث', live:'مباشر',
    wrong:'رمز غير صحيح · Wrong token',
    sent:'تم الإرسال ✅', rejected:'تم التجاهل', claimed_t:'تم الاستلام ✅', applied:'تم التطبيق ✅', taught:'تم حفظ المعلومة ✅', skipped:'تم التجاهل', resumed:'تم الاستئناف', err:'صار خطأ',
    needs_alert:'يبيك الحين', no_needs:'كل شي تمام · ما يبيك شي 🤍',
    occ_tonight:'الإشغال الليلة', rev_7:'إيراد ٧ أيام', rev_30:'إيراد ٣٠ يوم',
    pending_rep:'ردود معلّقة', open_esc:'تصعيدات', empty_units:'فاضية الليلة', active:'وحدات فعّالة',
    no_pending:'ما فيه ردود معلّقة 🎉', no_esc:'ما فيه تصعيدات 🎉',
    rep_send:'إرسال', rep_reject:'تجاهل', rep_edit_focus:'تعديل', rep_teach:'علّم',
    claim:'استلام', claim_ph:'اسمك…',
    drw_thread:'المحادثة', drw_context:'سياق الحجز', drw_reasoning:'تحليل المساعد', drw_draft:'مقترح الرد',
    drw_intent:'النوع', drw_confidence:'الثقة', drw_sentiment:'المشاعر', drw_reason:'السبب',
    drw_dates:'التواريخ', drw_nights:'الليالي', drw_total:'الإجمالي', drw_status:'الحالة',
    drw_confirmed:'مؤكد', drw_not_confirmed:'غير مؤكد',
    auto_msg:'تلقائي',
    teach_label:'علّم المساعد', teach_topic:'الموضوع (اختياري)', teach_fact:'المعلومة الصحيحة', teach_save:'حفظ',
    nights:'ليلة',
    sentiment_ok:'عادي', sentiment_upset:'منزعج',
    weak_days:'أضعف الأيام', strong_days:'أقوى الأيام',
    u_unit:'الوحدة', u_occ:'إشغال', u_adr:'سعر/ليلة', u_pace:'٣٠ي', u_reco:'التوصية',
    nav_home:'الرئيسية', nav_inbox:'الوارد', nav_today:'اليوم', nav_pricing:'تسعير', nav_more:'المزيد',
    apply_all_q:'تطبيق كل التغييرات على هالوحدة؟',
    no_data:'ما فيه بيانات', untilDays:'يوم',
    units_count:'وحدة'
  },
  en:{dir:'ltr',
    home:'Home', inbox:'Inbox', today:'Today', pricing:'Pricing', strat:'Strategies', rev:'Revenue', learn:'Learnings', log:'Activity', more:'More', clean:'Deep clean',
    clean_title:'🧹 Deep cleaning schedule',
    clean_sub:'Every unit gets a deep clean every 45-60 days. The schedule auto-fills and is confirmed at 9pm the night before.',
    clean_stat_total:'Total units', clean_stat_overdue:'Overdue', clean_stat_scheduled:'Scheduled', clean_stat_tomorrow:'Confirmed tomorrow',
    clean_link_title:'🔗 Share link for the cleaning company',
    clean_link_copy:'copy', clean_link_open:'open', clean_link_missing:'⚠ CLEANING_TOKEN env var not set in Railway — add it first to enable the link',
    clean_filter_all:'All', clean_filter_overdue:'Overdue', clean_filter_soon:'Soon (7 days)', clean_filter_unscheduled:'Unscheduled',
    clean_unit:'Unit', clean_last:'Last cleaned', clean_next:'Next', clean_status:'Status', clean_actions:'Actions',
    clean_mark_done:'mark done', clean_reschedule:'reschedule', clean_set_last:'edit last-cleaned date',
    clean_status_unscheduled:'unscheduled', clean_status_scheduled:'scheduled', clean_status_blocked:'confirmed + blocked', clean_status_pushed:'pushed',
    clean_days_ago:'days ago', clean_days_left:'days left',
    clean_overdue_lbl:'overdue',
    clean_modal_set_last:'Last deep-clean date', clean_modal_resched:'New cleaning date',
    clean_modal_save:'Save', clean_modal_cancel:'Cancel',
    clean_confirm_done:'Mark this unit as deep-cleaned today?',
    copied:'copied ✓',
    guests:'Guests', guests_title:'👤 Guest profiles',
    guests_sub:'Everyone who has interacted with us — repeats + VIPs surfaced',
    guests_stat_total:'Total', guests_stat_vip:'VIPs', guests_stat_repeat:'Returning (2+)',
    guests_filter_all:'All', guests_filter_vip:'VIPs only', guests_filter_repeat:'Returning',
    guests_search:'Search name/phone…',
    guest_name:'Name', guest_stays:'stays', guest_nights:'nights', guest_last:'last seen',
    guest_no_data:'No guests recorded yet', guest_vip_on:'⭐ VIP',
    guest_drw_stays:'Stays', guest_drw_summaries:'Conversation summaries',
    guest_drw_notes:'Internal notes (guest never sees these)', guest_drw_save:'Save',
    guest_drw_toggle_vip:'Toggle VIP',
    quality:'Cleaning quality', quality_title:'⭐ Cleaning quality',
    quality_sub:"Guest ratings on each unit's cleanliness — averages + comments",
    quality_stat_sent:'Requested', quality_stat_resp:'Responded', quality_stat_rate:'Response rate', quality_stat_avg:'Overall avg',
    quality_unit:'Unit', quality_avg:'Avg', quality_count:'Ratings', quality_recent:'Recent',
    quality_empty:'No ratings yet · the first one comes after a guest checks in post-deep-clean',
    quality_comments:'Comments',
    learn_title:'📚 What the assistant has learned',
    learn_sub:"Summaries the system distilled from your team's replies — you can edit or delete",
    learn_general:'General summary (applies to every unit)', learn_apt:'Per-apartment summaries',
    learn_empty:'No summary yet · needs more interactions',
    learn_empty_sel:'Pick an apartment on the right to view its summary',
    learn_distill:'Distill now', learn_edit:'Edit', learn_forget:'Delete', learn_save:'Save', learn_cancel:'Cancel',
    learn_last:'Last distill', learn_examples:'interactions', learn_search:'Search unit…',
    learn_no_apt:'No apartments with summaries yet', learn_confirm_forget:'Delete this summary?',
    learn_saved:'Saved ✅', learn_distilling:'Distilling now… may take 30s',
    learn_bootstrap:'Learn from history',
    learn_bootstrap_confirm:'This reads your last 300 Hostaway conversations and distills lessons per apartment. Takes 10-15 minutes in the background; Claude cost is modest. Continue?',
    learn_bootstrap_started:'Historical learning started · running in background',
    learn_bootstrap_running:'⏳ Reading past conversations now… may take 10-15 min. You can keep the dashboard open.',
    learn_bootstrap_done:'✅ Historical learning complete',
    learn_bootstrap_scanned:'conversations scanned', learn_bootstrap_pairs:'Q&A pairs extracted', learn_bootstrap_apts:'apartments distilled',
    learn_today:'Today', learn_yesterday:'Yesterday', learn_7d:'Last 7 days', learn_30d:'Last 30 days',
    learn_chart1:'Avg confidence (30d)', learn_chart2:'Auto-send rate (30d)',
    learn_chart3:'Daily reply volume', learn_chart4:'Escalation rate (30d)',
    learn_recent:'Recently learned', learn_recent_empty:'No activity in this window',
    learn_stat_replies:'Replies today', learn_stat_auto:'Auto-send rate', learn_stat_conf:'Avg confidence',
    learn_stat_esc:'Escalations today', learn_stat_vs_avg:'vs 7-day avg',
    learn_event_edited:'edited', learn_event_sent:'sent as-is', learn_event_auto:'auto', learn_event_via:'via',
    calendar:'Calendar', calendar_title:'📅 Revenue calendar · next 60 days',
    calendar_sub:'Day color = occupancy. Hover for detail · click to select a range for bulk apply',
    cal_event:'event', cal_avg:'avg price', cal_avail:'open', cal_occ:'booked',
    cal_no_events:'No Saudi events in this range',
    cal_events_legend:'Events',
    cal_pace:'Occupancy', cal_pct:'%',
    bulk_title:'⚡ Bulk apply to the selected range',
    bulk_from:'from', bulk_to:'to', bulk_pct:'percent', bulk_action:'action',
    bulk_raise:'raise', bulk_lower:'lower', bulk_apply:'apply to all',
    bulk_select_range:'Click days to select a range',
    bulk_confirm:'Sure? This will adjust prices on AVAILABLE nights only across every active unit (booked nights are untouched).',
    bulk_applied:'Done: {a} nights changed · {s} skipped',
    rev_pace:'30-day booking pace', rev_pace_sub:'How many of the next 30 nights are already booked',
    rev_kpi_mtd:'Revenue MTD', rev_kpi_occ:'Occupancy 30d', rev_kpi_adr:'Avg rate',
    rev_kpi_pace:'30d forward pace', rev_kpi_vs_prev:'vs prior period',
    rev_pace_total:'Total potential nights', rev_pace_booked:'booked', rev_pace_open:'open',
    rev_pace_proj:'Projected revenue', rev_pace_events:'Big upcoming dates',
    arrivals:'Upcoming arrivals', no_arrivals_window:'No arrivals in the next 36h',
    arr_signed:'signed', arr_unsigned:'unsigned', arr_in:'in', arr_hours:'h', arr_minutes:'min',
    arr_now:'now', arr_past:'past',
    urgent_title:'🚨 Needs you', urgent_none:'All clear ✓',
    urgent_esc:'open escalation', urgent_pending:'pending reply', urgent_unsigned:'unsigned · check-in soon',
    urgent_age:'ago', urgent_open:'open',
    refresh:'Refresh', theme:'Theme', logout:'Logout',
    today_h:'Today', today_date_sub:'',
    rev_card:'Monthly revenue', recent_h:'Recent activity', seeall:'See all →',
    inbox_sub:'Replies and escalations in one place',
    today_sub:'Current price + discount schedule for each empty unit',
    pr_sub:'45-day recommendations — click any unit for per-date detail',
    strat_sub:'Units being auto-priced — click for before/after per night',
    rev_sub:'12 months + salary cycle + unit performance',
    log_sub:'Everything the bot does',
    all_units:'All units', all_status:'All status', open:'Open', claimed:'Claimed',
    f_search:'Search · guest, message…', f_clear:'Clear',
    f_all:'All', f_replies:'Pending replies', f_esc:'Escalations', f_auto:'Auto',
    ib_no:'No items match this filter',
    discount_running:'Auto-discounts running', discount_paused:'Discounts paused until',
    pause_24:'Pause 24h', resume:'Resume',
    eu_now:'Current', eu_skip:'Skip discounts on this unit', eu_unskip:'Resume',  eu_skipped_until:'Skipped until',
    no_empty_tonight:'Every unit is booked tonight 🎉',
    tier_t1:'Midnight', tier_t2:'Noon', tier_t3:'Evening', tier_w:'Weekend',
    pr_total:'Total estimate', pr_empty:'No pricing opportunities right now',
    pr_apply:'Apply', pr_apply_all:'Apply all', pr_confirm:'Sure? This changes real prices.',
    pr_change:'change', pr_uplift:'Est. extra revenue', pr_conf:'Confidence',
    pr_d_date:'Date', pr_d_day:'Day', pr_d_cur:'Current', pr_d_new:'Proposed', pr_d_why:'Factors', pr_d_lead:'In',
    pr_why_month:'season', pr_why_pay:'payday', pr_why_week:'weekday', pr_d_days:'d',
    st_empty:'No strategies yet. Apply a price opportunity to start one.',
    st_running:'Running', st_done:'Finished', st_stop:'Stop', st_booked:'booked', st_open:'open', st_changes:'changes',
    st_d_date:'Date', st_d_day:'Day', st_d_start:'Start', st_d_cur:'Current', st_d_st:'Status', st_d_why:'Factors', st_d_chg:'Moves',
    st_d_booked:'Booked ✓', st_d_open:'Open',
    rev_month:'Monthly revenue', rev_sal:'Salary cycle', rev_units:'Unit performance', rev_no:'No data yet',
    log_empty:'No activity',
    fresh:'Updated', live:'live',
    wrong:'Wrong token',
    sent:'Sent ✅', rejected:'Dismissed', claimed_t:'Claimed ✅', applied:'Applied ✅', taught:'Saved ✅', skipped:'Skipped', resumed:'Resumed', err:'Something went wrong',
    needs_alert:'Needs you', no_needs:"You're all caught up 🤍",
    occ_tonight:'Occupied tonight', rev_7:'Revenue 7d', rev_30:'Revenue 30d',
    pending_rep:'Pending replies', open_esc:'Escalations', empty_units:'Empty tonight', active:'Active units',
    no_pending:'No pending replies 🎉', no_esc:'No open escalations 🎉',
    rep_send:'Send', rep_reject:'Dismiss', rep_edit_focus:'Edit', rep_teach:'Teach',
    claim:'Claim', claim_ph:'Your name…',
    drw_thread:'Conversation', drw_context:'Booking context', drw_reasoning:"Assistant's analysis", drw_draft:'Draft reply',
    drw_intent:'Intent', drw_confidence:'Confidence', drw_sentiment:'Sentiment', drw_reason:'Reason',
    drw_dates:'Dates', drw_nights:'Nights', drw_total:'Total', drw_status:'Status',
    drw_confirmed:'Confirmed', drw_not_confirmed:'Not confirmed',
    auto_msg:'auto',
    teach_label:'Teach the assistant', teach_topic:'Topic (optional)', teach_fact:'The correct fact', teach_save:'Save',
    nights:'nights',
    sentiment_ok:'ok', sentiment_upset:'upset',
    weak_days:'Weak days', strong_days:'Strong days',
    u_unit:'Unit', u_occ:'Occ', u_adr:'Rate/nt', u_pace:'30d', u_reco:'Action',
    nav_home:'Home', nav_inbox:'Inbox', nav_today:'Today', nav_pricing:'Pricing', nav_more:'More',
    apply_all_q:'Apply all changes for this unit?',
    no_data:'No data', untilDays:'d',
    units_count:'units'
  }
};

let L = localStorage.getItem('ouja_lang') || 'ar';
let theme = localStorage.getItem(TH) || 'auto';
let view = 'home';
const D = {};
let inboxFilter = {type:'all', unit:'', status:'', search:''};
let openInboxId = null;     // currently expanded inline item id
let drawerOpen = false;

/* ============================================================
   UTILS
   ============================================================ */
function t(){return T[L]}
function tok(){return localStorage.getItem(TK)||''}
function saveTok(){localStorage.setItem(TK, document.getElementById('tok').value.trim()); init()}
function logout(){localStorage.removeItem(TK); location.reload()}
function esc(s){return (s==null?'':String(s)).replace(/[<>&"']/g,function(c){return ({'<':'&lt;','>':'&gt;','&':'&amp;','"':'&quot;',"'":'&#39;'})[c]})}
function fmt(n){return (Math.round(n||0)).toLocaleString('en-US')}
function toast(m){const e=document.getElementById('toast');e.textContent=m;e.classList.add('show');clearTimeout(e._t);e._t=setTimeout(function(){e.classList.remove('show')},2200)}
function shortTime(s){return (s||'').replace('T',' ').slice(5,16)}
function dayTime(s){if(!s) return '';try{const d=new Date(s);return d.toLocaleString(L==='ar'?'ar-SA':'en-US',{day:'numeric',month:'short',hour:'2-digit',minute:'2-digit'})}catch(_){return shortTime(s)}}

async function api(path){
  const r = await fetch(path + (path.indexOf('?')>=0?'&':'?') + 'token=' + encodeURIComponent(tok()));
  if(r.status===401) throw 'unauthorized';
  return r.json();
}
async function post(path, body){
  const r = await fetch(path + '?token=' + encodeURIComponent(tok()), {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(body||{})});
  return r.json().catch(function(){return {}});
}

/* ============================================================
   THEME + LANG
   ============================================================ */
function applyTheme(){
  document.documentElement.setAttribute('data-theme', theme);
  const tBtn = document.getElementById('themeBtn'); if(tBtn) tBtn.textContent = theme==='dark'?'☀':(theme==='light'?'☾':'◐');
}
function toggleTheme(){theme = theme==='auto'?'light':(theme==='light'?'dark':'auto'); localStorage.setItem(TH, theme); applyTheme()}
function toggleLang(){L = L==='ar'?'en':'ar'; localStorage.setItem('ouja_lang', L); applyLang(); renderAll()}
function applyLang(){
  document.documentElement.dir = t().dir;
  document.documentElement.lang = L;
  const map = {
    t_home:'home', t_refresh:'refresh', t_theme:'theme',
    t_today_h:'today_h', t_rev_card:'rev_card', t_recent_h:'recent_h', t_seeall:'seeall',
    t_inbox:'inbox', t_inbox_sub:'inbox_sub',
    t_today:'today', t_today_sub:'today_sub',
    t_pricing:'pricing', t_pr_sub:'pr_sub', t_pr_list:'pr_list', t_pr_total:'pr_total',
    t_strat:'strat', t_strat_sub:'strat_sub',
    t_rev:'rev', t_rev_sub:'rev_sub', t_rev_month:'rev_month', t_rev_sal:'rev_sal', t_rev_units:'rev_units',
    t_log:'log', t_log_sub:'log_sub',
    t_learn:'learn_title', t_learn_sub:'learn_sub', t_learn_distill:'learn_distill',
    t_learn_general:'learn_general', t_learn_apt:'learn_apt',
    t_learn_chart1:'learn_chart1', t_learn_chart2:'learn_chart2',
    t_learn_chart3:'learn_chart3', t_learn_chart4:'learn_chart4',
    t_learn_recent:'learn_recent',
    t_arrivals:'arrivals',
    t_calendar:'calendar_title', t_calendar_sub:'calendar_sub',
    t_cal_events_legend:'cal_events_legend', t_bulk_title:'bulk_title',
    t_rev_pace:'rev_pace',
    t_clean:'clean_title', t_clean_sub:'clean_sub', t_clean_link_title:'clean_link_title',
    cf_all:'clean_filter_all', cf_overdue:'clean_filter_overdue', cf_soon:'clean_filter_soon', cf_unsch:'clean_filter_unscheduled',
    t_guests:'guests_title', t_guests_sub:'guests_sub',
    gf_all:'guests_filter_all', gf_vip:'guests_filter_vip', gf_repeat:'guests_filter_repeat',
    t_quality:'quality_title', t_quality_sub:'quality_sub',
    t_clear_filt:'f_clear'
  };
  for(const id in map){
    const el = document.getElementById(id);
    if(el && t()[map[id]] != null) el.textContent = t()[map[id]];
  }
  ['langBtn','dLangBtn','sLangBtn'].forEach(function(id){const e=document.getElementById(id); if(e) e.textContent = L==='ar'?'EN':'ع'});
  const dt = document.getElementById('dThemeBtn'); if(dt) dt.innerHTML = '◐ '+t().theme;
  buildSideNav(); buildBottomNav(); buildMoreNav(); buildInboxTabs();
  const mhT = document.getElementById('mhead_title');
  if(mhT) mhT.textContent = t()[view] || t().home;
}

/* ============================================================
   NAVIGATION
   ============================================================ */
const NAV = [
  {id:'home',    ic:'◇', tk:'home'},
  {id:'inbox',   ic:'✉', tk:'inbox', badge:'inbox'},
  {id:'today',   ic:'◎', tk:'today'},
  {id:'calendar',ic:'📅', tk:'calendar'},
  {id:'pricing', ic:'$', tk:'pricing', badge:'pricing'},
  {id:'strat',   ic:'⚡', tk:'strat'},
  {id:'clean',   ic:'🧹', tk:'clean', badge:'clean'},
  {id:'guests',  ic:'👤', tk:'guests'},
  {id:'quality', ic:'⭐', tk:'quality'},
  {id:'rev',     ic:'∿', tk:'rev'},
  {id:'learn',   ic:'📚', tk:'learn'},
  {id:'log',     ic:'≡', tk:'log'}
];
const MNAV = [
  {id:'home', ic:'◇', tk:'nav_home'},
  {id:'inbox', ic:'✉', tk:'nav_inbox', badge:'inbox'},
  {id:'today', ic:'◎', tk:'nav_today'},
  {id:'calendar', ic:'📅', tk:'calendar'},
  {id:'pricing', ic:'$', tk:'nav_pricing', badge:'pricing'},
  {id:'more', ic:'⋯', tk:'nav_more'}
];

function badgeCount(key){
  if(!key) return 0;
  const ib = D.inbox || {};
  if(key==='inbox') return (ib.replies?ib.replies.length:0) + (ib.escalations?ib.escalations.filter(function(e){return !e.claimed_by}).length:0);
  if(key==='pricing') return ((D.pr && D.pr.units) || []).length;
  if(key==='clean') return ((D.clean && D.clean.counts) || {}).overdue || 0;
  return 0;
}
function buildSideNav(){
  const el = document.getElementById('sideNav'); if(!el) return;
  el.innerHTML = NAV.map(function(n){
    const c = badgeCount(n.badge);
    return '<a class="item'+(view===n.id?' on':'')+'" onclick="go(\\''+n.id+'\\')"><span class="ic">'+n.ic+'</span><span>'+t()[n.tk]+'</span>'+(c>0?'<span class="badge">'+c+'</span>':'')+'</a>';
  }).join('');
}
function buildBottomNav(){
  const el = document.getElementById('bottomNav'); if(!el) return;
  el.innerHTML = MNAV.map(function(n){
    const c = badgeCount(n.badge);
    return '<button class="bn'+(view===n.id?' on':'')+'" onclick="go(\\''+n.id+'\\')"><span class="ic">'+n.ic+'</span><span>'+t()[n.tk]+'</span>'+(c>0?'<span class="badge">'+c+'</span>':'')+'</button>';
  }).join('');
}
function buildMoreNav(){
  const el = document.getElementById('moreNav'); if(!el) return;
  const items = [
    {id:'strat', tk:'strat'},
    {id:'rev', tk:'rev'},
    {id:'log', tk:'log'},
    {action:'theme', tk:'theme'},
    {action:'lang', tk:'EN/ع'},
    {action:'logout', tk:'logout'}
  ];
  el.innerHTML = '<div class="inbox-list">' + items.map(function(i){
    const label = i.tk==='EN/ع' ? 'English / عربي' : t()[i.tk];
    let click;
    if(i.id) click = "go('"+i.id+"')";
    else if(i.action==='theme') click = 'toggleTheme()';
    else if(i.action==='lang') click = 'toggleLang()';
    else if(i.action==='logout') click = 'logout()';
    return '<div class="ibox"><div class="ibox-row" style="cursor:pointer" onclick="'+click+'"><div class="ibox-main"><div class="ibox-who">'+label+'</div></div><span style="color:var(--mut)">←</span></div></div>';
  }).join('') + '</div>';
}
function buildInboxTabs(){
  const el = document.getElementById('ibFilterTabs'); if(!el) return;
  const tabs = [['all', t().f_all], ['replies', t().f_replies], ['esc', t().f_esc], ['auto', t().f_auto]];
  el.innerHTML = tabs.map(function(p){
    return '<button onclick="setIbFilter(\\''+p[0]+'\\')" class="'+(inboxFilter.type===p[0]?'on':'')+'">'+p[1]+'</button>';
  }).join('');
}
function setIbFilter(t_){inboxFilter.type=t_; buildInboxTabs(); renderInbox()}
function clearInboxFilters(){
  inboxFilter = {type:'all', unit:'', status:'', search:''};
  document.getElementById('ibFilterUnit').value=''; document.getElementById('ibFilterStatus').value=''; document.getElementById('ibFilterSearch').value='';
  buildInboxTabs(); renderInbox();
}

function go(id){
  view = id;
  document.querySelectorAll('.view').forEach(function(v){ v.classList.toggle('on', v.id === 'view_'+id) });
  buildSideNav(); buildBottomNav();
  const mhT = document.getElementById('mhead_title'); if(mhT) mhT.textContent = t()[id] || t().home;
  window.scrollTo({top:0});
  if(id==='today' && !D.tonight) loadTodayEmpty();
  if(id==='pricing' && !D.pr) loadPricing();
  if(id==='strat' && !D.strat) loadStrategies();
  if(id==='rev' && (!D.rev || D.rev.loading)) loadRevenue();
  if(id==='log') renderLog();
  if(id==='inbox') { renderInbox(); populateUnitFilter() }
  if(id==='learn') loadLearnings();
  if(id==='calendar') loadForwardCalendar();
  if(id==='clean') loadCleaning();
  if(id==='guests') loadGuests();
  if(id==='quality') loadQuality();
}

/* ============================================================
   INIT + LOAD
   ============================================================ */
async function init(){
  try{
    document.getElementById('lerr').textContent='';
    await api('/api/overview');
    document.getElementById('login').style.display='none';
    document.getElementById('app').style.display='block';
    applyTheme(); applyLang();
    await loadAll();
    setInterval(loadAll, 15000);
  }catch(e){
    document.getElementById('lerr').textContent = t().wrong;
  }
}
async function loadAll(){
  try{
    const r = await Promise.all([
      api('/api/overview'), api('/api/today'), api('/api/inbox'),
      api('/api/discount/status'), api('/api/log'), api('/api/autolog'),
      api('/api/revenue').catch(function(){return {loading:true}}),
      api('/api/home/urgent').catch(function(){return {items:[]}}),
      api('/api/home/arrivals?hours=36').catch(function(){return {items:[]}}),
    ]);
    D.ov=r[0]; D.today=r[1]; D.inbox=r[2]; D.disc=r[3];
    D.log=(r[4]||{}).items||[]; D.auto=(r[5]||{}).items||[]; D.rev=r[6];
    D.urgent=r[7]; D.arrivals=r[8];
    populateUnitFilter();
    renderAll();
  }catch(e){ if(e==='unauthorized') logout() }
}
async function loadPricing(){
  document.getElementById('prListBody').innerHTML = '<div class="empty sk">—</div>';
  try{ D.pr = await api('/api/pricing') }catch(_){ D.pr={loading:true} }
  renderPricing();
}
async function loadStrategies(){
  document.getElementById('stratListBody').innerHTML = '<div class="empty sk">—</div>';
  try{ D.strat = await api('/api/strategies') }catch(_){ D.strat={items:[]} }
  renderStrategies();
}
async function loadRevenue(){
  // Pull revenue + forward calendar in parallel so the new KPI cards and the
  // forward-pace summary paint together.
  try{
    const r = await Promise.all([
      api('/api/revenue'),
      api('/api/calendar/forward?days=30').catch(function(){return {days:[]}}),
    ]);
    D.rev = r[0];
    D.revCal30 = r[1];
  }catch(_){ D.rev = {loading:true}; D.revCal30 = {days:[]} }
  renderRevenueFull();
  renderRevKpis();
  renderRevPace();
}

function renderRevKpis(){
  const el = document.getElementById('revKpis'); if(!el) return;
  const rev = D.rev || {};
  const months = (rev.monthly || []).slice();
  // MTD vs prior month
  const this_m = months[months.length-1] || {rev:0};
  const prev_m = months.length > 1 ? months[months.length-2] : {rev:0};
  const delta_mtd = prev_m.rev ? Math.round((this_m.rev - prev_m.rev) / prev_m.rev * 100) : null;
  // 30-day forward pace
  const cal = (D.revCal30||{}).days || [];
  let booked = 0, total = 0, projRev = 0;
  cal.forEach(function(d){
    booked += d.occupied || 0;
    total += d.total || 0;
    if(d.avg_price && d.available != null && d.occupied != null){
      // already-booked nights * their price won't be in d.avg_price (it's only of available);
      // so project = booked nights * portfolio rough ADR (use overall_adr if visible)
    }
  });
  const pace_pct = total ? Math.round((booked / total) * 100) : 0;
  // occupancy + ADR from revenue endpoint
  const occ_units = (rev.units || []);
  const avg_occ = occ_units.length ? Math.round(occ_units.reduce(function(s,u){return s + (u.occ||0)},0) / occ_units.length) : 0;
  const avg_adr = occ_units.length ? Math.round(occ_units.reduce(function(s,u){return s + (u.adr||0)},0) / occ_units.length) : 0;

  const cards = [
    {ic:'$', cls:'g', val:fmt(this_m.rev||0)+' SAR', lbl:t().rev_kpi_mtd, delta:delta_mtd},
    {ic:'◌', cls:'b', val:avg_occ+'%', lbl:t().rev_kpi_occ},
    {ic:'∿', cls:'p', val:fmt(avg_adr)+' SAR', lbl:t().rev_kpi_adr},
    {ic:'🚀', cls:(pace_pct>=60?'g':(pace_pct>=40?'y':'r')), val:pace_pct+'%', lbl:t().rev_kpi_pace},
  ];
  el.innerHTML = cards.map(function(c){
    let dh = '';
    if(c.delta != null && c.delta !== 0){
      const cls = c.delta > 0 ? 'up' : 'dn';
      dh = '<span class="kpi-delta '+cls+'">'+(c.delta>0?'+':'')+c.delta+'%</span>';
    }
    return '<div class="kpi"><div class="kpi-head"><div class="kpi-ic '+c.cls+'">'+c.ic+'</div>'+dh+'</div>'
      + '<div class="kpi-val">'+c.val+'</div><div class="kpi-lbl">'+c.lbl+(c.delta!=null?(' · <span style="opacity:.7">'+t().rev_kpi_vs_prev+'</span>'):'')+'</div></div>';
  }).join('');
}

function renderRevPace(){
  const el = document.getElementById('revPaceBody'); if(!el) return;
  const sub = document.getElementById('revPaceSub'); if(sub) sub.textContent = t().rev_pace_sub;
  const cal = (D.revCal30||{}).days || [];
  if(!cal.length){ el.innerHTML = '<div class="empty">'+t().rev_no+'</div>'; return; }
  let booked = 0, open = 0, total = 0;
  let projRev = 0;
  // Estimate projected revenue: avg_price across days × open nights doesn't account for booked
  // night value; we approximate using overall ADR from units (better than nothing).
  const units = (D.rev||{}).units || [];
  const avg_adr = units.length ? units.reduce(function(s,u){return s+(u.adr||0)},0)/units.length : 0;
  const futureEvents = [];
  cal.forEach(function(d){
    total += d.total || 0;
    booked += d.occupied || 0;
    open += d.available || 0;
    if((d.events||[]).length){
      futureEvents.push({date:d.date, name:(d.events[0]||{}).name, pace:d.pace_pct});
    }
  });
  projRev = Math.round(booked * avg_adr);
  const pace = total ? Math.round((booked/total)*100) : 0;
  // Visual bar: booked vs open
  const barW = Math.max(2, pace);
  el.innerHTML =
    '<div style="display:flex;gap:14px;align-items:baseline;flex-wrap:wrap;margin-bottom:12px">'
    + '<div><div class="kpi-val" style="color:var(--gold)">'+booked+'<span style="font-size:14px;color:var(--mut);font-weight:500">/'+total+'</span></div><div class="muted">'+t().rev_pace_total+'</div></div>'
    + '<div><div class="kpi-val green">'+pace+'%</div><div class="muted">'+t().rev_pace_booked+'</div></div>'
    + '<div><div class="kpi-val">'+open+'</div><div class="muted">'+t().rev_pace_open+'</div></div>'
    + (avg_adr ? ('<div><div class="kpi-val" style="color:var(--blue)">~'+fmt(projRev)+' SAR</div><div class="muted">'+t().rev_pace_proj+'</div></div>') : '')
    + '</div>'
    + '<div style="background:var(--surface-2);border-radius:8px;height:14px;overflow:hidden;border:1px solid var(--line)"><div style="width:'+barW+'%;height:100%;background:linear-gradient(90deg,var(--gold),var(--gold-2))"></div></div>';
  if(futureEvents.length){
    el.innerHTML += '<div style="margin-top:14px"><div class="muted" style="margin-bottom:6px;font-weight:600;font-size:12px">'+t().rev_pace_events+'</div>'
      + futureEvents.slice(0,5).map(function(e){
        const cls = e.pace < 50 ? 'danger' : (e.pace < 75 ? 'warn' : 'ok');
        return '<div class="log-row"><div class="log-lts">'+e.date+'</div><div><b>'+esc(e.name)+'</b> <span class="pill '+cls+'">'+e.pace+'% booked</span></div></div>';
      }).join('') + '</div>';
  }
}
async function loadTodayEmpty(){
  const wrap = document.getElementById('emptyGridWrap');
  if(wrap) wrap.innerHTML = '<div class="empty sk">—</div>';
  try{ D.tonight = await api('/api/today/empty') }catch(_){ D.tonight={items:[]} }
  renderTodayEmpty();
}
async function refresh(){
  await loadAll();
  if(view==='today') await loadTodayEmpty();
  if(view==='pricing') await loadPricing();
  if(view==='strat') await loadStrategies();
  if(view==='rev') await loadRevenue();
}

/* ============================================================
   RENDER: top-level
   ============================================================ */
function renderAll(){
  renderFresh(); renderKpis(); renderNeedsBanner();
  renderUrgentStrip(); renderArrivalsTimeline();
  renderTodayHome(); renderRevCard(); renderRecent();
  // Don't blow away an expanded item's content on the 15-second auto-refresh —
  // re-rendering the whole list wipes the body div and the user just sees an
  // empty pane. The badges + KPIs are still updated above.
  if(!openInboxId) renderInbox();
  renderDiscountBanner();
  buildSideNav(); buildBottomNav();
}

function renderUrgentStrip(){
  const el = document.getElementById('urgentStrip');
  if(!el) return;
  const d = D.urgent || {items:[], counts:{}};
  const items = d.items || [];
  if(!items.length){
    el.innerHTML = '<div class="card" style="background:var(--green-soft);border-color:rgba(14,158,95,.18);text-align:center;padding:13px"><span style="color:var(--green);font-weight:600">✓ '+t().urgent_none+'</span></div>';
    return;
  }
  // Show up to 6 most urgent inline; collapse the rest
  const top = items.slice(0, 6);
  const more = items.length - top.length;
  const sevColor = function(s){ return s==='high' ? 'var(--red)' : (s==='med' ? 'var(--yellow)' : 'var(--blue)') };
  const sevBg = function(s){ return s==='high' ? 'var(--red-soft)' : (s==='med' ? 'var(--yellow-soft)' : 'var(--blue-soft)') };
  const kindLabel = function(k){
    if(k==='escalation') return t().urgent_esc;
    if(k==='pending_reply') return t().urgent_pending;
    if(k==='unsigned_agreement') return t().urgent_unsigned;
    return k;
  };
  const rows = top.map(function(it){
    const age = (it.age_min != null) ? '<span class="muted" style="font-size:11px">· '+it.age_min+'m '+t().urgent_age+'</span>' : '';
    const detail = it.detail ? '<div class="muted" style="font-size:11.5px;margin-top:3px">'+esc(it.detail)+'</div>' : '';
    return '<div style="display:flex;align-items:flex-start;gap:10px;padding:10px 12px;border-bottom:1px solid var(--line);cursor:pointer" onclick="go(\\''+(it.action_view||'inbox')+'\\')">'
      + '<div style="width:6px;align-self:stretch;background:'+sevColor(it.severity)+';border-radius:3px;flex-shrink:0"></div>'
      + '<div style="flex:1;min-width:0">'
        + '<div style="display:flex;justify-content:space-between;align-items:baseline;gap:8px;flex-wrap:wrap">'
          + '<span style="font-weight:600;font-size:13.5px">'+esc(it.title)+'</span>'
          + '<span class="pill" style="background:'+sevBg(it.severity)+';color:'+sevColor(it.severity)+'">'+kindLabel(it.kind)+'</span>'
        + '</div>'
        + '<div style="display:flex;justify-content:space-between;align-items:center;gap:8px;margin-top:2px">'
          + '<span class="muted" style="font-size:12px">'+esc(it.subtitle||'')+' '+age+'</span>'
        + '</div>'
        + detail
      + '</div>'
    + '</div>';
  }).join('');
  el.innerHTML = '<div class="card" style="padding:0;overflow:hidden">'
    + '<div class="card-head" style="padding:14px 16px;margin:0;border-bottom:1px solid var(--line)">'
      + '<span class="card-title">'+t().urgent_title+' <span class="pill danger">'+items.length+'</span></span>'
    + '</div>'
    + rows
    + (more > 0 ? '<div style="padding:10px;text-align:center;color:var(--mut);font-size:12px">+ '+more+'</div>' : '')
    + '</div>';
}

// ============== CLEANING QUALITY ==============
async function loadQuality(){
  document.getElementById('qualUnitsBody').innerHTML = '<div class="empty sk">—</div>';
  document.getElementById('qualCommentsBody').innerHTML = '<div class="empty sk">—</div>';
  try{ D.quality = await api('/api/cleaning/quality') }catch(_){ D.quality = {units:[], stats:{}} }
  const sub = document.getElementById('t_quality_sub'); if(sub) sub.textContent = t().quality_sub;
  renderQualityStats();
  renderQualityUnits();
  renderQualityComments();
}
function _stars(n){ if(n == null) return '—'; const k = Math.round(n); return '★'.repeat(k) + '☆'.repeat(5-k) + ' ' + n }
function renderQualityStats(){
  const el = document.getElementById('qualStats'); if(!el) return;
  const s = (D.quality||{}).stats || {};
  const cards = [
    {ic:'📤', cls:'b', val:s.sent||0, lbl:t().quality_stat_sent},
    {ic:'✓', cls:'g', val:s.responded||0, lbl:t().quality_stat_resp},
    {ic:'%', cls:'p', val:(s.response_rate||0)+'%', lbl:t().quality_stat_rate},
    {ic:'⭐', cls:'gold', val:s.overall_avg!=null?s.overall_avg:'—', lbl:t().quality_stat_avg},
  ];
  el.innerHTML = cards.map(function(c){
    return '<div class="kpi"><div class="kpi-head"><div class="kpi-ic '+c.cls+'">'+c.ic+'</div></div>'
      + '<div class="kpi-val">'+c.val+'</div><div class="kpi-lbl">'+c.lbl+'</div></div>';
  }).join('');
}
function renderQualityUnits(){
  const body = document.getElementById('qualUnitsBody'); if(!body) return;
  const units = ((D.quality||{}).units) || [];
  if(!units.length){ body.innerHTML = '<div class="empty">'+t().quality_empty+'</div>'; return; }
  let html = '<div style="overflow-x:auto"><table class="data"><thead><tr>'
    + '<th>'+t().quality_unit+'</th><th class="num">'+t().quality_avg+'</th>'
    + '<th class="num">'+t().quality_count+'</th><th>'+t().quality_recent+'</th></tr></thead><tbody>';
  for(const u of units){
    const avgCls = u.avg == null ? '' : (u.avg >= 4.5 ? 'ok' : u.avg >= 3.5 ? 'warn' : 'danger');
    const avgPill = u.avg == null ? '<span class="muted">—</span>'
                  : '<span class="pill '+avgCls+'">'+_stars(u.avg)+'</span>';
    const recent = u.recent && u.recent.length
      ? u.recent.slice(-3).map(function(r){return '★'.repeat(r.score)}).join(' · ') : '—';
    html += '<tr><td class="strong">'+esc(u.name||'—')+'</td>'
      + '<td class="num">'+avgPill+'</td>'
      + '<td class="num">'+u.count+'</td>'
      + '<td class="muted" style="font-size:11.5px">'+recent+'</td></tr>';
  }
  html += '</tbody></table></div>';
  body.innerHTML = html;
}
function renderQualityComments(){
  const body = document.getElementById('qualCommentsBody'); if(!body) return;
  const units = ((D.quality||{}).units) || [];
  const all = [];
  for(const u of units){
    for(const c of (u.comments || [])){
      all.push({unit:u.name, ts:c.ts, comment:c.comment, score:c.score});
    }
  }
  if(!all.length){ body.innerHTML = '<div class="empty">'+t().quality_empty+'</div>'; return; }
  all.sort(function(a,b){return (b.ts||'').localeCompare(a.ts||'')});
  body.innerHTML = all.slice(0,20).map(function(c){
    return '<div class="log-row"><div class="log-lic">★'+c.score+'</div>'
      + '<div class="log-lts">'+esc((c.ts||'').replace('T',' '))+'<br><b style="color:var(--text)">'+esc(c.unit||'')+'</b></div>'
      + '<div class="log-ltxt">'+esc(c.comment||'')+'</div></div>';
  }).join('');
}

// ============== GUEST PROFILES ==============
async function loadGuests(){
  document.getElementById('guestListBody').innerHTML = '<div class="empty sk">—</div>';
  try{ D.guests = await api('/api/guests') }catch(_){ D.guests = {items:[], counts:{}} }
  const sub = document.getElementById('t_guests_sub'); if(sub) sub.textContent = t().guests_sub;
  renderGuestStats();
  renderGuestList();
}
function renderGuestStats(){
  const el = document.getElementById('guestStats'); if(!el) return;
  const c = (D.guests||{}).counts || {};
  const cards = [
    {ic:'👤', cls:'b', val:c.total||0, lbl:t().guests_stat_total},
    {ic:'⭐', cls:'gold', val:c.vip||0, lbl:t().guests_stat_vip},
    {ic:'🔁', cls:'g', val:c.repeat||0, lbl:t().guests_stat_repeat},
  ];
  el.innerHTML = cards.map(function(c){
    return '<div class="kpi"><div class="kpi-head"><div class="kpi-ic '+c.cls+'">'+c.ic+'</div></div>'
      + '<div class="kpi-val">'+c.val+'</div><div class="kpi-lbl">'+c.lbl+'</div></div>';
  }).join('');
}
function renderGuestList(){
  const body = document.getElementById('guestListBody'); if(!body) return;
  const items = ((D.guests||{}).items) || [];
  const filt = (document.getElementById('guestFilter')||{}).value || 'all';
  const q = ((document.getElementById('guestSearch')||{}).value || '').toLowerCase();
  let f = items;
  if(filt === 'vip') f = items.filter(function(x){return x.vip});
  else if(filt === 'repeat') f = items.filter(function(x){return x.stays >= 2});
  if(q) f = f.filter(function(x){
    return (x.name||'').toLowerCase().indexOf(q) >= 0 || (x.phone||'').toLowerCase().indexOf(q) >= 0;
  });
  if(!f.length){ body.innerHTML = '<div class="empty">'+t().guest_no_data+'</div>'; return; }
  let html = '<div style="overflow-x:auto"><table class="data"><thead><tr>'
    + '<th>'+t().guest_name+'</th><th class="num">'+t().guest_stays+'</th>'
    + '<th class="num">'+t().guest_nights+'</th><th>'+t().guest_last+'</th><th></th></tr></thead><tbody>';
  for(const g of f){
    const vipTag = g.vip ? ' <span class="pill gold">'+t().guest_vip_on+'</span>' : '';
    html += '<tr style="cursor:pointer" onclick="openGuestDrawer(&#39;'+esc(g.key)+'&#39;)">'
      + '<td class="strong">'+esc(g.name||'—')+vipTag+(g.phone?' <span class="muted" style="font-size:11px">· '+esc(g.phone)+'</span>':'')+'</td>'
      + '<td class="num">'+g.stays+'</td><td class="num">'+g.nights+'</td>'
      + '<td class="muted" style="font-size:11.5px">'+esc((g.last_seen||'').replace('T',' ').slice(0,16))+'</td>'
      + '<td style="text-align:end"><span class="muted">←</span></td></tr>';
  }
  html += '</tbody></table></div>';
  body.innerHTML = html;
}

async function openGuestDrawer(key){
  openDrawer('—','');
  setDrawerBody('<div class="empty sk">—</div>');
  let p;
  try{ p = await api('/api/guests/detail?key='+encodeURIComponent(key)) }catch(_){ p = null }
  if(!p || p.error){ setDrawerBody('<div class="empty">⚠</div>'); return; }
  setDrawerTitle((p.names||['—'])[p.names.length-1]+(p.vip?' ⭐':''), (p.phone||'')+' · '+(p.email||''));
  const stays = (p.reservations||[]).slice().reverse();
  const sums = (p.summaries||[]).slice().reverse();
  let body = '<div class="strat-overview">'
    + '<div class="stat-mini"><div class="v g">'+(stays.length)+'</div><div class="l">'+t().guest_stays+'</div></div>'
    + '<div class="stat-mini"><div class="v">'+(p.total_nights||0)+'</div><div class="l">'+t().guest_nights+'</div></div>'
    + '<div class="stat-mini"><div class="v gold">'+(sums.length)+'</div><div class="l">'+t().guest_drw_summaries+'</div></div>'
    + '</div>';
  if(stays.length){
    body += '<div class="context-h">'+t().guest_drw_stays+'</div><div class="list">'
      + stays.map(function(r){
        return '<div class="list-row"><span class="l-name">'+esc(r.unit||'')+'</span>'
          + '<span class="l-val">'+(r.checkin||'')+' → '+(r.checkout||'')+'</span>'
          + '<span class="l-tag">'+r.nights+'n</span></div>';
      }).join('') + '</div>';
  }
  if(sums.length){
    body += '<div class="context-h">'+t().guest_drw_summaries+'</div>'
      + sums.map(function(s){
        return '<div class="needs-item-text" style="margin-bottom:6px">'
          + '<div class="muted" style="font-size:10.5px;margin-bottom:4px">'+esc(s.ts||'')+'</div>'
          + esc(s.text||'') + '</div>';
      }).join('');
  }
  body += '<div class="context-h">'+t().guest_drw_notes+'</div>'
    + '<textarea id="guestNotes" style="min-height:80px">'+esc(p.notes||'')+'</textarea>';
  setDrawerBody(body);
  setDrawerFoot(
      '<button class="btn ghost sm" onclick="toggleGuestVip(&#39;'+esc(key)+'&#39;)">⭐ '+t().guest_drw_toggle_vip+'</button>'
    + '<button class="btn primary sm" onclick="saveGuestNotes(&#39;'+esc(key)+'&#39;)">💾 '+t().guest_drw_save+'</button>'
  );
}
async function saveGuestNotes(key){
  const ta = document.getElementById('guestNotes');
  const r = await post('/api/guests/notes', {key:key, notes:(ta?ta.value:'')});
  if(r.ok){ toast('✓'); closeDrawer(); loadGuests(); } else toast(r.error||t().err);
}
async function toggleGuestVip(key){
  const r = await post('/api/guests/toggle-vip', {key:key});
  if(r.ok){ toast(r.vip?'⭐':'✓'); openGuestDrawer(key); loadGuests(); }
}

// ============== DEEP CLEAN SCHEDULE ==============
async function loadCleaning(){
  document.getElementById('cleanListBody').innerHTML = '<div class="empty sk">—</div>';
  try{ D.clean = await api('/api/cleaning/schedule') }catch(_){ D.clean = {items:[], counts:{}} }
  const sub = document.getElementById('t_clean_sub'); if(sub) sub.textContent = t().clean_sub;
  renderCleaningStats();
  renderCleaningWeek();
  renderCleaningVisual();
  renderCleaningLink();
  renderCleaningList();
}

function _ar_wd(n){ return ['الاثنين','الثلاثاء','الأربعاء','الخميس','الجمعة','السبت','الأحد'][n] }
function _en_wd(n){ return ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][n] }
function _wd(n){ return L==='ar' ? _ar_wd(n) : _en_wd(n) }

function renderCleaningWeek(){
  const el = document.getElementById('cleanWeek'); if(!el) return;
  const items = ((D.clean||{}).items)||[];
  const today = new Date((D.clean||{}).today || new Date().toISOString().slice(0,10));
  // Build a date -> items[] map
  const byDate = {};
  for(const it of items){
    if(!it.next_scheduled) continue;
    if(!byDate[it.next_scheduled]) byDate[it.next_scheduled] = [];
    byDate[it.next_scheduled].push(it);
  }
  let html = '<div style="display:grid;grid-template-columns:repeat(7,1fr);gap:6px">';
  for(let i = 0; i < 7; i++){
    const d = new Date(today); d.setDate(today.getDate() + i);
    const iso = d.toISOString().slice(0,10);
    const wdIdx = (d.getDay() + 6) % 7;   // Monday=0
    const isToday = i === 0;
    const isWeekend = wdIdx === 3 || wdIdx === 4;   // Thu/Fri
    const todays = byDate[iso] || [];
    const bg = todays.length ? 'var(--gold-soft)' : (isWeekend ? 'var(--surface-2)' : 'var(--surface)');
    const border = isToday ? '2px solid var(--gold)' : '1px solid var(--line)';
    let body = '<div class="muted" style="font-size:11px;text-align:center;padding:14px 4px">'
             + (isWeekend ? '🚫 نهاية أسبوع' : '—') + '</div>';
    if(todays.length){
      body = todays.map(function(it){
        const stCls = it.next_status === 'blocked' ? 'ok' : 'info';
        return '<div style="background:var(--surface);border-radius:6px;padding:5px 8px;margin-bottom:4px;font-size:11.5px;font-weight:600">'
          + esc(it.name) + ' <span class="pill ' + stCls + '" style="font-size:9.5px">' + (it.next_status === 'blocked' ? '🔒' : '📅') + '</span></div>';
      }).join('');
    }
    html += '<div style="background:' + bg + ';border:' + border + ';border-radius:10px;padding:10px 8px 8px">'
      + '<div style="font-size:10.5px;color:var(--mut);font-weight:600">' + _wd(wdIdx) + (isToday ? ' · اليوم' : '') + '</div>'
      + '<div style="font-family:var(--font-mono);font-weight:700;font-size:16px;margin-bottom:6px">' + d.getDate() + '/' + (d.getMonth()+1) + '</div>'
      + body
      + '</div>';
  }
  html += '</div>';
  el.innerHTML = html;
}

function renderCleaningVisual(){
  const el = document.getElementById('cleanVisualGrid'); if(!el) return;
  const items = ((D.clean||{}).items)||[];
  const today = new Date((D.clean||{}).today || new Date().toISOString().slice(0,10));
  const byDate = {};
  for(const it of items){
    if(!it.next_scheduled) continue;
    if(!byDate[it.next_scheduled]) byDate[it.next_scheduled] = [];
    byDate[it.next_scheduled].push(it);
  }
  // 60-day grid, 7 cols
  const wdLabels = ['الإث','الثل','الأر','الخم','الجم','السب','الأح'];
  const wdLabelsEn = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'];
  const wds = L==='ar' ? wdLabels : wdLabelsEn;
  let html = '<div class="calgrid" style="margin-bottom:6px">';
  for(let i = 0; i < 7; i++) html += '<div style="text-align:center;color:var(--mut);font-size:10.5px;font-weight:600;padding:3px 0">' + wds[i] + '</div>';
  html += '</div><div class="calgrid">';
  // pad to align under correct weekday (Mon=0)
  const firstWd = (today.getDay() + 6) % 7;
  for(let i = 0; i < firstWd; i++) html += '<div></div>';
  for(let i = 0; i < 60; i++){
    const d = new Date(today); d.setDate(today.getDate() + i);
    const iso = d.toISOString().slice(0,10);
    const wdIdx = (d.getDay() + 6) % 7;
    const isWE = wdIdx === 3 || wdIdx === 4;
    const todays = byDate[iso] || [];
    const has = todays.length > 0;
    const cls = has
      ? (todays[0].next_status === 'blocked' ? 'cal-high' : 'cal-mid')
      : (isWE ? '' : '');
    const isToday = i === 0;
    const evtLabel = has ? (todays.length === 1 ? todays[0].name.slice(0, 11) : (todays.length + ' وحدات')) : '';
    html += '<div class="calday ' + cls + (isWE ? ' weekend' : '') + (has ? ' evt' : '') + '"'
      + (isToday ? ' style="outline:2px solid var(--gold);outline-offset:-2px"' : '')
      + ' onclick="cleanDayDrill(&#39;' + iso + '&#39;)">'
      + '<div class="cd-dnum">' + d.getDate() + '</div>'
      + '<div class="cd-wd">' + wds[wdIdx].toLowerCase() + '</div>'
      + (has ? '<div class="cd-evt">' + esc(evtLabel) + '</div>' : '')
      + '</div>';
  }
  html += '</div>';
  // Legend
  html += '<div style="display:flex;gap:12px;flex-wrap:wrap;margin-top:12px;font-size:11px;color:var(--mut)">'
    + '<span><span style="display:inline-block;width:12px;height:12px;background:#dcf3e6;border-radius:3px;vertical-align:middle;border:1.5px solid var(--gold)"></span> مؤكدة + مقفلة</span>'
    + '<span><span style="display:inline-block;width:12px;height:12px;background:#faeed1;border-radius:3px;vertical-align:middle;border:1.5px solid var(--gold)"></span> مجدولة</span>'
    + '<span><span style="display:inline-block;width:12px;height:12px;background:var(--surface-2);border-radius:3px;vertical-align:middle"></span> فاضي</span>'
    + '<span><span style="display:inline-block;width:12px;height:12px;background:var(--surface-2);border-radius:3px;vertical-align:middle;border:1.5px solid var(--gold)"></span> اليوم</span>'
    + '</div>';
  el.innerHTML = html;
}

function cleanDayDrill(iso){
  const items = ((D.clean||{}).items||[]).filter(function(x){return x.next_scheduled === iso});
  if(!items.length){
    toast('ما فيه تنظيفات في ' + iso);
    return;
  }
  openDrawer('🧹 ' + iso, items.length + ' وحدة');
  setDrawerBody(
    items.map(function(it){
      const status = it.next_status === 'blocked' ? '✓ مؤكد + الكالندر مقفل' : '📅 مجدول (يتأكد ٩م الليلة قبل)';
      return '<div class="list-item">'
        + '<div class="top"><span class="name">'+esc(it.name)+'</span><span class="meta">'+(it.beds || '?')+' غرف</span></div>'
        + '<div class="muted">'+esc(it.area || '')+'</div>'
        + '<div style="margin-top:6px"><span class="pill '+(it.next_status==='blocked'?'ok':'info')+'">'+status+'</span></div>'
        + '<div class="actions"><button class="btn green xs" onclick="cleanMarkDone('+it.lid+');closeDrawer()">✓ علّم منجز</button>'
        + '<button class="btn ghost xs" onclick="cleanResched('+it.lid+');closeDrawer()">📅 إعادة جدولة</button></div>'
        + '</div>';
    }).join('')
  );
  setDrawerFoot('<button class="btn ghost sm" onclick="closeDrawer()">إغلاق</button>');
}

function renderCleaningStats(){
  const el = document.getElementById('cleanStats'); if(!el) return;
  const c = (D.clean||{}).counts || {};
  const cards = [
    {ic:'🏠', cls:'b', val:c.total||0, lbl:t().clean_stat_total},
    {ic:'⚠', cls:'r', val:c.overdue||0, lbl:t().clean_stat_overdue},
    {ic:'📅', cls:'g', val:c.scheduled||0, lbl:t().clean_stat_scheduled},
    {ic:'🔒', cls:'p', val:c.blocked_tomorrow||0, lbl:t().clean_stat_tomorrow},
  ];
  el.innerHTML = cards.map(function(c){
    return '<div class="kpi"><div class="kpi-head"><div class="kpi-ic '+c.cls+'">'+c.ic+'</div></div>'
      +'<div class="kpi-val">'+c.val+'</div><div class="kpi-lbl">'+c.lbl+'</div></div>';
  }).join('');
}

function renderCleaningLink(){
  const el = document.getElementById('cleanLinkBox'); if(!el) return;
  const d = D.clean || {};
  if(!d.have_token){
    el.innerHTML = '<div class="muted" style="background:var(--yellow-soft);padding:11px 14px;border-radius:8px;color:var(--yellow);font-weight:500">'+t().clean_link_missing+'</div>';
    return;
  }
  const fullUrl = location.origin + d.cleaning_url;
  el.innerHTML =
    '<div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">'
    + '<input value="'+esc(fullUrl)+'" readonly id="cleanLinkInput" style="flex:1;min-width:240px;font-family:var(--font-mono);font-size:12px;background:var(--surface-2)">'
    + '<button class="btn ghost sm" onclick="copyCleanLink()">📋 '+t().clean_link_copy+'</button>'
    + '<a href="'+esc(d.cleaning_url)+'" target="_blank" class="btn primary sm">🔗 '+t().clean_link_open+'</a>'
    + '</div>'
    + '<div class="muted" style="margin-top:8px;font-size:11.5px">شارك هذا الرابط مع شركة التنظيف فقط · يحدّث تلقائياً كل يوم</div>';
}

function copyCleanLink(){
  const inp = document.getElementById('cleanLinkInput'); if(!inp) return;
  inp.select(); document.execCommand('copy');
  toast(t().copied);
}

function _cleanStatusPill(s){
  const map = {
    unscheduled: ['muted', t().clean_status_unscheduled],
    scheduled:   ['info', t().clean_status_scheduled],
    blocked:     ['ok', t().clean_status_blocked],
    pushed:      ['warn', t().clean_status_pushed],
  };
  const m = map[s] || ['muted', s||'—'];
  return '<span class="pill '+m[0]+'">'+m[1]+'</span>';
}

function renderCleaningList(){
  const body = document.getElementById('cleanListBody'); if(!body) return;
  const items = ((D.clean||{}).items)||[];
  const filt = (document.getElementById('cleanFilter')||{}).value || 'all';
  const q = ((document.getElementById('cleanSearch')||{}).value || '').toLowerCase();
  const today = (D.clean||{}).today || '';
  let filtered = items;
  if(filt === 'overdue') filtered = items.filter(function(x){return x.overdue});
  else if(filt === 'soon') filtered = items.filter(function(x){return x.days_until_next != null && x.days_until_next <= 7});
  else if(filt === 'unscheduled') filtered = items.filter(function(x){return !x.next_scheduled});
  if(q) filtered = filtered.filter(function(x){return (x.name||'').toLowerCase().indexOf(q) >= 0});
  if(!filtered.length){ body.innerHTML = '<div class="empty">—</div>'; return; }
  let html = '<div style="overflow-x:auto"><table class="data"><thead><tr>'
    + '<th>'+t().clean_unit+'</th>'
    + '<th>'+t().clean_last+'</th>'
    + '<th>'+t().clean_next+'</th>'
    + '<th>'+t().clean_status+'</th>'
    + '<th>'+t().clean_actions+'</th>'
    + '</tr></thead><tbody>';
  for(const it of filtered){
    const last = it.last_done ? (it.last_done + ' <span class="muted" style="font-size:11px">· '+it.days_since_last+' '+t().clean_days_ago+'</span>') : '—';
    const nxt = it.next_scheduled ? (it.next_scheduled + ' <span class="muted" style="font-size:11px">· '+it.days_until_next+' '+t().clean_days_left+'</span>') : '—';
    const overdueTag = it.overdue ? ' <span class="pill danger" style="margin-inline-start:6px">'+t().clean_overdue_lbl+'</span>' : '';
    const bedTag = it.beds ? ' <span class="muted" style="font-size:11px">'+it.beds+' غرف</span>' : '';
    html += '<tr><td class="strong">'+esc(it.name)+bedTag+overdueTag+'</td>'
      + '<td>'+last+' <button class="btn ghost xs" onclick="cleanSetLast('+it.lid+')" title="'+t().clean_set_last+'">✎</button></td>'
      + '<td>'+nxt+' <button class="btn ghost xs" onclick="cleanResched('+it.lid+')" title="'+t().clean_reschedule+'">📅</button></td>'
      + '<td>'+_cleanStatusPill(it.next_status)+'</td>'
      + '<td><button class="btn green xs" onclick="cleanMarkDone('+it.lid+')">✓ '+t().clean_mark_done+'</button></td>'
      + '</tr>';
  }
  html += '</tbody></table></div>';
  body.innerHTML = html;
}

async function cleanMarkDone(lid){
  if(!confirm(t().clean_confirm_done)) return;
  const r = await post('/api/cleaning/mark-done', {lid:lid});
  if(r.ok){ toast('✓'); loadCleaning(); } else toast(r.error||t().err);
}
async function cleanResched(lid){
  const date = prompt(t().clean_modal_resched + ' (YYYY-MM-DD)');
  if(!date) return;
  const r = await post('/api/cleaning/reschedule', {lid:lid, date:date});
  if(r.ok){ toast('✓'); loadCleaning(); } else toast(r.error||t().err);
}
async function cleanSetLast(lid){
  const date = prompt(t().clean_modal_set_last + ' (YYYY-MM-DD)');
  if(!date) return;
  const r = await post('/api/cleaning/set-last', {lid:lid, date:date});
  if(r.ok){ toast('✓'); loadCleaning(); } else toast(r.error||t().err);
}

async function uploadCleaningCSV(ev){ return _uploadCleaning(ev, '/api/cleaning/import-csv') }
async function uploadCleaningXLSX(ev){ return _uploadCleaning(ev, '/api/cleaning/import-xlsx') }

async function _uploadCleaning(ev, endpoint){
  const file = ev.target.files && ev.target.files[0];
  if(!file) return;
  const fd = new FormData(); fd.append('file', file);
  toast('⏳ يستورد…');
  let r;
  try{
    const resp = await fetch(endpoint + '?token=' + encodeURIComponent(tok()), {method:'POST', body: fd});
    r = await resp.json();
  }catch(e){ toast('خطأ'); return; }
  ev.target.value = '';
  const box = document.getElementById('cleanImportResult');
  if(r.ok){
    const dated = (r.matched||[]).length;
    const unknown = (r.unknown||[]).length;
    const noDate = (r.no_date||[]).length;
    const bad = (r.unmatched||[]).length;
    box.style.display = 'block';
    box.innerHTML = '<div class="card" style="background:var(--green-soft);border-color:rgba(14,158,95,.2)">'
      + '<div style="color:var(--green);font-weight:700;margin-bottom:8px">✓ تم الاستيراد بنجاح</div>'
      + '<div style="display:flex;gap:18px;flex-wrap:wrap;font-size:13px">'
        + '<span><b class="mono">'+dated+'</b> وحدة بتاريخ محدّد</span>'
        + (unknown ? '<span><b class="mono">'+unknown+'</b> وحدة تاريخها غير معروف (تُعتبر متأخّرة لجدولة فورية)</span>' : '')
        + (noDate ? '<span><b class="mono">'+noDate+'</b> بدون تاريخ</span>' : '')
        + (bad ? '<span style="color:var(--red)"><b class="mono">'+bad+'</b> ما طابقت أي وحدة</span>' : '')
      + '</div>'
      + ((r.unmatched||[]).length ? '<div class="muted" style="margin-top:10px;font-size:11.5px">⚠ غير مطابقة: '+r.unmatched.map(function(x){return esc(x.name)}).join('، ')+'</div>' : '')
      + '</div>';
    loadCleaning();
  }else{
    box.style.display = 'block';
    box.innerHTML = '<div class="card" style="background:var(--red-soft);border-color:rgba(196,67,67,.2);color:var(--red)">⚠ '+esc(r.error||'فشل الاستيراد')+'</div>';
  }
}

// ============== CALENDAR (forward pace + bulk apply) ==============
let calSelect = {start:null, end:null};

async function loadForwardCalendar(){
  const el = document.getElementById('calendarGrid');
  if(el) el.innerHTML = '<div class="empty sk">—</div>';
  try{ D.cal = await api('/api/calendar/forward?days=60') }catch(_){ D.cal = {days:[], events:[]} }
  const sub = document.getElementById('t_calendar_sub'); if(sub) sub.textContent = t().calendar_sub;
  renderForwardCalendar();
  renderCalEvents();
  renderBulkForm();
}

function _calClass(p){
  if(p >= 90) return 'cal-full';
  if(p >= 70) return 'cal-high';
  if(p >= 40) return 'cal-mid';
  return 'cal-low';
}

function renderForwardCalendar(){
  const el = document.getElementById('calendarGrid');
  if(!el) return;
  const days = (D.cal||{}).days || [];
  if(!days.length){ el.innerHTML = '<div class="empty">'+t().rev_no+'</div>'; return; }
  // Group by weekday header row (Sat-Sun-Mon-... for ar / Mon-...-Sun for en)
  const wdLabelsAr = ['الإث','الثل','الأر','الخم','الجم','السب','الأح'];
  const wdLabelsEn = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'];
  const wdLabels = (L==='ar') ? wdLabelsAr : wdLabelsEn;
  // pad start so days align under their weekday header
  const first = days[0];
  const firstWd = first.weekday; // 0..6
  const pad = firstWd;
  let html = '<div class="calgrid" style="margin-bottom:6px">';
  for(let i=0;i<7;i++){ html += '<div style="text-align:center;color:var(--mut);font-size:11px;font-weight:600;padding:4px 0">'+wdLabels[i]+'</div>'; }
  html += '</div><div class="calgrid">';
  for(let i=0;i<pad;i++){ html += '<div></div>'; }
  for(const d of days){
    const cls = _calClass(d.pace_pct);
    const isEvt = (d.events||[]).length > 0;
    const isWE = d.is_weekend;
    const isSel = calSelect.start && calSelect.end && d.date >= calSelect.start && d.date <= calSelect.end;
    const dnum = parseInt(d.date.slice(8,10), 10);
    const evtName = isEvt ? esc((d.events[0]||{}).name||'') : '';
    const tip = d.date + ' · ' + t().cal_pace + ' ' + d.pace_pct + '% · ' + d.available + ' ' + t().cal_avail
              + (d.avg_price ? ' · ' + t().cal_avg + ' ' + d.avg_price + ' SAR' : '')
              + (isEvt ? ' · ' + d.events.map(function(e){return e.name}).join(', ') : '');
    html += '<div class="calday '+cls+(isEvt?' evt':'')+(isWE?' weekend':'')+(isSel?' sel':'')+'"'
      + ' title="'+tip+'"'
      + ' onclick="calClick(&#39;'+d.date+'&#39;)">'
      + '<div class="cd-dnum">'+dnum+'</div>'
      + '<div class="cd-pct">'+d.pace_pct+'%</div>'
      + '<div class="cd-wd">'+wdLabels[d.weekday].toLowerCase()+'</div>'
      + (evtName ? '<div class="cd-evt">'+evtName+'</div>' : '')
      + '</div>';
  }
  html += '</div>';
  // legend
  html += '<div style="display:flex;gap:14px;flex-wrap:wrap;margin-top:14px;font-size:11.5px;color:var(--mut)">'
    + '<span><span style="display:inline-block;width:14px;height:14px;background:#fae3e3;border-radius:3px;vertical-align:middle"></span> &lt;40%</span>'
    + '<span><span style="display:inline-block;width:14px;height:14px;background:#faeed1;border-radius:3px;vertical-align:middle"></span> 40-69%</span>'
    + '<span><span style="display:inline-block;width:14px;height:14px;background:#dcf3e6;border-radius:3px;vertical-align:middle"></span> 70-89%</span>'
    + '<span><span style="display:inline-block;width:14px;height:14px;background:#a3e0bd;border-radius:3px;vertical-align:middle"></span> 90%+</span>'
    + '<span><span style="display:inline-block;width:14px;height:14px;border:1.5px solid var(--gold);border-radius:3px;vertical-align:middle"></span> '+t().cal_event+'</span>'
    + '</div>';
  el.innerHTML = html;
}

function calClick(date){
  if(!calSelect.start || (calSelect.start && calSelect.end)){
    // start a new selection
    calSelect = {start:date, end:date};
  }else{
    // extend
    if(date < calSelect.start){ calSelect.end = calSelect.start; calSelect.start = date; }
    else                      { calSelect.end = date; }
  }
  renderForwardCalendar();
  renderBulkForm();
}

async function renderCalEvents(){
  const el = document.getElementById('calEventsBody');
  if(!el) return;
  // Pull the full editable events list (default + custom)
  try{ D.events = await api('/api/events') }catch(_){ D.events = {events:[]} }
  const all = (D.events||{}).events || [];
  if(!all.length){ el.innerHTML = '<div class="empty">'+t().cal_no_events+'</div>'; return; }
  // Group by source
  const today = new Date().toISOString().slice(0,10);
  const sorted = all.slice().sort(function(a,b){ return (a.start||'').localeCompare(b.start||'') });
  let html = '';
  for(const e of sorted){
    const active = e.start <= today && today <= e.end;
    const past = e.end < today;
    const boostPct = Math.round((e.boost-1)*100);
    const boostLbl = (boostPct >= 0 ? '+' : '') + boostPct + '%';
    const boostCls = boostPct > 0 ? 'ok' : (boostPct < 0 ? 'danger' : 'muted');
    const rangeLbl = e.start === e.end ? e.start : (e.start + ' → ' + e.end);
    const editable = e.source === 'custom';
    const dimStyle = past ? 'opacity:.45' : '';
    html += '<div class="log-row" style="grid-template-columns:auto 1fr auto;'+dimStyle+'">'
      + '<div class="log-lic">'+(active?'🟢':(past?'⚪':'🎉'))+'</div>'
      + '<div><div style="font-weight:600;font-size:13px">'+esc(e.name)+' <span class="pill '+boostCls+'">'+boostLbl+'</span>'+(editable?'':' <span class="muted" style="font-size:10px">· افتراضي</span>')+'</div>'
      + '<div class="muted" style="font-size:11.5px">'+rangeLbl+' · '+esc(e.kind||'')+'</div></div>'
      + '<div style="display:flex;gap:5px">'
        + (editable ? ('<button class="btn ghost xs" onclick="editEvent(&#39;'+e.id+'&#39;)">✎</button>'
                    + '<button class="btn red xs" onclick="deleteEvent(&#39;'+e.id+'&#39;)">🗑</button>') : '')
      + '</div></div>';
  }
  // Add-new form
  html += '<div style="border-top:1px solid var(--line);padding-top:12px;margin-top:12px">'
    + '<div style="font-weight:600;font-size:12px;color:var(--mut);margin-bottom:8px">+ أضف مناسبة جديدة</div>'
    + '<div id="evtFormHost"></div>'
    + '<button class="btn ghost sm" onclick="renderEventForm()">+ مناسبة جديدة</button>'
    + '</div>';
  el.innerHTML = html;
}

function renderEventForm(existing){
  const host = document.getElementById('evtFormHost');
  if(!host) return;
  const e = existing || {id:'', name:'', start:'', end:'', boost:1.2, kind:'custom'};
  host.innerHTML =
    '<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:8px">'
    + '<input id="evN" placeholder="اسم المناسبة" value="'+esc(e.name)+'" style="grid-column:1/-1">'
    + '<input id="evS" type="date" value="'+e.start+'">'
    + '<input id="evE" type="date" value="'+e.end+'">'
    + '<input id="evB" type="number" step="0.05" min="0.3" max="3.0" value="'+e.boost+'" placeholder="boost (1.4 = +40%)">'
    + '<input id="evK" placeholder="kind (eid/national/custom...)" value="'+esc(e.kind||'custom')+'">'
    + '</div>'
    + '<div style="display:flex;gap:5px;justify-content:flex-end">'
      + '<button class="btn ghost xs" onclick="document.getElementById(&#39;evtFormHost&#39;).innerHTML=&#39;&#39;">✕</button>'
      + '<button class="btn primary xs" onclick="saveEvent(&#39;'+e.id+'&#39;)">💾 حفظ</button>'
    + '</div>';
}

async function saveEvent(eid){
  const body = {
    id: eid || null,
    name: (document.getElementById('evN')||{}).value || '',
    start: (document.getElementById('evS')||{}).value || '',
    end: (document.getElementById('evE')||{}).value || '',
    boost: parseFloat((document.getElementById('evB')||{}).value || '1.0'),
    kind: (document.getElementById('evK')||{}).value || 'custom',
  };
  const r = await post('/api/events/save', body);
  if(r.ok){ toast('✅'); document.getElementById('evtFormHost').innerHTML=''; loadForwardCalendar(); }
  else toast(r.error || t().err);
}
function editEvent(eid){
  const e = ((D.events||{}).events||[]).find(function(x){return x.id === eid});
  if(e) renderEventForm(e);
}
async function deleteEvent(eid){
  if(!confirm('احذف المناسبة؟')) return;
  const r = await post('/api/events/delete', {id:eid});
  if(r.ok){ toast('🗑'); loadForwardCalendar(); }
}

async function renderBulkForm(){
  const el = document.getElementById('bulkForm');
  if(!el) return;
  const has = !!(calSelect.start && calSelect.end);
  if(!has){
    el.innerHTML = '<div class="empty">'+t().bulk_select_range+'</div>';
    return;
  }
  // Fetch units list once for the filter dropdowns
  if(!D.units){
    try{ D.units = await api('/api/units') }catch(_){ D.units = {units:[], beds:[], areas:[]} }
  }
  const beds = (D.units||{}).beds || [];
  const areas = (D.units||{}).areas || [];
  el.innerHTML =
    '<div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:12px">'
    + '<span class="muted">'+t().bulk_from+'</span>'
    + '<span class="mono" style="font-weight:600">'+calSelect.start+'</span>'
    + '<span class="muted">'+t().bulk_to+'</span>'
    + '<span class="mono" style="font-weight:600">'+calSelect.end+'</span>'
    + '<button class="btn ghost xs" onclick="calSelect={start:null,end:null};renderForwardCalendar();renderBulkForm()">✕</button>'
    + '</div>'
    + '<div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:10px">'
    + '<select id="bulkAction" style="width:auto;padding:7px 12px"><option value="raise">'+t().bulk_raise+'</option><option value="lower">'+t().bulk_lower+'</option></select>'
    + '<input id="bulkPct" type="number" min="1" max="80" value="10" style="width:90px;text-align:center"> %'
    + '</div>'
    + '<div class="muted" style="margin-bottom:6px;font-size:11.5px;font-weight:600">فلتر النطاق (اختياري):</div>'
    + '<div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:12px">'
    + '<select id="bulkBeds" style="width:auto;padding:7px 12px"><option value="">كل عدد الغرف</option>'
      + beds.map(function(b){return '<option value="'+b+'">'+b+' غرفة</option>'}).join('')
    + '</select>'
    + '<select id="bulkArea" style="width:auto;padding:7px 12px;max-width:180px"><option value="">كل المناطق</option>'
      + areas.map(function(a){return '<option value="'+esc(a)+'">'+esc(a)+'</option>'}).join('')
    + '</select>'
    + '<button class="btn ghost xs" onclick="toggleBulkUnits()">'+(bulkPickerOpen?'إخفاء قائمة الوحدات':'اختر وحدات محددة…')+'</button>'
    + '</div>'
    + '<div id="bulkUnitPicker" style="'+(bulkPickerOpen?'':'display:none')+'max-height:200px;overflow-y:auto;border:1px solid var(--line);border-radius:8px;padding:8px;margin-bottom:12px;background:var(--surface-2)">'
      + ((D.units||{}).units||[]).map(function(u){
          const checked = bulkSelectedUnits.indexOf(u.id) >= 0 ? 'checked' : '';
          return '<label style="display:flex;align-items:center;gap:8px;padding:4px 0;font-size:12.5px;cursor:pointer">'
            + '<input type="checkbox" value="'+u.id+'" '+checked+' onchange="toggleBulkUnit('+u.id+')">'
            + '<span>'+esc(u.name||('unit-'+u.id))+'</span>'
            + '<span class="muted" style="margin-inline-start:auto">'+(u.beds||'?')+' غرف · '+esc(u.area||'')+'</span>'
            + '</label>';
        }).join('')
    + '</div>'
    + '<button class="btn primary sm" onclick="doBulkApply()" style="width:100%">⚡ '+t().bulk_apply+'</button>'
    + '<div class="muted" style="margin-top:10px;font-size:11.5px">'+t().bulk_confirm+'</div>';
}

let bulkPickerOpen = false;
let bulkSelectedUnits = [];
function toggleBulkUnits(){ bulkPickerOpen = !bulkPickerOpen; renderBulkForm() }
function toggleBulkUnit(lid){
  const i = bulkSelectedUnits.indexOf(lid);
  if(i >= 0) bulkSelectedUnits.splice(i, 1); else bulkSelectedUnits.push(lid);
}

async function doBulkApply(){
  if(!calSelect.start || !calSelect.end) return;
  if(!confirm(t().bulk_confirm)) return;
  const action = document.getElementById('bulkAction').value;
  const pct = parseFloat(document.getElementById('bulkPct').value || '0');
  const beds = document.getElementById('bulkBeds').value;
  const area = document.getElementById('bulkArea').value;
  if(!pct || pct <= 0) return;
  const body = {
    start: calSelect.start, end: calSelect.end, percent: pct, action: action
  };
  if(bulkSelectedUnits.length) body.lids = bulkSelectedUnits.slice();
  if(beds) body.beds = parseInt(beds, 10);
  if(area) body.area = area;
  const r = await post('/api/pricing/bulk', body);
  if(r.ok){
    toast(t().bulk_applied.replace('{a}', r.applied).replace('{s}', r.skipped) + (r.dry_run?' (DRY-RUN)':''));
    calSelect = {start:null, end:null};
    bulkSelectedUnits = [];
    loadForwardCalendar();
  } else toast(r.error || t().err);
}

function renderArrivalsTimeline(){
  const el = document.getElementById('arrivalsTimeline');
  const cnt = document.getElementById('arrivalsCount');
  if(!el) return;
  const d = D.arrivals || {items:[]};
  const items = d.items || [];
  if(cnt) cnt.textContent = items.length ? ('· '+items.length) : '';
  if(!items.length){
    el.innerHTML = '<div class="empty">'+t().no_arrivals_window+'</div>';
    return;
  }
  el.innerHTML = items.map(function(a){
    const hrs = a.hours_until;
    let when;
    if(hrs < 0) when = '<span style="color:var(--mut)">'+t().arr_past+'</span>';
    else if(hrs < 1) when = '<span style="color:var(--red);font-weight:600">'+Math.round(hrs*60)+' '+t().arr_minutes+'</span>';
    else if(hrs < 24) when = '<span style="color:var(--yellow);font-weight:600">'+t().arr_in+' '+hrs.toFixed(1)+' '+t().arr_hours+'</span>';
    else when = '<span class="muted">'+t().arr_in+' '+Math.round(hrs/24)+' '+t().pr_d_days+'</span>';
    const signPill = a.signed
      ? '<span class="pill ok">✓ '+t().arr_signed+'</span>'
      : '<span class="pill danger">✗ '+t().arr_unsigned+'</span>';
    return '<div style="display:grid;grid-template-columns:auto 1fr auto;gap:10px;padding:11px 12px;border-bottom:1px solid var(--line);align-items:center">'
      + '<div style="font-family:var(--font-mono);font-size:12.5px;font-weight:600;color:var(--text-2);white-space:nowrap">'+esc(a.checkin_label)+'</div>'
      + '<div style="min-width:0"><div style="font-weight:600;font-size:13.5px">'+esc(a.guest)+' <span class="muted" style="font-weight:500;font-size:12px">· '+esc(a.unit)+'</span></div><div style="display:flex;gap:6px;margin-top:3px;flex-wrap:wrap">'+signPill+'<span class="muted" style="font-size:11px">'+a.nights+' '+t().nights+'</span></div></div>'
      + '<div style="font-size:12px;white-space:nowrap">'+when+'</div>'
      + '</div>';
  }).join('');
}

function renderFresh(){
  const ov = D.ov || {};
  const ready = ov.ready !== false;
  ['dot','sideDot'].forEach(function(id){ const el=document.getElementById(id); if(el) el.className='dot'+(ready?'':' warm') });
  const u = ov.updated ? new Date(ov.updated*1000) : null;
  const ts = u ? u.toLocaleTimeString(L==='ar'?'ar-SA':'en-US',{hour:'2-digit',minute:'2-digit'}) : '—';
  const txt = t().fresh+' '+ts+' · '+t().live;
  ['freshness','sideStatus'].forEach(function(id){ const el=document.getElementById(id); if(el) el.textContent = txt });
}

function populateUnitFilter(){
  const sel = document.getElementById('ibFilterUnit'); if(!sel) return;
  const cur = sel.value;
  const units = new Set();
  ((D.inbox||{}).replies||[]).forEach(function(r){ if(r.unit) units.add(r.unit) });
  ((D.inbox||{}).escalations||[]).forEach(function(e){ if(e.unit) units.add(e.unit) });
  sel.innerHTML = '<option value="">'+t().all_units+'</option>' +
    Array.from(units).sort().map(function(u){return '<option value="'+esc(u)+'">'+esc(u)+'</option>'}).join('');
  sel.value = cur;
}

/* ============================================================
   KPIs (HOME)
   ============================================================ */
function renderKpis(){
  const ov = D.ov || {}, td = D.today || {};
  const occN = td.occupied||0, occT = td.active||ov.active_units||0;
  const occPct = occT ? Math.round(occN/occT*100) : 0;
  const pending = ov.pending_cards||0, escs = ov.open_escalations||0, empty = td.empty_n||0;
  const k = [
    {ic:'◌', cls:'g', val:occN+'/'+occT, lbl:t().occ_tonight, sub:'<span class="pill '+(occPct>=85?'ok':(occPct>=60?'info':'warn'))+'">'+occPct+'%</span>'},
    {ic:'$', cls:'gold', val:fmt(ov.rev_7)+' SAR', lbl:t().rev_7},
    {ic:'∿', cls:'b', val:fmt(ov.rev_30)+' SAR', lbl:t().rev_30},
    {ic:'⌂', cls:(empty>0?'y':'g'), val:empty, lbl:t().empty_units, sub:(empty>0?'<a onclick="go(\\'today\\')" style="cursor:pointer;color:var(--gold);font-weight:600;font-size:11px">←</a>':'<span class="pill ok">✓</span>')},
    {ic:'💬', cls:(pending>0?'y':''), val:pending, lbl:t().pending_rep, sub:(pending>0?'<a onclick="go(\\'inbox\\')" style="cursor:pointer;color:var(--gold);font-weight:600;font-size:11px">←</a>':'<span class="muted">—</span>')},
    {ic:'🚨', cls:(escs>0?'r':''), val:escs, lbl:t().open_esc, sub:(escs>0?'<a onclick="go(\\'inbox\\')" style="cursor:pointer;color:var(--red);font-weight:600;font-size:11px">←</a>':'<span class="muted">—</span>'), vc:(escs>0?'red':'')}
  ];
  document.getElementById('kpis').innerHTML = k.map(function(x){
    return '<div class="kpi">'
      +'<div class="kpi-head"><div class="kpi-ic '+x.cls+'">'+x.ic+'</div>'+(x.sub||'')+'</div>'
      +'<div class="kpi-val '+(x.vc||'')+'">'+x.val+'</div>'
      +'<div class="kpi-lbl">'+x.lbl+'</div></div>';
  }).join('');
}

function renderNeedsBanner(){
  const ib = D.inbox||{replies:[],escalations:[]};
  const escs = (ib.escalations||[]).filter(function(e){return !e.claimed_by});
  const reps = ib.replies||[];
  const el = document.getElementById('needsBanner');
  if(escs.length===0 && reps.length===0){
    el.innerHTML = '<div class="card" style="background:var(--green-soft);border-color:rgba(14,158,95,.18);text-align:center;padding:14px"><span style="color:var(--green);font-weight:600">✓ '+t().no_needs+'</span></div>';
    return;
  }
  el.innerHTML = '<div class="card" style="background:var(--yellow-soft);border-color:rgba(201,150,23,.20);display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;padding:13px">'
    +'<div style="font-size:13px;color:var(--text)"><b style="color:var(--yellow)">⚠ '+t().needs_alert+':</b> '+(escs.length?escs.length+' '+t().open_esc:'')+(escs.length&&reps.length?' · ':'')+(reps.length?reps.length+' '+t().pending_rep:'')+'</div>'
    +'<button class="btn primary sm" onclick="go(\\'inbox\\')">←</button>'
    +'</div>';
}

function renderTodayHome(){
  const td = D.today||{};
  const dEl = document.getElementById('t_today_date');
  if(dEl && td.date){
    try{ dEl.textContent = new Date(td.date).toLocaleDateString(L==='ar'?'ar-SA':'en-US',{weekday:'short',day:'numeric',month:'short'}) }catch(_){}
  }
  const arr = td.arrivals||[], dep = td.departures||[], em = td.empty||[];
  const body = document.getElementById('todayBody');
  body.innerHTML =
    '<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px">'
      +'<div style="background:var(--green-soft);border-radius:8px;padding:11px;text-align:center"><div style="font-size:22px;font-weight:700;color:var(--green);font-family:var(--font-mono)">'+arr.length+'</div><div style="font-size:11px;color:var(--mut);margin-top:2px">🟢 وصول</div></div>'
      +'<div style="background:var(--blue-soft);border-radius:8px;padding:11px;text-align:center"><div style="font-size:22px;font-weight:700;color:var(--blue);font-family:var(--font-mono)">'+dep.length+'</div><div style="font-size:11px;color:var(--mut);margin-top:2px">🔵 مغادرة</div></div>'
      +'<div style="background:var(--gold-soft);border-radius:8px;padding:11px;text-align:center;cursor:pointer" onclick="go(\\'today\\')"><div style="font-size:22px;font-weight:700;color:var(--gold);font-family:var(--font-mono)">'+em.length+'</div><div style="font-size:11px;color:var(--mut);margin-top:2px">🏠 فاضية</div></div>'
    +'</div>';
}

function renderRevCard(){
  const rev = D.rev||{};
  const monthly = (rev.monthly||[]).slice(-12);
  const body = document.getElementById('revCardBody');
  if(!monthly.length){ body.innerHTML='<div class="empty">'+t().rev_no+'</div>'; return }
  const max = Math.max.apply(null, monthly.map(function(m){return m.rev}));
  const last = monthly[monthly.length-1] || {rev:0};
  const prev = monthly.length>1 ? monthly[monthly.length-2] : null;
  const delta = prev && prev.rev ? Math.round((last.rev-prev.rev)/prev.rev*100) : null;
  const sub = document.getElementById('revCardSub');
  if(sub) sub.innerHTML = delta!==null ? ('<span class="kpi-delta '+(delta>=0?'up':'dn')+'">'+(delta>=0?'+':'')+delta+'%</span>') : '';
  const bars = monthly.map(function(m){
    const h = Math.max(5, (m.rev/max)*110);
    return '<div class="bar-col"><div class="bar-tip">'+fmt(m.rev)+'</div><div class="bar" style="height:'+h+'px"></div><div class="bar-label">'+m.m.slice(5)+'</div></div>';
  }).join('');
  body.innerHTML = '<div class="bar-chart">'+bars+'</div>';
}

function renderRecent(){
  const items = (D.log||[]).slice(0,10);
  const body = document.getElementById('recentBody');
  if(!items.length){ body.innerHTML='<div class="empty">'+t().log_empty+'</div>'; return }
  const ic={guest:'💬',escalation:'🚨',pricing:'💰',report:'📊'};
  body.innerHTML = items.map(function(e){
    return '<div class="log-row"><div class="log-lic">'+(ic[e.cat]||'·')+'</div><div class="log-lts">'+esc(shortTime(e.ts))+'</div><div class="log-ltxt">'+esc(e.text)+'</div></div>';
  }).join('');
}

/* ============================================================
   INBOX (with filters + collapsible items)
   ============================================================ */
function filteredInboxItems(){
  const ib = D.inbox||{replies:[],escalations:[]};
  let items = [];
  const f = inboxFilter;
  const escs = (ib.escalations||[]).filter(function(e){return !e.claimed_by});
  const reps = ib.replies||[];
  if(f.type==='all' || f.type==='esc') escs.forEach(function(e){items.push({k:'esc', d:e})});
  if(f.type==='all' || f.type==='replies') reps.forEach(function(r){items.push({k:'rep', d:r})});
  if(f.type==='auto') (D.auto||[]).slice(0,50).forEach(function(a){items.push({k:'auto', d:a})});
  // filter by unit
  if(f.unit) items = items.filter(function(x){return (x.d.unit||'') === f.unit});
  // filter by search
  if(f.search){
    const q = f.search.toLowerCase();
    items = items.filter(function(x){
      const d = x.d;
      return ((d.guest||'')+(d.unit||'')+(d.guest_text||'')+(d.reply||'')+(d.draft||'')).toLowerCase().indexOf(q) >= 0;
    });
  }
  return items;
}

function renderInbox(){
  buildInboxTabs();
  const items = filteredInboxItems();
  const el = document.getElementById('inboxList');
  if(!items.length){ el.innerHTML='<div class="empty">'+t().ib_no+'</div>'; return }
  el.innerHTML = items.map(function(x){
    if(x.k==='auto') return renderAutoItem(x.d);
    return renderInboxItem(x.k, x.d);
  }).join('');
  // re-attach textarea values if user was editing
}

function renderInboxItem(k, d){
  // IDs are returned as STRINGS from the backend (Discord snowflakes overflow
  // JS number precision). Wrap them in HTML-encoded single quotes inside the
  // onclick attribute (`&#39;` decodes to ') so JS receives a string literal.
  const idAttr = String(d.id);
  const idJs = "&#39;" + idAttr + "&#39;";
  const isOpen = openInboxId === idAttr;
  const conf = d.confidence!==undefined && d.confidence!==null ? d.confidence : null;
  const confClass = conf===null?'':(conf>=85?'high':(conf>=60?'mid':'low'));
  const confChip = conf!==null && k==='rep' ? '<span class="ibox-conf '+confClass+'">'+conf+'%</span>' : '';
  return '<div class="ibox '+(k==='esc'?'escalation':'reply')+(isOpen?' open':'')+'" id="ib_'+idAttr+'">'
    + '<div class="ibox-row" onclick="toggleInbox('+idJs+')">'
    + '<div class="ibox-icon '+(k==='esc'?'esc':'rep')+'">'+(k==='esc'?'🚨':'💬')+'</div>'
    + '<div class="ibox-main"><div class="ibox-top"><span class="ibox-who">'+esc(d.guest||'')+'</span><span class="ibox-unit">'+esc(d.unit||'')+'</span></div><div class="ibox-preview">'+esc((d.guest_text||'').slice(0,160))+'</div></div>'
    + '<div class="ibox-meta">'+confChip+'<span class="ibox-time">'+esc(shortTime(d.time||''))+'</span></div>'
    + '<span class="ibox-expand">⌃</span>'
    + '</div>'
    + (isOpen ? '<div class="ibox-body" id="ibbody_'+idAttr+'"><div class="empty sk">—</div></div>' : '')
    + '</div>';
}

function renderAutoItem(a){
  return '<div class="ibox" style="border-inline-start:3px solid var(--green)">'
    + '<div class="ibox-row">'
    + '<div class="ibox-icon" style="background:var(--green-soft);color:var(--green)">⚡</div>'
    + '<div class="ibox-main">'
      + '<div class="ibox-top"><span class="ibox-who">'+esc(a.guest||'')+'</span><span class="ibox-unit">'+esc(a.unit||'')+'</span><span class="pill ok">'+(a.conf||0)+'%</span></div>'
      + '<div class="ibox-preview"><b style="color:var(--text-2)">Q:</b> '+esc((a.guest_text||'').slice(0,100))+' <b style="color:var(--gold);margin-inline-start:8px">A:</b> '+esc((a.reply||'').slice(0,100))+'</div>'
    + '</div>'
    + '<div class="ibox-meta"><span class="ibox-time">'+esc(shortTime(a.ts||''))+'</span></div>'
    + '</div></div>';
}

async function toggleInbox(id){
  id = String(id);          // always a string (Discord snowflake)
  if(openInboxId === id){
    openInboxId = null;
    const el = document.getElementById('ib_'+id);
    if(el){ el.classList.remove('open'); const b=document.getElementById('ibbody_'+id); if(b) b.remove() }
    return;
  }
  // close any other
  if(openInboxId){
    const prev = document.getElementById('ib_'+openInboxId);
    if(prev){ prev.classList.remove('open'); const pb=document.getElementById('ibbody_'+openInboxId); if(pb) pb.remove() }
  }
  openInboxId = id;
  // re-render to add body slot
  renderInbox();
  // fetch detail
  try{
    const det = await api('/api/inbox/detail?id='+encodeURIComponent(id));
    if(det && det.error){
      const b = document.getElementById('ibbody_'+id);
      if(b) b.innerHTML = '<div class="empty">⚠ '+esc(det.error)+'</div>';
      return;
    }
    renderInboxDetail(id, det);
  }catch(e){
    const b = document.getElementById('ibbody_'+id);
    if(b) b.innerHTML = '<div class="empty">⚠ '+(e==='unauthorized'?'unauthorized':'error')+'</div>';
  }
}

function renderInboxDetail(id, det){
  const b = document.getElementById('ibbody_'+id);
  if(!b) return;
  const k = det.kind;
  const isEsc = k==='escalation';

  // thread
  let thread = '<div class="empty muted" style="padding:14px">—</div>';
  if(det.thread && det.thread.length){
    thread = '<div class="thread">' + det.thread.map(function(m){
      const cls = m.from==='guest'?'g':(m.automated?'h auto':'h');
      const auto = m.automated ? ' <span class="bub-auto-tag">'+t().auto_msg+'</span>' : '';
      return '<div class="bub '+cls+'">'+
        '<div class="bub-meta">'+esc(shortTime(m.ts))+auto+'</div>'+
        '<div class="bub-tx">'+esc(m.text||'')+'</div></div>';
    }).join('') + '</div>';
  }

  // context box
  const checkin = det.checkin || '—', checkout = det.checkout || '—';
  const nights = det.nights || '—';
  const total = det.total_price ? fmt(det.total_price)+' SAR' : '—';
  const status = det.confirmed ? '<span class="pill ok">'+t().drw_confirmed+'</span>' : '<span class="pill warn">'+t().drw_not_confirmed+'</span>';
  const ctx = '<div class="context-box">'
    + '<div class="context-h">'+t().drw_context+'</div>'
    + '<div class="context-row"><span class="l">'+t().drw_status+'</span><span class="v">'+status+'</span></div>'
    + '<div class="context-row"><span class="l">'+t().drw_dates+'</span><span class="v">'+esc(checkin)+' → '+esc(checkout)+'</span></div>'
    + '<div class="context-row"><span class="l">'+t().drw_nights+'</span><span class="v">'+nights+'</span></div>'
    + '<div class="context-row"><span class="l">'+t().drw_total+'</span><span class="v">'+total+'</span></div>'
    + '</div>';

  // reasoning
  let reasoning = '';
  if(isEsc){
    reasoning = '<div class="reasoning-box"><div class="h">🧠 '+t().drw_reasoning+'</div>'
      + (det.reason?'<div class="reason-txt">⚠ '+esc(det.reason)+'</div>':'')
      + '</div>';
  }else if(det.intent || det.confidence!==null){
    const sent = det.sentiment==='upset' ? '<span class="pill danger">'+t().sentiment_upset+'</span>' : '<span class="pill ok">'+t().sentiment_ok+'</span>';
    reasoning = '<div class="reasoning-box"><div class="h">🧠 '+t().drw_reasoning+'</div>'
      + '<div class="reasoning-chips">'
        + (det.intent?'<span class="pill info">'+esc(det.intent)+'</span>':'')
        + (det.confidence!==null?'<span class="pill purple">'+t().drw_confidence+': '+det.confidence+'%</span>':'')
        + sent
      + '</div></div>';
  }

  // action row + draft for replies — quote id so JS gets a string literal
  const idJs = "&#39;" + id + "&#39;";
  let actions = '';
  if(isEsc){
    actions = '<div class="action-row"><input id="cn_'+id+'" placeholder="'+t().claim_ph+'"><button class="btn primary sm" onclick="doClaim('+idJs+')">🙋 '+t().claim+'</button></div>';
  }else{
    actions = '<div class="draft-label">✍ '+t().drw_draft+'</div>'
      + '<textarea id="ta_'+id+'" placeholder="'+t().drw_draft+'">'+esc(det.draft||'')+'</textarea>'
      + '<div class="action-row">'
        + '<button class="btn green sm" onclick="doSend('+idJs+')">✓ '+t().rep_send+'</button>'
        + '<button class="btn ghost sm" onclick="focusEdit('+idJs+')">✎ '+t().rep_edit_focus+'</button>'
        + '<button class="btn red sm" onclick="doReject('+idJs+')">✕ '+t().rep_reject+'</button>'
        + '<button class="btn ghost sm" onclick="toggleTeach('+idJs+')">🧠 '+t().rep_teach+'</button>'
      + '</div>'
      + '<div class="teach-form" id="teach_'+id+'">'
        + '<div style="font-size:11px;color:var(--purple);font-weight:700;text-transform:uppercase;letter-spacing:.4px;margin-bottom:7px">🧠 '+t().teach_label+'</div>'
        + '<input id="teachT_'+id+'" placeholder="'+t().teach_topic+'">'
        + '<textarea id="teachF_'+id+'" placeholder="'+t().teach_fact+'" style="min-height:60px"></textarea>'
        + '<div class="row"><button class="btn ghost xs" onclick="toggleTeach('+idJs+')">✕</button><button class="btn primary xs" onclick="doTeach('+idJs+')">💾 '+t().teach_save+'</button></div>'
      + '</div>';
  }

  b.innerHTML = '<div class="context-grid">'
    + '<div>'+thread+'</div>'
    + '<div>'+ctx+reasoning+'</div>'
    + '</div>'+actions;
}

function focusEdit(id){ id=String(id); const ta=document.getElementById('ta_'+id); if(ta){ ta.focus(); ta.setSelectionRange(ta.value.length, ta.value.length) } }
function toggleTeach(id){ id=String(id); const el=document.getElementById('teach_'+id); if(el) el.classList.toggle('open') }

/* ============================================================
   DISCOUNT BANNER + TODAY EMPTY GRID
   ============================================================ */
function renderDiscountBanner(){
  const el = document.getElementById('discountBanner');
  if(!el) return;
  const d = D.disc || {};
  if(d.paused){
    const u = (d.until_iso||'').replace('T',' ').slice(0,16);
    el.innerHTML = '<div class="discount-banner paused"><div class="info"><span class="pulse"></span><span class="txt">⏸ '+t().discount_paused+' <b>'+esc(u)+'</b></span></div><button class="btn green sm" onclick="doResume()">▶ '+t().resume+'</button></div>';
  }else{
    el.innerHTML = '<div class="discount-banner"><div class="info"><span class="pulse"></span><span class="txt">▶ '+t().discount_running+'</span></div><button class="btn ghost sm" onclick="doPause(24)">⏸ '+t().pause_24+'</button></div>';
  }
}

function renderTodayEmpty(){
  renderDiscountBanner();
  const wrap = document.getElementById('emptyGridWrap');
  const data = D.tonight || {items:[]};
  const items = data.items || [];
  if(!items.length){ wrap.innerHTML = '<div class="card empty"><span class="ic">🎉</span>'+t().no_empty_tonight+'</div>'; return }
  wrap.innerHTML = '<div class="empty-grid">' + items.map(renderEmptyUnitCard).join('') + '</div>';
}

function renderEmptyUnitCard(u){
  const skipped = u.skipped_until && u.skipped_until !== '';
  const tiers = u.tier_times || [];
  const now = new Date();
  const tierHtml = tiers.map(function(tt){
    const fire = new Date(); fire.setHours(tt.hour, tt.minute||0, 0, 0);
    const passed = fire < now;
    const isNext = (u.next && u.next.label === tt.label);
    const cls = isNext ? 'next' : (passed ? 'passed' : '');
    const lblMap = {T1:t().tier_t1, T2:t().tier_t2, T3:t().tier_t3, Weekend:t().tier_w};
    return '<div class="tier '+cls+'">'
      + '<span class="tlbl">'+(lblMap[tt.label]||tt.label)+'</span>'
      + '<span class="tprice">'+fmt(tt.price)+'</span>'
      + '<span class="tpct">−'+tt.pct+'%</span>'
      + '</div>';
  }).join('<span class="tline-arrow">›</span>');

  const skipMsg = skipped ? '<div style="font-size:11px;color:var(--yellow);margin-top:6px;font-weight:600">⏸ '+t().eu_skipped_until+' '+esc(u.skipped_until.replace('T',' ').slice(0,16))+'</div>' : '';

  const skipBtn = skipped
    ? '<button class="btn green xs" onclick="doUnskip('+u.lid+')">▶ '+t().eu_unskip+'</button>'
    : '<button class="btn ghost xs" onclick="doSkipUnit('+u.lid+')">⏸ '+t().eu_skip+'</button>';

  return '<div class="eu '+(skipped?'skipped':'')+'">'
    + '<div class="eu-top"><span class="eu-name">'+esc(u.name)+'</span>'+(u.paused_global?'<span class="pill warn">paused</span>':'')+'</div>'
    + '<div class="eu-now"><span class="lbl">'+t().eu_now+'</span><span class="v">'+fmt(u.price)+' SAR</span></div>'
    + '<div class="timeline">'+tierHtml+'</div>'
    + skipMsg
    + '<div class="eu-actions">'+skipBtn+'</div>'
    + '</div>';
}

/* ============================================================
   PRICING
   ============================================================ */
function renderPricing(){
  const d = D.pr; const tot = document.getElementById('prTotalBody');
  const body = document.getElementById('prListBody');
  if(!d || d.loading){ body.innerHTML='<div class="empty">…</div>'; tot.innerHTML=''; return }
  const units = d.units||[];
  document.getElementById('prListCount').textContent = units.length?'· '+units.length:'';
  tot.innerHTML = '<div style="display:flex;justify-content:space-between;align-items:center;gap:14px;flex-wrap:wrap">'
    + '<div><div class="muted">'+t().pr_uplift+'</div><div style="font-size:26px;font-weight:700;color:var(--gold);font-family:var(--font-mono);margin-top:3px">~'+fmt(d.total_uplift)+' SAR</div></div>'
    + '<div style="text-align:end"><div class="muted">'+units.length+' '+t().units_count+'</div></div>'
    + '</div>';
  if(!units.length){ body.innerHTML='<div class="empty">'+t().pr_empty+'</div>'; return }
  body.innerHTML = '<div class="inbox-list">' + units.map(function(u){
    const changes = (u.raise||0) + (u.drop||0);
    return '<div class="ibox" style="border-inline-start:3px solid var(--gold);cursor:pointer" onclick="openPriceDetail('+u.lid+')">'
      + '<div class="ibox-row" style="cursor:pointer">'
      + '<div class="ibox-icon rep">💰</div>'
      + '<div class="ibox-main"><div class="ibox-top"><span class="ibox-who">'+esc(u.name)+'</span></div><div class="ibox-preview">'+changes+' '+t().pr_change+' · '+t().pr_uplift+' ~'+fmt(u.uplift)+' SAR · '+t().pr_conf+' '+(u.confidence||0)+'%</div></div>'
      + '<div class="ibox-meta"><span class="ibox-conf high">~'+fmt(u.uplift)+'</span></div>'
      + '<span class="ibox-expand">←</span>'
      + '</div></div>';
  }).join('') + '</div>';
}

async function openPriceDetail(lid){
  openDrawer('—','');
  setDrawerBody('<div class="empty sk">—</div>');
  try{
    const d = await api('/api/pricing/detail?lid='+lid);
    renderPriceDetail(lid, d);
  }catch(_){
    setDrawerBody('<div class="empty">⚠ error</div>');
  }
}

function whyChipsPricing(r){
  const out = [];
  if(r.mi && Math.abs(r.mi-1)>=0.05) out.push('<span class="pill '+(r.mi>1?'ok':'danger')+'">'+t().pr_why_month+' ×'+r.mi+'</span>');
  if(r.di && Math.abs(r.di-1)>=0.05) out.push('<span class="pill '+(r.di>1?'ok':'danger')+'">'+t().pr_why_pay+' ×'+r.di+'</span>');
  if(r.wi && Math.abs(r.wi-1)>=0.05) out.push('<span class="pill '+(r.wi>1?'ok':'danger')+'">'+t().pr_why_week+' ×'+r.wi+'</span>');
  if(!out.length) out.push('<span class="pill muted">—</span>');
  return out.join(' ');
}

function renderPriceDetail(lid, d){
  setDrawerTitle(d.name||'—', t().pr_change+'s · '+(d.rows?d.rows.length:0));
  if(!d || !d.rows || !d.rows.length){
    setDrawerBody('<div class="empty">'+t().pr_empty+'</div>');
    setDrawerFoot('');
    return;
  }
  const rows = d.rows.map(function(r){
    const up = r.kind==='raise';
    const cur = r.current?fmt(r.current):'—';
    const newP = fmt(r.proposed);
    const arrow = up ? '↑' : '↓';
    const wd = L==='ar'?r.wd_ar:r.wd_en;
    return '<tr><td class="strong">'+r.date+'</td><td>'+wd+'</td><td class="num">'+cur+'</td>'
      + '<td><span class="pchange '+(up?'up':'dn')+'">'+arrow+' '+newP+'</span></td>'
      + '<td>'+whyChipsPricing(r)+'</td><td class="num">'+r.lead+t().pr_d_days+'</td></tr>';
  }).join('');
  setDrawerBody(
    '<div style="margin-bottom:12px;display:flex;justify-content:space-between;gap:10px;flex-wrap:wrap">'
    +'<div><span class="muted">'+t().pr_conf+': </span><b>'+(d.confidence||0)+'%</b></div>'
    +'<div><span class="muted">base: </span><b>~'+fmt(d.base||0)+' SAR</b></div>'
    +'</div>'
    + '<div style="overflow-x:auto"><table class="data">'
    + '<thead><tr><th>'+t().pr_d_date+'</th><th>'+t().pr_d_day+'</th><th class="num">'+t().pr_d_cur+'</th><th>'+t().pr_d_new+'</th><th>'+t().pr_d_why+'</th><th class="num">'+t().pr_d_lead+'</th></tr></thead>'
    + '<tbody>'+rows+'</tbody></table></div>'
  );
  setDrawerFoot('<button class="btn ghost sm" onclick="closeDrawer()">'+t().f_clear+'</button><button class="btn primary sm" onclick="doApplyFromDrawer('+lid+',this)">✓ '+t().pr_apply_all+'</button>');
}

async function doApplyFromDrawer(lid, btn){
  if(!confirm(t().pr_confirm)) return;
  btn.disabled = true; const o=btn.textContent; btn.textContent='…';
  const r = await post('/api/apply',{lid:lid});
  if(r.ok){ toast(t().applied+(r.dry_run?' (DRY-RUN)':'')+' · '+r.applied); closeDrawer(); loadPricing(); loadStrategies() }
  else toast(r.error || t().err);
  setTimeout(function(){btn.disabled=false; btn.textContent=o},900);
}

/* ============================================================
   STRATEGIES
   ============================================================ */
function renderStrategies(){
  const d = D.strat || {items:[]}; const items = d.items||[];
  const body = document.getElementById('stratListBody');
  if(!items.length){ body.innerHTML='<div class="empty"><span class="ic">⚡</span>'+t().st_empty+'</div>'; return }
  body.innerHTML = '<div class="inbox-list">' + items.map(function(s){
    const pct = s.total?Math.round(s.booked/s.total*100):0;
    const pill = s.active ? '<span class="pill ok">● '+t().st_running+'</span>' : '<span class="pill muted">'+t().st_done+'</span>';
    return '<div class="ibox" style="border-inline-start:3px solid '+(s.active?'var(--green)':'var(--mut)')+';cursor:pointer" onclick="openStrategyDetail('+s.lid+')">'
      + '<div class="ibox-row">'
      + '<div class="ibox-icon" style="background:'+(s.active?'var(--green-soft)':'var(--surface-2)')+';color:'+(s.active?'var(--green)':'var(--mut)')+'">⚡</div>'
      + '<div class="ibox-main"><div class="ibox-top"><span class="ibox-who">'+esc(s.name)+'</span>'+pill+'</div><div class="ibox-preview">'+s.booked+'/'+s.total+' '+t().st_booked+' · '+s.changes_total+' '+t().st_changes+(s.base?' · base ~'+fmt(s.base)+' SAR':'')+'</div></div>'
      + '<div class="ibox-meta"><span class="ibox-conf '+(pct>=50?'high':'mid')+'">'+pct+'%</span></div>'
      + '<span class="ibox-expand">←</span>'
      + '</div></div>';
  }).join('') + '</div>';
}

async function openStrategyDetail(lid){
  openDrawer('—','');
  setDrawerBody('<div class="empty sk">—</div>');
  try{
    const d = await api('/api/strategy?lid='+lid);
    renderStrategyDetail(lid, d);
  }catch(_){
    setDrawerBody('<div class="empty">⚠ error</div>');
  }
}

function renderStrategyDetail(lid, s){
  if(!s || !s.dates){
    setDrawerBody('<div class="empty">—</div>'); setDrawerFoot(''); return;
  }
  setDrawerTitle(s.name||'—', (s.active?'● '+t().st_running:t().st_done) + ' · ' + s.booked + '/' + s.total + ' '+t().st_booked);
  const pct = s.total?Math.round(s.booked/s.total*100):0;
  const overview = '<div class="strat-overview">'
    + '<div class="stat-mini"><div class="v g">'+s.booked+'/'+s.total+'</div><div class="l">'+t().st_booked+'</div></div>'
    + '<div class="stat-mini"><div class="v">'+s.open+'</div><div class="l">'+t().st_open+'</div></div>'
    + '<div class="stat-mini"><div class="v">'+s.changes_total+'</div><div class="l">'+t().st_changes+'</div></div>'
    + '<div class="stat-mini"><div class="v gold">'+fmt(s.base||0)+'</div><div class="l">base SAR</div></div>'
    + '</div>';

  const rows = s.dates.map(function(r){
    const wd = L==='ar'?r.wd_ar:r.wd_en;
    const changed = r.start !== r.cur;
    const arrow = r.cur > r.start ? '↑' : (r.cur < r.start ? '↓' : '·');
    const arrCls = r.cur > r.start ? 'up' : (r.cur < r.start ? 'dn' : '');
    const why = [];
    if(r.mi && Math.abs(r.mi-1)>=0.05) why.push('<span class="pill '+(r.mi>1?'ok':'danger')+'">'+t().pr_why_month+' ×'+r.mi+'</span>');
    if(r.di && Math.abs(r.di-1)>=0.05) why.push('<span class="pill '+(r.di>1?'ok':'danger')+'">'+t().pr_why_pay+' ×'+r.di+'</span>');
    if(r.wi && Math.abs(r.wi-1)>=0.05) why.push('<span class="pill '+(r.wi>1?'ok':'danger')+'">'+t().pr_why_week+' ×'+r.wi+'</span>');
    if(r.lead_factor && r.lead_factor < 1) why.push('<span class="pill warn">lead −'+Math.round((1-r.lead_factor)*100)+'%</span>');
    if(!why.length) why.push('<span class="pill muted">—</span>');
    const st = r.booked ? '<span class="pill ok">'+t().st_d_booked+'</span>' : '<span class="pill muted">'+t().st_d_open+'</span>';
    return '<tr><td class="strong">'+r.date+'</td><td>'+wd+'</td>'
      + '<td class="num"><span class="muted">'+fmt(r.start)+'</span></td>'
      + '<td><span class="pchange '+arrCls+'">'+arrow+' '+fmt(r.cur)+'</span></td>'
      + '<td>'+st+'</td><td class="num">'+(r.changes||0)+'</td>'
      + '<td>'+why.join(' ')+'</td></tr>';
  }).join('');

  setDrawerBody(overview
    + '<div style="overflow-x:auto"><table class="data">'
    + '<thead><tr><th>'+t().st_d_date+'</th><th>'+t().st_d_day+'</th><th class="num">'+t().st_d_start+'</th><th>'+t().st_d_cur+'</th><th>'+t().st_d_st+'</th><th class="num">'+t().st_d_chg+'</th><th>'+t().st_d_why+'</th></tr></thead>'
    + '<tbody>'+rows+'</tbody></table></div>'
  );
  setDrawerFoot((s.active ? '<button class="btn red sm" onclick="doStopStrategy('+lid+')">■ '+t().st_stop+'</button>':'') + '<button class="btn ghost sm" onclick="closeDrawer()">✕</button>');
}

/* ============================================================
   REVENUE FULL
   ============================================================ */
function renderRevenueFull(){
  const rev = D.rev||{};
  const monthly = (rev.monthly||[]).slice(-12);
  const mBody = document.getElementById('revMonthlyBody');
  if(monthly.length){
    const max = Math.max.apply(null, monthly.map(function(m){return m.rev}));
    const bars = monthly.map(function(m){
      const h = Math.max(5, (m.rev/max)*150);
      return '<div class="bar-col"><div class="bar-tip">'+fmt(m.rev)+' SAR</div><div class="bar" style="height:'+h+'px"></div><div class="bar-label">'+m.m.slice(5)+'</div></div>';
    }).join('');
    mBody.innerHTML = '<div class="bar-chart" style="height:170px">'+bars+'</div>';
  }else{ mBody.innerHTML='<div class="empty">'+t().rev_no+'</div>' }

  const sal = rev.salary || {};
  const sBody = document.getElementById('revSalaryBody');
  if(!Object.keys(sal).length){ sBody.innerHTML='<div class="empty">'+t().no_data+'</div>' }
  else{
    const doms = []; for(let i=1;i<=28;i++) doms.push(i);
    const vals = doms.map(function(d){return sal[d]||1});
    const max = Math.max.apply(null, vals);
    const bars = doms.map(function(d,i){
      const h = Math.max(4,(vals[i]/max)*110);
      const muted = vals[i]<0.95;
      return '<div class="bar-col"><div class="bar-tip">d'+d+': '+vals[i].toFixed(2)+'×</div><div class="bar '+(muted?'muted':'')+'" style="height:'+h+'px"></div><div class="bar-label">'+d+'</div></div>';
    }).join('');
    let extras = '';
    if(rev.weak) extras += '<div class="muted" style="margin-top:6px">🔻 '+t().weak_days+': '+rev.weak[0]+'–'+rev.weak[1]+'</div>';
    if(rev.strong) extras += '<div class="muted">🔺 '+t().strong_days+': '+rev.strong[0]+'–'+rev.strong[1]+'</div>';
    sBody.innerHTML = '<div class="bar-chart">'+bars+'</div>'+extras;
  }

  const units = ((rev.units)||[]).slice(0,40);
  const uBody = document.getElementById('revUnitsBody');
  if(!units.length){ uBody.innerHTML='<div class="empty">'+t().rev_no+'</div>'; return }
  uBody.innerHTML = '<div style="overflow-x:auto"><table class="data"><thead><tr><th>'+t().u_unit+'</th><th class="num">'+t().u_occ+'</th><th class="num">'+t().u_adr+'</th><th class="num">'+t().u_pace+'</th><th>'+t().u_reco+'</th></tr></thead><tbody>'
    + units.map(function(u){
      const tag = (u.reco==='raise'||u.reco==='raise_small') ? '<span class="pill ok">↑ '+esc(u.label||'')+'</span>'
        : (u.reco==='lower' ? '<span class="pill danger">↓ '+esc(u.label||'')+'</span>'
        : '<span class="pill muted">'+esc(u.label||'-')+'</span>');
      return '<tr><td class="strong">'+esc(u.name)+'</td><td class="num">'+u.occ+'%</td><td class="num">'+fmt(u.adr||0)+'</td><td class="num">'+u.pace+'%</td><td>'+tag+'</td></tr>';
    }).join('') + '</tbody></table></div>';
}

/* ============================================================
   LOG
   ============================================================ */
function renderLog(){
  const cat = (document.getElementById('logFilter')||{}).value || '';
  const items = (D.log||[]).filter(function(e){return !cat||e.cat===cat}).slice(0,200);
  const body = document.getElementById('logBody');
  if(!items.length){ body.innerHTML='<div class="empty">'+t().log_empty+'</div>'; return }
  const ic={guest:'💬',escalation:'🚨',pricing:'💰',report:'📊'};
  body.innerHTML = items.map(function(e){
    return '<div class="log-row"><div class="log-lic">'+(ic[e.cat]||'·')+'</div><div class="log-lts">'+esc(shortTime(e.ts))+'</div><div class="log-ltxt">'+esc(e.text)+'</div></div>';
  }).join('');
}

/* ============================================================
   LEARNINGS — view, edit, forget, distill-now
   ============================================================ */
let learnSelectedLid = null;
let learnEditingGeneral = false;
let learnEditingApt = false;

async function loadLearnings(){
  const list = document.getElementById('learnAptList');
  if(list) list.innerHTML = '<div class="empty sk">—</div>';
  // Fetch summaries + metrics + recent events in parallel so the page paints once.
  try{
    const r = await Promise.all([
      api('/api/learning/summary'),
      api('/api/metrics/daily?days=30').catch(function(){return {days:[]}}),
      api('/api/learning/today?days=1').catch(function(){return {apartments:[], total_events:0}}),
    ]);
    D.learn = r[0]; D.learnMetrics = r[1]; D.learnRecent = r[2];
  }catch(_){ D.learn = {} }
  // sync sub copy
  const sub = document.getElementById('t_learn_sub'); if(sub) sub.textContent = t().learn_sub;
  const eSel = document.getElementById('t_learn_empty_sel'); if(eSel) eSel.textContent = t().learn_empty_sel;
  renderLearnStats();
  renderLearnCharts();
  renderLearnRecent();
  renderLearnings();
}

async function loadLearnToday(){
  const sel = document.getElementById('learnWindow');
  const days = sel ? parseInt(sel.value || '1', 10) : 1;
  try{ D.learnRecent = await api('/api/learning/today?days='+days) }catch(_){ D.learnRecent = {apartments:[], total_events:0} }
  renderLearnRecent();
}

function _delta(a, b){
  if(b === 0 || b === null || b === undefined) return null;
  return Math.round(((a - b) / b) * 100);
}

function renderLearnStats(){
  const m = D.learnMetrics || {days:[]};
  const days = m.days || [];
  if(!days.length){
    document.getElementById('learnStats').innerHTML = '<div class="kpi" style="grid-column:1/-1"><div class="kpi-lbl">'+t().learn_recent_empty+'</div></div>';
    return;
  }
  const today = days[days.length-1] || {};
  const last7 = days.slice(-8, -1);   // 7 days before today
  const avg = function(arr, k){ if(!arr.length) return 0; return arr.reduce(function(s,x){return s+(x[k]||0)},0)/arr.length };
  const avg_replies = avg(last7, 'replies_total');
  const avg_auto    = avg(last7, 'auto_rate');
  const avg_conf    = avg(last7, 'avg_confidence');
  const avg_esc     = avg(last7, 'escalations_created');

  const cards = [
    {ic:'💬', cls:'b', val:today.replies_total||0, lbl:t().learn_stat_replies,
     delta:_delta(today.replies_total||0, avg_replies)},
    {ic:'⚡', cls:'g', val:(today.auto_rate||0)+'%', lbl:t().learn_stat_auto,
     delta:_delta(today.auto_rate||0, avg_auto)},
    {ic:'🎯', cls:'p', val:(today.avg_confidence||0)+'%', lbl:t().learn_stat_conf,
     delta:_delta(today.avg_confidence||0, avg_conf)},
    {ic:'🚨', cls:(today.escalations_created>0?'r':''), val:today.escalations_created||0, lbl:t().learn_stat_esc,
     delta:_delta(today.escalations_created||0, avg_esc), inverted:true},
  ];
  document.getElementById('learnStats').innerHTML = cards.map(function(c){
    let deltaHtml = '';
    if(c.delta !== null && c.delta !== 0){
      const positive = c.inverted ? c.delta < 0 : c.delta > 0;
      const cls = positive ? 'up' : 'dn';
      const sign = c.delta > 0 ? '+' : '';
      deltaHtml = '<span class="kpi-delta '+cls+'">'+sign+c.delta+'%</span>';
    }
    return '<div class="kpi"><div class="kpi-head"><div class="kpi-ic '+c.cls+'">'+c.ic+'</div>'+deltaHtml+'</div>'
      +'<div class="kpi-val">'+c.val+'</div><div class="kpi-lbl">'+c.lbl+' · <span style="opacity:.7">'+t().learn_stat_vs_avg+'</span></div></div>';
  }).join('');
}

function _lineChartSvg(values, color, suffix){
  if(!values || !values.length) return '<div class="empty">'+t().rev_no+'</div>';
  const w = 600, h = 140, pad = 28;
  const max = Math.max.apply(null, values.concat([1]));
  const min = 0;
  const range = max - min || 1;
  const dx = (w - pad*2) / Math.max(values.length - 1, 1);
  const pts = values.map(function(v, i){
    const x = pad + i * dx;
    const y = h - pad - ((v - min) / range) * (h - pad*2);
    return x.toFixed(1)+','+y.toFixed(1);
  }).join(' ');
  // area fill polygon
  const fillPts = pts + ' ' + (pad + (values.length-1)*dx).toFixed(1) + ',' + (h-pad) + ' ' + pad + ',' + (h-pad);
  const last = values[values.length-1];
  const lastX = pad + (values.length-1) * dx;
  const lastY = h - pad - ((last - min) / range) * (h - pad*2);
  // y-axis ticks
  const ticks = [0, Math.round(max/2), Math.round(max)];
  const tickHtml = ticks.map(function(tv){
    const y = h - pad - ((tv - min) / range) * (h - pad*2);
    return '<line x1="'+pad+'" y1="'+y+'" x2="'+(w-pad)+'" y2="'+y+'" stroke="var(--line)" stroke-dasharray="2,3"/>'
      + '<text x="'+(pad-4)+'" y="'+(y+3)+'" text-anchor="end" fill="var(--mut)" font-size="9" font-family="var(--font-mono)">'+tv+(suffix||'')+'</text>';
  }).join('');
  return '<svg viewBox="0 0 '+w+' '+h+'" preserveAspectRatio="xMidYMid meet" style="width:100%;height:160px;overflow:visible">'
    + tickHtml
    + '<polygon points="'+fillPts+'" fill="'+color+'" fill-opacity="0.12"/>'
    + '<polyline points="'+pts+'" fill="none" stroke="'+color+'" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>'
    + '<circle cx="'+lastX.toFixed(1)+'" cy="'+lastY.toFixed(1)+'" r="3.5" fill="'+color+'"/>'
    + '<text x="'+lastX+'" y="'+(lastY-8)+'" text-anchor="middle" fill="'+color+'" font-size="11" font-weight="700" font-family="var(--font-mono)">'+last+(suffix||'')+'</text>'
    + '</svg>';
}

function _barChartSvg(values, color){
  if(!values || !values.length) return '<div class="empty">'+t().rev_no+'</div>';
  const w = 600, h = 140, pad = 28;
  const max = Math.max.apply(null, values.concat([1]));
  const bw = (w - pad*2) / values.length;
  const bars = values.map(function(v, i){
    const bh = ((v) / max) * (h - pad*2);
    const x = pad + i * bw + 1;
    const y = h - pad - bh;
    return '<rect x="'+x.toFixed(1)+'" y="'+y.toFixed(1)+'" width="'+(bw-2).toFixed(1)+'" height="'+bh.toFixed(1)+'" fill="'+color+'" rx="2"/>';
  }).join('');
  return '<svg viewBox="0 0 '+w+' '+h+'" preserveAspectRatio="xMidYMid meet" style="width:100%;height:160px">'
    + bars
    + '<line x1="'+pad+'" y1="'+(h-pad)+'" x2="'+(w-pad)+'" y2="'+(h-pad)+'" stroke="var(--line)"/>'
    + '<text x="'+(pad-4)+'" y="'+(pad+3)+'" text-anchor="end" fill="var(--mut)" font-size="9" font-family="var(--font-mono)">'+max+'</text>'
    + '</svg>';
}

function renderLearnCharts(){
  const m = D.learnMetrics || {days:[]};
  const days = m.days || [];
  const confs = days.map(function(d){return d.avg_confidence || 0});
  const autos = days.map(function(d){return d.auto_rate || 0});
  const vols  = days.map(function(d){return d.replies_total || 0});
  const escs  = days.map(function(d){return d.escalation_rate || 0});
  document.getElementById('learnChartConf').innerHTML = _lineChartSvg(confs, '#6D58C2', '%');
  document.getElementById('learnChartAuto').innerHTML = _lineChartSvg(autos, '#0E9E5F', '%');
  document.getElementById('learnChartVol').innerHTML  = _barChartSvg(vols, '#A37728');
  document.getElementById('learnChartEsc').innerHTML  = _lineChartSvg(escs, '#C44343', '%');
}

function renderLearnRecent(){
  const d = D.learnRecent || {apartments:[], total_events:0};
  const body = document.getElementById('learnRecentBody');
  if(!d.total_events){ body.innerHTML = '<div class="empty">'+t().learn_recent_empty+'</div>'; return; }
  body.innerHTML = d.apartments.map(function(a){
    const events = a.events.slice(0, 6).map(function(e){
      const tag = e.via === 'auto' ? t().learn_event_auto
                : (e.was_edited ? t().learn_event_edited : t().learn_event_sent);
      const tagCls = e.via === 'auto' ? 'ok' : (e.was_edited ? 'warn' : 'info');
      return '<div class="log-row" style="grid-template-columns:auto 1fr;align-items:flex-start">'
        + '<span class="log-lts">'+esc(shortTime(e.ts))+' <span class="pill '+tagCls+'">'+tag+'</span></span>'
        + '<span class="log-ltxt"><b>س:</b> '+esc(e.guest_question||'')+'<br><b>ر:</b> '+esc(e.final_reply||'')+'</span>'
        + '</div>';
    }).join('');
    const more = a.events.length > 6 ? '<div class="muted" style="text-align:center;padding:5px">+ '+(a.events.length-6)+'</div>' : '';
    return '<div class="card" style="margin-bottom:10px;background:var(--surface-2)"><div style="font-weight:700;margin-bottom:8px;display:flex;justify-content:space-between"><span>'+esc(a.unit)+'</span><span class="pill muted">'+a.count+'</span></div>'+events+more+'</div>';
  }).join('');
}

function _fmtDistillTime(ts){
  if(!ts) return '—';
  try{ return new Date(ts*1000).toLocaleString(L==='ar'?'ar-SA':'en-US',{month:'short',day:'numeric',hour:'2-digit',minute:'2-digit'}) }
  catch(_){ return '—' }
}

function renderLearnings(){
  const d = D.learn || {};
  // ---- general ----
  const gen = d.general || {};
  const genBody = document.getElementById('learnGeneralBody');
  const genActs = document.getElementById('genActions');
  if(learnEditingGeneral){
    genBody.innerHTML = '<textarea id="learnGenTa" style="min-height:200px;width:100%;line-height:1.7">'+esc(gen.summary||'')+'</textarea>';
    genActs.innerHTML = '<button class="btn ghost sm" onclick="learnEditingGeneral=false;renderLearnings()">✕ '+t().learn_cancel+'</button>'
      + '<button class="btn primary sm" onclick="saveGeneralSummary()">💾 '+t().learn_save+'</button>';
  }else if(gen.summary){
    genBody.innerHTML = '<div style="white-space:pre-wrap;line-height:1.75;font-size:13px">'+esc(gen.summary)+'</div>'
      + '<div class="muted" style="margin-top:10px;font-size:11px">⏱ '+t().learn_last+': '+_fmtDistillTime(gen.last_distilled)+' · '+(gen.examples_count||0)+' '+t().learn_examples+'</div>';
    genActs.innerHTML = '<button class="btn ghost sm" onclick="learnEditingGeneral=true;renderLearnings()">✎ '+t().learn_edit+'</button>'
      + '<button class="btn red sm" onclick="forgetLearning(&#39;general&#39;,null)">🗑 '+t().learn_forget+'</button>';
  }else{
    genBody.innerHTML = '<div class="empty">'+t().learn_empty+'</div>';
    genActs.innerHTML = '<button class="btn ghost sm" onclick="learnEditingGeneral=true;renderLearnings()">✎ '+t().learn_edit+'</button>';
  }

  // ---- apartments list ----
  const apts = d.apartments || [];
  const q = (document.getElementById('learnSearch')||{}).value || '';
  const filtered = q ? apts.filter(function(a){return (a.unit||'').toLowerCase().indexOf(q.toLowerCase())>=0}) : apts;
  const listEl = document.getElementById('learnAptList');
  if(!filtered.length){
    listEl.innerHTML = '<div class="empty">'+t().learn_no_apt+'</div>';
  }else{
    listEl.innerHTML = filtered.map(function(a){
      const on = learnSelectedLid === a.lid ? ' style="background:var(--gold-tint);border-radius:6px"' : '';
      return '<div class="list-row" onclick="selectLearnApt('+a.lid+')" style="cursor:pointer;padding:9px 10px"'+on+'>'
        + '<span class="l-name">'+esc(a.unit||('unit-'+a.lid))+'</span>'
        + '<span class="l-tag muted">'+(a.examples_count||0)+'</span>'
        + '</div>';
    }).join('');
  }

  // ---- selected apartment ----
  renderLearnAptDetail();
}

function selectLearnApt(lid){
  learnSelectedLid = lid;
  learnEditingApt = false;
  renderLearnings();
}

function renderLearnAptDetail(){
  const el = document.getElementById('learnAptDetail');
  if(!el) return;
  const apts = (D.learn||{}).apartments || [];
  const apt = apts.find(function(a){return a.lid === learnSelectedLid});
  if(!apt){
    el.innerHTML = '<div class="empty">'+t().learn_empty_sel+'</div>';
    return;
  }
  if(learnEditingApt){
    el.innerHTML = '<div class="card-head"><span class="card-title">'+esc(apt.unit||'')+'</span>'
      + '<div class="card-actions">'
        + '<button class="btn ghost sm" onclick="learnEditingApt=false;renderLearnings()">✕ '+t().learn_cancel+'</button>'
        + '<button class="btn primary sm" onclick="saveAptSummary('+apt.lid+')">💾 '+t().learn_save+'</button>'
      + '</div></div>'
      + '<textarea id="learnAptTa" style="min-height:340px;width:100%;line-height:1.7">'+esc(apt.summary||'')+'</textarea>';
  }else{
    el.innerHTML = '<div class="card-head"><span class="card-title">'+esc(apt.unit||'')+'</span>'
      + '<div class="card-actions">'
        + '<button class="btn ghost sm" onclick="learnEditingApt=true;renderLearnings()">✎ '+t().learn_edit+'</button>'
        + '<button class="btn red sm" onclick="forgetLearning(&#39;apartment&#39;,'+apt.lid+')">🗑 '+t().learn_forget+'</button>'
      + '</div></div>'
      + '<div style="white-space:pre-wrap;line-height:1.75;font-size:13px;max-height:540px;overflow-y:auto">'+esc(apt.summary||'')+'</div>'
      + '<div class="muted" style="margin-top:10px;font-size:11px">⏱ '+t().learn_last+': '+_fmtDistillTime(apt.last_distilled)+' · '+(apt.examples_count||0)+' '+t().learn_examples+'</div>';
  }
}

async function saveGeneralSummary(){
  const ta = document.getElementById('learnGenTa');
  if(!ta) return;
  const r = await post('/api/learning/edit', {scope:'general', summary:ta.value});
  if(r.ok){ toast(t().learn_saved); learnEditingGeneral=false; loadLearnings(); }
  else toast(r.error || t().err);
}
async function saveAptSummary(lid){
  const ta = document.getElementById('learnAptTa');
  if(!ta) return;
  const r = await post('/api/learning/edit', {scope:'apartment', lid:lid, summary:ta.value});
  if(r.ok){ toast(t().learn_saved); learnEditingApt=false; loadLearnings(); }
  else toast(r.error || t().err);
}
async function forgetLearning(scope, lid){
  if(!confirm(t().learn_confirm_forget)) return;
  const body = {scope:scope}; if(lid) body.lid = lid;
  const r = await post('/api/learning/forget', body);
  if(r.ok){
    if(scope==='apartment' && lid===learnSelectedLid) learnSelectedLid = null;
    loadLearnings();
  } else toast(r.error || t().err);
}
async function distillLearningsNow(){
  const btn = document.getElementById('learnDistillBtn');
  if(btn){ btn.disabled = true; btn.textContent = '⏳'; }
  toast(t().learn_distilling);
  try{ await post('/api/learning/distill', {}); }catch(_){}
  await loadLearnings();
  if(btn){ btn.disabled = false; btn.innerHTML = '↻ '+t().learn_distill; }
}

let _bootstrapPoll = null;
async function bootstrapLearnings(){
  if(!confirm(t().learn_bootstrap_confirm)) return;
  const btn = document.getElementById('learnBootstrapBtn');
  if(btn){ btn.disabled = true; btn.innerHTML = '⏳'; }
  const r = await post('/api/learning/bootstrap', {limit_conversations:300});
  if(!r.ok){ toast(r.error || t().err); if(btn){ btn.disabled=false; btn.innerHTML='📥 '+t().learn_bootstrap; } return; }
  toast(t().learn_bootstrap_started);
  showBootstrapStatus({running:true});
  // poll status every 8s
  if(_bootstrapPoll) clearInterval(_bootstrapPoll);
  _bootstrapPoll = setInterval(pollBootstrap, 8000);
  pollBootstrap();
}
async function pollBootstrap(){
  try{
    const s = await api('/api/learning/bootstrap/status');
    showBootstrapStatus(s);
    if(!s.running){
      clearInterval(_bootstrapPoll); _bootstrapPoll = null;
      const btn = document.getElementById('learnBootstrapBtn');
      if(btn){ btn.disabled = false; btn.innerHTML = '📥 '+t().learn_bootstrap; }
      await loadLearnings();
    }
  }catch(_){}
}
function showBootstrapStatus(s){
  const el = document.getElementById('bootstrapStatus');
  if(!el) return;
  if(!s || (!s.running && !s.result && !s.error)){ el.style.display='none'; return; }
  el.style.display='block';
  if(s.running){
    el.innerHTML = '<div class="card" style="background:var(--gold-tint);border-color:var(--gold)"><span style="color:var(--gold);font-weight:600">'+t().learn_bootstrap_running+'</span></div>';
  }else if(s.error){
    el.innerHTML = '<div class="card" style="background:var(--red-soft);border-color:var(--red)"><span style="color:var(--red)">⚠ '+esc(s.error)+'</span></div>';
  }else if(s.result){
    const r = s.result;
    el.innerHTML = '<div class="card" style="background:var(--green-soft);border-color:var(--green)">'
      + '<div style="color:var(--green);font-weight:700;margin-bottom:8px">'+t().learn_bootstrap_done+'</div>'
      + '<div style="display:flex;gap:18px;flex-wrap:wrap;font-size:13px">'
      + '<span><b class="mono">'+r.conversations_scanned+'</b> '+t().learn_bootstrap_scanned+'</span>'
      + '<span><b class="mono">'+r.pairs_extracted+'</b> '+t().learn_bootstrap_pairs+'</span>'
      + '<span><b class="mono">'+r.apartments_distilled+'</b> '+t().learn_bootstrap_apts+'</span>'
      + '</div></div>';
  }
}

/* ============================================================
   DRAWER
   ============================================================ */
function openDrawer(title, sub){
  document.getElementById('drwTitle').textContent = title;
  document.getElementById('drwSub').textContent = sub||'';
  document.getElementById('drawer').classList.add('open');
  document.getElementById('drawerBg').classList.add('show');
  drawerOpen = true;
}
function setDrawerTitle(title, sub){
  document.getElementById('drwTitle').textContent = title;
  document.getElementById('drwSub').textContent = sub||'';
}
function setDrawerBody(html){ document.getElementById('drwBody').innerHTML = html }
function setDrawerFoot(html){
  const f = document.getElementById('drwFoot');
  if(!html){ f.style.display='none'; f.innerHTML='' } else { f.style.display='flex'; f.innerHTML = html }
}
function closeDrawer(){
  document.getElementById('drawer').classList.remove('open');
  document.getElementById('drawerBg').classList.remove('show');
  drawerOpen = false;
}

/* ============================================================
   ACTIONS
   ============================================================ */
async function doSend(id){
  id = String(id);
  const ta = document.getElementById('ta_'+id);
  const text = ta?ta.value:'';
  const r = await post('/api/send',{id:id, text:text});
  if(r.ok){ toast(t().sent); openInboxId=null; loadAll() } else toast(r.error||t().err);
}
async function doReject(id){
  id = String(id);
  await post('/api/reject',{id:id});
  toast(t().rejected); openInboxId=null; loadAll();
}
async function doClaim(id){
  id = String(id);
  const inEl = document.getElementById('cn_'+id);
  const n = inEl ? inEl.value : '';
  const r = await post('/api/claim',{id:id, name:n});
  if(r.ok){ toast(t().claimed_t); openInboxId=null; loadAll() } else toast(r.error||t().err);
}
async function doTeach(id){
  id = String(id);
  const topic = (document.getElementById('teachT_'+id)||{}).value || '';
  const fact = (document.getElementById('teachF_'+id)||{}).value || '';
  if(!fact.trim()){ toast(t().err); return }
  const r = await post('/api/teach',{topic:topic, fact:fact});
  if(r.ok){ toast(t().taught); toggleTeach(id) } else toast(r.error||t().err);
}
async function doStopStrategy(lid){
  await post('/api/strategy/stop',{lid:lid});
  toast('■'); loadStrategies(); closeDrawer();
}
async function doPause(hours){
  const r = await post('/api/discount/pause',{hours:hours});
  if(r.ok){ D.disc = r; renderDiscountBanner(); if(view==='today') renderTodayEmpty() }
}
async function doResume(){
  const r = await post('/api/discount/resume',{});
  if(r.ok){ D.disc = r; renderDiscountBanner(); if(view==='today') renderTodayEmpty() }
}
async function doSkipUnit(lid){
  const r = await post('/api/discount/skip-unit',{lid:lid, hours:24});
  if(r.ok){ toast(t().skipped); loadTodayEmpty() }
}
async function doUnskip(lid){
  const r = await post('/api/discount/unskip-unit',{lid:lid});
  if(r.ok){ toast(t().resumed); loadTodayEmpty() }
}

/* ============================================================
   BOOT
   ============================================================ */
applyTheme();
if(tok()) init();
</script>
</body>
</html>"""

CLEANING_HTML = """<!doctype html>
<html lang="ar" dir="rtl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<title>عوجا · جدول التنظيف العميق</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@400;500;600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0;-webkit-font-smoothing:antialiased}
:root{
  --bg:#FAFAF7;--surface:#FFFFFF;--surface-2:#F5F2EC;--line:#E8E2D5;
  --text:#1A1815;--text-2:#544D43;--mut:#A09989;
  --gold:#A37728;--gold-2:#8B6320;--gold-soft:#F4EBD5;
  --green:#0E9E5F;--green-soft:#DCF3E6;
  --red:#C44343;--red-soft:#FAE3E3;
  --blue:#2F6FD0;--blue-soft:#E0EBFA;
  --r:14px;--sh:0 2px 8px rgba(26,24,21,.06);
}
html,body{font-family:'IBM Plex Sans Arabic','Inter',-apple-system,system-ui,sans-serif;background:var(--bg);color:var(--text);min-height:100vh;font-size:14px;line-height:1.6}
body{padding:0 0 60px}
header{background:linear-gradient(135deg,var(--gold),var(--gold-2));color:#fff;padding:36px 22px 30px;text-align:center;box-shadow:var(--sh)}
header h1{font-size:24px;font-weight:700;letter-spacing:.3px;margin-bottom:6px}
header h2{font-size:13px;font-weight:500;opacity:.85;letter-spacing:.4px}
header .lang-en{margin-top:14px;padding:8px 18px;background:rgba(255,255,255,.18);border-radius:10px;display:inline-block;font-size:11.5px;font-weight:500;line-height:1.65;backdrop-filter:blur(4px)}
.wrap{max-width:880px;margin:0 auto;padding:22px 18px}
.stats{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:22px}
.stat{background:var(--surface);border:1px solid var(--line);border-radius:var(--r);padding:14px;text-align:center;box-shadow:var(--sh)}
.stat .v{font-size:24px;font-weight:700;color:var(--gold);letter-spacing:-.5px}
.stat .l{font-size:11.5px;color:var(--mut);margin-top:4px}
.day-card{background:var(--surface);border:1px solid var(--line);border-radius:var(--r);margin-bottom:14px;overflow:hidden;box-shadow:var(--sh)}
.day-head{padding:12px 16px;background:var(--gold-soft);border-bottom:1px solid var(--line);display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}
.day-head .date{font-weight:700;font-size:15px;color:var(--text)}
.day-head .wd{color:var(--mut);font-size:12px;font-weight:500}
.day-head .days-away{background:var(--surface);padding:3px 10px;border-radius:8px;font-size:11px;color:var(--text-2);font-weight:600}
.day-head .today-pill{background:var(--green);color:#fff;padding:4px 12px;border-radius:8px;font-size:11.5px;font-weight:700}
.day-body{padding:14px 16px}
.unit-name{font-size:17px;font-weight:700;color:var(--text);margin-bottom:8px}
.unit-meta{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:12px}
.tag{display:inline-block;background:var(--surface-2);padding:4px 11px;border-radius:7px;font-size:11.5px;color:var(--text-2);font-weight:500;border:1px solid var(--line)}
.tag.green{background:var(--green-soft);color:var(--green);border-color:rgba(14,158,95,.18)}
.tag.blue{background:var(--blue-soft);color:var(--blue);border-color:rgba(47,111,208,.18)}
.tag.gold{background:var(--gold-soft);color:var(--gold);border-color:rgba(163,119,40,.18);font-weight:700}
.guide-btn{display:inline-flex;align-items:center;gap:8px;background:var(--gold);color:#fff;padding:10px 16px;border-radius:10px;font-weight:600;font-size:13px;text-decoration:none;margin-top:6px;transition:.15s}
.guide-btn:hover{background:var(--gold-2);transform:translateY(-1px)}
.empty{text-align:center;color:var(--mut);padding:40px 18px;background:var(--surface);border:1px dashed var(--line);border-radius:var(--r)}
.empty .ic{font-size:42px;display:block;margin-bottom:10px;opacity:.4}
.section-title{font-size:13px;font-weight:700;color:var(--text-2);text-transform:uppercase;letter-spacing:.5px;margin:18px 0 10px;padding:0 4px}
.section-title:first-child{margin-top:0}
.err{background:var(--red-soft);color:var(--red);padding:14px;border-radius:var(--r);text-align:center;font-weight:600}
.last-cleaned{font-size:11px;color:var(--mut);margin-top:8px;display:flex;align-items:center;gap:6px}
.dot{width:8px;height:8px;border-radius:50%;background:var(--green);box-shadow:0 0 0 3px rgba(14,158,95,.18);display:inline-block}
.status-pill{display:inline-flex;align-items:center;gap:5px;padding:3px 10px;border-radius:7px;font-size:11px;font-weight:600;background:var(--blue-soft);color:var(--blue)}
.status-pill.blocked{background:var(--green-soft);color:var(--green)}
@media (max-width:600px){
  header h1{font-size:20px}
  .stats{grid-template-columns:repeat(3,1fr);gap:7px}
  .stat .v{font-size:20px}
  .unit-name{font-size:15px}
}
</style></head>
<body>
<header>
  <h1>👋 أهلاً بشركاء النجاح</h1>
  <h2>جدول التنظيف العميق · عوجا</h2>
  <div class="lang-en">Welcome our success partners<br>Deep cleaning schedule · Ouja Residence</div>
</header>
<div class="wrap" id="root">
  <div class="empty"><span class="ic">⏳</span>جاري التحميل...<br><span style="font-size:11.5px">Loading...</span></div>
</div>
<script>
function esc(s){return (s==null?'':String(s)).replace(/[<>&]/g,function(c){return ({'<':'&lt;','>':'&gt;','&':'&amp;'})[c]})}
function wdAr(d){return ['الإثنين','الثلاثاء','الأربعاء','الخميس','الجمعة','السبت','الأحد'][d]}
function fmtDate(s){try{const dt=new Date(s);return dt.toLocaleDateString('ar-SA',{day:'numeric',month:'long',year:'numeric'})}catch(_){return s}}

(async function(){
  const root = document.getElementById('root');
  const token = new URLSearchParams(location.search).get('token') || '';
  if(!token){ root.innerHTML = '<div class="err">⚠ رابط غير مكتمل · Missing access token</div>'; return; }
  let data;
  try{
    const r = await fetch('/api/cleaning/public?token=' + encodeURIComponent(token));
    if(r.status === 401){ root.innerHTML = '<div class="err">⚠ رابط غير صحيح · Invalid access</div>'; return; }
    data = await r.json();
  }catch(e){
    root.innerHTML = '<div class="err">⚠ تعذّر التحميل · Could not load schedule</div>'; return;
  }
  const items = data.items || [];
  const today = data.today;
  // group by date
  const byDate = {};
  for(const it of items){
    if(!byDate[it.date]) byDate[it.date] = [];
    byDate[it.date].push(it);
  }
  const todayItems = items.filter(function(x){return x.date === today});
  const upcoming = items.filter(function(x){return x.date > today});
  const total = items.length;
  const this7 = items.filter(function(x){
    const d = new Date(x.date), tdy = new Date(today);
    return (d - tdy) / 86400000 <= 7 && d >= tdy;
  }).length;

  let html = '<div class="stats">'
    + '<div class="stat"><div class="v">'+todayItems.length+'</div><div class="l">اليوم · today</div></div>'
    + '<div class="stat"><div class="v">'+this7+'</div><div class="l">٧ أيام · 7 days</div></div>'
    + '<div class="stat"><div class="v">'+total+'</div><div class="l">المجموع · total</div></div>'
    + '</div>';

  function renderDay(d_iso, list){
    const dt = new Date(d_iso);
    const wd = wdAr(dt.getDay() === 0 ? 6 : dt.getDay() - 1);
    const isToday = d_iso === today;
    const daysAway = Math.round((new Date(d_iso) - new Date(today)) / 86400000);
    let body = '';
    for(const u of list){
      const lastDone = u.last_done ? '<div class="last-cleaned"><span class="dot"></span>آخر تنظيف عميق: ' + fmtDate(u.last_done) + ' · last deep clean</div>' : '';
      const guide = u.guide_url ? '<a class="guide-btn" href="' + esc(u.guide_url) + '" target="_blank">📖 دليل الوصول · arrival guide</a>' : '';
      const stPill = u.status === 'blocked' ? '<span class="status-pill blocked">🔒 مؤكد · confirmed</span>' : '<span class="status-pill">📅 مجدول · scheduled</span>';
      body += '<div class="day-body">'
        + '<div class="unit-name">'+esc(u.name)+'</div>'
        + '<div class="unit-meta">'
          + stPill
          + (u.beds ? '<span class="tag gold">'+u.beds+' غرف · bedrooms</span>' : '')
          + (u.baths ? '<span class="tag">'+u.baths+' حمام · bath</span>' : '')
          + (u.area ? '<span class="tag blue">📍 '+esc(u.area)+'</span>' : '')
        + '</div>'
        + lastDone
        + guide
        + '</div>';
    }
    const todayTag = isToday ? '<span class="today-pill">اليوم · today</span>'
                             : '<span class="days-away">بعد '+daysAway+' يوم · in '+daysAway+'d</span>';
    return '<div class="day-card">'
      + '<div class="day-head"><div><div class="date">'+fmtDate(d_iso)+'</div><div class="wd">'+wd+'</div></div>'+todayTag+'</div>'
      + body
      + '</div>';
  }

  if(todayItems.length){
    html += '<div class="section-title">🌟 تنظيف اليوم · today\\'s clean</div>' + renderDay(today, todayItems);
  }
  if(upcoming.length){
    html += '<div class="section-title">📅 القادم · upcoming</div>';
    const grouped = {};
    for(const u of upcoming){ if(!grouped[u.date]) grouped[u.date] = []; grouped[u.date].push(u); }
    for(const d of Object.keys(grouped).sort()){ html += renderDay(d, grouped[d]); }
  }
  if(!items.length){
    html += '<div class="empty"><span class="ic">✓</span>ما فيه تنظيفات مجدولة حالياً<br><span style="font-size:11.5px">No scheduled cleanings yet</span></div>';
  }
  root.innerHTML = html;
})();
</script>
</body>
</html>"""

CLEAN_FEEDBACK_HTML = """<!doctype html>
<html lang="ar" dir="rtl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<title>تقييم نظافة الشقة · عوجا</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@400;500;600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{--bg:#FAFAF7;--surface:#FFFFFF;--gold:#A37728;--gold-2:#8B6320;--text:#1A1815;--mut:#A09989;--green:#0E9E5F;--line:#E8E2D5}
html,body{font-family:'IBM Plex Sans Arabic','Inter',sans-serif;background:var(--bg);min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:24px;color:var(--text)}
.card{max-width:440px;width:100%;background:var(--surface);border-radius:18px;box-shadow:0 12px 40px rgba(26,24,21,.08);padding:30px 24px;text-align:center}
.brand{font-size:26px;font-weight:700;color:var(--gold);margin-bottom:4px}
h1{font-size:18px;font-weight:600;margin:14px 0 6px;color:var(--text);line-height:1.5}
.unit{font-size:14px;color:var(--mut);margin-bottom:24px}
.stars{display:flex;justify-content:center;gap:8px;margin:18px 0}
.star{font-size:44px;cursor:pointer;color:#E0DCD2;transition:.12s;-webkit-tap-highlight-color:transparent}
.star:hover,.star.on{color:var(--gold);transform:scale(1.08)}
textarea{width:100%;border:1px solid var(--line);border-radius:10px;padding:11px;font-family:inherit;font-size:14px;min-height:80px;margin:14px 0;resize:vertical}
button{background:linear-gradient(135deg,var(--gold),var(--gold-2));color:#fff;border:none;padding:13px 26px;border-radius:11px;font-size:15px;font-weight:600;cursor:pointer;width:100%;transition:.15s}
button:hover{filter:brightness(1.06)}
button:disabled{opacity:.5;cursor:default}
.thanks{padding:36px 20px;text-align:center}
.thanks .ic{font-size:56px;display:block;margin-bottom:10px}
.thanks h2{font-size:20px;font-weight:700;color:var(--green);margin-bottom:6px}
.thanks p{color:var(--mut);font-size:14px}
.err{color:#C44343;padding:18px;text-align:center}
</style></head>
<body>
<div class="card" id="root">
  <div style="text-align:center;padding:30px">⏳ يحمّل…</div>
</div>
<script>
let chosen = 0;
function setStar(n){
  chosen = n;
  document.querySelectorAll('.star').forEach(function(el, i){
    el.classList.toggle('on', i < n);
  });
  document.getElementById('sendBtn').disabled = (n < 1);
}
async function submit(){
  const btn = document.getElementById('sendBtn');
  btn.disabled = true; btn.textContent = '⏳';
  const id = new URLSearchParams(location.search).get('id');
  const comment = (document.getElementById('cmt')||{}).value || '';
  const r = await fetch('/api/clean-feedback?id=' + encodeURIComponent(id), {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({id:id, score:chosen, comment:comment})
  });
  const j = await r.json().catch(function(){return {}});
  const root = document.getElementById('root');
  if(j.ok){
    root.innerHTML = '<div class="thanks"><span class="ic">🤍</span><h2>شكراً لك</h2><p>تقييمك وصل لفريقنا · Your rating has been received</p></div>';
  }else{
    root.innerHTML = '<div class="err">⚠ تعذّر الإرسال · ' + (j.error || 'try again') + '</div>';
  }
}
(async function(){
  const id = new URLSearchParams(location.search).get('id');
  if(!id){ document.getElementById('root').innerHTML = '<div class="err">⚠ رابط غير صحيح</div>'; return; }
  let data;
  try{
    const r = await fetch('/api/clean-feedback?id=' + encodeURIComponent(id));
    data = await r.json();
  }catch(e){ document.getElementById('root').innerHTML = '<div class="err">⚠ خطأ</div>'; return; }
  if(data.error){ document.getElementById('root').innerHTML = '<div class="err">⚠ ' + data.error + '</div>'; return; }
  if(data.already_done){
    document.getElementById('root').innerHTML = '<div class="thanks"><span class="ic">✓</span><h2>تم التقييم مسبقاً</h2><p>شكراً لك · Already received</p></div>';
    return;
  }
  document.getElementById('root').innerHTML =
    '<div class="brand">عوجا</div>'
    + '<h1>كيف لقيت نظافة الشقة لما دخلت؟<br><span style="color:var(--mut);font-size:14px;font-weight:400">How clean was the apartment on check-in?</span></h1>'
    + '<div class="unit">' + (data.unit || '') + '</div>'
    + '<div class="stars">'
      + [1,2,3,4,5].map(function(n){ return '<span class="star" onclick="setStar(' + n + ')">★</span>' }).join('')
    + '</div>'
    + '<textarea id="cmt" placeholder="ملاحظة (اختياري) · optional comment"></textarea>'
    + '<button id="sendBtn" disabled onclick="submit()">إرسال · Send</button>';
})();
</script>
</body>
</html>"""

def _dash_auth(request):
    return bool(DASHBOARD_TOKEN) and (
        request.query.get("token") or request.headers.get("X-Token", "")) == DASHBOARD_TOKEN

def _json(data, status=200):
    return web.json_response(data, status=status,
                             dumps=lambda o: json.dumps(o, ensure_ascii=False))

def _cache_get(key):
    hit = _dash_cache.get(key)
    return hit[0] if hit else None

def _live_counts():
    return {"pending_cards": len(_pending_replies),
            "open_escalations": sum(1 for e in _escalations.values() if not e.get("claimed_by"))}

def fetch_inhouse(day):
    """Reservations overlapping `day` (arrived on/before, departing on/after) — a precise,
    bounded query (~one row per occupied/turning unit) so occupancy is always accurate."""
    try:
        data = api_get("/reservations", params={
            "arrivalEndDate": day.isoformat(),          # arrivalDate <= today
            "departureStartDate": day.isoformat(),      # departureDate >= today
            "limit": 200})
        return data.get("result", []) or []
    except Exception as e:
        print("in-house fetch error:", e)
        return []

def _compute_today():
    """The morning cockpit: arrivals, departures, who's empty tonight, same-day turnovers,
    and the revenue on the table. Occupancy comes from a targeted in-house query (accurate)."""
    today = datetime.now(TZ).date()
    listings = get_listings_map()
    rows = fetch_inhouse(today)
    if not rows:                                            # fallback to history if the query fails
        rows = [r for r in get_reservations_cached() if _res_realized(r)]
    print(f"today: in-house query returned {len(rows)} reservations")

    def nm(r):
        return listings.get(r.get("listingMapId")) or r.get("listingName") or f"unit-{r.get('listingMapId')}"

    arrivals, departures = [], []
    occ_lids, dep_lids, arr_lids = set(), set(), set()
    for r in rows:
        if not _res_realized(r):
            continue
        a, d = _parse_date(r.get("arrivalDate")), _parse_date(r.get("departureDate"))
        lid = r.get("listingMapId")
        if a == today:
            arrivals.append({"guest": r.get("guestName") or "Guest", "unit": nm(r),
                             "nights": _res_nights(r), "checkout": d.isoformat() if d else ""})
            arr_lids.add(lid)
        if d == today:
            departures.append({"guest": r.get("guestName") or "Guest", "unit": nm(r)})
            dep_lids.add(lid)
        if a and d and a <= today < d:
            occ_lids.add(lid)
    units = {u["id"]: u["name"] for u in _catalog_units if u.get("id")}
    active = len(units) or len(occ_lids)
    empty = [{"unit": units.get(lid, str(lid)), "lid": lid}
             for lid in units if lid not in occ_lids]
    tight = [units.get(lid, str(lid)) for lid in (dep_lids & arr_lids)]      # same-day turnover
    prices = [u.get("price") for u in _catalog_units if u.get("price")]
    avg = round(sum(prices) / len(prices)) if prices else 0
    return {"date": today.isoformat(), "active": active, "occupied": len(occ_lids),
            "arrivals": sorted(arrivals, key=lambda x: x["unit"]),
            "departures": sorted(departures, key=lambda x: x["unit"]),
            "empty": sorted(empty, key=lambda x: x["unit"])[:90], "empty_n": len(empty),
            "tight": tight, "avg": avg, "risk": len(empty) * avg,
            "pending_cards": len(_pending_replies),
            "open_escalations": sum(1 for e in _escalations.values() if not e.get("claimed_by"))}

def _compute_overview():
    today = datetime.now(TZ).date()
    reservations = get_reservations_cached()
    listings = get_listings_map()
    factors = compute_demand_factors(reservations)
    lw = compute_last_week(reservations, listings, factors)
    nights, _ = _explode_nights(reservations)
    d30 = today - timedelta(days=30)
    rev30 = sum(nl for _, d, nl in nights if d30 <= d < today)
    n30 = sum(1 for _, d, _2 in nights if d30 <= d < today)
    active = len(_catalog_units) or len(set(lid for lid, _, _2 in nights))
    occ30 = (n30 / (active * 30)) if active else 0
    ci = sum(1 for r in reservations if _res_realized(r) and _parse_date(r.get("arrivalDate")) == today)
    co = sum(1 for r in reservations if _res_realized(r) and _parse_date(r.get("departureDate")) == today)
    return {"active_units": active, "occ_30": round(occ30 * 100), "rev_30": round(rev30),
            "rev_7": round(lw["tot_rev"]), "occ_7": round(lw["occ"] * 100),
            "missed_7": round(lw["tot_missed"]), "pending_cards": len(_pending_replies),
            "open_escalations": sum(1 for e in _escalations.values() if not e.get("claimed_by")),
            "checkins_today": ci, "checkouts_today": co}

def _compute_revenue():
    reservations = get_reservations_cached()
    listings = get_listings_map()
    rep = compute_revenue_report(reservations, listings)
    nights, _ = _explode_nights(reservations)
    series = {}
    for _, d, nl in nights:
        series[f"{d.year}-{d.month:02d}"] = series.get(f"{d.year}-{d.month:02d}", 0) + nl
    monthly = [{"m": k, "rev": round(series[k])} for k in sorted(series)[-12:]]
    return {"monthly": monthly,
            "seasonality": [{"name": m["name"], "adr": round(m["adr"]), "index": round(m["index"], 2)}
                            for m in rep["months"]],
            "salary": {str(k): round(v, 2) for k, v in rep["salary"]["dom_index"].items()},
            "weak": rep["salary"]["weak_window"], "strong": rep["salary"]["strong_window"],
            "units": [{"name": u["name"], "occ": round(u["occ90"] * 100),
                       "adr": round(u["adr"]) if u["adr"] else 0, "pace": round((u["pace30"] or 0) * 100),
                       "reco": u["reco"], "label": u["label"]} for u in rep["units"]]}

def _compute_pricing():
    reservations = get_reservations_cached()
    if not _catalog_units:
        load_catalog(True)
    factors = compute_demand_factors(reservations)
    per_unit, _ = compute_price_opportunities(factors, _catalog_units)
    _last_price_changes.clear()
    _last_price_detail.clear()
    _WD_AR = ["الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
    _WD_EN = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    units = []
    for lid, p in sorted(per_unit.items(), key=lambda kv: kv[1]["uplift"], reverse=True):
        if not p["raise"] and not p["drop"]:
            continue
        _last_price_changes[lid] = _unit_changes(p)
        drows = []
        for r, kind in [(x, "raise") for x in p["raise"]] + [(x, "drop") for x in p["drop"]]:
            d = _parse_date(r["date"])
            mi = round(factors["month_index"].get(d.month, 1), 2) if d else 1
            di = round(factors["dom_index"].get(d.day, 1), 2) if d else 1
            wi = round(factors["dow_index"].get(d.weekday(), 1), 2) if d else 1
            drows.append({"date": r["date"], "wd_ar": _WD_AR[r["wd"]], "wd_en": _WD_EN[r["wd"]],
                          "current": r["current"], "proposed": (r["target"] if kind == "raise" else r["clear"]),
                          "kind": kind, "lead": r["lead"], "mi": mi, "di": di, "wi": wi})
        drows.sort(key=lambda x: x["date"])
        _last_price_detail[lid] = {"name": p["name"], "base": p.get("base", 0),
                                   "confidence": p.get("confidence", 50), "rows": drows}
        units.append({"lid": lid, "name": p["name"], "raise": len(p["raise"]), "drop": len(p["drop"]),
                      "uplift": round(p["uplift"]), "confidence": p.get("confidence", 50),
                      "base": p.get("base", 0)})
    return {"total_uplift": round(sum(p["uplift"] for p in per_unit.values())), "units": units[:40]}

async def _api_overview(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    data = dict(_cache_get("overview") or {})
    data.update(_live_counts())                 # these two are always real-time & cheap
    data["ready"] = _cache_get("overview") is not None
    data["updated"] = _dash_cache.get("overview", (None, 0))[1]
    return _json(data)

async def _api_revenue(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    d = _cache_get("revenue")
    return _json(d if d else {"loading": True})

async def _api_pricing(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    d = _cache_get("pricing")
    return _json(d if d else {"loading": True})

async def _api_log(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    cat = request.query.get("cat", "")
    items = [e for e in reversed(_activity) if not cat or e["cat"] == cat][:200]
    return _json({"items": items})

async def _api_autolog(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    return _json({"items": list(_auto_replies)[:200]})

async def _api_today(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    d = dict(_cache_get("today") or {})
    d.update(_live_counts())                 # escalations + pending always real-time
    d["ready"] = _cache_get("today") is not None
    return _json(d if d.get("ready") else {"loading": True, **_live_counts()})

async def _api_inbox(request):
    """Live mirror of what the team is acting on: pending replies + open escalations.

    NOTE: Discord message IDs (snowflakes) are 18-19 digit ints. JavaScript's
    Number type only preserves precision up to 2^53 (16 digits), so we ALWAYS
    return IDs as strings — otherwise JS rounds off the last 2 digits and every
    detail/send/reject/claim lookup hits a 404."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    replies = []
    for mid, d in list(_pending_replies.items()):
        it = d.get("item", {})
        replies.append({"id": str(mid), "guest": it.get("guest", "Guest"), "unit": it.get("unit", ""),
                        "guest_text": (it.get("guest_text") or "")[:600],
                        "thread": (it.get("history") or "")[:2500],
                        "time": it.get("last_time", ""),
                        "confidence": d.get("confidence"),
                        "draft": (d.get("draft") or "")[:1200]})
    escs = []
    for eid, e in list(_escalations.items()):
        escs.append({"id": str(eid), "guest": e.get("guest", ""), "unit": e.get("unit", ""),
                     "reason": (e.get("reason") or "")[:400], "guest_text": (e.get("guest_text") or "")[:400],
                     "time": e.get("last_ping") and datetime.fromtimestamp(e["last_ping"], TZ).isoformat(timespec="minutes") or "",
                     "claimed_by": e.get("claimed_by")})
    return _json({"replies": replies, "escalations": escs})

async def _read_body(request):
    try:
        return await request.json()
    except Exception:
        return {}

async def _api_send(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    try:
        mid = int(b.get("id"))
    except Exception:
        return _json({"error": "bad id"}, 400)
    if mid in _replied_msgs:
        return _json({"error": "already handled"}, 409)
    data = _pending_replies.pop(mid, None)
    if not data:
        return _json({"error": "not found / already handled"}, 409)
    _replied_msgs.add(mid)
    item = data["item"]
    original_draft = data.get("draft") or ""
    reply = (b.get("text") or original_draft).strip()
    try:
        await asyncio.to_thread(send_guest_message, item["conversation_id"], reply,
                                item.get("comm_type", "email"))
        # learning: dashboard send — if the user changed the text in the textarea
        # vs the original draft, that's a correction signal (was_edited=True via diff)
        record_learning(item, original_draft, reply,
                        via="dashboard_send", approver="(dashboard)")
        log_event("guest", f"رد (من اللوحة) · {item.get('guest','')} · {item.get('unit','')}")
        return _json({"ok": True})
    except Exception as e:
        return _json({"error": str(e)}, 500)

async def _api_reject(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    try:
        mid = int(b.get("id"))
    except Exception:
        return _json({"error": "bad id"}, 400)
    _replied_msgs.add(mid)
    data = _pending_replies.pop(mid, None)
    if data:
        log_event("guest", f"تجاهل رد (من اللوحة) · {data.get('item',{}).get('guest','')}")
    return _json({"ok": True})

async def _api_claim(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    try:
        eid = int(b.get("id"))
    except Exception:
        return _json({"error": "bad id"}, 400)
    name = (b.get("name") or "الفريق").strip()
    e = _escalations.get(eid)
    if not e:
        return _json({"error": "not found"}, 409)
    if e.get("claimed_by"):
        return _json({"error": f"already claimed by {e['claimed_by']}"}, 409)
    e["claimed_by"] = name
    metric_bump("escalations_resolved")
    if e.get("conversation_id"):
        _claimed_convos.add(e["conversation_id"])
    log_event("escalation", f"استلام تصعيد (من اللوحة) بواسطة {name} · {e.get('unit','')}")
    return _json({"ok": True})

async def _api_apply(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    try:
        lid = int(b.get("lid"))
    except Exception:
        return _json({"error": "bad lid"}, 400)
    changes = _last_price_changes.get(lid)
    if not changes:                                    # fall back to the detail rows
        det = _last_price_detail.get(lid)
        if det and det.get("rows"):
            changes = [{"date": r["date"], "price": r["proposed"], "kind": r["kind"]}
                       for r in det["rows"]]
    if not changes:
        return _json({"error": "no pending changes (refresh pricing first)"}, 409)
    applied, skipped, results = await asyncio.to_thread(apply_price_changes, lid, changes)
    _last_price_changes.pop(lid, None)
    if PRICING_STRATEGY_ENABLED and activate_strategy(lid, applied):
        asyncio.create_task(_kick_strategy(lid))   # run one pass now so the page isn't blank
    name = next((u["name"] for u in _catalog_units if u.get("id") == lid), str(lid))
    log_event("pricing", f"طبّق {applied} سعر (من اللوحة) · {name}"
              + (" (DRY-RUN)" if PRICE_APPLY_DRYRUN else ""))
    return _json({"ok": True, "applied": applied, "skipped": skipped,
                  "dry_run": PRICE_APPLY_DRYRUN, "results": results})

async def _handle_dashboard(request):
    return web.Response(text=DASHBOARD_HTML, content_type="text/html")

async def _api_pricing_detail(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    try:
        lid = int(request.query.get("lid"))
    except Exception:
        return _json({"error": "bad lid"}, 400)
    return _json(_last_price_detail.get(lid) or {"rows": []})

def _strategy_price(base, d, factors):
    """Best-number price for one night: demand target × Saudi-event boost,
    stepped down as the night nears while empty."""
    mi = factors["month_index"].get(d.month, 1)
    di = factors["dom_index"].get(d.day, 1)
    wi = factors["dow_index"].get(d.weekday(), 1)
    ev = event_boost_for_date(d)           # 1.4-1.6 on Eid / National Day, 0.75 during Ramadan
    target = max(0.6 * base, min(2.2 * base, base * mi * di * wi * ev))
    lead = (d - datetime.now(TZ).date()).days
    f = 0.80 if lead <= 2 else 0.86 if lead <= 5 else 0.92 if lead <= 10 else 0.97 if lead <= 20 else 1.0
    return int(round(max(0.6 * base, target * f)))

def activate_strategy(lid, applied=0):
    """Start (or refresh) an active pricing strategy for a unit from its latest detail rows."""
    det = _last_price_detail.get(lid)
    if not det or not det.get("rows"):
        return False
    dates = {}
    for r in det["rows"]:
        p = r["proposed"]
        dates[r["date"]] = {"start": p, "cur": p, "booked": False, "changes": 0,
                            "last": datetime.now(TZ).isoformat(timespec="minutes")}
    prev = _pricing_strategies.get(lid, {})
    _pricing_strategies[lid] = {"name": det.get("name", str(lid)), "base": det.get("base", 0),
                                "started": prev.get("started") or datetime.now(TZ).isoformat(timespec="minutes"),
                                "updated": time.time(), "active": True, "dates": dates,
                                "applied_start": applied, "changes_total": prev.get("changes_total", 0),
                                "dry_at_start": PRICE_APPLY_DRYRUN}
    return True

async def _kick_strategy(lid):
    """Run a single optimization pass for one unit immediately (called right after Apply)."""
    try:
        strat = _pricing_strategies.get(lid)
        if not strat:
            return
        factors = await asyncio.to_thread(lambda: compute_demand_factors(get_reservations_cached()))
        await asyncio.to_thread(_run_strategy_unit, lid, strat, factors, datetime.now(TZ).date())
    except Exception as e:
        print(f"kick strategy {lid}:", e)

@tasks.loop(minutes=PRICING_STRATEGY_MIN)
async def pricing_strategy_loop():
    """Keep re-optimizing every active unit's open nights toward the best numbers."""
    if not (PRICING_STRATEGY_ENABLED and any(s.get("active") for s in _pricing_strategies.values())):
        return
    try:
        factors = await asyncio.to_thread(lambda: compute_demand_factors(get_reservations_cached()))
    except Exception as e:
        print("strategy factors error:", e)
        return
    today = datetime.now(TZ).date()
    for lid, strat in list(_pricing_strategies.items()):
        if not strat.get("active"):
            continue
        try:
            await asyncio.to_thread(_run_strategy_unit, lid, strat, factors, today)
        except Exception as e:
            print(f"strategy unit {lid} error:", e)

def _run_strategy_unit(lid, strat, factors, today):
    dates = strat["dates"]
    future = [d for d in dates if (_parse_date(d) or today) >= today]
    if not future:
        strat["active"] = False
        return
    ds = sorted(_parse_date(d) for d in future)
    booked_now = {}
    try:
        cal = api_get(f"/listings/{lid}/calendar",
                      params={"startDate": ds[0].isoformat(), "endDate": ds[-1].isoformat()})
        for day in (cal.get("result") or []):
            booked_now[day.get("date")] = (int(day.get("isAvailable", 0) or 0) != 1
                                           or bool(day.get("reservationId")))
    except Exception as e:
        print("strategy calendar error:", e)
        return
    base = strat.get("base") or factors["overall_adr"]
    for dstr in future:
        d = _parse_date(dstr)
        if d < today:
            continue
        rec = dates[dstr]
        if booked_now.get(dstr):                       # it sold — success, stop managing it
            if not rec["booked"]:
                rec["booked"] = True
                rec["last"] = datetime.now(TZ).isoformat(timespec="minutes")
                log_event("pricing", f"استراتيجية · انحجزت ليلة {dstr} · {strat['name']} ✅")
            continue
        want = _strategy_price(base, d, factors)
        cur = rec.get("cur") or want
        if cur and abs(want - cur) / cur >= 0.03:       # only meaningful moves
            if not PRICE_APPLY_DRYRUN:
                try:
                    api_put(f"/listings/{lid}/calendar",
                            {"startDate": dstr, "endDate": dstr, "isAvailable": 1,
                             "price": want, "note": f"ouja-orig:{want}"})
                except Exception as e:
                    print(f"strategy put {lid} {dstr}:", e)
                    continue
            rec["cur"] = want
            rec["changes"] = rec.get("changes", 0) + 1
            rec["last"] = datetime.now(TZ).isoformat(timespec="minutes")
            strat["changes_total"] = strat.get("changes_total", 0) + 1
    strat["updated"] = time.time()
    if all(rec["booked"] or (_parse_date(d) and _parse_date(d) < today) for d, rec in dates.items()):
        strat["active"] = False

def _strategy_view(lid):
    """Strategy detail with per-night before/after + the 'why' (month/dom/dow factors +
    lead-time discount) so the dashboard can elaborate each price decision."""
    s = _pricing_strategies.get(lid)
    if not s:
        return {"active": False, "dates": []}
    # compute factors once so we can attach per-night reasoning to the rows
    try:
        factors = compute_demand_factors(get_reservations_cached())
    except Exception:
        factors = {"month_index": {}, "dom_index": {}, "dow_index": {}, "overall_adr": 0}
    today = datetime.now(TZ).date()
    _WD_AR = ["الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
    _WD_EN = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    rows = []
    for d_str, r in s["dates"].items():
        d = _parse_date(d_str)
        if not d:
            continue
        lead = (d - today).days
        mi = round(factors["month_index"].get(d.month, 1), 2) if d else 1
        di = round(factors["dom_index"].get(d.day, 1), 2) if d else 1
        wi = round(factors["dow_index"].get(d.weekday(), 1), 2) if d else 1
        lead_factor = (0.80 if lead <= 2 else 0.86 if lead <= 5 else
                       0.92 if lead <= 10 else 0.97 if lead <= 20 else 1.0)
        rows.append({
            "date": d_str, "wd_ar": _WD_AR[d.weekday()], "wd_en": _WD_EN[d.weekday()],
            "lead": lead, "start": r["start"], "cur": r["cur"],
            "booked": r["booked"], "changes": r.get("changes", 0),
            "last": r.get("last", ""),
            "mi": mi, "di": di, "wi": wi, "lead_factor": lead_factor,
        })
    rows.sort(key=lambda x: x["date"])
    booked = sum(1 for r in rows if r["booked"])
    return {"name": s.get("name"), "base": s.get("base"), "active": s.get("active", False),
            "started": s.get("started"), "updated": s.get("updated", 0),
            "interval": PRICING_STRATEGY_MIN, "dry_run": PRICE_APPLY_DRYRUN,
            "applied_start": s.get("applied_start", 0), "changes_total": s.get("changes_total", 0),
            "total": len(rows), "booked": booked, "open": len(rows) - booked, "dates": rows}

def _strategies_list():
    """Summary of every strategy (active + finished) for the standing Strategies page."""
    out = []
    for lid, s in _pricing_strategies.items():
        dates = s.get("dates", {})
        booked = sum(1 for r in dates.values() if r.get("booked"))
        out.append({"lid": lid, "name": s.get("name", str(lid)), "base": s.get("base", 0),
                    "active": s.get("active", False), "started": s.get("started"),
                    "updated": s.get("updated", 0), "total": len(dates), "booked": booked,
                    "open": len(dates) - booked, "applied_start": s.get("applied_start", 0),
                    "changes_total": s.get("changes_total", 0),
                    "dry": s.get("dry_at_start", PRICE_APPLY_DRYRUN)})
    # active first, then most-recently-updated
    out.sort(key=lambda x: (not x["active"], -(x["updated"] or 0)))
    return out

async def _api_strategies(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    return _json({"items": _strategies_list(), "enabled": PRICING_STRATEGY_ENABLED,
                  "interval": PRICING_STRATEGY_MIN, "dry_run": PRICE_APPLY_DRYRUN})

async def _api_strategy(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    try:
        lid = int(request.query.get("lid"))
    except Exception:
        return _json({"error": "bad lid"}, 400)
    return _json(_strategy_view(lid))

async def _api_strategy_stop(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    try:
        lid = int(b.get("lid"))
    except Exception:
        return _json({"error": "bad lid"}, 400)
    s = _pricing_strategies.get(lid)
    if s:
        s["active"] = False
        log_event("pricing", f"إيقاف استراتيجية · {s.get('name', lid)}")
    return _json({"ok": True})

async def _api_discount_status(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    return _json(discount_pause_status())

async def _api_discount_pause(request):
    """POST {hours: N (default 24)} — pause discount tiers for N hours."""
    global _discount_paused_until
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    try:
        hours = float(b.get("hours", 24))
    except Exception:
        hours = 24.0
    hours = max(0.5, min(168.0, hours))               # clamp 30 min … 7 days
    _discount_paused_until = time.time() + hours * 3600
    await asyncio.to_thread(persist_state)            # survive a redeploy
    log_event("pricing", f"إيقاف الخصومات لمدة {hours:g} ساعة")
    return _json({"ok": True, **discount_pause_status()})

async def _api_discount_resume(request):
    global _discount_paused_until
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    _discount_paused_until = 0
    await asyncio.to_thread(persist_state)
    log_event("pricing", "استئناف الخصومات التلقائية")
    return _json({"ok": True, **discount_pause_status()})

# ---- per-unit discount skip (hold price on a specific apartment) ----
async def _api_unit_skip(request):
    """POST {lid, hours: N (default 24)} — skip discounts on this unit for N hours."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    try:
        lid = int(b.get("lid"))
        hours = float(b.get("hours", 24))
    except Exception:
        return _json({"error": "bad params"}, 400)
    hours = max(0.5, min(168.0, hours))
    _unit_discount_skip[lid] = time.time() + hours * 3600
    await asyncio.to_thread(persist_state)
    name = next((u["name"] for u in _catalog_units if u.get("id") == lid), str(lid))
    log_event("pricing", f"تجاهل خصم {hours:g}س · {name}")
    return _json({"ok": True, "lid": lid, "until_iso": unit_skip_until_iso(lid)})

async def _api_unit_unskip(request):
    """POST {lid} — clear the skip on this unit."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    try:
        lid = int(b.get("lid"))
    except Exception:
        return _json({"error": "bad lid"}, 400)
    _unit_discount_skip.pop(lid, None)
    await asyncio.to_thread(persist_state)
    name = next((u["name"] for u in _catalog_units if u.get("id") == lid), str(lid))
    log_event("pricing", f"إلغاء تجاهل الخصم · {name}")
    return _json({"ok": True})

async def _api_today_empty(request):
    """Per-unit empty-tonight detail: current price, scheduled tier prices, skip status."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    items = await asyncio.to_thread(compute_tonight_empty)
    return _json({"items": items, "weekend": datetime.now(TZ).weekday() in WEEKEND_DAYS,
                  "paused": is_discount_paused(),
                  "paused_until_iso": discount_pause_status().get("until_iso", "")})

async def _api_inbox_detail(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    try:
        item_id = int(request.query.get("id"))
    except Exception:
        return _json({"error": "bad id"}, 400)
    detail = await asyncio.to_thread(get_inbox_item_detail, item_id)
    if not detail:
        return _json({"error": "not found"}, 404)
    return _json(detail)

async def _api_teach(request):
    """POST {topic, fact} — save a fact to the #knowledge Discord channel (mirrors the
    Discord 'علّم' modal). Future drafts pick it up after the next knowledge refresh."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    topic = (b.get("topic") or "").strip()
    fact = (b.get("fact") or "").strip()
    if not fact:
        return _json({"error": "fact required"}, 400)
    line = f"**{topic}**: {fact}" if topic else fact
    guild = bot.get_guild(GUILD_ID)
    if guild is None:
        return _json({"error": "discord guild not ready"}, 503)
    ok = await save_fact(guild, line)
    if ok:
        log_event("guest", f"تعلّم معلومة جديدة (من اللوحة): {(topic or fact)[:80]}")
    return _json({"ok": bool(ok)})

# ---- self-learning views ----
async def _api_learning_summary(request):
    """Return distilled summaries. Optional ?lid= for one apartment, else all."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    lid = request.query.get("lid")
    if lid:
        try:
            lid_i = int(lid)
        except Exception:
            return _json({"error": "bad lid"}, 400)
        apt = _apartment_learnings.get(lid_i) or {}
        return _json({"apartment": apt, "general": _general_learnings})
    # all apartments
    return _json({
        "general": _general_learnings,
        "apartments": [{"lid": k, **v} for k, v in
                       sorted(_apartment_learnings.items(),
                              key=lambda kv: -(kv[1].get("last_distilled") or 0))],
        "log_size": len(_learning_log),
    })

async def _api_learning_log(request):
    """Last N learning log entries (most recent first)."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    try:
        n = int(request.query.get("limit", "50"))
    except Exception:
        n = 50
    n = max(1, min(500, n))
    items = list(_learning_log)[-n:][::-1]
    return _json({"items": items, "total": len(_learning_log)})

async def _api_learning_forget(request):
    """POST {lid?, scope?: 'apartment'|'general'|'all'} — drop a learned summary
    so the bot stops citing it. Useful when a summary went wrong."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    scope = (b.get("scope") or "").strip()
    if scope == "general":
        _general_learnings["summary"] = ""
        _general_learnings["examples_count"] = 0
        log_event("guest", "نسيت الملخص العام للتعلّم")
        await asyncio.to_thread(persist_state)
        return _json({"ok": True})
    if scope == "apartment":
        try:
            lid = int(b.get("lid"))
        except Exception:
            return _json({"error": "bad lid"}, 400)
        prev = _apartment_learnings.pop(lid, None)
        log_event("guest", f"نسيت الملخص الخاص بـ {(prev or {}).get('unit','وحدة')}")
        await asyncio.to_thread(persist_state)
        return _json({"ok": True})
    if scope == "all":
        _apartment_learnings.clear()
        _general_learnings["summary"] = ""
        _general_learnings["examples_count"] = 0
        log_event("guest", "نسيت كل ملخصات التعلّم")
        await asyncio.to_thread(persist_state)
        return _json({"ok": True})
    return _json({"error": "scope must be apartment|general|all"}, 400)

async def _api_learning_distill_now(request):
    """Force a re-distillation right now (useful for testing or after a new wave of edits)."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    await asyncio.to_thread(distill_learnings)
    return _json({"ok": True,
                  "apartments_with_summary": len(_apartment_learnings),
                  "general_examples": _general_learnings.get("examples_count", 0)})

_bootstrap_state = {"running": False, "started": 0, "finished": 0, "result": None, "error": ""}

async def _api_learning_bootstrap(request):
    """POST {limit_conversations?: int (50..1000), force?: bool} — one-shot historical
    seeding. Walks the most-recent N Hostaway conversations, extracts team replies,
    distills per-apartment summaries. Runs in the background — poll
    /api/learning/bootstrap/status for progress."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    if _bootstrap_state["running"]:
        return _json({"error": "bootstrap already running",
                      "started": _bootstrap_state["started"]}, 409)
    b = await _read_body(request)
    try:
        limit = int(b.get("limit_conversations", 300))
    except Exception:
        limit = 300
    limit = max(50, min(1000, limit))
    async def _run():
        _bootstrap_state.update({"running": True, "started": time.time(),
                                 "finished": 0, "result": None, "error": ""})
        try:
            res = await asyncio.to_thread(bootstrap_learnings_from_history, limit)
            _bootstrap_state["result"] = res
            await asyncio.to_thread(persist_state)
        except Exception as e:
            _bootstrap_state["error"] = str(e)
            print("bootstrap_learnings error:", e)
        finally:
            _bootstrap_state["running"] = False
            _bootstrap_state["finished"] = time.time()
    asyncio.create_task(_run())
    return _json({"ok": True, "started": True, "limit_conversations": limit,
                  "note": "Running in background; poll /api/learning/bootstrap/status"})

async def _api_learning_bootstrap_status(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    return _json(_bootstrap_state)

async def _api_clean_quality_summary(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    return _json(cleaning_quality_summary())

async def _api_clean_feedback_get(request):
    """Public — fetches the rating context for a token (so the page can show
    which unit they're rating). No auth required; the token IS the auth."""
    tok = request.query.get("id", "")
    fb = _cleaning_feedback.get(tok)
    if not fb:
        return _json({"error": "not found"}, 404)
    return _json({"unit": fb.get("unit"), "guest": fb.get("guest", ""),
                  "already_done": fb.get("score") is not None})

async def _api_clean_feedback_submit(request):
    """Public — POST {id, score(1-5), comment?}. Idempotent: first submission wins."""
    b = await _read_body(request)
    tok = (b.get("id") or "").strip()
    fb = _cleaning_feedback.get(tok)
    if not fb:
        return _json({"error": "not found"}, 404)
    if fb.get("score") is not None:
        return _json({"ok": True, "already": True})
    try:
        score = int(b.get("score"))
    except Exception:
        return _json({"error": "bad score"}, 400)
    if score < 1 or score > 5:
        return _json({"error": "score 1-5"}, 400)
    fb["score"] = score
    fb["comment"] = (b.get("comment") or "")[:600]
    fb["ts_done"] = datetime.now(TZ).isoformat(timespec="minutes")
    await asyncio.to_thread(persist_state)
    log_event("guest", f"تقييم نظافة جديد · {fb.get('unit','')} · {score}⭐")
    return _json({"ok": True})

async def _handle_clean_feedback_page(request):
    return web.Response(text=CLEAN_FEEDBACK_HTML, content_type="text/html")

async def _api_guests_list(request):
    """Return guest profiles for the dashboard. Filters/sorts in the client."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    out = []
    for k, p in _guest_profiles.items():
        out.append({
            "key": k,
            "name": (p.get("names") or ["—"])[-1],
            "names": p.get("names", []),
            "phone": p.get("phone", ""),
            "email": p.get("email", ""),
            "vip": p.get("vip", False),
            "stays": len(p.get("reservations", [])),
            "nights": p.get("total_nights", 0),
            "first_seen": p.get("first_seen"),
            "last_seen": p.get("last_seen"),
            "summaries_count": len(p.get("summaries", [])),
        })
    out.sort(key=lambda x: (not x["vip"], -(x["stays"] or 0)))
    return _json({"items": out, "counts": {
        "total": len(out),
        "vip": sum(1 for x in out if x["vip"]),
        "repeat": sum(1 for x in out if x["stays"] >= 2),
    }})

async def _api_guest_detail(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    key = request.query.get("key", "")
    p = _guest_profiles.get(key)
    if not p:
        return _json({"error": "not found"}, 404)
    return _json(p)

async def _api_guest_notes(request):
    """POST {key, notes} — set free-form notes on a guest profile."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    key = b.get("key", "")
    p = _guest_profiles.get(key)
    if not p:
        return _json({"error": "not found"}, 404)
    p["notes"] = (b.get("notes") or "")[:1000]
    await asyncio.to_thread(persist_state)
    return _json({"ok": True})

async def _api_guest_toggle_vip(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    p = _guest_profiles.get(b.get("key", ""))
    if not p:
        return _json({"error": "not found"}, 404)
    p["vip"] = not p.get("vip", False)
    await asyncio.to_thread(persist_state)
    return _json({"ok": True, "vip": p["vip"]})

async def _api_cleaning_schedule(request):
    """Dashboard view of every unit's deep-clean state."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    listings = get_listings_map() or {}
    today = datetime.now(TZ).date()
    if not _catalog_units:
        await asyncio.to_thread(load_catalog, True)
    out = []
    for lid, name in listings.items():
        if lid not in _deep_clean_state:
            _dc_init(lid)
        s = _deep_clean_state[lid]
        u = next((c for c in _catalog_units if c.get("id") == lid), {})
        last = _parse_date(s.get("last_done"))
        nxt = _parse_date(s.get("next_scheduled"))
        days_since = (today - last).days if last else None
        days_until = (nxt - today).days if nxt else None
        # overdue if days_since > MAX
        overdue = bool(days_since and days_since > DEEPCLEAN_MAX_DAYS)
        out.append({
            "lid": lid, "name": name,
            "beds": u.get("beds"), "area": u.get("area") or u.get("neighbourhood"),
            "last_done": s.get("last_done"),
            "next_scheduled": s.get("next_scheduled"),
            "next_status": s.get("next_status"),
            "days_since_last": days_since, "days_until_next": days_until,
            "overdue": overdue,
            "history": (s.get("history") or [])[-5:],
            "notes": s.get("notes", ""),
        })
    out.sort(key=lambda x: (-1 if x["overdue"] else 0, -(x["days_since_last"] or 0)))
    counts = {
        "total": len(out),
        "overdue": sum(1 for x in out if x["overdue"]),
        "scheduled": sum(1 for x in out if x["next_scheduled"]),
        "blocked_tomorrow": sum(1 for x in out
                                if x["next_scheduled"] == (today + timedelta(days=1)).isoformat()
                                and x["next_status"] == "blocked"),
    }
    return _json({"items": out, "today": today.isoformat(), "counts": counts,
                  "cleaning_url": ("/cleaning?token=" + CLEANING_TOKEN) if CLEANING_TOKEN else "",
                  "have_token": bool(CLEANING_TOKEN)})

async def _api_cleaning_mark_done(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    try:
        lid = int(b.get("lid"))
    except Exception:
        return _json({"error": "bad lid"}, 400)
    ok = await asyncio.to_thread(mark_deep_clean_done, lid, b.get("date"), b.get("notes", ""))
    await asyncio.to_thread(persist_state)
    name = next((u["name"] for u in _catalog_units if u.get("id") == lid), str(lid))
    log_event("pricing", f"تنظيف عميق · {name}: تم الإنجاز ✓")
    return _json({"ok": ok})

async def _api_cleaning_reschedule(request):
    """POST {lid, date} — manual override of next scheduled date."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    try:
        lid = int(b.get("lid"))
    except Exception:
        return _json({"error": "bad lid"}, 400)
    date = b.get("date")
    if not _parse_date(date):
        return _json({"error": "bad date"}, 400)
    if lid not in _deep_clean_state:
        _dc_init(lid)
    _deep_clean_state[lid]["next_scheduled"] = date
    _deep_clean_state[lid]["next_status"] = "scheduled"
    await asyncio.to_thread(persist_state)
    return _json({"ok": True})

async def _api_cleaning_import_csv(request):
    """POST multipart-form with field 'file' = a CSV with two columns: name,date.
    Header row optional. Date format: YYYY-MM-DD (or DD/MM/YYYY also accepted).
    Names are matched against listings by normalized comparison
    (strip 'Ouja|', lower-case, collapse whitespace)."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    try:
        reader = await request.multipart()
        field = await reader.next()
        if not field or field.name != "file":
            return _json({"error": "expected 'file' multipart field"}, 400)
        raw = (await field.read()).decode("utf-8-sig", errors="ignore")
    except Exception as e:
        return _json({"error": f"upload error: {e}"}, 400)

    import csv as _csv, io as _io
    listings = get_listings_map() or {}
    norm_to_lid = {norm_unit(name): lid for lid, name in listings.items()}
    rows = list(_csv.reader(_io.StringIO(raw)))
    if not rows:
        return _json({"error": "empty CSV"}, 400)
    # Skip the header row if first row's "date" cell doesn't parse as a date
    start = 0
    if rows[0]:
        test_date = (rows[0][-1] or "").strip()
        if not _parse_any_date(test_date):
            start = 1

    matched, unmatched, no_date = [], [], []
    for r in rows[start:]:
        if not r or len(r) < 2:
            continue
        raw_name = (r[0] or "").strip()
        raw_date = (r[-1] or "").strip()
        if not raw_name:
            continue
        d = _parse_any_date(raw_date)
        if not d:
            no_date.append({"name": raw_name, "raw_date": raw_date})
            continue
        n = norm_unit(raw_name)
        lid = norm_to_lid.get(n)
        if not lid:
            # fuzzy fallback: contains-match
            for k, v in norm_to_lid.items():
                if n and (n in k or k in n) and len(n) >= 3:
                    lid = v
                    break
        if not lid:
            unmatched.append({"name": raw_name})
            continue
        if lid not in _deep_clean_state:
            _dc_init(lid)
        _deep_clean_state[lid]["last_done"] = d.isoformat()
        _deep_clean_state[lid]["next_scheduled"] = None
        _deep_clean_state[lid]["next_status"] = "unscheduled"
        matched.append({"lid": lid, "name": listings[lid], "date": d.isoformat()})
    await asyncio.to_thread(persist_state)
    # rerun scheduler so the new dates flow into next_scheduled
    await asyncio.to_thread(schedule_deep_cleans)
    log_event("pricing", f"تنظيف عميق · استورد {len(matched)} وحدة من CSV "
                          f"(غير مطابقة: {len(unmatched)}, بدون تاريخ: {len(no_date)})")
    return _json({"ok": True, "matched": matched, "unmatched": unmatched, "no_date": no_date})

def _parse_any_date(s):
    """Tolerant date parser: accepts YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY, MM/DD/YYYY."""
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            pass
    return None

# ---- Smart Arabic free-form date parser (for the cleaning-baseline import) ----
_AR_NUM_MAP = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
_AR_MONTHS = {
    "يناير": 1, "كانون الثاني": 1,
    "فبراير": 2, "شباط": 2,
    "مارس": 3, "آذار": 3, "اذار": 3,
    "أبريل": 4, "ابريل": 4, "نيسان": 4,
    "مايو": 5, "أيار": 5, "ايار": 5,
    "يونيو": 6, "حزيران": 6,
    "يوليو": 7, "تموز": 7,
    "أغسطس": 8, "اغسطس": 8, "آب": 8,
    "سبتمبر": 9, "أيلول": 9, "ايلول": 9,
    "أكتوبر": 10, "اكتوبر": 10, "تشرين الأول": 10,
    "نوفمبر": 11, "تشرين الثاني": 11,
    "ديسمبر": 12, "كانون الأول": 12,
}
# Phrases that mean "we don't know" / "never done". Caller treats these as the
# unit being maximally overdue (forces immediate scheduling).
_AR_UNKNOWN_PHRASES = ("مدري", "ماسوينا", "ما سوينا", "ما تم", "?", "n/a", "—")

def parse_ar_freeform_date(s, today=None):
    """Returns (date | None, was_unknown:bool). Handles ISO, DD/MM/YYYY,
    "قبل X يوم/اسبوع/شهر[ ونص]", Arabic month names with Arabic or western
    numerals, and uncertainty phrases like "مدري"/"ماسوينا" → returns (None, True)."""
    if s is None:
        return None, False
    today = today or datetime.now(TZ).date()
    if isinstance(s, datetime):
        return s.date(), False
    if isinstance(s, date) and not isinstance(s, datetime):
        return s, False
    raw = str(s).strip()
    if not raw:
        return None, False
    s = raw.translate(_AR_NUM_MAP).lower().strip()
    if any(s == u or s.startswith(u) for u in _AR_UNKNOWN_PHRASES if u):
        return None, True
    d = _parse_any_date(s)
    if d:
        return d, False
    # relative day phrases
    m = re.search(r"قبل\s*(\d+)\s*(?:يوم|ايام|أيام)", s)
    if m: return today - timedelta(days=int(m.group(1))), False
    if "قبل اسبوعين" in s or "قبل أسبوعين" in s: return today - timedelta(days=14), False
    if "قبل اسبوع" in s or "قبل أسبوع" in s: return today - timedelta(days=7), False
    m = re.search(r"قبل\s*(\d+)\s*(?:اسابيع|أسابيع)", s)
    if m: return today - timedelta(weeks=int(m.group(1))), False
    if "قبل شهرين ونص" in s or "قبل شهرين ونصف" in s: return today - timedelta(days=75), False
    if "قبل شهر ونص" in s or "قبل شهر ونصف" in s: return today - timedelta(days=45), False
    if "قبل شهرين" in s: return today - timedelta(days=60), False
    if "قبل شهر" in s: return today - timedelta(days=30), False
    m = re.search(r"قبل\s*(\d+)\s*(?:اشهر|أشهر|شهور)", s)
    if m: return today - timedelta(days=30 * int(m.group(1))), False
    # Arabic month name with day number
    for name, num in _AR_MONTHS.items():
        if name in s:
            digs = re.findall(r"\d+", s)
            if digs:
                day_n = int(digs[0])
                if 1 <= day_n <= 31:
                    try:
                        d = date(today.year, num, day_n)
                        if (d - today).days > 180:
                            d = date(today.year - 1, num, day_n)
                        return d, False
                    except ValueError:
                        pass
    return None, False

# Add a 'date' alias at module level since we use it in the helper above
from datetime import date  # noqa: E402

async def _api_cleaning_import_xlsx(request):
    """POST multipart 'file' = .xlsx with one row per apartment.
    Auto-detects which columns hold the apartment name and the last-clean date
    (any cell in a row that parses as an Arabic-or-ISO date counts as the date,
    any cell that matches a catalog name as the name). 'مدري' / 'ماسوينا' / blank
    are treated as 'overdue from day one' so the scheduler picks them first."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    try:
        import openpyxl as _xl, io as _io
    except Exception:
        return _json({"error": "openpyxl not installed"}, 500)
    try:
        reader = await request.multipart()
        field = await reader.next()
        if not field or field.name != "file":
            return _json({"error": "expected 'file' multipart field"}, 400)
        blob = await field.read()
    except Exception as e:
        return _json({"error": f"upload error: {e}"}, 400)
    try:
        wb = _xl.load_workbook(_io.BytesIO(blob), data_only=True)
    except Exception as e:
        return _json({"error": f"not a valid xlsx: {e}"}, 400)

    listings = get_listings_map() or {}
    norm_to_lid = {norm_unit(n): lid for lid, n in listings.items()}
    today = datetime.now(TZ).date()
    # treat "unknown" as 90 days ago so the scheduler picks them immediately
    UNKNOWN_FALLBACK = today - timedelta(days=90)

    matched, unmatched, unknown_dates = [], [], []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            # Find name-cell + date-cell heuristically
            name_cell, date_cell, parsed_date, was_unknown = None, None, None, False
            for cell in row:
                if isinstance(cell, str) and len(cell.strip()) >= 2:
                    norm = norm_unit(cell)
                    if norm in norm_to_lid:
                        name_cell = cell; continue
                d, unk = parse_ar_freeform_date(cell, today=today)
                if d:
                    date_cell = cell; parsed_date = d
                if unk:
                    was_unknown = True; date_cell = cell
            # Fallback: if no exact match, try fuzzy contains on string cells
            if not name_cell:
                for cell in row:
                    if isinstance(cell, str) and len(cell.strip()) >= 2:
                        norm = norm_unit(cell)
                        for k in norm_to_lid:
                            if norm and (norm in k or k in norm) and abs(len(norm) - len(k)) <= 6:
                                name_cell = cell
                                break
                        if name_cell:
                            break
            if not name_cell:
                continue
            lid = norm_to_lid.get(norm_unit(name_cell))
            if not lid:
                for k, v in norm_to_lid.items():
                    n = norm_unit(name_cell)
                    if n and (n in k or k in n) and abs(len(n) - len(k)) <= 6:
                        lid = v; break
            if not lid:
                unmatched.append({"name": name_cell, "raw_date": str(date_cell or "")})
                continue
            if lid not in _deep_clean_state:
                _dc_init(lid)
            if parsed_date:
                _deep_clean_state[lid]["last_done"] = parsed_date.isoformat()
                _deep_clean_state[lid]["next_scheduled"] = None
                _deep_clean_state[lid]["next_status"] = "unscheduled"
                matched.append({"lid": lid, "name": listings[lid],
                                "raw_date": str(date_cell), "parsed": parsed_date.isoformat()})
            elif was_unknown:
                _deep_clean_state[lid]["last_done"] = UNKNOWN_FALLBACK.isoformat()
                _deep_clean_state[lid]["next_scheduled"] = None
                _deep_clean_state[lid]["next_status"] = "unscheduled"
                _deep_clean_state[lid]["notes"] = (
                    "تاريخ آخر تنظيف غير معروف (مدري/ماسوينا) — تمت معاملته كأنه قبل ٩٠ يوم لضمان جدولة فورية."
                )
                unknown_dates.append({"lid": lid, "name": listings[lid],
                                      "raw_date": str(date_cell)})
    # Re-run the scheduler so the freshly-imported dates yield real next dates
    await asyncio.to_thread(persist_state)
    await asyncio.to_thread(schedule_deep_cleans)
    log_event("pricing",
              f"تنظيف عميق · استيراد XLSX: {len(matched)} تاريخ + "
              f"{len(unknown_dates)} غير معروف · {len(unmatched)} غير مطابق")
    return _json({"ok": True, "matched": matched, "unknown": unknown_dates,
                  "unmatched": unmatched})

async def _api_cleaning_set_last(request):
    """POST {lid, date} — set last_done (used to import the April 27 baseline file)."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    try:
        lid = int(b.get("lid"))
    except Exception:
        return _json({"error": "bad lid"}, 400)
    date = b.get("date")
    if not _parse_date(date):
        return _json({"error": "bad date"}, 400)
    if lid not in _deep_clean_state:
        _dc_init(lid)
    _deep_clean_state[lid]["last_done"] = date
    _deep_clean_state[lid]["next_scheduled"] = None
    _deep_clean_state[lid]["next_status"] = "unscheduled"
    await asyncio.to_thread(persist_state)
    return _json({"ok": True})

async def _api_cleaning_public(request):
    """Public data for the cleaning company. Auth via CLEANING_TOKEN query param,
    not the dashboard token — so the link can be shared with the cleaners safely."""
    token = request.query.get("token", "")
    if not CLEANING_TOKEN or token != CLEANING_TOKEN:
        return _json({"error": "unauthorized"}, 401)
    if not _catalog_units:
        await asyncio.to_thread(load_catalog, True)
    listings = get_listings_map() or {}
    today = datetime.now(TZ).date()
    out = []
    for lid, s in _deep_clean_state.items():
        if not s.get("next_scheduled"):
            continue
        sd = _parse_date(s["next_scheduled"])
        if not sd or sd < today or (sd - today).days > 45:
            continue
        u = next((c for c in _catalog_units if c.get("id") == lid), {})
        out.append({
            "name": listings.get(lid, str(lid)),
            "date": s["next_scheduled"],
            "status": s.get("next_status", "scheduled"),
            "beds": u.get("beds"), "baths": u.get("baths"),
            "area": u.get("area") or u.get("neighbourhood"),
            "last_done": s.get("last_done"),
            "guide_url": await asyncio.to_thread(get_guide_url, lid),
        })
    out.sort(key=lambda x: x["date"])
    return _json({"items": out, "today": today.isoformat()})

async def _handle_cleaning_page(request):
    return web.Response(text=CLEANING_HTML, content_type="text/html")

async def _api_events_list(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    return _json({"events": _all_events()})

async def _api_events_save(request):
    """POST {id?, name, start, end, boost, kind?} — add a new custom event or
    update an existing one (matched by id). Default events are read-only."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    name = (b.get("name") or "").strip()
    start = (b.get("start") or "").strip()
    end = (b.get("end") or "").strip()
    try:
        boost = float(b.get("boost", 1.0))
    except Exception:
        boost = 1.0
    kind = (b.get("kind") or "custom").strip()[:30]
    if not name or not _parse_date(start) or not _parse_date(end):
        return _json({"error": "name + valid start + end required"}, 400)
    if end < start:
        return _json({"error": "end must be >= start"}, 400)
    boost = max(0.3, min(3.0, boost))     # clamp to sane range
    eid = b.get("id")
    if eid:                                # update existing
        for e in _custom_events:
            if e.get("id") == eid:
                e.update({"name": name, "start": start, "end": end,
                          "boost": boost, "kind": kind})
                break
        else:
            return _json({"error": "id not found"}, 404)
    else:                                  # create
        new_id = f"c{int(time.time()*1000)}"
        _custom_events.append({"id": new_id, "name": name, "start": start,
                               "end": end, "boost": boost, "kind": kind})
    # invalidate forward cache so the calendar reflects the new event
    _forward_cache["data"] = None; _forward_cache["ts"] = 0
    await asyncio.to_thread(persist_state)
    log_event("pricing", f"حدث مخصص: {name} ({start} → {end}) × {boost}")
    return _json({"ok": True})

async def _api_events_delete(request):
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    eid = b.get("id")
    if not eid:
        return _json({"error": "id required"}, 400)
    before = len(_custom_events)
    _custom_events[:] = [e for e in _custom_events if e.get("id") != eid]
    if len(_custom_events) == before:
        return _json({"error": "id not found"}, 404)
    _forward_cache["data"] = None; _forward_cache["ts"] = 0
    await asyncio.to_thread(persist_state)
    return _json({"ok": True})

async def _api_units_list(request):
    """Return active units with the fields the bulk-apply filter needs."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    if not _catalog_units:
        await asyncio.to_thread(load_catalog, True)
    out = [{"id": u.get("id"), "name": u.get("name"), "beds": u.get("beds"),
            "area": u.get("area"), "neighbourhood": u.get("neighbourhood"),
            "capacity": u.get("capacity")}
           for u in _catalog_units if u.get("id")]
    # also surface available area + bedroom values for picker dropdowns
    beds = sorted({u["beds"] for u in out if u.get("beds")})
    areas = sorted({(u.get("neighbourhood") or u.get("area") or "").strip()
                    for u in out if (u.get("neighbourhood") or u.get("area"))})
    return _json({"units": out, "beds": beds, "areas": [a for a in areas if a]})

async def _api_calendar_forward(request):
    """Aggregate per-date forward calendar: occupancy + avg price + Saudi events."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    try:
        days = int(request.query.get("days", "60"))
    except Exception:
        days = 60
    days = max(7, min(120, days))
    data = await asyncio.to_thread(get_forward_calendar, days)
    return _json({"days": data, "events": SAUDI_EVENTS})

async def _api_pricing_bulk(request):
    """POST {start, end, percent, action:'raise'|'lower', only_available?:bool=True,
            lids?:[int], beds?:int, area?:str} — bulk adjust prices in a date range
    with optional filtering by specific unit ids, bedroom count, and/or area
    substring. Only adjusts currently-available unbooked nights. Honors
    PRICE_APPLY_DRYRUN. Writes the ouja-orig: note so discount tiers respect
    the new anchor."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    start = _parse_date(b.get("start"))
    end = _parse_date(b.get("end"))
    try:
        pct = float(b.get("percent", 0))
        action = (b.get("action") or "raise").lower()
    except Exception:
        return _json({"error": "bad params"}, 400)
    if not start or not end or end < start or pct <= 0 or pct > 80:
        return _json({"error": "bad date range or percent"}, 400)
    factor = (1 + pct / 100) if action == "raise" else (1 - pct / 100)
    only_available = b.get("only_available", True)
    # ---- apply filter to the listing set ----
    f_lids = b.get("lids") or []
    f_beds = b.get("beds")
    f_area = (b.get("area") or "").strip().lower()
    if not _catalog_units:
        await asyncio.to_thread(load_catalog, True)
    candidates = _catalog_units or []
    if f_lids:
        wanted = set(int(x) for x in f_lids)
        candidates = [u for u in candidates if u.get("id") in wanted]
    if f_beds:
        try:
            bv = int(f_beds)
            candidates = [u for u in candidates if u.get("beds") == bv]
        except Exception:
            pass
    if f_area:
        candidates = [u for u in candidates
                      if (u.get("area") or "").lower().find(f_area) >= 0
                      or (u.get("neighbourhood") or "").lower().find(f_area) >= 0]
    listings = [u["id"] for u in candidates if u.get("id")]
    if not listings:
        return _json({"error": "filter matched zero units"}, 400)

    def _adjust_one(lid):
        applied, skipped = 0, 0
        try:
            cal = api_get(f"/listings/{lid}/calendar",
                          params={"startDate": start.isoformat(), "endDate": end.isoformat()})
            for day in (cal.get("result") or []):
                d_iso = day.get("date")
                if not d_iso:
                    continue
                available = int(day.get("isAvailable", 0) or 0) == 1 and not day.get("reservationId")
                if only_available and not available:
                    skipped += 1
                    continue
                cur = day.get("price")
                if not isinstance(cur, (int, float)) or cur <= 0:
                    skipped += 1
                    continue
                new_price = int(round(float(cur) * factor))
                if new_price == int(cur):
                    skipped += 1
                    continue
                if not PRICE_APPLY_DRYRUN:
                    api_put(f"/listings/{lid}/calendar",
                            {"startDate": d_iso, "endDate": d_iso,
                             "isAvailable": 1, "price": new_price,
                             "note": f"ouja-orig:{new_price}"})
                applied += 1
        except Exception as e:
            print(f"bulk_apply error ({lid}):", e)
        return applied, skipped

    tot_applied, tot_skipped = 0, 0
    if listings:
        with ThreadPoolExecutor(max_workers=INTEL_PARALLEL) as ex:
            for f in as_completed([ex.submit(_adjust_one, lid) for lid in listings]):
                a, s = f.result()
                tot_applied += a; tot_skipped += s
    scope_label = (f"{len(listings)} وحدة" if f_lids or f_beds or f_area else "كل الوحدات")
    log_event("pricing",
              f"تطبيق جماعي {action} {pct}% · {start.isoformat()}→{end.isoformat()} · "
              f"{scope_label} · {tot_applied} ليلة" + (" (DRY-RUN)" if PRICE_APPLY_DRYRUN else ""))
    # Invalidate the forward-calendar cache so the page reflects new prices immediately
    _forward_cache["data"] = None
    _forward_cache["ts"] = 0
    return _json({"ok": True, "applied": tot_applied, "skipped": tot_skipped,
                  "dry_run": PRICE_APPLY_DRYRUN, "factor": factor})

async def _api_home_urgent(request):
    """Operational urgency feed for the home page — high-severity items first."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    items = await asyncio.to_thread(compute_urgent_now)
    return _json({"items": items, "counts": {
        "high": sum(1 for i in items if i.get("severity") == "high"),
        "med":  sum(1 for i in items if i.get("severity") == "med"),
        "low":  sum(1 for i in items if i.get("severity") == "low"),
    }})

async def _api_home_arrivals(request):
    """Next-N-hours arrivals with per-guest status (signed/code)."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    try:
        h = int(request.query.get("hours", "36"))
    except Exception:
        h = 36
    items = await asyncio.to_thread(compute_arrivals_with_status, max(6, min(72, h)))
    return _json({"items": items, "window_hours": h})

async def _api_metrics_daily(request):
    """Return up to N days of daily counters + derived rates so the dashboard can
    chart the assistant's improvement over time. ?days=30 by default."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    try:
        days = int(request.query.get("days", "30"))
    except Exception:
        days = 30
    days = max(1, min(120, days))
    today = datetime.now(TZ).date()
    out = []
    for i in range(days - 1, -1, -1):
        d = today - timedelta(days=i)
        k = d.isoformat()
        row = _daily_metrics.get(k, _new_day_row())
        total = row.get("replies_total", 0)
        manual_or_edit = row.get("replies_manual", 0) + row.get("replies_edited", 0) + row.get("replies_dashboard", 0)
        auto_rate = round((row.get("replies_auto", 0) / total) * 100) if total else 0
        edit_rate = round((row.get("replies_edited", 0) / manual_or_edit) * 100) if manual_or_edit else 0
        drafts = row.get("drafts_made", 0)
        esc_rate = round((row.get("escalations_created", 0) / drafts) * 100) if drafts else 0
        avg_conf = round((row["confidence_sum"] / row["confidence_count"]) * 100) if row.get("confidence_count") else 0
        out.append({
            "date": k,
            "replies_total": total,
            "replies_auto": row.get("replies_auto", 0),
            "replies_manual": row.get("replies_manual", 0),
            "replies_edited": row.get("replies_edited", 0),
            "replies_dashboard": row.get("replies_dashboard", 0),
            "drafts_made": drafts,
            "escalations_created": row.get("escalations_created", 0),
            "escalations_resolved": row.get("escalations_resolved", 0),
            "auto_rate": auto_rate,         # % of replies that auto-sent
            "edit_rate": edit_rate,         # % of human-touched replies that needed editing
            "escalation_rate": esc_rate,    # % of drafts that escalated
            "avg_confidence": avg_conf,     # avg confidence across drafts that day
            "apartments_touched": len(row.get("apartments_touched", [])),
            "topics": row.get("topics", {}),
        })
    return _json({"days": out})

async def _api_learning_today(request):
    """Return today's learning events grouped by apartment so the page can show
    'what changed today' instead of just the static summary text."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    try:
        window_days = int(request.query.get("days", "1"))
    except Exception:
        window_days = 1
    window_days = max(1, min(30, window_days))
    cutoff = datetime.now(TZ) - timedelta(days=window_days)
    cutoff_iso = cutoff.isoformat(timespec="seconds")
    by_apt = defaultdict(list)
    for e in reversed(_learning_log):
        ts = e.get("ts", "")
        if ts < cutoff_iso:
            break
        unit = e.get("unit") or "—"
        by_apt[unit].append({
            "ts": ts, "via": e.get("via", ""),
            "guest_question": (e.get("guest_question") or "")[:240],
            "final_reply": (e.get("final_reply") or "")[:240],
            "was_edited": e.get("was_edited", False),
            "diff_ratio": e.get("diff_ratio", 0),
        })
    items = sorted([{"unit": u, "count": len(es), "events": es} for u, es in by_apt.items()],
                   key=lambda x: -x["count"])
    return _json({"window_days": window_days, "apartments": items,
                  "total_events": sum(x["count"] for x in items)})

async def _api_learning_edit(request):
    """POST {scope:'apartment'|'general', lid?, summary} — directly overwrite
    a distilled summary. Useful when the owner wants to add a fact the bot
    hasn't learned yet, or fix something the distiller got wrong. The edited
    summary sticks until the next distill — which only fires when there are
    new examples — so an edit is effectively permanent unless real new traffic
    overwrites it."""
    if not _dash_auth(request):
        return _json({"error": "unauthorized"}, 401)
    b = await _read_body(request)
    scope = (b.get("scope") or "").strip()
    summary = (b.get("summary") or "").strip()
    if scope == "general":
        _general_learnings["summary"] = summary
        _general_learnings["last_distilled"] = time.time()
        if not _general_learnings.get("examples_count"):
            _general_learnings["examples_count"] = 1   # mark as set
        log_event("guest", "تعديل يدوي للملخص العام")
        await asyncio.to_thread(persist_state)
        return _json({"ok": True})
    if scope == "apartment":
        try:
            lid = int(b.get("lid"))
        except Exception:
            return _json({"error": "bad lid"}, 400)
        existing = _apartment_learnings.get(lid) or {}
        unit_name = existing.get("unit") or next(
            (u["name"] for u in _catalog_units if u.get("id") == lid), str(lid))
        _apartment_learnings[lid] = {
            "summary": summary,
            "last_distilled": time.time(),
            "examples_count": existing.get("examples_count") or 1,
            "unit": unit_name,
        }
        log_event("guest", f"تعديل يدوي لملخص {unit_name}")
        await asyncio.to_thread(persist_state)
        return _json({"ok": True})
    return _json({"error": "scope must be apartment|general"}, 400)

@tasks.loop(minutes=DASH_REFRESH_MIN)
async def dashboard_cache_loop():
    """Pre-compute heavy analytics in the background so the dashboard serves instantly."""
    if not (DASHBOARD_ENABLED and DASHBOARD_TOKEN):
        return
    try:
        if not _catalog_units:
            await asyncio.to_thread(load_catalog, True)
        ov = await asyncio.to_thread(_compute_overview)
        if ov.get("active_units") or ov.get("rev_30") or ov.get("rev_7"):  # don't cache empty
            _dash_cache["overview"] = (ov, time.time())
        rv = await asyncio.to_thread(_compute_revenue)
        if rv.get("monthly") or rv.get("units"):
            _dash_cache["revenue"] = (rv, time.time())
        td = await asyncio.to_thread(_compute_today)
        if td.get("active"):
            _dash_cache["today"] = (td, time.time())
        last = _dash_cache.get("pricing")
        if (not last) or (time.time() - last[1] > 1800):   # pricing is calendar-heavy: every ~30 min
            pr = await asyncio.to_thread(_compute_pricing)
            _dash_cache["pricing"] = (pr, time.time())
        print(f"dashboard cache warmed · units={ov.get('active_units')} rev30={ov.get('rev_30')}")
    except Exception as e:
        print("dashboard cache error:", e)

async def start_web_server():
    """Run a tiny HTTP server so Hostaway can push new-message events to us."""
    global _web_runner
    if _web_runner is not None or not _HAS_AIOHTTP:
        return
    app = web.Application()
    app.router.add_get("/", _handle_health)                 # health check / browser test
    app.router.add_post("/hook/{secret}", _handle_hook)     # Hostaway posts here
    app.router.add_get("/hook/{secret}", _handle_health)    # so you can open it in a browser
    if DASHBOARD_ENABLED:
        app.router.add_get("/dashboard", _handle_dashboard)
        app.router.add_get("/api/overview", _api_overview)
        app.router.add_get("/api/revenue", _api_revenue)
        app.router.add_get("/api/pricing", _api_pricing)
        app.router.add_get("/api/log", _api_log)
        app.router.add_get("/api/inbox", _api_inbox)
        app.router.add_get("/api/autolog", _api_autolog)
        app.router.add_get("/api/today", _api_today)
        app.router.add_get("/api/pricing/detail", _api_pricing_detail)
        app.router.add_get("/api/strategy", _api_strategy)
        app.router.add_get("/api/strategies", _api_strategies)
        app.router.add_post("/api/strategy/stop", _api_strategy_stop)
        app.router.add_get("/api/discount/status", _api_discount_status)
        app.router.add_post("/api/discount/pause", _api_discount_pause)
        app.router.add_post("/api/discount/resume", _api_discount_resume)
        app.router.add_post("/api/discount/skip-unit", _api_unit_skip)
        app.router.add_post("/api/discount/unskip-unit", _api_unit_unskip)
        app.router.add_get("/api/today/empty", _api_today_empty)
        app.router.add_get("/api/inbox/detail", _api_inbox_detail)
        app.router.add_post("/api/teach", _api_teach)
        app.router.add_get("/api/learning/summary", _api_learning_summary)
        app.router.add_get("/api/learning/log", _api_learning_log)
        app.router.add_post("/api/learning/forget", _api_learning_forget)
        app.router.add_post("/api/learning/distill", _api_learning_distill_now)
        app.router.add_post("/api/learning/edit", _api_learning_edit)
        app.router.add_post("/api/learning/bootstrap", _api_learning_bootstrap)
        app.router.add_get("/api/learning/bootstrap/status", _api_learning_bootstrap_status)
        app.router.add_get("/api/metrics/daily", _api_metrics_daily)
        app.router.add_get("/api/learning/today", _api_learning_today)
        app.router.add_get("/api/home/urgent", _api_home_urgent)
        app.router.add_get("/api/home/arrivals", _api_home_arrivals)
        app.router.add_get("/api/calendar/forward", _api_calendar_forward)
        app.router.add_post("/api/pricing/bulk", _api_pricing_bulk)
        app.router.add_get("/api/events", _api_events_list)
        app.router.add_post("/api/events/save", _api_events_save)
        app.router.add_post("/api/events/delete", _api_events_delete)
        app.router.add_get("/api/units", _api_units_list)
        app.router.add_get("/api/guests", _api_guests_list)
        app.router.add_get("/api/cleaning/quality", _api_clean_quality_summary)
        # Public no-auth feedback page + endpoints (token IS the auth)
        app.router.add_get("/clean-feedback", _handle_clean_feedback_page)
        app.router.add_get("/api/clean-feedback", _api_clean_feedback_get)
        app.router.add_post("/api/clean-feedback", _api_clean_feedback_submit)
        app.router.add_get("/api/guests/detail", _api_guest_detail)
        app.router.add_post("/api/guests/notes", _api_guest_notes)
        app.router.add_post("/api/guests/toggle-vip", _api_guest_toggle_vip)
        app.router.add_get("/api/cleaning/schedule", _api_cleaning_schedule)
        app.router.add_post("/api/cleaning/mark-done", _api_cleaning_mark_done)
        app.router.add_post("/api/cleaning/reschedule", _api_cleaning_reschedule)
        app.router.add_post("/api/cleaning/set-last", _api_cleaning_set_last)
        app.router.add_post("/api/cleaning/import-csv", _api_cleaning_import_csv)
        app.router.add_post("/api/cleaning/import-xlsx", _api_cleaning_import_xlsx)
        # Public cleaning-company page (gated by CLEANING_TOKEN, not the dashboard token)
        app.router.add_get("/cleaning", _handle_cleaning_page)
        app.router.add_get("/api/cleaning/public", _api_cleaning_public)
        app.router.add_post("/api/send", _api_send)
        app.router.add_post("/api/reject", _api_reject)
        app.router.add_post("/api/claim", _api_claim)
        app.router.add_post("/api/apply", _api_apply)
    _web_runner = web.AppRunner(app)
    await _web_runner.setup()
    site = web.TCPSite(_web_runner, "0.0.0.0", WEB_PORT)
    await site.start()
    print(f"web server listening on :{WEB_PORT}  (webhook path: /hook/{WEBHOOK_SECRET})")

def _as_int(v):
    """Hostaway sometimes returns message/reservation IDs as strings; coerce safely.
    Falsy or unparseable values become 0 so '<=' comparisons stay valid."""
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0

def _cohost_responded(esc):
    """True if a human teammate replied to the guest directly since the escalation —
    an outbound message NEWER than the trigger that we didn't send ourselves."""
    cid = esc.get("conversation_id")
    if not cid:
        return False
    try:
        msgs = sorted((api_get(f"/conversations/{cid}/messages").get("result") or []),
                      key=lambda m: _as_int(m.get("id")))
    except Exception:
        return False
    base = _as_int(esc.get("last_msg_id"))
    acks = esc.get("acks", [])
    for m in msgs:
        if _as_int(m.get("id")) <= base:
            continue                                   # not newer than the escalation trigger
        if _msg_is_inbound(m):
            continue                                   # guest's own message
        body = (m.get("body") or "").strip()
        if not body or _looks_automated(body):
            continue
        if any(a and a[:40] in body for a in acks):
            continue                                   # this is one of our own ack messages
        return True                                    # a human/co-host reply we didn't send
    return False

async def _resolve_escalation(mid, esc, reason="رد عليه أحد المضيفين"):
    """Stop nagging: mark resolved, free the conversation, and update the Discord card."""
    cid = esc.get("conversation_id")
    if cid:
        _claimed_convos.add(cid)
    _escalations.pop(mid, None)
    metric_bump("escalations_resolved")
    log_event("escalation", f"أُغلق تلقائياً ({reason}) · {esc.get('guest','')} · {esc.get('unit','')}")
    try:
        ch = bot.get_channel(esc.get("channel_id"))
        if ch:
            msg = await ch.fetch_message(mid)
            done = ClaimView()
            for c in done.children:
                c.disabled = True
            emb = msg.embeds[0] if msg.embeds else discord.Embed()
            emb.color = 0x3BA55D
            emb.add_field(name="✅ تم الإغلاق تلقائياً", value=reason, inline=False)
            await msg.edit(content=f"✅ تم التعامل معه — {reason}", embed=emb, view=done)
    except Exception as e:
        print("resolve-escalation edit error:", e)

@tasks.loop(minutes=1)
async def escalation_reping_loop():
    """Re-ping the operation team about any escalation that hasn't been claimed yet —
    but first auto-resolve any escalation where a co-host already replied (and stop nagging)."""
    if not _escalations:
        return
    guild = bot.get_guild(GUILD_ID)
    if guild is None:
        return
    op_role = find_operation_role(guild)
    mention = op_role.mention if op_role else f"@{OPERATION_ROLE_NAME}"
    now = time.time()
    for mid, esc in list(_escalations.items()):
        if esc.get("claimed_by"):
            continue
        # --- auto-close if a teammate already answered the guest directly ---
        if await asyncio.to_thread(_cohost_responded, esc):
            await _resolve_escalation(mid, esc)
            continue
        if esc["attempts"] >= ESCALATION_MAX_PINGS:
            # we've nagged the max number of times and nobody picked it up — stop
            # silently sitting on the dashboard counter; mark it as exhausted so the
            # KPI reflects reality and the card carries a clear visual state.
            await _resolve_escalation(mid, esc, reason=f"انتهت محاولات التنبيه ({ESCALATION_MAX_PINGS}× بدون استلام)")
            continue
        if (now - esc["last_ping"]) < ESCALATION_REPING_MIN * 60:
            continue
        ch = bot.get_channel(esc["channel_id"])
        if ch is None:
            continue
        try:
            ref = discord.MessageReference(message_id=mid, channel_id=esc["channel_id"],
                                           fail_if_not_exists=False)
            await ch.send(f"{mention} ⏰ لسه ما تم استلام التصعيد — **{esc['guest']} · {esc['unit']}**. "
                          f"اضغط 🙋 أخذ المهمة.",
                          reference=ref, allowed_mentions=discord.AllowedMentions(roles=True))
            esc["last_ping"] = now
            esc["attempts"] += 1
        except Exception as e:
            print("reping error:", e)

@tasks.loop(minutes=KNOWLEDGE_REFRESH_MIN)
async def knowledge_loop():
    guild = bot.get_guild(GUILD_ID)
    if guild is not None:
        await load_knowledge(guild)
    await asyncio.to_thread(load_catalog)   # refreshes only if >1h old

@tasks.loop(minutes=LEARNING_DISTILL_MIN)
async def learning_distillation_loop():
    """Re-summarise recent send/edit events per-apartment + general. Cheap if no
    new entries since the last distill (LEARNING_MIN_NEW_EXAMPLES gate)."""
    try:
        await asyncio.to_thread(distill_learnings)
    except Exception as e:
        print("learning_distillation_loop error:", e)

@tasks.loop(minutes=AGREEMENT_REMINDER_POLL_MIN)
async def agreement_reminder_loop():
    """Every N minutes, check today's arrivals: if check-in is within the lead-time
    window and the agreement isn't signed yet, re-send the signing link with an
    explanation that the door code is gated on the signature."""
    try:
        await asyncio.to_thread(check_agreement_reminders)
    except Exception as e:
        print("agreement_reminder_loop error:", e)

@tasks.loop(hours=4)
async def deepclean_schedule_loop():
    """Top up the deep-clean schedule for any unit that doesn't have a next date."""
    try:
        await asyncio.to_thread(schedule_deep_cleans)
    except Exception as e:
        print("deepclean_schedule_loop error:", e)

@tasks.loop(time=dt_time(hour=DEEPCLEAN_CONFIRM_HOUR, tzinfo=TZ))
async def deepclean_confirm_loop():
    """At 9pm Riyadh: confirm tomorrow's deep cleans (block calendar) or push them."""
    try:
        await asyncio.to_thread(confirm_tomorrow_deepcleans)
    except Exception as e:
        print("deepclean_confirm_loop error:", e)

@tasks.loop(time=dt_time(hour=WORK_START_HOUR, tzinfo=TZ))
async def offhours_ack_reset_loop():
    """At the start of every working day, clear the set of conversations that already
    got an off-hours auto-ack so they can be re-acked next time we go offline."""
    _offhours_acked_convos.clear()

@tasks.loop(time=dt_time(hour=WILT_HOUR, tzinfo=TZ))
async def wilt_loop():
    """Post the daily WILT (What I Learned Today) at WILT_HOUR Riyadh."""
    try:
        await post_wilt()
    except Exception as e:
        print("wilt_loop error:", e)

@tasks.loop(time=dt_time(hour=WORK_START_HOUR, minute=2, tzinfo=TZ))
async def morning_escalation_reminder_loop():
    """First thing each working day, post a summary of every unclaimed escalation
    into #escalations so the team starts the day looking at the right stack."""
    try:
        guild = bot.get_guild(GUILD_ID)
        if guild is None:
            return
        open_escs = [(mid, e) for mid, e in _escalations.items() if not e.get("claimed_by")]
        if not open_escs:
            return
        category = await get_assistant_category(guild)
        ch = await ensure_channel(guild, ESCALATION_CHANNEL, category)
        if ch is None:
            return
        op_role = find_operation_role(guild)
        mention = op_role.mention if op_role else f"@{OPERATION_ROLE_NAME}"
        lines = []
        for mid, e in open_escs[:25]:
            age_min = int((time.time() - (e.get("last_ping") or time.time())) / 60)
            lines.append(f"• **{e.get('guest','ضيف')}** · {e.get('unit','')} · من {age_min} دقيقة")
        embed = discord.Embed(
            title="☀ صباح الخير — تصعيدات بانتظار الاستلام",
            description=f"{len(open_escs)} تصعيد مفتوح. تحت قائمة بأول {min(25,len(open_escs))}:\n\n"
                        + "\n".join(lines),
            color=GOLD,
        )
        embed.set_footer(text="افتح أي تصعيد فوق ↑ واضغط 🙋 أخذ المهمة")
        await ch.send(content=f"{mention} ☀", embed=embed,
                      allowed_mentions=discord.AllowedMentions(roles=True))
        log_event("escalation", f"تذكير الصباح · {len(open_escs)} تصعيد مفتوح")
    except Exception as e:
        print("morning_escalation_reminder_loop error:", e)

@tasks.loop(minutes=30)
async def cleaning_feedback_loop():
    """Twice an hour: check today's arrivals and dispatch a feedback request
    to any guest whose check-in passed >= CLEAN_FEEDBACK_DELAY_HOURS ago."""
    try:
        await asyncio.to_thread(check_cleaning_feedback_requests)
    except Exception as e:
        print("cleaning_feedback_loop error:", e)

@tasks.loop(minutes=GUEST_SUMMARY_REFRESH_MIN)
async def guest_summary_loop():
    """Walk recent conversations and append a Claude-generated 2-3 line summary
    to the matching guest profile, capped to last 5 summaries per profile."""
    if not ANTHROPIC_API_KEY or not _guest_profiles:
        return
    try:
        # Look at the last 50 send events to find recent conversation activity
        recent = list(_learning_log)[-50:]
        by_conv = {}
        for e in recent:
            cid = e.get("conversation_id")
            if cid:
                by_conv.setdefault(cid, []).append(e)
        for cid, events in by_conv.items():
            text_blob = "\n".join(
                "Guest: " + (e.get("guest_question","") or "") + "\nHost: " + (e.get("final_reply","") or "")
                for e in events[-6:]
            )
            # find the matching profile by guest name from the most recent event
            last = events[-1]
            name = last.get("guest", "")
            key = _profile_key(name)
            if not key or key not in _guest_profiles:
                continue
            p = _guest_profiles[key]
            # skip if we already summarised this conversation recently
            if any(s.get("conversation_id") == cid for s in p.get("summaries", [])):
                continue
            summary = await asyncio.to_thread(_summarise_conversation_for_profile, p, cid, text_blob)
            if summary:
                p.setdefault("summaries", []).append({
                    "ts": datetime.now(TZ).isoformat(timespec="minutes"),
                    "conversation_id": cid,
                    "text": summary.strip()[:600],
                })
                p["summaries"] = p["summaries"][-5:]
    except Exception as e:
        print("guest_summary_loop error:", e)

def load_state():
    """Restore in-memory state from the volume so nothing is lost across restarts/redeploys."""
    global _assistant_seen, _pending_replies, _escalations, _esc_ack_count, _claimed_convos
    global _price_opps, _discount_paused_until, _unit_discount_skip, _agreement_reminded
    try:
        _assistant_seen = _BoundedSet(_load_json("seen.json", []), maxlen=20000)
        _pending_replies = {int(k): v for k, v in _load_json("pending.json", {}).items()}
        _escalations = {int(k): v for k, v in _load_json("escalations.json", {}).items()}
        _esc_ack_count = {int(k): v for k, v in _load_json("ack_count.json", {}).items()}
        _claimed_convos = set(int(x) for x in _load_json("claimed.json", []))
        _activity.clear()
        _activity.extend(_load_json("activity.json", []))
        _auto_replies.clear()
        _auto_replies.extend(_load_json("auto_replies.json", []))
        _pricing_strategies.clear()
        _pricing_strategies.update({int(k): v for k, v in _load_json("strategies.json", {}).items()})
        _price_opps = {int(k): v for k, v in _load_json("price_opps.json", {}).items()}
        _discount_paused_until = float(_load_json("discount_pause.json", 0) or 0)
        _unit_discount_skip = {int(k): float(v) for k, v in _load_json("unit_discount_skip.json", {}).items()
                               if float(v) > time.time()}   # drop expired entries on boot
        _learning_log.clear()
        _learning_log.extend(_load_json("learning_log.json", []))
        _apartment_learnings.clear()
        _apartment_learnings.update({int(k): v for k, v in _load_json("apartment_learnings.json", {}).items()})
        _general_learnings.update(_load_json("general_learnings.json", {"summary": "", "last_distilled": 0, "examples_count": 0}))
        _agreement_reminded = set(int(x) for x in _load_json("agreement_reminded.json", []) if str(x).strip())
        _daily_metrics.clear()
        _daily_metrics.update(_load_json("daily_metrics.json", {}))
        _custom_events.clear()
        _custom_events.extend(_load_json("custom_events.json", []))
        _deep_clean_state.clear()
        _deep_clean_state.update({int(k): v for k, v in _load_json("deep_clean.json", {}).items()})
        _guest_profiles.clear()
        _guest_profiles.update(_load_json("guest_profiles.json", {}))
        _cleaning_feedback.clear()
        _cleaning_feedback.update(_load_json("cleaning_feedback.json", {}))
        _cleaning_feedback_sent.clear()
        _cleaning_feedback_sent.update(int(x) for x in _load_json("cleaning_feedback_sent.json", []) if str(x).strip())
        if _assistant_seen or _pending_replies or _escalations:
            print(f"state: restored {len(_assistant_seen)} seen · {len(_pending_replies)} cards · "
                  f"{len(_escalations)} escalations · {len(_claimed_convos)} claimed · "
                  f"{len(_price_opps)} price cards")
    except Exception as e:
        print("state load error:", e)

def persist_state():
    _save_json("seen.json", list(_assistant_seen))
    _save_json("pending.json", {str(k): v for k, v in _pending_replies.items()})
    _save_json("escalations.json", {str(k): v for k, v in _escalations.items()})
    _save_json("ack_count.json", {str(k): v for k, v in _esc_ack_count.items()})
    _save_json("claimed.json", list(_claimed_convos))
    _save_json("activity.json", list(_activity))
    _save_json("auto_replies.json", list(_auto_replies))
    _save_json("strategies.json", {str(k): v for k, v in _pricing_strategies.items()})
    _save_json("price_opps.json", {str(k): v for k, v in _price_opps.items()})
    _save_json("discount_pause.json", _discount_paused_until)
    _save_json("unit_discount_skip.json", {str(k): v for k, v in _unit_discount_skip.items()})
    _save_json("learning_log.json", list(_learning_log))
    _save_json("apartment_learnings.json", {str(k): v for k, v in _apartment_learnings.items()})
    _save_json("general_learnings.json", _general_learnings)
    _save_json("agreement_reminded.json", list(_agreement_reminded))
    # cap to last 120 days to keep the JSON small
    if len(_daily_metrics) > 120:
        keep = sorted(_daily_metrics.keys())[-120:]
        for k in list(_daily_metrics.keys()):
            if k not in keep:
                _daily_metrics.pop(k, None)
    _save_json("daily_metrics.json", _daily_metrics)
    _save_json("custom_events.json", _custom_events)
    _save_json("deep_clean.json", {str(k): v for k, v in _deep_clean_state.items()})
    _save_json("guest_profiles.json", _guest_profiles)
    _save_json("cleaning_feedback.json", _cleaning_feedback)
    _save_json("cleaning_feedback_sent.json", list(_cleaning_feedback_sent))

@tasks.loop(seconds=60)
async def persist_loop():
    await asyncio.to_thread(persist_state)

# ==================== Weekly Revenue Report (recommend-only) ====================
# Pulls full booking history from Hostaway and produces a weekly Discord report:
#  (1) per-unit raise/hold/lower recommendation (optimizing TOTAL REVENUE),
#  (2) which months pay best/worst, (3) the real intra-month "salary cycle" dip.
# It NEVER changes prices — it only recommends. All numbers are computed in code.

def fetch_all_reservations(max_pages=None):
    """Paginate the full reservation history from Hostaway."""
    max_pages = max_pages or REVENUE_MAX_PAGES
    out, offset, limit = [], 0, 100
    for _ in range(max_pages):
        try:
            data = api_get("/reservations", params={"limit": limit, "offset": offset})
        except Exception as e:
            print("reservations fetch error:", e)
            break
        rows = data.get("result", []) or []
        if not rows:
            break
        out.extend(rows)
        if len(rows) < limit:
            break
        offset += limit
    if REVENUE_DEBUG and out:
        print("reservation sample keys:", sorted(out[0].keys()))
    print(f"revenue: fetched {len(out)} reservations")
    return out

def _res_nights(r):
    n = r.get("nights")
    if isinstance(n, int) and n > 0:
        return n
    ci, co = _parse_date(r.get("arrivalDate")), _parse_date(r.get("departureDate"))
    return (co - ci).days if (ci and co and co > ci) else 0

def _res_revenue(r):
    """Best-effort gross booking revenue (consistent across channels). Tune via REVENUE_DEBUG."""
    for k in ("totalPrice", "baseRate", "ownerPayout", "hostPayout", "price"):
        v = r.get(k)
        if isinstance(v, (int, float)) and v > 0:
            return float(v)
    return 0.0

def _res_realized(r):
    return (r.get("status") or "").lower() in CONFIRMED_STATUSES

def _explode_nights(reservations):
    """Return (nights, arrivals): nights = list of (listing_id, date, nightly_price);
    arrivals = list of (listing_id, checkin_date)."""
    nights, arrivals = [], []
    for r in reservations:
        if not _res_realized(r):
            continue
        ci = _parse_date(r.get("arrivalDate"))
        n = _res_nights(r)
        if not ci or n <= 0:
            continue
        lid = r.get("listingMapId")
        nightly = _res_revenue(r) / n if n else 0
        arrivals.append((lid, ci))
        for i in range(n):
            nights.append((lid, ci + timedelta(days=i), nightly))
    return nights, arrivals

def _unit_reco(pace30, occ90, adr, adr_median):
    """Revenue-max heuristic -> (key, label_ar, reason_ar, suggested_pct)."""
    if pace30 is None:
        return ("watch", "🟡 راقب", "بيانات غير كافية", 0)
    p = round(pace30 * 100)
    if pace30 >= 0.85:
        return ("raise", "🔼 ارفع", f"يمتلئ بسرعة (محجوز {p}% من الـ٣٠ يوم الجاية) — فيه مجال ترفع", 12)
    if pace30 >= 0.65:
        return ("raise_small", "🔼 ارفع بسيط", f"طلب قوي (pace {p}%)", 6)
    if pace30 >= 0.40:
        if adr and adr_median and adr < 0.9 * adr_median and (occ90 or 0) >= 0.6:
            return ("raise_small", "🔼 ارفع بسيط", "سعرك أقل من متوسط المحفظة وإشغالك جيد", 5)
        return ("hold", "⏸️ ثابت", f"وضع متوازن (pace {p}%)", 0)
    if adr and adr_median and adr > 1.1 * adr_median:
        return ("lower", "🔽 خفّض", f"طلب ضعيف وسعرك أعلى من المتوسط (pace {p}%)", -12)
    return ("lower", "🔽 خفّض", f"طلب ضعيف للـ٣٠ يوم الجاية (pace {p}%)", -8)

_AR_MONTHS = ["", "يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
              "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]

def compute_revenue_report(reservations, listings_map):
    """Crunch the history into per-unit recs + seasonality + salary-cycle. Pure/no I/O."""
    today = datetime.now(TZ).date()
    w_start = today - timedelta(days=REVENUE_WINDOW_DAYS)
    n30 = today + timedelta(days=30)
    y_start = today - timedelta(days=365)
    nights, arrivals = _explode_nights(reservations)

    # ---- per-unit: trailing-window occupancy/ADR + forward 30-day pace ----
    sold90, rev90, booked30 = defaultdict(int), defaultdict(float), defaultdict(int)
    for lid, d, nightly in nights:
        if w_start <= d < today:
            sold90[lid] += 1
            rev90[lid] += nightly
        if today <= d < n30:
            booked30[lid] += 1
    units = []
    adrs = []
    for lid in set(list(sold90) + list(booked30)):
        s = sold90[lid]
        adr = (rev90[lid] / s) if s else None
        if adr:
            adrs.append(adr)
    adr_median = sorted(adrs)[len(adrs) // 2] if adrs else None
    all_lids = set(list(sold90) + list(booked30))
    for lid in all_lids:
        s = sold90[lid]
        occ90 = s / REVENUE_WINDOW_DAYS
        adr = (rev90[lid] / s) if s else None
        pace30 = booked30[lid] / 30
        revpar = (rev90[lid] / REVENUE_WINDOW_DAYS) if REVENUE_WINDOW_DAYS else 0
        key, label, reason, pct = _unit_reco(pace30, occ90, adr, adr_median)
        units.append({
            "lid": lid, "name": listings_map.get(lid) or f"unit-{lid}",
            "occ90": occ90, "adr": adr, "revpar": revpar, "pace30": pace30,
            "reco": key, "label": label, "reason": reason, "pct": pct,
        })
    units.sort(key=lambda u: (u["pace30"] or 0), reverse=True)

    # ---- seasonality: avg nightly (ADR) + volume by month-of-year ----
    m_rev, m_nights = defaultdict(float), defaultdict(int)
    for lid, d, nightly in nights:
        m_rev[d.month] += nightly
        m_nights[d.month] += 1
    total_n = sum(m_nights.values())
    overall_adr = (sum(m_rev.values()) / total_n) if total_n else 0
    months = []
    for m in range(1, 13):
        if m_nights[m]:
            adr_m = m_rev[m] / m_nights[m]
            months.append({"m": m, "name": _AR_MONTHS[m], "adr": adr_m,
                           "nights": m_nights[m],
                           "index": (adr_m / overall_adr) if overall_adr else 1})
    months.sort(key=lambda x: x["index"], reverse=True)

    # ---- salary cycle: arrivals per day-of-month (last 365d), normalized by occurrences ----
    occ_count = defaultdict(int)        # how many calendar dates had each day-of-month
    d = y_start
    while d < today:
        occ_count[d.day] += 1
        d += timedelta(days=1)
    arr_by_dom = defaultdict(int)
    for lid, ci in arrivals:
        if y_start <= ci < today:
            arr_by_dom[ci.day] += 1
    rates = {dom: (arr_by_dom[dom] / occ_count[dom]) for dom in range(1, 32) if occ_count.get(dom)}
    mean_rate = (sum(arr_by_dom.values()) / sum(occ_count[d] for d in range(1, 32) if occ_count.get(d))) \
        if arr_by_dom else 0
    dom_index = {dom: (rates[dom] / mean_rate if mean_rate else 1) for dom in rates}
    # find the weakest/strongest contiguous run within days 1..28 (avoid month-end sparsity:
    # months with 30/31 days mean day 29-31 have ~half the samples and skew the index)
    weak_days = [dom for dom in range(1, 29) if dom_index.get(dom, 1) < 0.92]
    strong_days = [dom for dom in range(1, 29) if dom_index.get(dom, 1) > 1.08]

    def _run(days):
        if not days:
            return None
        best = cur = [days[0]]
        for x in days[1:]:
            if x == cur[-1] + 1:
                cur.append(x)
            else:
                if len(cur) > len(best):
                    best = cur
                cur = [x]
        if len(cur) > len(best):
            best = cur
        return (best[0], best[-1])

    salary = {
        "dom_index": dom_index,
        "weak_window": _run(weak_days),
        "strong_window": _run(strong_days),
        "have_data": bool(arr_by_dom),
    }
    return {"units": units, "months": months, "salary": salary,
            "overall_adr": overall_adr, "total_nights": total_n,
            "active": len(all_lids)}

def _pct(x):
    return f"{round((x - 1) * 100):+d}%"

def build_revenue_embeds(rep):
    """Turn the computed report into Discord embeds + a detail CSV (bytes)."""
    units = rep["units"]
    raises = [u for u in units if u["reco"] in ("raise", "raise_small")]
    lowers = [u for u in units if u["reco"] == "lower"]
    holds = [u for u in units if u["reco"] == "hold"]
    port_occ = (sum(u["occ90"] for u in units) / len(units)) if units else 0
    port_pace = (sum((u["pace30"] or 0) for u in units) / len(units)) if units else 0

    e1 = discord.Embed(
        title="📊 تقرير الإيرادات الأسبوعي · عوجا",
        description=(f"الوحدات الفعّالة: **{rep['active']}** · إشغال آخر {REVENUE_WINDOW_DAYS} يوم: "
                     f"**{round(port_occ*100)}%** · متوسط الحجز للـ٣٠ يوم الجاية: **{round(port_pace*100)}%** · "
                     f"متوسط السعر/الليلة: **{round(rep['overall_adr'])} ر.س**\n"
                     f"الهدف: **أعلى إيراد إجمالي** · *توصيات فقط — أنت تقرر التطبيق*"),
        color=GOLD)
    if raises:
        txt = "\n".join(f"• **{u['name']}** {u['label']} ~{u['pct']:+d}% — {u['reason']}"
                        for u in raises[:12])
        e1.add_field(name=f"🔼 ارفع السعر ({len(raises)})", value=txt[:1024], inline=False)
    if lowers:
        txt = "\n".join(f"• **{u['name']}** {u['pct']:+d}% — {u['reason']}" for u in lowers[:12])
        e1.add_field(name=f"🔽 خفّض السعر ({len(lowers)})", value=txt[:1024], inline=False)
    e1.add_field(name="⏸️ ثابت", value=f"{len(holds)} وحدة وضعها متوازن", inline=False)
    e1.set_footer(text="التفاصيل الكاملة لكل وحدة في الملف المرفق · الأسعار قبل الضريبة ورسوم المنصة")

    e2 = discord.Embed(title="🗓️ أقوى وأضعف الشهور (حسب متوسط السعر)", color=GOLD)
    if rep["months"]:
        top = rep["months"][:3]
        bot_ = rep["months"][-3:]
        e2.add_field(name="الأقوى",
                     value="\n".join(f"• {m['name']}: {round(m['adr'])} ر.س/ليلة ({_pct(m['index'])} عن المعدل)"
                                     for m in top) or "—", inline=False)
        e2.add_field(name="الأضعف",
                     value="\n".join(f"• {m['name']}: {round(m['adr'])} ر.س/ليلة ({_pct(m['index'])} عن المعدل)"
                                     for m in reversed(bot_)) or "—", inline=False)
        e2.set_footer(text="ارفع الأسعار الأساسية في الشهور القوية، وخفّفها في الضعيفة")

    s = rep["salary"]
    e3 = discord.Embed(title="💸 دورة الراتب داخل الشهر", color=GOLD)
    if s["have_data"]:
        parts = []
        ww, sw = s["weak_window"], s["strong_window"]
        if ww:
            a, b = ww
            avgw = sum(s["dom_index"].get(d, 1) for d in range(a, b + 1)) / (b - a + 1)
            parts.append(f"📉 الطلب يضعف في أيام **{a}–{b}** ({_pct(avgw)} عن المعدل) — اقتراح: خفّض "
                         f"~{min(12, round((1-avgw)*100))}% على الليالي اللي تقع في هالأيام.")
        if sw:
            a, b = sw
            avgs = sum(s["dom_index"].get(d, 1) for d in range(a, b + 1)) / (b - a + 1)
            parts.append(f"📈 الطلب يرتفع حول أيام **{a}–{b}** ({_pct(avgs)} عن المعدل) — اقتراح: ثبّت أو "
                         f"ارفع ~{min(12, round((avgs-1)*100))}% على هالأيام.")
        parts.append("ملاحظتك كانت: ضعف ٢٠–٢٤ والراتب يوم ٢٥ — فوق تشوف اللي طلعته بياناتك فعلياً.")
        e3.description = "\n\n".join(parts)
    else:
        e3.description = "ما فيه بيانات وصول كافية في آخر سنة لتحليل دورة الراتب بعد."

    # ---- detail CSV (every unit) ----
    rows = ["unit,occupancy_90d_%,ADR_SAR,RevPAR_SAR,pace_30d_%,recommendation,suggested_%"]
    for u in sorted(rep["units"], key=lambda x: (x["reco"] != "raise", x["name"])):
        rows.append(",".join([
            '"' + str(u["name"]).replace('"', "'") + '"',
            str(round(u["occ90"] * 100)),
            str(round(u["adr"])) if u["adr"] else "",
            str(round(u["revpar"])),
            str(round((u["pace30"] or 0) * 100)),
            u["reco"], str(u["pct"]),
        ]))
    csv_bytes = ("\n".join(rows)).encode("utf-8-sig")
    return [e1, e2, e3], csv_bytes

async def post_revenue_report():
    if not ASSISTANT_ENABLED:
        pass  # report doesn't require the assistant; keep running regardless
    guild = bot.get_guild(GUILD_ID)
    if guild is None:
        return
    channel = await ensure_channel(guild, REVENUE_CHANNEL, await get_assistant_category(guild))
    if channel is None:
        return
    reservations = await asyncio.to_thread(fetch_all_reservations)
    if not reservations:
        await channel.send("⚠️ ما قدرت أجيب بيانات الحجوزات من Hostaway لهذا التقرير.")
        return
    listings_map = await asyncio.to_thread(get_listings_map)
    rep = await asyncio.to_thread(compute_revenue_report, reservations, listings_map)
    embeds, csv_bytes = build_revenue_embeds(rep)
    file = discord.File(io.BytesIO(csv_bytes), filename="ouja_revenue_detail.csv")
    await channel.send(embeds=embeds, file=file)
    print(f"revenue: posted weekly report ({rep['active']} units)")
    log_event("report", f"تقرير الإيرادات الأسبوعي · {rep['active']} وحدة")

_revenue_last_date = None

@tasks.loop(minutes=30)
async def revenue_loop():
    """Fire once on the configured weekday + hour (Riyadh)."""
    global _revenue_last_date
    now = datetime.now(TZ)
    if now.weekday() != REVENUE_REPORT_DOW or now.hour != REVENUE_REPORT_HOUR:
        return
    if _revenue_last_date == now.date():
        return
    _revenue_last_date = now.date()
    try:
        await post_revenue_report()
    except Exception as e:
        print("revenue_loop error:", e)

# ==================== Per-date pricing opportunities + last-week review ====================
_res_cache = {"data": None, "ts": 0}

def get_reservations_cached(ttl=1800):
    """Cache the full reservation pull so multiple reports in a short window reuse it."""
    if _res_cache["data"] is not None and (time.time() - _res_cache["ts"]) < ttl:
        return _res_cache["data"]
    data = fetch_all_reservations()
    _res_cache["data"], _res_cache["ts"] = data, time.time()
    return data

def fetch_calendar_days(listing_id, start, end):
    try:
        data = api_get(f"/listings/{listing_id}/calendar",
                       params={"startDate": start.isoformat(), "endDate": end.isoformat()})
        return data.get("result", []) or []
    except Exception as e:
        print(f"calendar fetch error ({listing_id}):", e)
        return []

def compute_demand_factors(reservations):
    """Build pricing multipliers from real history: per-unit achieved ADR + month / day-of-month
    / day-of-week demand indices (all relative to the portfolio average)."""
    today = datetime.now(TZ).date()
    y_start = today - timedelta(days=365)
    nights, arrivals = _explode_nights(reservations)
    u_rev, u_n = defaultdict(float), defaultdict(int)
    m_rev, m_n = defaultdict(float), defaultdict(int)
    w_rev, w_n = defaultdict(float), defaultdict(int)
    for lid, d, nightly in nights:
        u_rev[lid] += nightly
        u_n[lid] += 1
        m_rev[d.month] += nightly
        m_n[d.month] += 1
        w_rev[d.weekday()] += nightly
        w_n[d.weekday()] += 1
    unit_adr = {lid: (u_rev[lid] / u_n[lid]) for lid in u_n if u_n[lid]}
    tot_n = sum(u_n.values())
    overall = (sum(u_rev.values()) / tot_n) if tot_n else 0
    month_index = {m: ((m_rev[m] / m_n[m]) / overall if overall and m_n[m] else 1) for m in range(1, 13)}
    dow_index = {w: ((w_rev[w] / w_n[w]) / overall if overall and w_n[w] else 1) for w in range(7)}
    occ_count = defaultdict(int)
    dd = y_start
    while dd < today:
        occ_count[dd.day] += 1
        dd += timedelta(days=1)
    arr = defaultdict(int)
    for lid, ci in arrivals:
        if y_start <= ci < today:
            arr[ci.day] += 1
    denom = sum(occ_count[d] for d in range(1, 32) if occ_count.get(d))
    mean = (sum(arr.values()) / denom) if denom else 0
    dom_index = {d: ((arr[d] / occ_count[d]) / mean if mean and occ_count.get(d) else 1)
                 for d in range(1, 32)}
    return {"unit_adr": unit_adr, "overall_adr": overall, "month_index": month_index,
            "dow_index": dow_index, "dom_index": dom_index, "unit_nights": dict(u_n)}

def compute_price_opportunities(factors, units, horizon=None):
    """For each active unit's UNBOOKED future nights, compute a demand-informed target price
    and compare to the current calendar price. Returns (per_unit, all_rows)."""
    horizon = horizon or PRICE_OPP_HORIZON
    today = datetime.now(TZ).date()
    end = today + timedelta(days=horizon)
    overall = factors["overall_adr"]
    per_unit = {}
    all_rows = []
    for u in units:
        lid, name = u.get("id"), u.get("name")
        if not lid:
            continue
        base = factors["unit_adr"].get(lid) or u.get("price") or overall
        if not base:
            continue
        pu = per_unit.setdefault(lid, {"name": name, "raise": [], "drop": [], "uplift": 0.0,
                                       "base": round(base), "n_hist": factors.get("unit_nights", {}).get(lid, 0),
                                       "checked": 0, "gap_sum": 0.0})
        for day in fetch_calendar_days(lid, today, end):
            d = _parse_date(day.get("date"))
            if not d:
                continue
            available = int(day.get("isAvailable", 0) or 0) == 1 and not day.get("reservationId")
            if not available:
                continue                                   # booked/blocked -> can't reprice
            pu["checked"] += 1
            cur = day.get("price")
            cur = float(cur) if isinstance(cur, (int, float)) and cur > 0 else None
            mi = factors["month_index"].get(d.month, 1)
            di = factors["dom_index"].get(d.day, 1)
            wi = factors["dow_index"].get(d.weekday(), 1)
            ev = event_boost_for_date(d)
            target = max(0.6 * base, min(2.2 * base, base * mi * di * wi * ev))
            lead = (d - today).days
            clear = target * (0.88 if lead <= 7 else (0.95 if lead <= 14 else 1.0))
            ev_names = [e["name"] for e in events_for_date(d)]
            row = {"lid": lid, "name": name, "date": d.isoformat(), "wd": d.weekday(),
                   "current": round(cur) if cur else None, "target": round(target),
                   "clear": round(clear), "lead": lead, "reco": "ok",
                   "event": ev_names[0] if ev_names else None, "event_boost": round(ev, 2)}
            if cur and cur < 0.90 * target:
                row["reco"] = "raise"
                pu["uplift"] += (target - cur)
                pu["gap_sum"] += (target - cur) / cur
                pu["raise"].append(row)
            elif cur and lead <= 10 and cur > 1.12 * clear:
                row["reco"] = "drop"
                pu["drop"].append(row)
            all_rows.append(row)
    # ---- neutral confidence score per unit (data volume + signal strength + corroboration) ----
    for pu in per_unit.values():
        nr, nd = len(pu["raise"]), len(pu["drop"])
        avg_gap = (pu["gap_sum"] / nr) if nr else 0.0          # average price gap (fraction)
        data_conf = min(1.0, pu["n_hist"] / 60.0)             # 60+ nights of history saturates
        signal_conf = max(min(1.0, avg_gap / 0.30), 0.6 if nd else 0.0)
        count_conf = min(1.0, (nr + nd) / 12.0)
        conf = 100 * (0.45 * data_conf + 0.35 * signal_conf + 0.20 * count_conf)
        pu["confidence"] = int(max(40, min(90, round(conf))))
        pu["avg_gap"] = avg_gap
    return per_unit, all_rows

_AR_WD = ["الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]

# pending price-change cards: discord_message_id -> {listing_id, name, changes:[{date,price,kind}]}
_price_opps = {}

def apply_price_changes(listing_id, changes):
    """Write the approved nightly prices to the Hostaway calendar. Re-verifies the dates are
    still available first (one range read), then PUTs each. Honors PRICE_APPLY_DRYRUN.
    Returns (applied, skipped, results) where results = [{date, kind, price, status}]."""
    if not changes:
        return (0, 0, [])
    dates = sorted(d for d in (_parse_date(c["date"]) for c in changes) if d)
    available = set()
    try:
        cal = api_get(f"/listings/{listing_id}/calendar",
                      params={"startDate": dates[0].isoformat(), "endDate": dates[-1].isoformat()})
        for day in (cal.get("result") or []):
            if int(day.get("isAvailable", 0) or 0) == 1 and not day.get("reservationId"):
                dd = _parse_date(day.get("date"))
                if dd:
                    available.add(dd.isoformat())
    except Exception as e:
        print("price re-verify error:", e)
        available = None                       # couldn't verify -> best effort, apply anyway
    applied, skipped, results = 0, 0, []
    for c in changes:
        price = int(round(c["price"]))
        if available is not None and c["date"] not in available:
            skipped += 1                        # got booked / blocked since the report
            results.append({"date": c["date"], "kind": c.get("kind"), "price": price, "status": "booked"})
            continue
        if PRICE_APPLY_DRYRUN:
            print(f"[DRY-RUN] would set {listing_id} {c['date']} -> {price} ر.س")
            applied += 1
            results.append({"date": c["date"], "kind": c.get("kind"), "price": price, "status": "dry"})
            continue
        try:
            api_put(f"/listings/{listing_id}/calendar",
                    {"startDate": c["date"], "endDate": c["date"],
                     "isAvailable": 1, "price": price,
                     "note": f"ouja-orig:{price}"})   # anchor for the discount tiers later
            applied += 1
            results.append({"date": c["date"], "kind": c.get("kind"), "price": price, "status": "applied"})
        except Exception as e:
            print(f"apply price error ({listing_id} {c['date']}):", e)
            skipped += 1
            results.append({"date": c["date"], "kind": c.get("kind"), "price": price, "status": "error"})
    return applied, skipped, results

def _unit_changes(pu):
    """Flatten a unit's raise/drop rows into apply-ready change dicts."""
    out = []
    for r in pu["raise"]:
        out.append({"date": r["date"], "price": r["target"], "kind": "raise"})
    for r in pu["drop"]:
        out.append({"date": r["date"], "price": r["clear"], "kind": "drop"})
    return out

def build_price_opp_summary(per_unit, all_rows):
    total_uplift = round(sum(p["uplift"] for p in per_unit.values()))
    n_raise = sum(len(p["raise"]) for p in per_unit.values())
    n_drop = sum(len(p["drop"]) for p in per_unit.values())
    e = discord.Embed(
        title="💰 فرص التسعير لكل ليلة · عوجا",
        description=(f"فحصت الليالي المتاحة للـ{PRICE_OPP_HORIZON} يوم الجاية.\n"
                     f"🔼 ليالي ناقصة تسعير: **{n_raise}** · 🔽 ليالي قريبة تحتاج تخفيض لتمتلئ: **{n_drop}**\n"
                     f"💵 إيراد إضافي تقديري لو رفعت الناقصة: **~{total_uplift} ر.س**\n"
                     f"تحت تلقى بطاقة لكل وحدة — اضغط **✅ طبّق** عشان أغيّر الأسعار فعلياً، أو **❌ تجاهل**."
                     + ("\n⚠️ **وضع التجربة (DRY-RUN)** شغّال — ما راح يتغيّر شي فعلي."
                        if PRICE_APPLY_DRYRUN else "")),
        color=GOLD)
    out = ["unit,date,weekday,current_SAR,target_SAR,clearing_SAR,lead_days,reco"]
    for r in all_rows:
        if r["reco"] == "ok":
            continue
        out.append(",".join(['"' + str(r["name"]).replace('"', "'") + '"', r["date"],
                              str(r["wd"]), str(r["current"] or ""), str(r["target"]),
                              str(r["clear"]), str(r["lead"]), r["reco"]]))
    return e, ("\n".join(out)).encode("utf-8-sig")

def _conf_bar(pct):
    filled = round(pct / 10)
    return "█" * filled + "░" * (10 - filled)

def build_unit_card(pu):
    """An embed for one unit's opportunities (shown with Apply/Skip buttons)."""
    nr, nd = len(pu["raise"]), len(pu["drop"])
    conf = pu.get("confidence", 50)
    desc = []
    if nr:
        desc.append(f"🔼 **رفع {nr} ليلة** (إيراد إضافي ~{round(pu['uplift'])} ر.س)")
    if nd:
        desc.append(f"🔽 **خفض {nd} ليلة** (ليالي قريبة فاضية، عشان تمتلئ)")
    e = discord.Embed(title=str(pu["name"])[:256], description="\n".join(desc), color=GOLD)
    lines = []
    for r in sorted(pu["raise"], key=lambda x: (x["target"] - (x["current"] or 0)), reverse=True)[:8]:
        lines.append(f"🔼 {r['date']} ({_AR_WD[r['wd']]}): {r['current']} → ~{r['target']} ر.س")
    if nr > 8:
        lines.append(f"… +{nr-8} ليالي رفع أخرى")
    for r in sorted(pu["drop"], key=lambda x: x["lead"])[:5]:
        lines.append(f"🔽 {r['date']} ({_AR_WD[r['wd']]}, بعد {r['lead']} يوم): {r['current']} → ~{r['clear']} ر.س")
    if nd > 5:
        lines.append(f"… +{nd-5} ليالي خفض أخرى")
    if lines:
        e.add_field(name="التفاصيل", value="\n".join(lines)[:1024], inline=False)
    # ---- neutral "why" / facts ----
    facts = [f"• متوسط سعرك الفعلي تاريخياً: ~{pu.get('base', 0)} ر.س",
             f"• فحصت {pu.get('checked', 0)} ليلة متاحة في الـ{PRICE_OPP_HORIZON} يوم الجاية"]
    if nr:
        facts.append(f"• {nr} ليلة سعرها الحالي أقل من المتوقّع (فرق متوسط +{round(pu.get('avg_gap',0)*100)}%)")
    if nd:
        facts.append(f"• {nd} ليلة قريبة لا زالت فاضية وسعرها أعلى من سعر التصريف")
    facts.append("• الأساس: متوسط سعر وحدتك × الشهر × اليوم بالشهر × يوم الأسبوع — من بياناتك أنت")
    e.add_field(name="📊 ليه؟ (حقائق)", value="\n".join(facts)[:1024], inline=False)
    e.add_field(name="🎯 نسبة الثقة في التوصية",
                value=f"`{_conf_bar(conf)}` **{conf}%**\nمحسوبة من حجم بياناتك + قوة فرق السعر — مو ضمان بيع.",
                inline=False)
    e.set_footer(text="✅ طبّق = أغيّر هالأسعار في تقويمك · الأسعار قبل الضريبة ورسوم المنصة")
    return e

async def post_price_opportunities():
    guild = bot.get_guild(GUILD_ID)
    if guild is None:
        return
    channel = await ensure_channel(guild, PRICE_OPP_CHANNEL, await get_assistant_category(guild))
    if channel is None:
        return
    if not _catalog_units:
        await asyncio.to_thread(load_catalog, True)
    reservations = await asyncio.to_thread(get_reservations_cached)
    if not reservations or not _catalog_units:
        await channel.send("⚠️ ما قدرت أجهّز فرص التسعير — بيانات الحجوزات أو قائمة الوحدات ناقصة.")
        return
    factors = await asyncio.to_thread(compute_demand_factors, reservations)
    per_unit, rows = await asyncio.to_thread(compute_price_opportunities, factors, _catalog_units)
    summary, csv_bytes = build_price_opp_summary(per_unit, rows)
    file = discord.File(io.BytesIO(csv_bytes), filename="ouja_price_opportunities.csv")
    await channel.send(embed=summary, file=file)
    # one action card per unit (top by uplift), each with Apply/Skip buttons
    ranked = sorted([p for p in per_unit.values() if p["raise"] or p["drop"]],
                    key=lambda p: p["uplift"], reverse=True)
    posted = 0
    for pu in ranked[:PRICE_OPP_MAX_CARDS]:
        changes = _unit_changes(pu)
        if not changes:
            continue
        msg = await channel.send(embed=build_unit_card(pu), view=PriceApplyView())
        _price_opps[msg.id] = {"listing_id": pu["lid"], "name": pu["name"], "changes": changes}
        posted += 1
    print(f"price-opp: posted summary + {posted} action cards ({len(per_unit)} units checked)")
    log_event("pricing", f"فرص التسعير · {posted} بطاقة وحدة")

def compute_last_week(reservations, listings_map, factors):
    """Actual performance of the last 7 days vs the prior 7, per unit + portfolio."""
    today = datetime.now(TZ).date()
    w0, wp = today - timedelta(days=7), today - timedelta(days=14)
    nights, _ = _explode_nights(reservations)
    cur = defaultdict(lambda: {"n": 0, "rev": 0.0})
    prev = defaultdict(lambda: {"n": 0, "rev": 0.0})
    for lid, d, nightly in nights:
        if w0 <= d < today:
            cur[lid]["n"] += 1
            cur[lid]["rev"] += nightly
        elif wp <= d < w0:
            prev[lid]["n"] += 1
            prev[lid]["rev"] += nightly
    lids = set(list(cur) + list(prev) + [u["id"] for u in _catalog_units if u.get("id")])
    units = []
    for lid in lids:
        n, rev = cur[lid]["n"], cur[lid]["rev"]
        adr = (rev / n) if n else (factors["unit_adr"].get(lid) or factors["overall_adr"])
        empty = max(0, 7 - n)
        missed = empty * (adr or 0)
        units.append({"lid": lid, "name": listings_map.get(lid) or f"unit-{lid}",
                      "n": n, "rev": rev, "occ": n / 7, "adr": adr, "empty": empty,
                      "missed": missed, "prev_n": prev[lid]["n"], "prev_rev": prev[lid]["rev"]})
    tot_rev = sum(u["rev"] for u in units)
    tot_n = sum(u["n"] for u in units)
    tot_prev = sum(u["prev_rev"] for u in units)
    avail = len(units) * 7
    return {"units": units, "tot_rev": tot_rev, "tot_nights": tot_n,
            "occ": (tot_n / avail) if avail else 0, "tot_prev": tot_prev,
            "tot_missed": sum(u["missed"] for u in units), "w0": w0, "today": today}

def build_last_week_message(rep):
    delta = rep["tot_rev"] - rep["tot_prev"]
    arrow = "🟢▲" if delta >= 0 else "🔴▼"
    pct = (round(delta / rep["tot_prev"] * 100) if rep["tot_prev"] else 0)
    e = discord.Embed(
        title="📅 مراجعة الأسبوع الماضي · عوجا",
        description=(f"من {rep['w0']} إلى {rep['today']}\n"
                     f"الإيراد: **{round(rep['tot_rev'])} ر.س**  {arrow} {pct:+d}% عن الأسبوع قبله\n"
                     f"الإشغال: **{round(rep['occ']*100)}%** · ليالي مباعة: **{rep['tot_nights']}**\n"
                     f"💸 إيراد ضائع تقديري من الليالي الفاضية: **~{round(rep['tot_missed'])} ر.س**"),
        color=GOLD)
    worst = sorted([u for u in rep["units"] if u["empty"] > 0],
                   key=lambda u: u["missed"], reverse=True)[:8]
    if worst:
        e.add_field(
            name="أكثر وحدات فيها ليالي فاضية (فرصة ضائعة)",
            value="\n".join(f"• **{u['name']}**: {u['empty']} ليالي فاضية · "
                            f"~{round(u['missed'])} ر.س ضائعة · إشغال {round(u['occ']*100)}%"
                            for u in worst)[:1024],
            inline=False)
    best = sorted(rep["units"], key=lambda u: u["rev"], reverse=True)[:5]
    if best and best[0]["rev"]:
        e.add_field(name="الأعلى إيراداً هالأسبوع",
                    value="\n".join(f"• **{u['name']}**: {round(u['rev'])} ر.س · "
                                    f"إشغال {round(u['occ']*100)}%" for u in best if u["rev"])[:1024],
                    inline=False)
    e.set_footer(text="«الإيراد الضائع» = ليالي فاضية × متوسط سعر الوحدة · تقديري")
    return e

async def post_last_week_review():
    guild = bot.get_guild(GUILD_ID)
    if guild is None:
        return
    channel = await ensure_channel(guild, WEEKLY_REVIEW_CHANNEL, await get_assistant_category(guild))
    if channel is None:
        return
    if not _catalog_units:
        await asyncio.to_thread(load_catalog, True)
    reservations = await asyncio.to_thread(get_reservations_cached)
    if not reservations:
        await channel.send("⚠️ ما قدرت أجيب بيانات الحجوزات لمراجعة الأسبوع.")
        return
    listings_map = await asyncio.to_thread(get_listings_map)
    factors = await asyncio.to_thread(compute_demand_factors, reservations)
    rep = await asyncio.to_thread(compute_last_week, reservations, listings_map, factors)
    await channel.send(embed=build_last_week_message(rep))
    print("last-week-review: posted")
    log_event("report", "مراجعة الأسبوع الماضي")

_price_opp_last = None
_review_last = None

@tasks.loop(minutes=30)
async def price_opp_loop():
    global _price_opp_last
    now = datetime.now(TZ)
    if now.weekday() != PRICE_OPP_DOW or now.hour != PRICE_OPP_HOUR or _price_opp_last == now.date():
        return
    _price_opp_last = now.date()
    try:
        await post_price_opportunities()
    except Exception as e:
        print("price_opp_loop error:", e)

@tasks.loop(minutes=30)
async def weekly_review_loop():
    global _review_last
    now = datetime.now(TZ)
    if now.weekday() != WEEKLY_REVIEW_DOW or now.hour != WEEKLY_REVIEW_HOUR or _review_last == now.date():
        return
    _review_last = now.date()
    try:
        await post_last_week_review()
    except Exception as e:
        print("weekly_review_loop error:", e)


@bot.event
async def on_message(message):
    """Knowledge-channel feedback loop. Any team message in #knowledge gets:
       1) 👀 reaction immediately (acknowledgement of receipt)
       2) appended to the in-memory facts (apartment-scoped if a unit name is
          mentioned, otherwise general)
       3) ✅ reaction once persisted
       4) a one-line reply confirming the scope so the team knows exactly
          where it landed."""
    try:
        # Ignore the bot itself and any other bots
        if message.author.bot:
            return
        # Only act on the configured knowledge channel
        if not message.channel or getattr(message.channel, "name", "") != KNOWLEDGE_CHANNEL:
            return
        body = (message.content or "").strip()
        if not body or body.startswith(("/", "!")):
            return
        # 👀 — we see you
        try:
            await message.add_reaction("👀")
        except Exception:
            pass
        # Decide scope
        lid, unit_name = _detect_apartment_in_text(body)
        # Apply in-memory immediately so the very next draft can use it.
        # A full reload (load_knowledge) happens periodically + after this
        # handler to keep things in sync with Discord.
        if lid:
            _knowledge_apartment_facts.setdefault(int(lid), []).append(body)
            _knowledge_apartment_facts[int(lid)] = _knowledge_apartment_facts[int(lid)][-30:]
        else:
            globals()["_knowledge_text"] = (
                (_knowledge_text + "\n- " + body) if _knowledge_text else ("- " + body)
            )[-8000:]
        # Trigger a fresh load so the canonical source matches (and we
        # de-dupe in case the same fact was already there)
        try:
            await load_knowledge(message.guild)
        except Exception as e:
            print("knowledge reload error:", e)
        # ✅ — saved
        try:
            await message.add_reaction("✅")
        except Exception:
            pass
        # Reply confirming scope
        try:
            if lid and unit_name:
                await message.reply(
                    f"✅ تم! حفظت هذي المعلومة لـ **{unit_name}** فقط — راح أستخدمها "
                    f"كمصدر حقيقة في كل رد يخص هذي الوحدة من الحين فصاعداً.",
                    mention_author=False,
                )
            else:
                await message.reply(
                    "✅ تم! حفظت هذي المعلومة في **الذاكرة العامة** — راح تنطبق على كل الوحدات. "
                    "لو تبيها تخص شقة معيّنة، اكتب اسم الشقة داخل الرسالة (مثلاً: "
                    "*'F2: الواي فاي اسمه Ouja-F2'* أو *'في A12 - النرجس فيه باركن إضافي تحت'*).",
                    mention_author=False,
                )
            log_event("guest",
                      f"معرفة جديدة في #knowledge · " + (unit_name or "عام") + " · " + body[:60])
        except Exception as e:
            print("knowledge reply error:", e)
    finally:
        # CRITICAL: keep prefix commands (!ouja ...) working
        try:
            await bot.process_commands(message)
        except Exception:
            pass

@bot.event
async def on_ready():
    load_state()                       # restore seen/cards/escalations from the volume FIRST
    bot.add_view(CleaningDoneView())   # re-bind button handlers after a restart
    bot.add_view(ClaimView())          # re-bind escalation claim buttons after a restart
    bot.add_view(ApproveView())        # re-bind guest-reply approval buttons after a restart
    bot.add_view(PriceApplyView())     # re-bind price apply/skip buttons after a restart
    # On the FIRST start only (no saved state yet), mark the whole current inbox as
    # already-seen so we never replay old backlog. After this, ONLY genuinely new message
    # IDs get a card — and persistence keeps _assistant_seen across restarts, so redeploys
    # don't re-baseline and don't miss messages that arrive during a restart.
    # (No timestamp guessing here — that was reading Hostaway times in the wrong timezone
    #  and making fresh messages look hours old, which replayed old chats and skipped new ones.)
    if ASSISTANT_ENABLED and not ASSISTANT_TEST and not _assistant_seen:
        try:
            base = await asyncio.to_thread(fetch_new_guest_messages, set(), False)
            for it in base:
                _assistant_seen.add(it["message_id"])
            await asyncio.to_thread(persist_state)   # save immediately so it survives a restart
            print(f"assistant: baselined {len(_assistant_seen)} existing conversations "
                  f"— from now on only NEW messages get cards")
        except Exception as e:
            print("assistant baseline error:", e)
    print(f"Logged in as {bot.user}. Watching for checkouts every {POLL_MINUTES} min.")
    print(f"Weekday tiers (Riyadh): {DISCOUNT_TIER1_PERCENT:.0f}% at {DISCOUNT_TIER1_HOUR:02d}:00, "
          f"{DISCOUNT_TIER2_PERCENT:.0f}% at {DISCOUNT_TIER2_HOUR:02d}:00, "
          f"{DISCOUNT_TIER3_PERCENT:.0f}% at {DISCOUNT_TIER3_HOUR:02d}:00 "
          f"{'(DRY-RUN)' if DISCOUNT_DRY_RUN else '(LIVE)'}")
    print(f"Weekend rule (Thu/Fri): {WEEKEND_DISCOUNT_PERCENT:.0f}% at "
          f"{WEEKEND_DISCOUNT_HOUR:02d}:{WEEKEND_DISCOUNT_MINUTE:02d} only")
    print(f"Heads-up preview at {HEADS_UP_HOUR:02d}:00 -> #{HEADS_UP_CHANNEL}")
    print(f"Cleaning reminders: every {REMINDER_SLOW_MIN} min "
          f"{REMINDER_START_HOUR:02d}:00–{REMINDER_FAST_HOUR:02d}:00, then every "
          f"{REMINDER_FAST_MIN} min until {REMINDER_END_HOUR:02d}:00")
    print(f"AI assistant: {'ON' if ASSISTANT_ENABLED else 'OFF'} · model={CLAUDE_MODEL} · "
          f"premium={CLAUDE_MODEL_PREMIUM} · "
          f"auto-send simple={'ON' if ASSISTANT_AUTO else 'OFF'} · "
          f"escalation-ack={'ON' if ASSISTANT_ESC_ACK else 'OFF'} -> #{ASSISTANT_CHANNEL}")
    if not poll_loop.is_running():
        poll_loop.start()
    if not reminder_loop.is_running():
        reminder_loop.start()
    if not discount_tier1_loop.is_running():
        discount_tier1_loop.start()
    if not discount_tier2_loop.is_running():
        discount_tier2_loop.start()
    if not discount_tier3_loop.is_running():
        discount_tier3_loop.start()
    if not discount_weekend_loop.is_running():
        discount_weekend_loop.start()
    if not headsup_loop.is_running():
        headsup_loop.start()
    if ASSISTANT_ENABLED:
        guild0 = bot.get_guild(GUILD_ID)
        if guild0 is not None:
            await load_knowledge(guild0)
            try:                                   # make the auto-reply audit room exist up front
                cat = await get_assistant_category(guild0)
                ch = await ensure_channel(guild0, AUTO_REPLY_CHANNEL, cat)
                if ch is not None and not [m async for m in ch.history(limit=1)]:
                    note = ("📝 **سجل الردود التلقائية** — كل رد يرسله المساعد فيصل تلقائياً بيتسجّل هنا.\n"
                            f"التشغيل التلقائي الحين: **{'مفعّل ✅' if ASSISTANT_AUTO else 'متوقف — فعّله بـ ASSISTANT_AUTO=1'}**"
                            if not ASSISTANT_AUTO else
                            "📝 **سجل الردود التلقائية** — كل رد يرسله المساعد فيصل تلقائياً بيتسجّل هنا.")
                    await ch.send(note)
            except Exception as e:
                print("auto-reply room setup:", e)
        await asyncio.to_thread(load_catalog, True)
    if not assistant_loop.is_running():
        assistant_loop.start()
    if not escalation_reping_loop.is_running():
        escalation_reping_loop.start()
    if not persist_loop.is_running():
        persist_loop.start()
    if DASHBOARD_ENABLED and DASHBOARD_TOKEN and not dashboard_cache_loop.is_running():
        dashboard_cache_loop.start()   # warms the cache immediately, then every few minutes
    if PRICING_STRATEGY_ENABLED and not pricing_strategy_loop.is_running():
        pricing_strategy_loop.start()
    if not revenue_loop.is_running():
        revenue_loop.start()
    if not price_opp_loop.is_running():
        price_opp_loop.start()
    if not weekly_review_loop.is_running():
        weekly_review_loop.start()
    if WEBHOOKS_ENABLED and _HAS_AIOHTTP and _web_runner is None:
        try:
            await start_web_server()
        except Exception as e:
            print("web server start error:", e)
    elif WEBHOOKS_ENABLED and not _HAS_AIOHTTP:
        print("webhooks: aiohttp not installed — add 'aiohttp' to requirements.txt to enable")
    if ASSISTANT_ENABLED and not knowledge_loop.is_running():
        knowledge_loop.start()
    if ASSISTANT_ENABLED and not learning_distillation_loop.is_running():
        learning_distillation_loop.start()
    if AGREEMENT_REMINDER_ENABLED and not agreement_reminder_loop.is_running():
        agreement_reminder_loop.start()
    if DEEPCLEAN_ENABLED and not deepclean_schedule_loop.is_running():
        deepclean_schedule_loop.start()
    if DEEPCLEAN_ENABLED and not deepclean_confirm_loop.is_running():
        deepclean_confirm_loop.start()
    if not offhours_ack_reset_loop.is_running():
        offhours_ack_reset_loop.start()
    if not morning_escalation_reminder_loop.is_running():
        morning_escalation_reminder_loop.start()
    if WILT_ENABLED and not wilt_loop.is_running():
        wilt_loop.start()
    if WILT_TEST:
        print("WILT_TEST=1 — posting WILT now (test run)")
        try:
            await post_wilt()
        except Exception as e:
            print("test WILT error:", e)
    if ASSISTANT_ENABLED and not guest_summary_loop.is_running():
        guest_summary_loop.start()
    if CLEAN_FEEDBACK_ENABLED and not cleaning_feedback_loop.is_running():
        cleaning_feedback_loop.start()
    if HEADS_UP_TEST:
        print("HEADS_UP_TEST=1 — posting a heads-up preview now (test run)")
        try:
            items, tomorrow, weekend = await asyncio.to_thread(compute_headsup)
            await post_headsup(items, tomorrow, weekend)
        except Exception as e:
            print("test heads-up error:", e)
    if REVENUE_TEST:
        print("REVENUE_TEST=1 — building the weekly revenue report now (test run)")
        try:
            await post_revenue_report()
        except Exception as e:
            print("test revenue report error:", e)
    if PRICE_OPP_TEST:
        print("PRICE_OPP_TEST=1 — building price opportunities now (test run)")
        try:
            await post_price_opportunities()
        except Exception as e:
            print("test price-opp error:", e)
    if WEEKLY_REVIEW_TEST:
        print("WEEKLY_REVIEW_TEST=1 — building last-week review now (test run)")
        try:
            await post_last_week_review()
        except Exception as e:
            print("test last-week-review error:", e)
    if DISCOUNT_TEST not in ("0", "", "false", "False", "no"):
        try:
            pct = (DISCOUNT_TIER1_PERCENT if DISCOUNT_TEST in ("1", "true", "True", "yes")
                   else float(DISCOUNT_TEST))
        except ValueError:
            pct = DISCOUNT_TIER1_PERCENT
        print(f"DISCOUNT_TEST — running a {pct:.0f}% tier now "
              f"({'DRY-RUN' if DISCOUNT_DRY_RUN else 'LIVE'})")
        await _run_tier(pct, "Test (startup)")

if __name__ == "__main__":
    missing = [k for k in ("HOSTAWAY_ACCOUNT_ID", "HOSTAWAY_API_KEY", "DISCORD_TOKEN", "DISCORD_GUILD_ID")
               if not os.environ.get(k)]
    if missing:
        raise SystemExit("Missing environment variables: " + ", ".join(missing))
    bot.run(DISCORD_TOKEN)
